#!/usr/bin/env python3
"""
Mana PDF Excel AI Ultimate System
Gemini + OCR + マルチAI協調 + 自動化強化版
"""

import os
import json
import asyncio
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Callable
import fitz  # PyMuPDF
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import google.generativeai as genai
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
import httpx
import schedule
import time

# ログ設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/root/logs/mana_pdf_excel_ai_ultimate.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("ManaPDFExcelAIUltimate")


class ManaPDFExcelAIUltimate:
    """PDF→Excel AI統合システム完全版"""

    def __init__(self, progress_callback: Optional[Callable[[Dict[str, Any]], None]] = None):
        self.name = "Mana PDF Excel AI Ultimate"
        self.version = "2.0.0"
        self.stats = {
            'total_conversions': 0,
            'successful_conversions': 0,
            'failed_conversions': 0,
            'total_pages_processed': 0,
            'total_tables_extracted': 0,
            'ai_analysis_count': 0,
            'automation_runs': 0
        }

        # 進捗コールバック（WebSocket用）
        self.progress_callback = progress_callback

        # AI設定
        self.setup_ai_models()

        # 出力ディレクトリ
        self.output_dir = Path("/root/excel_output_ultimate")
        self.output_dir.mkdir(exist_ok=True)

        logger.info(f"🚀 {self.name} v{self.version} 初期化完了")

    def setup_ai_models(self):
        """AI モデル初期化"""
        try:
            # Gemini API設定
            api_key = os.getenv('GEMINI_API_KEY')
            if api_key:
                genai.configure(api_key=api_key)
                self.gemini_model = genai.GenerativeModel('gemini-1.5-pro')
                logger.info("✅ Gemini API初期化完了")
            else:
                logger.warning("⚠️ Gemini API Key未設定")
                self.gemini_model = None

            # OpenAI API設定
            self.openai_api_key = os.getenv('OPENAI_API_KEY')
            if self.openai_api_key:
                logger.info("✅ OpenAI API初期化完了")
            else:
                logger.warning("⚠️ OpenAI API Key未設定")

            # Ollama（ローカル）設定
            self.ollama_endpoint = os.getenv(
                'OLLAMA_ENDPOINT', 'http://localhost:11434')
            # 軽量モデルを優先（まずは単独で検証）
            self.ollama_models = [
                os.getenv('OLLAMA_MODEL_PRIMARY', 'phi3:mini')]

        except Exception as e:
            logger.error(f"❌ AI初期化エラー: {e}")
            self.gemini_model = None
            self.openai_api_key = None

        # OCR設定（オプション）
        self.easyocr_reader = None
        try:
            import easyocr
            self.easyocr_reader = easyocr.Reader(
                ['ja', 'en'], gpu=False, verbose=False)
            logger.info("✅ EasyOCR初期化完了")
        except ImportError:
            logger.warning("⚠️ EasyOCR未インストール（OCR機能は使用できません）")
        except Exception as e:
            logger.warning(f"⚠️ EasyOCR初期化エラー: {e}")

    async def analyze_with_gemini(self, text: str, context: str = "", max_retries: int = 3) -> Dict[str, Any]:
        """Gemini AIでテキスト解析（リトライ対応）"""
        if not self.gemini_model:
            return {"success": False, "error": "Gemini API未設定"}

        for attempt in range(max_retries):
            try:
                prompt = f"""
                以下のPDFから抽出されたテキストを分析し、構造化データとして整理してください。

                コンテキスト: {context}

                テキスト:
                {text[:4000]}  # トークン制限対応

                以下の形式でJSON出力してください:
                {{
                    "document_type": "文書タイプ",
                    "key_entities": ["重要なエンティティ"],
                    "structured_data": {{
                        "tables": [],
                        "key_values": {{}},
                        "summary": "要約"
                    }},
                    "confidence": 0.95
                }}
                """

                response = await asyncio.to_thread(
                    self.gemini_model.generate_content, prompt
                )

                # JSON解析
                result_text = response.text
                if result_text.startswith('```json'):
                    result_text = result_text[7:-3]
                elif result_text.startswith('```'):
                    result_text = result_text[3:-3]

                result = json.loads(result_text)
                result["success"] = True

                self.stats['ai_analysis_count'] += 1
                logger.info("✅ Gemini解析完了")
                return result

            except Exception as e:
                if attempt < max_retries - 1:
                    wait_time = 2 ** attempt  # 指数バックオフ
                    logger.warning(
                        f"⚠️ Gemini解析リトライ ({attempt + 1}/{max_retries}): {e}")
                    await asyncio.sleep(wait_time)
                else:
                    logger.error(f"❌ Gemini解析エラー（全{max_retries}回リトライ失敗）: {e}")
                    return {"success": False, "error": str(e)}

    async def analyze_with_openai(self, text: str, context: str = "", max_retries: int = 3) -> Dict[str, Any]:
        """OpenAI GPTでテキスト解析（リトライ対応）"""
        if not self.openai_api_key:
            return {"success": False, "error": "OpenAI API未設定"}

        for attempt in range(max_retries):
            try:
                headers = {
                    "Authorization": f"Bearer {self.openai_api_key}",
                    "Content-Type": "application/json"
                }

                # 無料寄りのモデル/環境変数で上書き可能
                openai_model = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
                data = {
                    "model": openai_model,
                    "messages": [
                        {
                            "role": "system",
                            "content": "PDFから抽出されたテキストを構造化データとして分析してください。"
                        },
                        {
                            "role": "user",
                            "content": f"コンテキスト: {context}\n\nテキスト: {text[:4000]}"
                        }
                    ],
                    "max_tokens": 2000,
                    "temperature": 0.1
                }

                response = await asyncio.to_thread(
                    requests.post,
                    "https://api.openai.com/v1/chat/completions",
                    headers=headers,
                    json=data
                )

                if response.status_code == 200:
                    result_data = response.json()
                    content = result_data['choices'][0]['message']['content']

                    # JSON解析
                    try:
                        result = json.loads(content)
                    except Exception:
                        result = {"analysis": content, "confidence": 0.8}

                    result["success"] = True
                    self.stats['ai_analysis_count'] += 1
                    logger.info("✅ OpenAI解析完了")
                    return result
                else:
                    # 429エラー（レート制限）の場合はリトライ
                    if response.status_code == 429 and attempt < max_retries - 1:
                        wait_time = 2 ** attempt
                        logger.warning(
                            f"⚠️ OpenAIレート制限、リトライ ({attempt + 1}/{max_retries})")
                        await asyncio.sleep(wait_time)
                        continue
                    return {"success": False, "error": f"API Error: {response.status_code}"}

            except Exception as e:
                if attempt < max_retries - 1:
                    wait_time = 2 ** attempt
                    logger.warning(
                        f"⚠️ OpenAI解析リトライ ({attempt + 1}/{max_retries}): {e}")
                    await asyncio.sleep(wait_time)
                else:
                    logger.error(f"❌ OpenAI解析エラー（全{max_retries}回リトライ失敗）: {e}")
                    return {"success": False, "error": str(e)}

    async def analyze_with_ollama(self, text: str, context: str = "", model: str = "phi3:mini") -> Dict[str, Any]:
        """OllamaローカルLLMでテキスト解析"""
        try:
            prompt = (
                "以下のPDFテキストを分析し、JSONで出力。\n"
                "キー: document_type, key_entities(配列), summary, confidence(0-1).\n\n"
                f"コンテキスト: {context}\n\nテキスト: {text[:1500]}"
            )
            # stream=False で単発JSON応答を取得
            payload = {"model": model, "prompt": prompt,
                       "stream": False, "options": {"temperature": 0.0}}
            async with httpx.AsyncClient(timeout=60) as client:
                resp = await client.post(f"{self.ollama_endpoint}/api/generate", json=payload)
                if resp.status_code != 200:
                    return {"success": False, "error": f"Ollama HTTP {resp.status_code}"}
                data = resp.json() if 'application/json' in resp.headers.get('content-type', '') else None
                # generate(stream=False) は {response: "...", done: true, ...}
                text_out = (data or {}).get('response', resp.text)
            # JSON抽出
            import re as _re
            import json as _json
            m = _re.search(r"\{[\s\S]*\}", text_out)
            if m:
                try:
                    obj = _json.loads(m.group(0))
                    obj.setdefault('confidence', obj.get('confidence', 0.6))
                    obj['success'] = True
                    return obj
                except Exception:
                    pass
            return {"success": True, "summary": text_out[:500], "confidence": 0.5}
        except Exception as e:
            logger.error(f"❌ Ollama解析エラー: {e}")
            return {"success": False, "error": str(e)}

    def extract_with_ocr(self, page, page_num: int) -> str:
        """OCRでテキスト抽出（EasyOCR使用）"""
        if not self.easyocr_reader:
            return ""

        try:
            from PIL import Image
            import numpy as np

            # ページを画像に変換（300 DPI）
            zoom = 300 / 72
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

            # EasyOCRでテキスト抽出
            img_array = np.array(img)
            results = self.easyocr_reader.readtext(img_array)

            # 結果をテキストに変換（位置順に並べ替え）
            text_items = []
            for detection in results:
                bbox, text, confidence = detection
                if confidence > 0.5:  # 信頼度50%以上
                    # 上から下へ、左から右へ並べ替え
                    y_center = sum([point[1] for point in bbox]) / len(bbox)
                    x_center = sum([point[0] for point in bbox]) / len(bbox)
                    text_items.append((y_center, x_center, text))

            # 位置順にソートしてテキストを結合
            text_items.sort(key=lambda x: (int(x[0] / 50), x[1]))  # 行ごとにグループ化
            extracted_text = '\n'.join([item[2] for item in text_items])

            return extracted_text

        except Exception as e:
            logger.warning(f"⚠️ OCR抽出エラー: {e}")
            return ""

    def extract_pdf_content(self, pdf_path: str, use_ocr: bool = False) -> Dict[str, Any]:
        """PDF内容抽出（マルチエンジン + OCR対応）"""
        content = {
            'text': '',
            'tables': [],
            'images': [],
            'metadata': {},
            'pages': []
        }

        try:
            # PyMuPDFで基本抽出
            doc = fitz.open(pdf_path)
            content['metadata'] = doc.metadata

            for page_num in range(len(doc)):
                page = doc[page_num]
                page_content = {
                    'page_number': page_num + 1,
                    'text': page.get_text(),
                    'images': [],
                    'tables': []
                }

                # テキストがない場合はOCRを試す
                if not page_content['text'].strip() and use_ocr:
                    try:
                        ocr_text = self.extract_with_ocr(page, page_num)
                        if ocr_text:
                            page_content['text'] = ocr_text
                            logger.info(f"📄 ページ{page_num + 1}: OCRでテキスト抽出成功")
                    except Exception as e:
                        logger.warning(f"⚠️ ページ{page_num + 1}: OCR抽出エラー: {e}")

                # 画像抽出
                image_list = page.get_images()
                for img_index, img in enumerate(image_list):
                    xref = img[0]
                    pix = fitz.Pixmap(doc, xref)
                    if pix.n - pix.alpha < 4:  # GRAY or RGB
                        img_data = pix.tobytes("png")
                        page_content['images'].append({
                            'index': img_index,
                            'data': img_data,
                            'width': pix.width,
                            'height': pix.height
                        })

                content['pages'].append(page_content)
                content['text'] += page_content['text'] + '\n'

            doc.close()

            # pdfplumberで表抽出
            with pdfplumber.open(pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    tables = page.extract_tables()
                    if tables:
                        content['pages'][page_num]['tables'] = tables
                        content['tables'].extend(tables)

            logger.info(
                f"✅ PDF内容抽出完了: {len(content['pages'])}ページ, {len(content['tables'])}表")
            return content

        except Exception as e:
            logger.error(f"❌ PDF抽出エラー: {e}")
            return content

    async def multi_ai_analysis(self, content: Dict[str, Any]) -> Dict[str, Any]:
        """マルチAI協調解析"""
        logger.info("🤖 マルチAI協調解析開始")

        # 並列で複数AI実行
        tasks = []

        if self.gemini_model:
            tasks.append(self.analyze_with_gemini(content['text'], "PDF解析"))

        if self.openai_api_key:
            tasks.append(self.analyze_with_openai(content['text'], "PDF解析"))
        # Ollama（ローカル）
        for m in [mm for mm in self.ollama_models if mm]:
            tasks.append(self.analyze_with_ollama(
                content['text'], "PDF解析", model=m))

        if not tasks:
            return {"success": False, "error": "AI API未設定"}

        # 並列実行
        results = await asyncio.gather(*tasks, return_exceptions=True)

        # 結果統合
        combined_result = {
            "success": True,
            "ai_results": [],
            "consensus": {},
            "confidence": 0.0
        }

        valid_results = []
        for result in results:
            if isinstance(result, dict) and result.get("success"):
                valid_results.append(result)
                combined_result["ai_results"].append(result)

        if valid_results:
            # 信頼度計算
            total_confidence = sum(r.get("confidence", 0.5)
                                   for r in valid_results)
            combined_result["confidence"] = total_confidence / \
                len(valid_results)

            # コンセンサス生成
            combined_result["consensus"] = self.generate_consensus(
                valid_results)

        logger.info(
            f"✅ マルチAI解析完了: {len(valid_results)}AI, 信頼度: {combined_result['confidence']:.2f}")
        return combined_result

    def generate_consensus(self, results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """AI結果のコンセンサス生成"""
        consensus = {
            "document_type": "",
            "key_entities": [],
            "structured_data": {},
            "recommendations": []
        }

        # 文書タイプの投票
        doc_types = [r.get("document_type", "")
                     for r in results if r.get("document_type")]
        if doc_types:
            consensus["document_type"] = max(
                set(doc_types), key=doc_types.count)

        # エンティティ統合
        all_entities = []
        for result in results:
            entities = result.get("key_entities", [])
            if isinstance(entities, list):
                all_entities.extend(entities)

        consensus["key_entities"] = list(set(all_entities))

        # 構造化データ統合
        consensus["structured_data"] = {
            "tables": [],
            "key_values": {},
            "summary": ""
        }

        summaries = [r.get("summary", "") for r in results if r.get("summary")]
        if summaries:
            consensus["structured_data"]["summary"] = summaries[0]  # 最初の要約を使用

        return consensus

    def create_enhanced_excel(self, content: Dict[str, Any],
                              ai_analysis: Dict[str, Any],
                              pdf_path: str) -> str:
        """強化版Excel生成"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_name = Path(pdf_path).stem
        excel_path = self.output_dir / f"{pdf_name}_AI強化版_{timestamp}.xlsx"

        wb = Workbook()

        # スタイル定義
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid")
        ai_fill = PatternFill(start_color="E8F4FD",
                              end_color="E8F4FD", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # 概要シート
        ws_summary = wb.active
        ws_summary.title = "📊 AI解析概要"

        summary_data = [
            ["🎯 文書情報", ""],
            ["ファイル名", Path(pdf_path).name],
            ["処理日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["総ページ数", len(content['pages'])],
            ["抽出テーブル数", len(content['tables'])],
            ["", ""],
            ["🤖 AI解析結果", ""],
            ["文書タイプ", ai_analysis.get("consensus", {}).get(
                "document_type", "不明")],
            ["信頼度", f"{ai_analysis.get('confidence', 0):.2%}"],
            ["AI数", len(ai_analysis.get("ai_results", []))],
            ["", ""],
            ["📈 統計情報", ""],
            ["総文字数", len(content['text'])],
            ["画像数", sum(len(p['images']) for p in content['pages'])],
            ["AI解析回数", self.stats['ai_analysis_count']]
        ]

        for row_idx, (label, value) in enumerate(summary_data, 1):
            ws_summary[f"A{row_idx}"] = label
            ws_summary[f"B{row_idx}"] = value

            # ヘッダースタイル
            if label.endswith("情報") or label.endswith("結果"):
                ws_summary[f"A{row_idx}"].font = header_font
                ws_summary[f"A{row_idx}"].fill = header_fill
                ws_summary[f"B{row_idx}"].fill = ai_fill

        # AI解析結果シート
        if ai_analysis.get("ai_results"):
            ws_ai = wb.create_sheet("🤖 AI解析詳細")

            ai_data = [
                ["AI", "文書タイプ", "信頼度", "要約"],
            ]

            for result in ai_analysis["ai_results"]:
                ai_data.append([
                    "Gemini" if "gemini" in str(result).lower() else "OpenAI",
                    result.get("document_type", ""),
                    f"{result.get('confidence', 0):.2%}",
                    result.get("summary", "")[:100] + "..."
                ])

            for row_idx, row_data in enumerate(ai_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_ai.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border

                    if row_idx == 1:  # ヘッダー
                        cell.font = header_font
                        cell.fill = header_fill

        # ページ別シート
        for page in content['pages']:
            page_num = page['page_number']
            ws_page = wb.create_sheet(f"📄 ページ_{page_num}")

            # テキスト内容
            ws_page["A1"] = f"ページ {page_num} テキスト内容"
            ws_page["A1"].font = header_font
            ws_page["A1"].fill = header_fill

            text_lines = page['text'].split('\n')
            for row_idx, line in enumerate(text_lines[:50], 2):  # 最大50行
                if line.strip():
                    ws_page[f"A{row_idx}"] = line.strip()

            # 表がある場合
            if page['tables']:
                table_start_row = len(text_lines) + 5
                for table_idx, table in enumerate(page['tables']):
                    ws_page[f"A{table_start_row}"] = f"表 {table_idx + 1}"
                    ws_page[f"A{table_start_row}"].font = header_font
                    ws_page[f"A{table_start_row}"].fill = ai_fill

                    for row_idx, row in enumerate(table):
                        for col_idx, cell in enumerate(row):
                            if cell:
                                ws_page.cell(
                                    row=table_start_row + row_idx + 1,
                                    column=col_idx + 1,
                                    value=str(cell)
                                ).border = border

        # 列幅自動調整
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(excel_path)
        logger.info(f"✅ 強化版Excel生成完了: {excel_path}")
        return str(excel_path)

    async def convert_pdf_to_excel_ai(self, pdf_path: str,
                                      use_multi_ai: bool = True,
                                      use_ocr: bool = True) -> Dict[str, Any]:
        """PDF→Excel AI統合変換（OCR対応）"""
        logger.info(f"🚀 AI統合変換開始: {Path(pdf_path).name}")
        start_time = time.time()

        def send_progress(progress: int, message: str, status: str = "info"):
            """進捗通知"""
            if self.progress_callback:
                self.progress_callback({
                    "type": "progress",
                    "progress": progress,
                    "message": message,
                    "status": status
                })

        try:
            # 1. PDF内容抽出（OCR対応）
            send_progress(10, "📄 PDF内容抽出中...")
            logger.info("📄 PDF内容抽出中...")
            content = self.extract_pdf_content(pdf_path, use_ocr=use_ocr)

            # テキストがない場合はOCRを試す
            if not content['text'].strip() and use_ocr and self.easyocr_reader:
                send_progress(20, "🔍 OCRでテキスト抽出中...")
                logger.info("🔍 OCRでテキスト抽出中...")
                content = self.extract_pdf_content(pdf_path, use_ocr=True)

            # 2. AI解析
            ai_analysis = {"success": False}
            if use_multi_ai:
                send_progress(50, "🤖 マルチAI解析中...")
                logger.info("🤖 マルチAI解析中...")
                ai_analysis = await self.multi_ai_analysis(content)
                send_progress(80, "✅ AI解析完了")
            else:
                send_progress(60, "📊 データ構造化中...")

            # 3. 強化版Excel生成
            send_progress(85, "📊 強化版Excel生成中...")
            logger.info("📊 強化版Excel生成中...")
            excel_path = self.create_enhanced_excel(
                content, ai_analysis, pdf_path)
            send_progress(95, "✅ Excel生成完了")

            # 統計更新
            self.stats['total_conversions'] += 1
            self.stats['successful_conversions'] += 1
            self.stats['total_pages_processed'] += len(content['pages'])
            self.stats['total_tables_extracted'] += len(content['tables'])

            processing_time = time.time() - start_time

            result = {
                'success': True,
                'pdf_file': Path(pdf_path).name,
                'excel_file': Path(excel_path).name,
                'excel_path': excel_path,
                'total_pages': len(content['pages']),
                'total_tables': len(content['tables']),
                'ai_analysis': ai_analysis,
                'processing_time': processing_time,
                'file_size': os.path.getsize(excel_path),
                'stats': self.stats.copy()
            }

            send_progress(100, f"✅ 変換完了: {processing_time:.2f}秒", "success")
            logger.info(f"✅ AI統合変換完了: {processing_time:.2f}秒")
            return result

        except Exception as e:
            logger.error(f"❌ 変換エラー: {e}")
            self.stats['failed_conversions'] += 1
            send_progress(0, f"❌ 変換エラー: {str(e)}", "error")

            return {
                'success': False,
                'error': str(e),
                'stats': self.stats.copy()
            }

    async def batch_convert_ai(self, pdf_paths: List[str],
                               use_multi_ai: bool = True) -> Dict[str, Any]:
        """AI統合バッチ変換"""
        logger.info(f"🚀 AI統合バッチ変換開始: {len(pdf_paths)}ファイル")
        batch_start_time = time.time()

        results = []
        successful = 0
        failed = 0

        # 並列処理（最大5ファイル同時）
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = [
                executor.submit(asyncio.run, self.convert_pdf_to_excel_ai(
                    pdf_path, use_multi_ai))
                for pdf_path in pdf_paths
            ]

            for future in as_completed(futures):
                try:
                    result = future.result()
                    results.append(result)

                    if result['success']:
                        successful += 1
                    else:
                        failed += 1

                except Exception as e:
                    logger.error(f"❌ バッチ処理エラー: {e}")
                    failed += 1
                    results.append({
                        'success': False,
                        'error': str(e)
                    })

        batch_end_time = time.time()
        total_time = batch_end_time - batch_start_time

        batch_result = {
            'batch_start_time': datetime.fromtimestamp(batch_start_time).isoformat(),
            'batch_end_time': datetime.fromtimestamp(batch_end_time).isoformat(),
            'total_files': len(pdf_paths),
            'successful': successful,
            'failed': failed,
            'total_processing_time': total_time,
            'average_time_per_file': total_time / len(pdf_paths) if pdf_paths else 0,
            'results': results,
            'stats': self.stats.copy()
        }

        # バッチレポート保存
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = self.output_dir / f"batch_report_ai_{timestamp}.json"
        with open(report_path, 'w', encoding='utf-8') as f:
            json.dump(batch_result, f, ensure_ascii=False, indent=2)

        logger.info(f"✅ AI統合バッチ変換完了: {successful}/{len(pdf_paths)}成功")
        return batch_result

    def get_stats(self) -> Dict[str, Any]:
        """統計情報取得"""
        return {
            'system_info': {
                'name': self.name,
                'version': self.version,
                'ai_models': {
                    'gemini': self.gemini_model is not None,
                    'openai': self.openai_api_key is not None
                }
            },
            'conversion_stats': self.stats.copy(),
            'output_directory': str(self.output_dir)
        }

    def setup_automation(self, schedule_config: Dict[str, Any]):
        """自動化設定"""
        logger.info("⏰ 自動化設定開始")

        # スケジュール設定例
        if schedule_config.get('daily_conversion'):
            schedule.every().day.at("09:00").do(self.daily_automation)

        if schedule_config.get('hourly_monitoring'):
            schedule.every().hour.do(self.monitoring_check)

        logger.info("✅ 自動化設定完了")

    def daily_automation(self):
        """日次自動化処理"""
        logger.info("🌅 日次自動化処理開始")
        self.stats['automation_runs'] += 1

        # 監視ディレクトリのPDFファイル処理
        watch_dir = Path("/root/automation_input")
        if watch_dir.exists():
            pdf_files = list(watch_dir.glob("*.pdf"))
            if pdf_files:
                asyncio.run(self.batch_convert_ai([str(f) for f in pdf_files]))

    def monitoring_check(self):
        """監視チェック"""
        logger.info("👁️ 監視チェック実行")
        # システム状態確認、ログ分析等


# グローバルインスタンス
_ultimate_converter = None


def get_ultimate_converter(progress_callback: Optional[Callable[[Dict[str, Any]], None]] = None) -> ManaPDFExcelAIUltimate:
    """Ultimate Converter取得"""
    global _ultimate_converter
    if _ultimate_converter is None or progress_callback is not None:
        _ultimate_converter = ManaPDFExcelAIUltimate(
            progress_callback=progress_callback)
    return _ultimate_converter


# Trinity統合API
async def pdf_to_excel_ai(pdf_path: str, use_multi_ai: bool = True, use_ocr: bool = True) -> Dict[str, Any]:
    """PDF→Excel AI統合変換（Trinity用API、OCR対応）"""
    converter = get_ultimate_converter()
    return await converter.convert_pdf_to_excel_ai(pdf_path, use_multi_ai, use_ocr=use_ocr)


async def batch_pdf_to_excel_ai(pdf_paths: List[str], use_multi_ai: bool = True) -> Dict[str, Any]:
    """AI統合バッチ変換（Trinity用API）"""
    converter = get_ultimate_converter()
    return await converter.batch_convert_ai(pdf_paths, use_multi_ai)


def get_ai_conversion_stats() -> Dict[str, Any]:
    """AI変換統計取得（Trinity用API）"""
    converter = get_ultimate_converter()
    return converter.get_stats()


if __name__ == "__main__":
    # テスト実行
    print("🧪 Mana PDF Excel AI Ultimate テスト")
    print("=" * 60)

    converter = get_ultimate_converter()
    stats = converter.get_stats()

    print(f"📊 システム名: {stats['system_info']['name']}")
    print(f"📦 バージョン: {stats['system_info']['version']}")
    print("🤖 AI モデル:")
    for model, enabled in stats['system_info']['ai_models'].items():
        status = "✅" if enabled else "❌"
        print(f"  {status} {model}")

    print("\n📈 統計情報:")
    for key, value in stats['conversion_stats'].items():
        print(f"  • {key}: {value}")

    print(f"\n📁 出力ディレクトリ: {stats['output_directory']}")
    print("\n✅ AI統合システム準備完了！")
    print("\nTrinityからの使用例:")
    print("```python")
    print("from mana_pdf_excel_ai_ultimate import pdf_to_excel_ai")
    print("")
    print("result = await pdf_to_excel_ai('/path/to/file.pdf', use_multi_ai=True)")
    print("if result['success']:")
    print("    print(f\"✅ {result['excel_file']}\")")
    print("    print(f\"🤖 AI信頼度: {result['ai_analysis']['confidence']:.2%}\")")
