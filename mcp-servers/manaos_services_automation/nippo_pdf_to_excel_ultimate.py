#!/usr/bin/env python3
"""
日報PDF→Excel変換システム (究極版)
Google Drive OCR + Google Docs API (表構造取得) + Gemini (賢い整形)

特徴：
- Google Drive OCRで高精度OCR（無料）
- Google Docs APIで表構造を完全取得
- Geminiで賢く整形・補完
- 最高精度 × 無料
"""

import os
import sys
from pathlib import Path
import logging
from datetime import datetime
import time
import json

try:
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
    import google.generativeai as genai
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError as e:
    print(f"❌ 必要なライブラリが不足: {e}")
    sys.exit(1)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/root/logs/nippo_pdf_to_excel_ultimate.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("NippoPDFToExcelUltimate")


class NippoPDFToExcelUltimate:
    """日報PDF→Excel変換 (究極版)"""
    
    def __init__(self):
        self.output_dir = Path('/root/daily_reports/excel_ultimate')
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Google Drive & Docs API
        self.drive_service = None
        self.docs_service = None
        self.temp_folder_id = None
        self.initialize_google_apis()
        
        # Gemini API
        self.gemini_api_key = os.getenv("GEMINI_API_KEY", "AIzaSyC5SA9-bXkB_nlAux9Fve47HvvDm1jx0Y0")
        genai.configure(api_key=self.gemini_api_key)
        self.gemini_model = genai.GenerativeModel('gemini-2.5-flash')
        logger.info("✅ Gemini API初期化完了")
    
    def initialize_google_apis(self):
        """Google APIs初期化"""
        try:
            token_path = '/root/token.json'
            
            with open(token_path, 'r') as f:
                token_data = json.load(f)
            
            creds = Credentials(
                token=token_data.get('token'),
                refresh_token=token_data.get('refresh_token'),
                token_uri=token_data.get('token_uri', 'https://oauth2.googleapis.com/token'),
                client_id=token_data.get('client_id'),
                client_secret=token_data.get('client_secret'),
                scopes=token_data.get('scopes', ['https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/documents'])
            )
            
            if creds.expired and creds.refresh_token:
                creds.refresh(Request())
            
            self.drive_service = build('drive', 'v3', credentials=creds)
            self.docs_service = build('docs', 'v1', credentials=creds)
            logger.info("✅ Google Drive & Docs API接続成功")
            
            self.temp_folder_id = self.get_or_create_folder('日報OCR処理用_一時')
            
        except Exception as e:
            logger.error(f"❌ Google APIs初期化エラー: {e}")
            sys.exit(1)
    
    def get_or_create_folder(self, folder_name: str):
        """フォルダ取得/作成"""
        try:
            query = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
            results = self.drive_service.files().list(q=query, fields='files(id)').execute()
            files = results.get('files', [])
            
            if files:
                return files[0]['id']
            
            file_metadata = {'name': folder_name, 'mimeType': 'application/vnd.google-apps.folder'}
            folder = self.drive_service.files().create(body=file_metadata, fields='id').execute()
            return folder.get('id')
            
        except Exception as e:
            logger.error(f"フォルダ作成エラー: {e}")
            return None
    
    def upload_pdf_with_ocr(self, pdf_path: str):
        """PDF→Google Docs (OCR)"""
        try:
            file_metadata = {
                'name': os.path.basename(pdf_path),
                'parents': [self.temp_folder_id],
                'mimeType': 'application/vnd.google-apps.document'
            }
            
            media = MediaFileUpload(pdf_path, mimetype='application/pdf', resumable=True)
            file = self.drive_service.files().create(
                body=file_metadata, media_body=media, fields='id', ocrLanguage='ja'
            ).execute()
            
            time.sleep(3)  # OCR処理待ち
            return file.get('id')
            
        except Exception as e:
            logger.error(f"アップロードエラー: {e}")
            return None
    
    def extract_document_content(self, doc_id: str):
        """Google Docsのコンテンツを取得（表含む）"""
        try:
            document = self.docs_service.documents().get(documentId=doc_id).execute()
            
            # テキストと表を抽出
            content = document.get('body', {}).get('content', [])
            
            tables = []
            all_text = []
            
            for element in content:
                # テーブル要素
                if 'table' in element:
                    table_data = self.extract_table(element['table'])
                    if table_data:
                        tables.append(table_data)
                
                # パラグラフ要素
                if 'paragraph' in element:
                    paragraph_text = self.extract_paragraph_text(element['paragraph'])
                    if paragraph_text:
                        all_text.append(paragraph_text)
            
            full_text = '\n'.join(all_text)
            
            return {'tables': tables, 'text': full_text}
            
        except Exception as e:
            logger.error(f"ドキュメント抽出エラー: {e}")
            return None
    
    def extract_table(self, table_element):
        """Google Docsの表を抽出"""
        try:
            table_data = []
            
            for row in table_element.get('tableRows', []):
                row_data = []
                for cell in row.get('tableCells', []):
                    cell_text = []
                    for content in cell.get('content', []):
                        if 'paragraph' in content:
                            text = self.extract_paragraph_text(content['paragraph'])
                            if text:
                                cell_text.append(text)
                    row_data.append(' '.join(cell_text))
                table_data.append(row_data)
            
            return table_data
            
        except Exception as e:
            logger.warning(f"表抽出エラー: {e}")
            return None
    
    def extract_paragraph_text(self, paragraph):
        """パラグラフからテキスト抽出"""
        try:
            text_parts = []
            for element in paragraph.get('elements', []):
                if 'textRun' in element:
                    text = element['textRun'].get('content', '')
                    text_parts.append(text)
            return ''.join(text_parts).strip()
        except Exception:
            return ''
    
    def delete_temp_file(self, file_id: str):
        """一時ファイル削除"""
        try:
            self.drive_service.files().delete(fileId=file_id).execute()
        except IOError:
            pass
    
    def gemini_enhance_data(self, tables, text):
        """Geminiでデータを賢く整形"""
        try:
            # テーブルデータをテキスト化
            table_text = ""
            for i, table in enumerate(tables, 1):
                table_text += f"\n【表{i}】\n"
                for row in table:
                    table_text += " | ".join(row) + "\n"
            
            prompt = f"""
以下は日報/見積書のOCRデータです。
表データとテキストから、重要な情報を抽出してJSON配列で出力してください。

【表データ】
{table_text}

【テキストデータ】
{text[:1000]}

【出力形式】
[
  {{"項目": "項目名", "値": "値"}},
  ...
]

【抽出項目】
- 日付、番号、会社名、住所、担当者、連絡先
- 商品情報、金額、税金、合計
- その他重要な情報

JSON配列のみ出力（説明不要）：
"""
            
            logger.info("🤖 Gemini賢い整形中...")
            response = self.gemini_model.generate_content(prompt)
            response_text = response.text.strip()
            
            if '```' in response_text:
                response_text = response_text.split('```')[1]
                if response_text.startswith('json'):
                    response_text = response_text[4:]
                response_text = response_text.strip()
            
            data = json.loads(response_text)
            logger.info(f"✅ Gemini整形成功: {len(data)}項目")
            
            return data
            
        except Exception as e:
            logger.error(f"Gemini整形エラー: {e}")
            return None
    
    def create_excel_from_tables(self, tables, df, output_path: str, pdf_name: str):
        """Excelファイル作成（表構造保持版）"""
        try:
            wb = openpyxl.Workbook()
            
            # 表がある場合は表構造を保持
            if tables:
                # 各表を別シートに
                for i, table_data in enumerate(tables, 1):
                    if i == 1:
                        ws = wb.active
                        ws.title = f'表{i}'
                    else:
                        ws = wb.create_sheet(f'表{i}')
                    
                    # 表データをそのまま書き込み
                    for r_idx, row_data in enumerate(table_data, start=1):
                        for c_idx, cell_value in enumerate(row_data, start=1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=cell_value)
                            cell.font = Font(name='メイリオ', size=10)
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                            cell.border = Border(
                                left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin')
                            )
                            
                            # 1行目をヘッダーとして強調
                            if r_idx == 1:
                                cell.font = Font(name='メイリオ', size=10, bold=True, color='FFFFFF')
                                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            elif r_idx % 2 == 0:
                                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                    
                    # 列幅自動調整
                    for col in ws.columns:
                        max_length = max((len(str(cell.value or '')) for cell in col), default=0)
                        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 60)
            
            # Gemini整形データも追加
            if not df.empty:
                ws_gemini = wb.create_sheet('Gemini整形データ')
                ws = ws_gemini
            else:
                ws = wb.active
                ws.title = '日報データ'
            
            # ヘッダー
            for c_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=1, column=c_idx, value=col_name)
                cell.font = Font(name='メイリオ', size=12, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='medium'), right=Side(style='medium'),
                    top=Side(style='medium'), bottom=Side(style='medium')
                )
            
            # データ
            for r_idx, row in enumerate(df.itertuples(index=False), start=2):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.font = Font(name='メイリオ', size=11)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    if r_idx % 2 == 0:
                        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # 列幅調整
            for col in ws.columns:
                max_length = max((len(str(cell.value or '')) for cell in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 3, 60)
            
            # サマリーシート
            summary_ws = wb.create_sheet('変換情報', 0)
            summary_data = [
                ['項目', '値'],
                ['元PDFファイル', pdf_name],
                ['変換日時', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['抽出項目数', len(df)],
                ['システム', '日報PDF→Excel 究極版'],
                ['OCR', 'Google Drive OCR (無料)'],
                ['表構造取得', 'Google Docs API'],
                ['AI整形', 'Gemini 2.5 Flash (フリーティア)']
            ]
            
            for r_idx, row_data in enumerate(summary_data, start=1):
                for c_idx, value in enumerate(row_data, start=1):
                    cell = summary_ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.font = Font(name='メイリオ', size=11)
                    if r_idx == 1:
                        cell.font = Font(name='メイリオ', size=11, bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
            
            summary_ws.column_dimensions['A'].width = 25
            summary_ws.column_dimensions['B'].width = 50
            
            wb.save(output_path)
            logger.info("✅ Excel作成完了")
            return True
            
        except Exception as e:
            logger.error(f"Excel作成エラー: {e}")
            return False
    
    def convert_pdf(self, pdf_path: str):
        """PDFをExcelに変換"""
        pdf_path = Path(pdf_path)
        
        if not pdf_path.exists():
            return {'success': False, 'error': 'File not found'}
        
        print(f"\n{'='*70}")
        print(f"📄 【究極版】変換: {pdf_path.name}")
        print(f"{'='*70}")
        
        file_id = None
        
        try:
            # 1. Google Drive OCR
            print("☁️  STEP 1: Google Drive OCR（無料・高精度）...")
            file_id = self.upload_pdf_with_ocr(str(pdf_path))
            if not file_id:
                return {'success': False, 'error': 'OCR failed'}
            print("   ✅ OCR完了")
            
            # 2. Google Docsから表構造とテキストを取得
            print("📋 STEP 2: Google Docs表構造取得...")
            doc_content = self.extract_document_content(file_id)
            if not doc_content:
                return {'success': False, 'error': 'Content extraction failed'}
            
            tables = doc_content['tables']
            text = doc_content['text']
            
            print(f"   検出表数: {len(tables)}")
            print(f"   テキスト: {len(text)}文字")
            
            # 3. Geminiで賢く整形
            print("🤖 STEP 3: Gemini賢い整形（フリーティア）...")
            structured_data = self.gemini_enhance_data(tables, text)
            
            if not structured_data:
                return {'success': False, 'error': 'Gemini failed'}
            
            print(f"   ✅ 整形完了: {len(structured_data)}項目")
            
            # 4. DataFrame変換
            print("📊 STEP 4: DataFrame変換...")
            
            # データをフラット化（配列やオブジェクトを文字列化）
            flattened_data = []
            for item in structured_data:
                flat_item = {}
                for key, value in item.items():
                    if isinstance(value, (list, dict)):
                        flat_item[key] = json.dumps(value, ensure_ascii=False)
                    else:
                        flat_item[key] = value
                flattened_data.append(flat_item)
            
            df = pd.DataFrame(flattened_data)
            
            # 5. Excel作成
            print("📝 STEP 5: Excel作成...")
            output_filename = f"{pdf_path.stem}_ultimate_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = self.output_dir / output_filename
            
            if self.create_excel_from_tables(tables, df, str(output_path), pdf_path.name):
                print("\n✅ 変換完了！")
                print(f"📁 {output_path}")
                print(f"{'='*70}\n")
                
                return {'success': True, 'output_file': str(output_path), 'items': len(df)}
            
            return {'success': False, 'error': 'Excel creation failed'}
            
        except Exception as e:
            logger.error(f"変換エラー: {e}")
            return {'success': False, 'error': str(e)}
        
        finally:
            if file_id:
                print("🗑️  一時ファイル削除...")
                self.delete_temp_file(file_id)
    
    def batch_convert(self, pdf_dir: str, limit: int = None):
        """一括変換"""
        pdf_dir = Path(pdf_dir)
        pdf_files = list(pdf_dir.glob('*.pdf')) + list(pdf_dir.glob('*.PDF'))
        
        if limit:
            pdf_files = pdf_files[:limit]
        
        print(f"\n{'='*70}")
        print("📁 【究極版】一括変換")
        print(f"{'='*70}")
        print(f"ファイル数: {len(pdf_files)}")
        print(f"出力先: {self.output_dir}\n")
        
        results = {'total': len(pdf_files), 'success': 0, 'failed': 0}
        
        for i, pdf_file in enumerate(pdf_files, 1):
            result = self.convert_pdf(str(pdf_file))
            
            if result['success']:
                results['success'] += 1
            else:
                results['failed'] += 1
            
            if i < len(pdf_files):
                time.sleep(2)  # API制限対策
        
        print(f"\n{'='*70}")
        print("📊 【究極版】一括変換完了")
        print(f"{'='*70}")
        print(f"総数: {results['total']}")
        print(f"✅ 成功: {results['success']}")
        print(f"❌ 失敗: {results['failed']}")
        print(f"📁 {self.output_dir}")
        print(f"{'='*70}\n")


def main():
    if len(sys.argv) < 2:
        print("使い方:")
        print("  python3 nippo_pdf_to_excel_ultimate.py [PDF]")
        print("  python3 nippo_pdf_to_excel_ultimate.py [フォルダ] [件数]")
        sys.exit(1)
    
    converter = NippoPDFToExcelUltimate()
    target = Path(sys.argv[1])
    
    if target.is_file():
        converter.convert_pdf(str(target))
    elif target.is_dir():
        limit = int(sys.argv[2]) if len(sys.argv) > 2 else None
        converter.batch_convert(str(target), limit=limit)


if __name__ == '__main__':
    main()

