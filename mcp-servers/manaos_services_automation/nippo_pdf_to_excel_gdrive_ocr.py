#!/usr/bin/env python3
"""
日報PDF→Excel変換システム (Google Drive OCR版)
Google Drive APIのOCR機能でPDFをテキスト抽出してExcel化

特徴：
- Google Drive OCRの高精度認識（無料）
- 表構造の保持
- 日本語の高精度認識
- Excel表形式で出力
"""

import os
import sys
from pathlib import Path
import logging
from datetime import datetime
import time
import re

# 必要なライブラリ
try:
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
    import io
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError as e:
    print(f"❌ 必要なライブラリが不足: {e}")
    print("インストール: pip3 install google-api-python-client google-auth pandas openpyxl")
    sys.exit(1)

# ログ設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/root/logs/nippo_pdf_to_excel_gdrive_ocr.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("NippoPDFToExcelGDriveOCR")


class NippoPDFToExcelGDriveOCR:
    """日報PDF→Excel変換システム (Google Drive OCR版)"""
    
    def __init__(self):
        self.output_dir = Path('/root/daily_reports/excel_gdrive_ocr')
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        self.service = None
        self.temp_folder_id = None
        self.initialize_drive()
    
    def initialize_drive(self):
        """Google Drive API初期化"""
        try:
            token_path = '/root/token.json'
            
            if not os.path.exists(token_path):
                logger.error("❌ /root/token.jsonが見つかりません")
                sys.exit(1)
            
            # トークンデータ読み込み
            import json
            with open(token_path, 'r') as f:
                token_data = json.load(f)
            
            creds = Credentials(
                token=token_data.get('token'),
                refresh_token=token_data.get('refresh_token'),
                token_uri=token_data.get('token_uri', 'https://oauth2.googleapis.com/token'),
                client_id=token_data.get('client_id'),
                client_secret=token_data.get('client_secret'),
                scopes=token_data.get('scopes', ['https://www.googleapis.com/auth/drive'])
            )
            
            # トークン更新が必要な場合
            if creds.expired and creds.refresh_token:
                creds.refresh(Request())
            
            self.service = build('drive', 'v3', credentials=creds)
            logger.info("✅ Google Drive API接続成功")
            
            # 一時フォルダ作成
            self.temp_folder_id = self.get_or_create_folder('日報OCR処理用_一時')
            
        except Exception as e:
            logger.error(f"❌ Google Drive初期化エラー: {e}")
            sys.exit(1)
    
    def get_or_create_folder(self, folder_name: str):
        """フォルダを取得または作成"""
        try:
            # 既存フォルダを検索
            query = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
            
            results = self.service.files().list(  # type: ignore[union-attr]
                q=query,
                spaces='drive',
                fields='files(id, name)'
            ).execute()
            
            files = results.get('files', [])
            
            if files:
                return files[0]['id']
            
            # フォルダを作成
            file_metadata = {
                'name': folder_name,
                'mimeType': 'application/vnd.google-apps.folder'
            }
            
            folder = self.service.files().create(  # type: ignore[union-attr]
                body=file_metadata,
                fields='id'
            ).execute()
            
            return folder.get('id')
            
        except Exception as e:
            logger.error(f"フォルダ作成エラー: {e}")
            return None
    
    def upload_pdf_with_ocr(self, pdf_path: str):
        """PDFをアップロードしてOCR実行"""
        try:
            file_name = os.path.basename(pdf_path)
            
            file_metadata = {
                'name': file_name,
                'parents': [self.temp_folder_id],
                'mimeType': 'application/vnd.google-apps.document'  # Google Docs形式に変換
            }
            
            media = MediaFileUpload(
                pdf_path,
                mimetype='application/pdf',
                resumable=True
            )
            
            # OCR付きアップロード
            file = self.service.files().create(  # type: ignore[union-attr]
                body=file_metadata,
                media_body=media,
                fields='id, name',
                ocrLanguage='ja'  # 日本語OCR
            ).execute()
            
            file_id = file.get('id')
            logger.info(f"✅ PDFアップロード＆OCR完了: {file_id}")
            
            # OCR処理完了を待つ
            time.sleep(2)
            
            return file_id
            
        except Exception as e:
            logger.error(f"PDFアップロードエラー: {e}")
            return None
    
    def export_google_doc_as_text(self, file_id: str):
        """Google DocsをテキストとしてエクスポートR"""
        try:
            # プレーンテキストでエクスポート
            request = self.service.files().export_media(  # type: ignore[union-attr]
                fileId=file_id,
                mimeType='text/plain'
            )
            
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            
            done = False
            while not done:
                status, done = downloader.next_chunk()
            
            # テキスト取得
            text = fh.getvalue().decode('utf-8')
            
            return text
            
        except Exception as e:
            logger.error(f"テキストエクスポートエラー: {e}")
            return None
    
    def delete_temp_file(self, file_id: str):
        """一時ファイルを削除"""
        try:
            self.service.files().delete(fileId=file_id).execute()  # type: ignore[union-attr]
            logger.info(f"🗑️  一時ファイル削除: {file_id}")
        except Exception as e:
            logger.warning(f"一時ファイル削除エラー: {e}")
    
    def text_to_dataframe(self, text: str):
        """OCRテキストをDataFrameに変換（賢い表形式化）"""
        try:
            # 行ごとに分割
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            
            if not lines:
                return None
            
            # 【方法1】項目:値のパターンを検出
            structured_data = []
            current_item = None
            current_value = []
            
            for line in lines:
                # 「項目:値」または「項目：値」のパターン
                if ':' in line or '：' in line:
                    # 前の項目を保存
                    if current_item:
                        structured_data.append({
                            '項目': current_item,
                            '値': ' '.join(current_value)
                        })
                    
                    # 新しい項目
                    parts = re.split(r'[:：]', line, maxsplit=1)
                    if len(parts) == 2:
                        current_item = parts[0].strip()
                        current_value = [parts[1].strip()]
                    else:
                        current_item = line
                        current_value = []
                else:
                    # 継続行
                    if current_item:
                        current_value.append(line)
                    else:
                        # 項目なしの行は「内容」として追加
                        structured_data.append({
                            '項目': '内容',
                            '値': line
                        })
            
            # 最後の項目を保存
            if current_item:
                structured_data.append({
                    '項目': current_item,
                    '値': ' '.join(current_value)
                })
            
            # 構造化データがある場合
            if structured_data and len(structured_data) > 3:
                df = pd.DataFrame(structured_data)
                return df
            
            # 【方法2】タブ・スペース区切りで表を検出
            rows = []
            for line in lines:
                # タブ区切り
                if '\t' in line:
                    cells = [cell.strip() for cell in line.split('\t')]
                    if len(cells) > 1:
                        rows.append(cells)
                # 複数スペース（2個以上）で区切り
                elif re.search(r'\s{2,}', line):
                    cells = [cell.strip() for cell in re.split(r'\s{2,}', line) if cell.strip()]
                    if len(cells) > 1:
                        rows.append(cells)
            
            # 表形式データがある場合
            if rows and len(rows) > 3:
                # 列数を統一
                max_cols = max(len(row) for row in rows)
                normalized_rows = []
                for row in rows:
                    while len(row) < max_cols:
                        row.append('')
                    normalized_rows.append(row[:max_cols])
                
                df = pd.DataFrame(normalized_rows)
                return df
            
            # 【方法3】フォールバック：全行を1列に
            df = pd.DataFrame({'内容': lines})
            
            return df
            
        except Exception as e:
            logger.error(f"DataFrame変換エラー: {e}")
            return None
    
    def create_excel(self, df, output_path: str, pdf_name: str):
        """DataFrameからExcelファイルを作成"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '日報データ'  # type: ignore[union-attr]
            
            # データを書き込み
            for r_idx, row in enumerate(df.itertuples(index=False), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)  # type: ignore[union-attr]
                    cell.font = Font(name='メイリオ', size=11)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # 1行目を強調
                    if r_idx == 1:
                        cell.font = Font(name='メイリオ', size=11, bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif r_idx % 2 == 0:
                        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # 列幅自動調整
            for column in ws.columns:  # type: ignore[union-attr]
                max_length = 0
                column_letter = column[0].column_letter  # type: ignore
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width  # type: ignore[union-attr]
            
            # サマリーシート
            summary_ws = wb.create_sheet('変換情報', 0)
            summary_data = [
                ['項目', '値'],
                ['元PDFファイル', pdf_name],
                ['変換日時', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['抽出行数', len(df)],
                ['抽出列数', len(df.columns)],
                ['システム', '日報PDF→Excel変換 (Google Drive OCR版)'],
                ['OCR', 'Google Drive OCR (無料)']
            ]
            
            for r_idx, row_data in enumerate(summary_data, start=1):
                for c_idx, value in enumerate(row_data, start=1):
                    cell = summary_ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.font = Font(name='メイリオ', size=11)
                    if r_idx == 1:
                        cell.font = Font(name='メイリオ', size=11, bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            summary_ws.column_dimensions['A'].width = 20
            summary_ws.column_dimensions['B'].width = 50
            
            # 保存
            wb.save(output_path)
            logger.info(f"✅ Excel作成完了: {output_path}")
            
            return True
            
        except Exception as e:
            logger.error(f"Excel作成エラー: {e}")
            return False
    
    def convert_pdf(self, pdf_path: str):
        """PDFをExcelに変換（メイン処理）"""
        pdf_path = Path(pdf_path)  # type: ignore
        
        if not pdf_path.exists():  # type: ignore
            logger.error(f"PDFファイルが見つかりません: {pdf_path}")
            return {'success': False, 'error': 'File not found'}
        
        print(f"\n{'='*60}")
        print(f"📄 変換開始: {pdf_path.name}")  # type: ignore
        print(f"{'='*60}")
        
        file_id = None
        
        try:
            # 1. PDFアップロード＆OCR
            print("☁️  STEP 1: Google DriveにアップロードしてOCR実行...")
            file_id = self.upload_pdf_with_ocr(str(pdf_path))
            
            if not file_id:
                return {'success': False, 'error': 'PDF upload failed'}
            
            # 2. テキスト抽出
            print("📝 STEP 2: OCRテキスト抽出...")
            text = self.export_google_doc_as_text(file_id)
            
            if not text:
                return {'success': False, 'error': 'Text export failed'}
            
            print(f"   抽出文字数: {len(text)}文字")
            
            # 3. DataFrame変換
            print("📊 STEP 3: DataFrame変換...")
            df = self.text_to_dataframe(text)
            
            if df is None or df.empty:
                return {'success': False, 'error': 'DataFrame conversion failed'}
            
            print(f"   データ: {len(df)}行 x {len(df.columns)}列")
            
            # 4. Excel作成
            print("📝 STEP 4: Excel作成...")
            output_filename = f"{pdf_path.stem}_gdrive_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"  # type: ignore
            output_path = self.output_dir / output_filename
            
            success = self.create_excel(df, str(output_path), pdf_path.name)  # type: ignore
            
            if success:
                print("\n✅ 変換完了！")
                print(f"📁 出力: {output_path}")
                print(f"{'='*60}\n")
                
                return {
                    'success': True,
                    'output_file': str(output_path),
                    'rows': len(df),
                    'cols': len(df.columns)
                }
            else:
                return {'success': False, 'error': 'Excel creation failed'}
            
        except Exception as e:
            logger.error(f"変換エラー: {e}")
            return {'success': False, 'error': str(e)}
        
        finally:
            # 一時ファイル削除
            if file_id:
                print("🗑️  一時ファイル削除中...")
                self.delete_temp_file(file_id)
    
    def batch_convert(self, pdf_dir: str, limit: int = None):  # type: ignore
        """フォルダ内の全PDFを一括変換"""
        pdf_dir = Path(pdf_dir)  # type: ignore
        
        if not pdf_dir.exists():  # type: ignore
            logger.error(f"フォルダが見つかりません: {pdf_dir}")
            return
        
        pdf_files = list(pdf_dir.glob('*.pdf'))  # type: ignore
        pdf_files.extend(list(pdf_dir.glob('*.PDF')))  # type: ignore
        
        if not pdf_files:
            logger.error("PDFファイルが見つかりません")
            return
        
        # 件数制限
        if limit:
            pdf_files = pdf_files[:limit]
        
        print(f"\n{'='*60}")
        print("📁 一括変換開始 (Google Drive OCR)")
        print(f"{'='*60}")
        print(f"PDFファイル数: {len(pdf_files)}")
        print(f"出力先: {self.output_dir}\n")
        
        results = {'total': len(pdf_files), 'success': 0, 'failed': 0}
        
        for i, pdf_file in enumerate(pdf_files, 1):
            print(f"[{i}/{len(pdf_files)}] {pdf_file.name}")
            
            result = self.convert_pdf(str(pdf_file))
            
            if result['success']:
                results['success'] += 1
            else:
                results['failed'] += 1
                print(f"   ❌ 失敗: {result.get('error', '不明')}\n")
        
        # サマリー
        print(f"\n{'='*60}")
        print("📊 一括変換完了 (Google Drive OCR)")
        print(f"{'='*60}")
        print(f"総ファイル数: {results['total']}")
        print(f"✅ 成功: {results['success']}")
        print(f"❌ 失敗: {results['failed']}")
        print(f"📁 出力先: {self.output_dir}")
        print(f"{'='*60}\n")


def main():
    """メイン処理"""
    
    if len(sys.argv) < 2:
        print("使い方:")
        print("  python3 nippo_pdf_to_excel_gdrive_ocr.py [PDFファイル]")
        print("  python3 nippo_pdf_to_excel_gdrive_ocr.py [フォルダ] [件数制限]")
        print("\n例:")
        print("  python3 nippo_pdf_to_excel_gdrive_ocr.py test.pdf")
        print("  python3 nippo_pdf_to_excel_gdrive_ocr.py /path/to/pdfs 10")
        sys.exit(1)
    
    converter = NippoPDFToExcelGDriveOCR()
    
    target = sys.argv[1]
    target_path = Path(target)
    
    if target_path.is_file():
        converter.convert_pdf(target)
    elif target_path.is_dir():
        limit = int(sys.argv[2]) if len(sys.argv) > 2 else None
        converter.batch_convert(target, limit=limit)  # type: ignore
    else:
        print(f"❌ パスが見つかりません: {target}")
        sys.exit(1)


if __name__ == '__main__':
    main()

