#!/usr/bin/env python3
"""
PDF→Excel変換 EasyOCR版（高精度・2025年10月31日）
EasyOCRを使って高精度なOCR変換を実現
"""

import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from statistics import median
from typing import Any, Dict, List
import shutil
import re

try:
    import camelot  # optional: text-based table extraction
except ImportError:
    camelot = None

try:
    import pdfplumber  # optional: text-based table extraction
except ImportError:
    pdfplumber = None

try:
    import fitz  # PyMuPDF
    from PIL import Image
    import numpy as np
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import easyocr
except ImportError as e:
    print(f"❌ 必要なライブラリが不足: {e}")
    print("インストール: pip3 install easyocr PyMuPDF Pillow numpy openpyxl")
    sys.exit(1)

NUMERIC_SIMPLE_PATTERN = re.compile(r'^-?\d+(?:\.\d+)?$')

class EasyOCRPDFToExcel:
    """EasyOCR PDF→Excel変換システム"""

    def __init__(self):
        print("🚀 EasyOCR PDF→Excel変換システム初期化中...")
        print("   （2025年10月31日版）\n")

        # EasyOCR初期化（日本語+英語）
        print("📚 EasyOCR初期化中（初回はモデルダウンロードに時間がかかります）...")
        try:
            self.easyocr_reader = easyocr.Reader(['ja', 'en'], gpu=False, verbose=False)
            print("✅ EasyOCR初期化完了\n")
        except Exception as e:
            print(f"❌ EasyOCR初期化エラー: {e}")
            sys.exit(1)

        # 出力ディレクトリ
        self.output_dir = Path("/root/excel_output_easyocr")
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # ハイブリッド抽出用のオプションツール可否
        self.camelot_available = camelot is not None
        self.pdfplumber_available = pdfplumber is not None
        self.latest_page_analysis: List[Dict[str, Any]] = []

    def pdf_to_high_res_image(self, pdf_path: str, page_num: int = 0, dpi: int = 300):
        """PDFを高解像度画像に変換"""
        try:
            doc = fitz.open(pdf_path)
            page = doc[page_num]

            # 高解像度で画像化
            zoom = dpi / 72
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)

            # PIL Imageに変換
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

            doc.close()
            return img

        except Exception as e:
            print(f"❌ PDF→画像変換エラー: {e}")
            return None

    def extract_with_easyocr(self, image):
        """EasyOCRで高精度テキスト抽出"""
        try:
            print("🔍 EasyOCRでテキスト抽出中...")

            # PIL Imageをnumpy arrayに変換
            img_array = np.array(image)

            # OCR実行
            results = self.easyocr_reader.readtext(img_array)

            # 結果をテキストに変換（位置順に並べ替え）
            text_items = []
            for detection in results:
                bbox = detection[0]  # バウンディングボックス
                text = detection[1]  # テキスト
                confidence = detection[2]  # 信頼度

                # 信頼度が低いものは除外
                if confidence > 0.5:
                    # y座標（上から下）でソートするための基準点
                    y_center = sum([point[1] for point in bbox]) / len(bbox)
                    text_items.append((y_center, text))

            # y座標でソート（上から下へ）
            text_items.sort(key=lambda x: x[0])

            # テキストを行ごとに連結
            text_lines = [item[1] for item in text_items]

            extracted_text = '\n'.join(text_lines)
            print(f"✅ 抽出完了（{len(text_lines)}行、{len(extracted_text)}文字）")

            return extracted_text

        except Exception as e:
            print(f"❌ EasyOCR抽出エラー: {e}")
            import traceback
            traceback.print_exc()
            return None

    # --- ハイブリッド抽出向けユーティリティ ----------------------------------

    def analyze_page_content(self, page) -> Dict[str, Any]:
        """ページ内容を解析し、テキスト主体かどうかを判定"""
        try:
            text = page.get_text("text") or ""
            words = page.get_text("words") or []
        except Exception:
            text = ""
            words = []

        char_count = len(text.strip())
        word_count = len(words)
        image_count = len(page.get_images(full=True))

        # ページサイズから文字密度をざっくり推定
        try:
            page_area = float(page.rect.width * page.rect.height)
        except Exception:
            page_area = 1.0

        text_density = char_count / page_area if page_area > 0 else 0.0

        # テキスト主体判定（閾値は経験則）
        is_text_based = (
            char_count >= 200
            or (word_count >= 30 and text_density > 0.0005)
        )

        return {
            "char_count": char_count,
            "word_count": word_count,
            "image_count": image_count,
            "text_density": text_density,
            "is_text_based": is_text_based
        }

    def convert_pdf_to_structured_excel(
        self,
        pdf_path: str,
        dpi: int = 300,
        confidence_threshold: float = 0.3,
        row_threshold: float = 8.0,  # より厳密に（10.0 → 8.0）
        cluster_threshold: float = 20.0,  # より厳密に（25.0 → 20.0）
        max_columns: int = 18
    ):
        """PDFからOCR結果を用いて表構造を推定し、Excelを生成する"""

        import gc
        try:
            import torch
        except ImportError:
            torch = None

        pdf_path_obj = Path(pdf_path)
        if not pdf_path_obj.exists():
            print(f"❌ PDFファイルが見つかりません: {pdf_path}")
            return None

        try:
            doc = fitz.open(str(pdf_path_obj))
        except Exception as e:
            print(f"❌ PDF読み込みエラー: {e}")
            return None

        total_pages = len(doc)
        if total_pages == 0:
            print("⚠️ ページが存在しないPDFです")
            doc.close()
            return None

        workbook = openpyxl.Workbook()
        default_sheet = workbook.active

        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_font = Font(bold=True, size=14, color='FFFFFF')
        header_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        header_font = Font(bold=True, size=11)

        def build_rows_from_results(ocr_results):
            entries = []
            for detection in ocr_results:
                if not detection or len(detection) < 3:
                    continue
                bbox, text, conf = detection[0], detection[1], detection[2]
                if conf < confidence_threshold:
                    continue
                text = (text or "").strip()
                if not text:
                    continue
                xs = [point[0] for point in bbox]
                ys = [point[1] for point in bbox]
                x_center = sum(xs) / len(xs)
                y_center = sum(ys) / len(ys)
                x_left = min(xs)
                x_right = max(xs)
                width = max(x_right - x_left, 1.0)
                entries.append({
                    'text': text,
                    'conf': conf,
                    'x': x_center,
                    'y': y_center,
                    'x_left': x_left,
                    'x_right': x_right,
                    'width': width
                })

            entries.sort(key=lambda item: item['y'])

            rows = []
            for entry in entries:
                target_row = None
                for row in rows:
                    if abs(row['y'] - entry['y']) <= row_threshold:
                        target_row = row
                        break
                if target_row is None:
                    target_row = {'y': entry['y'], 'items': []}
                    rows.append(target_row)
                else:
                    total = len(target_row['items'])
                    target_row['y'] = (target_row['y'] * total + entry['y']) / (total + 1)

                target_row['items'].append(entry)

            for row in rows:
                row['items'].sort(key=lambda item: item['x'])

            rows.sort(key=lambda row: row['y'])
            return rows

        def determine_column_centers(rows):
            if not rows:
                return []

            # 表の行（データが2つ以上ある行）のx座標を優先的に使用
            x_positions = []
            for row in rows:
                if len(row['items']) >= 2:  # 表の行
                    for item in row['items']:
                        # 幅が十分あるアイテムのみを使用（ノイズ除去）
                        if item['width'] >= 10:
                            x_positions.append(item['x'])

            # 表の行がない場合は全アイテムを使用
            if not x_positions:
                x_positions = [item['x'] for row in rows for item in row['items'] if item['width'] >= 8]

            if not x_positions:
                return []

            x_positions.sort()
            clusters = []
            current_cluster = [x_positions[0]]

            # より厳密なクラスタリング
            for position in x_positions[1:]:
                if position - current_cluster[-1] <= cluster_threshold:
                    current_cluster.append(position)
                else:
                    # クラスタが3つ以上のデータを含む場合のみ列として認識
                    if len(current_cluster) >= 3:
                        clusters.append(current_cluster)
                    current_cluster = [position]

            if len(current_cluster) >= 3:
                clusters.append(current_cluster)

            if clusters:
                # クラスタの中央値を列の中心とする
                centers = [median(cluster) for cluster in clusters]
            else:
                # クラスタが形成できない場合は、頻出するx座標を使用
                from collections import Counter
                x_counter = Counter(round(pos, 0) for pos in x_positions)
                # 3回以上出現するx座標を列の中心とする
                frequent_x = [x for x, count in x_counter.items() if count >= 3]
                if frequent_x:
                    centers = sorted(frequent_x)
                else:
                    centers = sorted(set(round(pos, 1) for pos in x_positions))

            centers = sorted(centers)[:max_columns]
            return centers

        def build_column_bounds(centers):
            bounds = []
            if not centers:
                return bounds
            count = len(centers)
            for idx, center in enumerate(centers):
                if count == 1:
                    bounds.append((center - 200, center + 200))
                    continue
                if idx == 0:
                    next_center = centers[idx + 1]
                    gap = (next_center - center) / 2
                    bounds.append((center - max(gap, 80), center + gap))
                elif idx == count - 1:
                    prev_center = centers[idx - 1]
                    gap = (center - prev_center) / 2
                    bounds.append((center - gap, center + max(gap, 80)))
                else:
                    prev_center = centers[idx - 1]
                    next_center = centers[idx + 1]
                    bounds.append(((prev_center + center) / 2, (center + next_center) / 2))
            return bounds

        def assign_rows_to_columns(rows, centers):
            assignments = []
            column_usage = defaultdict(int)
            column_bounds = build_column_bounds(centers)

            for row in rows:
                row_map = {}
                for item in row['items']:
                    if not centers:
                        col_idx = len(row_map)
                    else:
                        distances = [abs(item['x'] - center) for center in centers]
                        closest_idx = distances.index(min(distances))
                        min_distance = min(distances)

                        # より柔軟な距離判定（幅に応じて調整、より寛容に）
                        allowed_distance = max(50.0, item['width'] * 0.8)  # 40.0 → 50.0, 0.6 → 0.8

                        # 列の境界を考慮しつつ、より積極的に配置
                        if column_bounds:
                            left, right = column_bounds[closest_idx]
                            # 境界内にある場合はそのまま使用
                            if left <= item['x'] <= right:
                                col_idx = closest_idx
                            # 境界外でも許容距離内なら使用
                            elif min_distance <= allowed_distance:
                                col_idx = closest_idx
                            else:
                                # 境界内の列を探す
                                fallback_idx = None
                                for idx, (bound_left, bound_right) in enumerate(column_bounds):
                                    if bound_left <= item['x'] <= bound_right:
                                        fallback_idx = idx
                                        break

                                if fallback_idx is not None:
                                    col_idx = fallback_idx
                                else:
                                    # 最も近い列を使用（より寛容に）
                                    if min_distance <= allowed_distance * 1.8:  # 1.5 → 1.8
                                        col_idx = closest_idx
                                    else:
                                        # それでも配置（空セルを減らすため）
                                        col_idx = closest_idx
                        else:
                            # 境界がない場合は最も近い列を使用
                            col_idx = closest_idx

                    if col_idx in row_map:
                        # 既にデータがある場合、より積極的に別の列を探す
                        placed = False
                        if centers:
                            # まず近くの空き列を探す
                            offset = 1
                            max_offset = min(3, len(centers) // 2)  # 最大3列まで探索
                            while offset <= max_offset:
                                candidates = [col_idx - offset, col_idx + offset]
                                # 距離が近い順に試す
                                candidates.sort(key=lambda c: abs(c - col_idx))
                                for candidate in candidates:
                                    if 0 <= candidate < len(centers) and candidate not in row_map:
                                        # 新しい列が元の列に近いか確認
                                        if abs(centers[candidate] - item['x']) <= allowed_distance * 2:
                                            col_idx = candidate
                                            placed = True
                                            break
                                if placed:
                                    break
                                offset += 1

                        # 空き列が見つからない場合の処理
                        if not centers or not placed:
                            # より積極的に空き列を探す（列数を超えても配置）
                            if centers:
                                # 列の範囲を超えて配置（最大列数+3まで）
                                max_col_to_try = len(centers) + 3
                                for try_col in range(len(centers), max_col_to_try):
                                    if try_col not in row_map:
                                        col_idx = try_col
                                        placed = True
                                        break

                            if not placed:
                                # それでも見つからない場合は改行で追加（最後の手段）
                                existing_value = row_map.get(col_idx)
                                if existing_value:
                                    row_map[col_idx] = f"{existing_value}\n{item['text']}"
                                else:
                                    row_map[col_idx] = item['text']
                                column_usage[col_idx] += 1
                                continue

                    row_map[col_idx] = item['text']
                    column_usage[col_idx] += 1

                assignments.append(row_map)

            max_used_index = 0
            for row_map in assignments:
                if row_map:
                    max_used_index = max(max_used_index, max(row_map.keys()))

            total_columns = max(len(centers), max_used_index + 1 if assignments else len(centers))
            return assignments, column_usage, total_columns

        def determine_columns_to_keep(assignments, column_usage, total_columns, header_columns):
            if total_columns == 0:
                return []

            candidate_columns = []
            for col_idx in range(total_columns):
                usage = column_usage.get(col_idx, 0)
                if usage > 1 or col_idx in header_columns:
                    candidate_columns.append(col_idx)

            if not candidate_columns:
                candidate_columns = sorted(column_usage.keys(), key=lambda idx: (-column_usage[idx], idx))

            candidate_columns = candidate_columns[:max_columns]
            return sorted(set(candidate_columns))

        def build_matrix(assignments, columns_to_keep):
            matrix_with_index = []
            for idx, row_map in enumerate(assignments):
                row_values = []
                for col_idx in columns_to_keep:
                    value = row_map.get(col_idx)
                    if isinstance(value, str):
                        value = value.strip()
                    row_values.append(value if value not in ("", None) else None)
                matrix_with_index.append((idx, row_values))

            filtered_rows = []
            for source_idx, row_values in matrix_with_index:
                if any(value is not None and str(value).strip() for value in row_values):
                    filtered_rows.append((source_idx, row_values))

            return filtered_rows

        def detect_header_row(filtered_rows, header_candidate_idx):
            if not filtered_rows:
                return None
            if header_candidate_idx is None:
                return 0
            for filtered_idx, (source_idx, _) in enumerate(filtered_rows):
                if source_idx >= header_candidate_idx:
                    return filtered_idx
            return 0

        def parse_numeric(value_text):
            normalized = value_text.replace(',', '').replace(' ', '')
            normalized = normalized.replace('−', '-').replace('ー', '-').replace('―', '-')
            if NUMERIC_SIMPLE_PATTERN.match(normalized):
                try:
                    return float(normalized)
                except ValueError:
                    return None
            return None

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{pdf_path_obj.stem}_structured_{timestamp}.xlsx"
        output_path = self.output_dir / output_filename

        page_analysis_results = []

        hybrid_results: List[Dict[str, Any]] = []

        for page_index in range(total_pages):
            page = doc.load_page(page_index)
            analysis = self.analyze_page_content(page)
            analysis["page_index"] = page_index + 1
            page_analysis_results.append(analysis)

            page_number = page_index + 1
            print_prefix = f"ページ{page_number}"

            if analysis.get("is_text_based"):
                print(f"📄 {print_prefix}: テキスト主体 (文字数={analysis['char_count']}, 画像数={analysis['image_count']})")
            else:
                print(f"🖼️ {print_prefix}: 画像主体 (文字数={analysis['char_count']}, 画像数={analysis['image_count']})")

            # --- テキスト主体ページの処理 -------------------------------------
            structured_table = None
            structured_tables = []
            extraction_method = ""

            if analysis.get("is_text_based"):
                if self.camelot_available:
                    try:
                        tables = camelot.read_pdf(
                            str(pdf_path_obj),
                            pages=str(page_number),
                            flavor="lattice",
                            strip_text='\n'
                        )
                        if tables.n > 0:
                            structured_tables = [table.df for table in tables]
                            structured_table = structured_tables[0]
                            extraction_method = "camelot_lattice"
                            print(f"   ✅ Camelot(lattice)で抽出成功: {tables.n}件の表")
                        else:
                            print("   ⚠️ Camelot(lattice)では抽出できませんでした")
                    except Exception as e:
                        print(f"   ⚠️ Camelot(lattice)エラー: {e}")

                # latticeで何も取れなければstream試行
                if structured_table is None and self.camelot_available:
                    try:
                        tables = camelot.read_pdf(
                            str(pdf_path_obj),
                            pages=str(page_number),
                            flavor="stream",
                            strip_text='\n'
                        )
                        if tables.n > 0:
                            structured_tables = [table.df for table in tables]
                            structured_table = structured_tables[0]
                            extraction_method = "camelot_stream"
                            print(f"   ✅ Camelot(stream)で抽出成功: {tables.n}件の表")
                        else:
                            print("   ⚠️ Camelot(stream)でも抽出できませんでした")
                    except Exception as e:
                        print(f"   ⚠️ Camelot(stream)エラー: {e}")

                if structured_table is None and self.pdfplumber_available:
                    try:
                        with pdfplumber.open(str(pdf_path_obj)) as pdf_plumber_doc:
                            pdf_page = pdf_plumber_doc.pages[page_index]
                            tables = pdf_page.extract_tables()
                            if tables:
                                structured_tables = tables
                                structured_table = structured_tables[0]
                                extraction_method = "pdfplumber"
                                print(f"   ✅ pdfplumberで抽出成功: {len(tables)}件の表")
                            else:
                                print("   ⚠️ pdfplumberでも抽出できませんでした")
                    except Exception as e:
                        print(f"   ⚠️ pdfplumberエラー: {e}")

            # --- 画像主体、またはテキスト抽出失敗時はOCR処理 ------------------
            use_easyocr = False
            image = None
            ocr_results = []

            if structured_table is None:
                use_easyocr = True
                print("   ▶️ EasyOCRで抽出を実施 (テキスト抽出失敗または画像主体)")

                mat = fitz.Matrix(dpi / 72, dpi / 72)
                pix = page.get_pixmap(matrix=mat)
                image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

                # より積極的なリサイズ（メモリ節約）
                max_dim = max(image.width, image.height)
                if max_dim > 2800:
                    scale = 2800 / max_dim
                    if scale < 1.0:
                        new_width = max(1, int(image.width * scale))
                        new_height = max(1, int(image.height * scale))
                        image = image.resize((new_width, new_height), Image.BICUBIC)
                        print(f"   画像をリサイズ: {image.width}x{image.height}")

                max_slice_width = 1200
                overlap = 20
                combined_results = []

                if image.width > max_slice_width:
                    start_x = 0
                    while start_x < image.width:
                        end_x = min(image.width, start_x + max_slice_width)
                        crop = image.crop((start_x, 0, end_x, image.height))
                        crop_array = np.array(crop)

                        gc.collect()
                        if torch is not None and torch.cuda.is_available():
                            torch.cuda.empty_cache()

                        slice_results = self.easyocr_reader.readtext(
                            crop_array,
                            detail=1,
                            mag_ratio=1.0,
                            canvas_size=1200,
                            batch_size=1
                        )

                        for detection in slice_results:
                            if not detection or len(detection) < 3:
                                continue
                            bbox, text, conf = detection[0], detection[1], detection[2]
                            adjusted_bbox = [[point[0] + start_x, point[1]] for point in bbox]
                            combined_results.append((adjusted_bbox, text, conf))

                        if end_x == image.width:
                            break
                        start_x = max(0, end_x - overlap)

                    ocr_results = combined_results
                else:
                    img_array = np.array(image)

                    gc.collect()
                    if torch is not None and torch.cuda.is_available():
                        torch.cuda.empty_cache()

                    ocr_results = self.easyocr_reader.readtext(
                        img_array,
                        detail=1,
                        mag_ratio=1.0,
                        canvas_size=1200,
                        batch_size=1
                    )

            image_info = None
            if image is not None:
                image_info = {"width": image.width, "height": image.height}

            hybrid_results.append({
                "page_index": page_number,
                "analysis": analysis,
                "extraction_method": extraction_method if structured_table is not None else "easyocr",
                "structured_table": structured_table,
                "structured_tables": structured_tables,
                "use_easyocr": use_easyocr,
                "ocr_results": ocr_results,
                "image_info": image_info
            })

            # OCR結果を処理する際は一時的に使用した画像を解放
            image = None
            # シートの用意
            if page_index == 0:
                ws = default_sheet
            else:
                ws = workbook.create_sheet()

            ws.title = f"ページ{page_number}"

            # テキストベースの表が抽出できた場合はそのまま出力
            if structured_table is not None and not use_easyocr:
                title_cell = ws['A1']
                method_label = extraction_method.upper() if extraction_method else "TEXT"
                table_count = len(structured_tables) if structured_tables else 1
                title_cell.value = f"ページ {page_number} - {method_label} テーブル ({table_count}件)"
                title_cell.font = title_font
                title_cell.fill = title_fill
                title_cell.alignment = Alignment(horizontal='center', vertical='center')

                current_start_row = 3
                table_index = 1

                for table_data in structured_tables or [structured_table]:
                    if not table_data:
                        continue

                    if hasattr(table_data, "values"):
                        table_rows = table_data.values.tolist()
                    else:
                        table_rows = table_data

                    if not table_rows:
                        continue

                    max_cols = max(len(row) for row in table_rows if row) if table_rows else 0
                    if max_cols == 0:
                        continue

                    # 各テーブルごとのタイトル行
                    table_title_cell = ws.cell(row=current_start_row, column=1)
                    table_title_cell.value = f"テーブル{table_index}"
                    table_title_cell.font = header_font
                    table_title_cell.fill = header_fill
                    table_title_cell.alignment = Alignment(horizontal='left', vertical='center')

                    current_start_row += 1

                    for row in table_rows:
                        for col_idx, value in enumerate(row, start=1):
                            cell = ws.cell(row=current_start_row, column=col_idx)
                            cell.value = value if value not in ("", None) else None
                            cell.border = border_style
                            if current_start_row == 4:  # 最初のテーブルのヘッダーとみなす
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            else:
                                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        current_start_row += 1

                    current_start_row += 1  # テーブル間の空行
                    table_index += 1

                max_used_col = ws.max_column
                for col_idx in range(1, max_used_col + 1):
                    col_letter = get_column_letter(col_idx)
                    max_length = 0
                    for row_idx in range(1, ws.max_row + 1):
                        value = ws.cell(row=row_idx, column=col_idx).value
                        if value:
                            max_length = max(max_length, len(str(value)))
                    ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

                ws.freeze_panes = ws.cell(row=4, column=1)
                continue

            # ここからはOCRベースでの出力
            rows = build_rows_from_results(ocr_results)
            column_centers = determine_column_centers(rows)
            assignments, column_usage, total_columns = assign_rows_to_columns(rows, column_centers)

            header_candidate_idx = None
            for idx, row_map in enumerate(assignments):
                if len([v for v in row_map.values() if v and str(v).strip()]) >= 3:
                    header_candidate_idx = idx
                    break

            header_columns = set(assignments[header_candidate_idx].keys()) if header_candidate_idx is not None else set()
            columns_to_keep = determine_columns_to_keep(assignments, column_usage, total_columns, header_columns)

            filtered_rows = build_matrix(assignments, columns_to_keep)
            header_row_index = detect_header_row(filtered_rows, header_candidate_idx)

            if not filtered_rows:
                ws['A1'] = f"ページ {page_number} - データが検出できませんでした"
                ws['A1'].font = title_font
                continue

            total_active_columns = len(columns_to_keep)
            title_cell = ws['A1']
            title_cell.value = f"ページ {page_number} - OCR構造化テーブル"
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            if total_active_columns > 1:
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_active_columns)

            data_start_row = 3
            column_widths = [10] * total_active_columns
            merge_candidates = defaultdict(list)

            for relative_row_idx, (source_idx, row_values) in enumerate(filtered_rows):
                excel_row = data_start_row + relative_row_idx
                is_header = header_row_index is not None and relative_row_idx == header_row_index

                for col_offset, value in enumerate(row_values, start=1):
                    if value is None:
                        continue

                    display_text = value if isinstance(value, str) else str(value)
                    display_text = display_text.strip()
                    if not display_text:
                        continue

                    numeric_value = parse_numeric(display_text)
                    cell = ws.cell(row=excel_row, column=col_offset)

                    if numeric_value is not None:
                        cell.value = numeric_value
                        if numeric_value.is_integer():
                            cell.number_format = '#,##0'
                        else:
                            cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.value = display_text
                        if is_header:
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                    if is_header:
                        cell.font = header_font
                        cell.fill = header_fill

                    cell.border = border_style

                    max_line_length = max(len(line.strip()) for line in display_text.split('\n') if line.strip()) if display_text else 0
                    column_widths[col_offset - 1] = min(max(column_widths[col_offset - 1], max_line_length + 2), 60)

                    merge_candidates[col_offset - 1].append((relative_row_idx, display_text, numeric_value is None))

                ws.row_dimensions[excel_row].height = 18

            for idx, width in enumerate(column_widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width

            merge_ranges = set()
            for col_idx, column_values in merge_candidates.items():
                start_idx = None
                previous_value = None
                for row_idx, display_text, is_text in column_values:
                    # 数値はマージしない
                    if not is_text:
                        if start_idx is not None and row_idx - start_idx >= 1:
                            merge_ranges.add((col_idx, start_idx, row_idx - 1))
                        start_idx = None
                        previous_value = None
                        continue

                    # 空のテキストはマージしない
                    if not display_text:
                        if start_idx is not None and row_idx - start_idx >= 1:
                            merge_ranges.add((col_idx, start_idx, row_idx - 1))
                        start_idx = None
                        previous_value = None
                        continue

                    # ヘッダー行はマージしない（または特別な処理）
                    is_header_row = header_row_index is not None and row_idx == header_row_index
                    if is_header_row:
                        if start_idx is not None and row_idx - start_idx >= 1:
                            merge_ranges.add((col_idx, start_idx, row_idx - 1))
                        start_idx = None
                        previous_value = None
                        continue

                    # 短いテキスト（1-2文字）はマージしない（誤認識の可能性）
                    if len(display_text) <= 2:
                        if start_idx is not None and row_idx - start_idx >= 1:
                            merge_ranges.add((col_idx, start_idx, row_idx - 1))
                        start_idx = None
                        previous_value = None
                        continue

                    if previous_value is None:
                        start_idx = row_idx
                        previous_value = display_text
                    elif previous_value == display_text:
                        continue  # 同じ値が続いている
                    else:
                        # 値が変わったので、前の範囲をマージ
                        if start_idx is not None and row_idx - start_idx >= 1:
                            merge_ranges.add((col_idx, start_idx, row_idx - 1))
                        start_idx = row_idx
                        previous_value = display_text

                # 最後の範囲を処理
                if start_idx is not None and column_values:
                    last_row_idx = column_values[-1][0]
                    # 最小2行以上でマージ（1行だけではマージしない）
                    if last_row_idx - start_idx >= 1:
                        merge_ranges.add((col_idx, start_idx, last_row_idx))

            for col_idx, start_rel, end_rel in merge_ranges:
                if end_rel <= start_rel:
                    continue
                start_excel_row = data_start_row + start_rel
                end_excel_row = data_start_row + end_rel
                if end_excel_row - start_excel_row < 1:
                    continue
                col_letter = get_column_letter(col_idx + 1)
                ws.merge_cells(f"{col_letter}{start_excel_row}:{col_letter}{end_excel_row}")
                merged_cell = ws.cell(row=start_excel_row, column=col_idx + 1)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            ws.freeze_panes = ws.cell(row=data_start_row, column=1)

        doc.close()
        self.latest_page_analysis = page_analysis_results
        self.latest_hybrid_results = hybrid_results

        try:
            workbook.save(output_path)
            print(f"✅ 構造化Excel生成完了: {output_path}")

            # X280デスクトップへ自動コピー
            try:
                desktop_dir = Path('/home/mana/Desktop/HybridPDF_Excel')
                desktop_dir.mkdir(parents=True, exist_ok=True)
                desktop_path = desktop_dir / Path(output_path).name
                shutil.copy2(output_path, desktop_path)
                print(f"   📁 X280デスクトップにコピー: {desktop_path}")
            except Exception as copy_error:
                print(f"   ⚠️ デスクトップへのコピーに失敗: {copy_error}")

            return str(output_path)
        except Exception as e:
            print(f"❌ Excel保存エラー: {e}")
            return None

    def text_to_excel(self, text: str, base_name: str):
        """テキストをExcelに変換"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "EasyOCR抽出結果"

            # ヘッダー
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            ws.append(["行番号", "内容"])
            ws['A1'].font = header_font
            ws['A1'].fill = header_fill
            ws['A1'].alignment = header_alignment
            ws['B1'].font = header_font
            ws['B1'].fill = header_fill
            ws['B1'].alignment = header_alignment

            # データ行
            lines = [line.strip() for line in text.split('\n') if line.strip()]

            for i, line in enumerate(lines, 1):
                ws.append([i, line])

            # 列幅調整
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 120

            # 保存
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_filename = f"{base_name}_easyocr_{timestamp}.xlsx"
            excel_path = self.output_dir / excel_filename

            wb.save(excel_path)

            try:
                desktop_dir = Path('/home/mana/Desktop/HybridPDF_Excel')
                desktop_dir.mkdir(parents=True, exist_ok=True)
                desktop_path = desktop_dir / excel_path.name
                shutil.copy2(excel_path, desktop_path)
                print(f"   📁 X280デスクトップにコピー: {desktop_path}")
            except Exception as copy_error:
                print(f"   ⚠️ デスクトップへのコピーに失敗: {copy_error}")

            return str(excel_path)

        except Exception as e:
            print(f"❌ Excel生成エラー: {e}")
            import traceback
            traceback.print_exc()
            return None

    def convert_pdf(self, pdf_path: str, mode: str = "auto"):
        """PDFをEasyOCRで変換（構造化優先、フォールバックあり）"""
        pdf_path = Path(pdf_path)

        if not pdf_path.exists():
            print(f"❌ PDFファイルが見つかりません: {pdf_path}")
            return None

        print("=" * 60)
        print(f"📄 EasyOCR変換開始: {pdf_path.name}")
        print("=" * 60)

        mode = (mode or "auto").lower()
        structured_allowed = mode in ("structured", "auto")
        text_allowed = mode in ("text", "auto")

        try:
            structured_result = None
            if structured_allowed:
                print("\n📊 Step 1: 構造化Excel生成を試行中...")
                structured_result = self.convert_pdf_to_structured_excel(str(pdf_path))
                if structured_result:
                    print("\n" + "=" * 60)
                    print("✅ 構造化Excel変換完了！")
                    print("=" * 60)
                    print(f"📊 Excelファイル: {structured_result}")
                    print("=" * 60)
                    return structured_result
                elif mode == "structured":
                    print("❌ 構造化Excelの生成に失敗しました")
                    return None
                else:
                    print("⚠️ 構造化Excel生成に失敗したため、テキストモードにフォールバックします。")

            if not text_allowed:
                return None

            print("\n📸 Step 2: PDFを高解像度画像に変換中...")
            image = self.pdf_to_high_res_image(str(pdf_path), page_num=0, dpi=300)

            if not image:
                print("❌ 画像変換に失敗しました")
                return None

            print(f"✅ 画像変換完了（解像度: {image.size[0]}x{image.size[1]}）")

            print("\n🔍 Step 3: EasyOCRで高精度抽出中...")
            text = self.extract_with_easyocr(image)

            if not text:
                print("❌ テキスト抽出に失敗しました")
                return None

            print("\n📊 Step 4: テキストExcelファイル生成中...")
            excel_path = self.text_to_excel(text, pdf_path.stem)

            if excel_path:
                print("\n" + "=" * 60)
                print("✅ EasyOCRテキスト変換完了！")
                print("=" * 60)
                print(f"📊 Excelファイル: {excel_path}")
                print(f"📝 抽出テキスト: {len(text)}文字")
                print("=" * 60)
                return excel_path

            print("❌ Excel生成に失敗しました")
            return None

        except Exception as e:
            print(f"❌ 変換エラー: {e}")
            import traceback
            traceback.print_exc()
            return None

def main():
    if len(sys.argv) < 2:
        print("使い方:")
        print("  python3 easyocr_pdf_to_excel.py [PDFファイルパス] [mode(optional)]")
        print("\n例:")
        print("  python3 easyocr_pdf_to_excel.py /path/to/file.pdf")
        print("  mode: auto (既定) / structured / text")
        sys.exit(1)

    pdf_path = sys.argv[1]
    mode = sys.argv[2] if len(sys.argv) > 2 else "auto"

    converter = EasyOCRPDFToExcel()
    excel_path = converter.convert_pdf(pdf_path, mode=mode)

    if excel_path:
        print(f"\n🎉 変換完了！ファイル: {excel_path}")
    else:
        print("\n❌ 変換に失敗しました")

if __name__ == "__main__":
    main()









