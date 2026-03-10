#!/usr/bin/env python3
"""
日報PDF→Excel変換システム (Gemini Vision版)
Google Gemini Vision APIで画像から表データを抽出してExcel化

特徴：
- Gemini Visionの高精度画像認識
- 表構造の完全理解
- JSON形式でデータ抽出
- Excel表形式で出力
"""

import os
import sys
from pathlib import Path
import logging
from datetime import datetime
import json
import time

# 必要なライブラリ
try:
    import fitz  # PyMuPDF
    from PIL import Image
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import google.generativeai as genai
except ImportError as e:
    print(f"❌ 必要なライブラリが不足: {e}")
    print("インストール: pip3 install PyMuPDF Pillow pandas openpyxl google-generativeai")
    sys.exit(1)

# ログ設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/root/logs/nippo_pdf_to_excel_gemini.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("NippoPDFToExcelGemini")


class NippoPDFToExcelGemini:
    """日報PDF→Excel変換システム (Gemini Vision版)"""
    
    def __init__(self, api_key: str = None):  # type: ignore
        self.output_dir = Path('/root/daily_reports/excel_gemini')
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Gemini API設定
        self.api_key = api_key or os.getenv("GEMINI_API_KEY", "AIzaSyC5SA9-bXkB_nlAux9Fve47HvvDm1jx0Y0")
        
        if not self.api_key:
            logger.error("❌ GEMINI_API_KEYが設定されていません")
            sys.exit(1)
        
        genai.configure(api_key=self.api_key)
        
        # Gemini Visionモデル（Gemini 2.5 Flash）
        self.model = genai.GenerativeModel('gemini-2.5-flash')
        logger.info("✅ Gemini Vision API初期化完了")
    
    def pdf_to_image(self, pdf_path: str, page_num: int = 0, dpi: int = 200):
        """PDFを画像に変換"""
        try:
            doc = fitz.open(pdf_path)
            page = doc[page_num]
            
            # 画像化（DPI 200で十分）
            zoom = dpi / 72
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            
            # PIL Imageに変換
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)  # type: ignore
            
            doc.close()
            return img
            
        except Exception as e:
            logger.error(f"PDF→画像変換エラー: {e}")
            return None
    
    def extract_table_data_with_gemini(self, image):
        """Gemini Visionで画像から表データを抽出"""
        try:
            # 超詳細プロンプト（表構造保持版）
            prompt = """
この画像は日報または見積書のPDFです。
画像内の表を、セル配置をそのまま保持してJSON形式で出力してください。

【最重要指示】
この画像の表を、PDFと全く同じレイアウト（行・列の配置）で出力してください。

【出力形式】
{
  "表データ": [
    ["列1ヘッダー", "列2ヘッダー", "列3ヘッダー", ...],
    ["行1列1の値", "行1列2の値", "行1列3の値", ...],
    ["行2列1の値", "行2列2の値", "行2列3の値", ...],
    ...
  ]
}

【認識精度の指示】
1. 文字は100%正確に（漢字、固有名詞、数値）
2. 会社名、人名、地名は絶対に間違えない
3. 数値はカンマ区切り含めて正確に
4. 空のセルは空文字列 ""
5. 結合セルは最初のセルに値を入れる

【重要】
- 表の行・列の配置を絶対に崩さない
- PDFの表と全く同じ構造にする
- 表が複数ある場合は、"表1", "表2"と分ける
- 説明文・コードブロック不要、JSON のみ

JSONを出力：
"""
            
            # Gemini Visionで解析
            logger.info("🤖 Gemini Vision解析中...")
            response = self.model.generate_content([prompt, image])
            
            # レスポンステキスト取得
            response_text = response.text.strip()
            
            # JSONコードブロックを削除
            if response_text.startswith('```'):
                response_text = response_text.split('```')[1]
                if response_text.startswith('json'):
                    response_text = response_text[4:]
                response_text = response_text.strip()
            
            # JSON解析
            data = json.loads(response_text)
            
            # 新フォーマット対応
            if isinstance(data, dict):
                if '表データ' in data:
                    table_data = data['表データ']
                    logger.info(f"✅ データ抽出成功: {len(table_data)}行")
                    return table_data
                elif '表1' in data:
                    # 複数表の場合、最初の表を使用
                    table_data = data['表1']
                    logger.info(f"✅ データ抽出成功: {len(table_data)}行")
                    return table_data
            
            # 旧フォーマット（配列）の場合
            if isinstance(data, list):
                logger.info(f"✅ データ抽出成功: {len(data)}行")
                return data
            
            logger.error(f"不明なデータ形式: {type(data)}")
            return None
            
        except json.JSONDecodeError as e:
            logger.error(f"JSON解析エラー: {e}")
            logger.error(f"レスポンス: {response_text[:500]}")  # type: ignore[possibly-unbound]
            
            # フォールバック：テキストとして保存
            return [{"内容": response.text}]  # type: ignore[possibly-unbound]
            
        except Exception as e:
            logger.error(f"Gemini解析エラー: {e}")
            return None
    
    def data_to_dataframe(self, data):
        """JSON配列をDataFrameに変換"""
        try:
            if not data:
                return None
            
            # 2次元配列（表形式）の場合
            if isinstance(data, list) and len(data) > 0 and isinstance(data[0], list):
                # 表データをそのままDataFrameに
                df = pd.DataFrame(data[1:], columns=data[0] if data else None)
            else:
                # 通常の配列の場合
                df = pd.DataFrame(data)
            
            # 空の列を削除
            df = df.dropna(how='all', axis=1)
            
            # 空の行を削除
            df = df.dropna(how='all', axis=0)
            
            return df
            
        except Exception as e:
            logger.error(f"DataFrame変換エラー: {e}")
            return None
    
    def create_excel(self, df, output_path: str, pdf_name: str):
        """DataFrameからExcelファイルを作成"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '日報データ'  # type: ignore[union-attr]
            
            # ヘッダー行を書き込み
            for c_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=1, column=c_idx, value=col_name)  # type: ignore[union-attr]
                cell.font = Font(name='メイリオ', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # データ行を書き込み
            for r_idx, row in enumerate(df.itertuples(index=False), start=2):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)  # type: ignore[union-attr]
                    cell.font = Font(name='メイリオ', size=11)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # 交互に背景色
                    if r_idx % 2 == 0:
                        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # 列幅自動調整
            for column in ws.columns:  # type: ignore[union-attr]
                max_length = 0
                column_letter = column[0].column_letter  # type: ignore
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width  # type: ignore[union-attr]
            
            # サマリーシート追加
            summary_ws = wb.create_sheet('変換情報', 0)
            summary_data = [
                ['項目', '値'],
                ['元PDFファイル', pdf_name],
                ['変換日時', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['抽出行数', len(df)],
                ['抽出列数', len(df.columns)],
                ['システム', '日報PDF→Excel変換 (Gemini Vision版)'],
                ['AI', 'Google Gemini 1.5 Flash']
            ]
            
            for r_idx, row_data in enumerate(summary_data, start=1):
                for c_idx, value in enumerate(row_data, start=1):
                    cell = summary_ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.font = Font(name='メイリオ', size=11)
                    if r_idx == 1:
                        cell.font = Font(name='メイリオ', size=11, bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            summary_ws.column_dimensions['A'].width = 20
            summary_ws.column_dimensions['B'].width = 50
            
            # 保存
            wb.save(output_path)
            logger.info(f"✅ Excel作成完了: {output_path}")
            
            return True
            
        except Exception as e:
            logger.error(f"Excel作成エラー: {e}")
            return False
    
    def convert_pdf(self, pdf_path: str):
        """PDFをExcelに変換（メイン処理）"""
        pdf_path = Path(pdf_path)  # type: ignore
        
        if not pdf_path.exists():  # type: ignore
            logger.error(f"PDFファイルが見つかりません: {pdf_path}")
            return {'success': False, 'error': 'File not found'}
        
        print(f"\n{'='*60}")
        print(f"📄 変換開始: {pdf_path.name}")  # type: ignore
        print(f"{'='*60}")
        
        try:
            # 1. PDFを画像に変換
            print("🖼️  STEP 1: PDF→画像変換...")
            image = self.pdf_to_image(str(pdf_path), dpi=200)
            
            if image is None:
                return {'success': False, 'error': 'Image conversion failed'}
            
            print(f"   画像サイズ: {image.size[0]} x {image.size[1]}")
            
            # 2. Gemini Visionでデータ抽出
            print("🤖 STEP 2: Gemini Vision解析...")
            data = self.extract_table_data_with_gemini(image)
            
            if not data:
                return {'success': False, 'error': 'Data extraction failed'}
            
            print(f"   抽出データ: {len(data)}行")
            
            # 3. DataFrameに変換
            print("📊 STEP 3: DataFrame変換...")
            
            # データフラット化
            import json as json_module
            
            # 2次元配列（表形式）の場合はそのまま
            if isinstance(data, list) and len(data) > 0 and isinstance(data[0], list):
                df = self.data_to_dataframe(data)
            else:
                # オブジェクト配列の場合はフラット化
                flattened_data = []
                for item in data:
                    flat_item = {}
                    for key, value in item.items():
                        if isinstance(value, (list, dict)):
                            flat_item[key] = json_module.dumps(value, ensure_ascii=False)
                        else:
                            flat_item[key] = value
                    flattened_data.append(flat_item)
                
                df = self.data_to_dataframe(flattened_data)
            
            if df is None or df.empty:
                return {'success': False, 'error': 'DataFrame conversion failed'}
            
            print(f"   データ: {len(df)}行 x {len(df.columns)}列")
            print(f"   列名: {list(df.columns)[:5]}..." if len(df.columns) > 5 else f"   列名: {list(df.columns)}")
            
            # 4. Excel作成
            print("📝 STEP 4: Excel作成...")
            output_filename = f"{pdf_path.stem}_gemini_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"  # type: ignore
            output_path = self.output_dir / output_filename
            
            success = self.create_excel(df, str(output_path), pdf_path.name)  # type: ignore
            
            if success:
                print("\n✅ 変換完了！")
                print(f"📁 出力: {output_path}")
                print(f"{'='*60}\n")
                
                return {
                    'success': True,
                    'output_file': str(output_path),
                    'rows': len(df),
                    'cols': len(df.columns)
                }
            else:
                return {'success': False, 'error': 'Excel creation failed'}
            
        except Exception as e:
            logger.error(f"変換エラー: {e}")
            return {'success': False, 'error': str(e)}
    
    def batch_convert(self, pdf_dir: str, limit: int = None, delay: int = 2):  # type: ignore
        """フォルダ内の全PDFを一括変換"""
        pdf_dir = Path(pdf_dir)  # type: ignore
        
        if not pdf_dir.exists():  # type: ignore
            logger.error(f"フォルダが見つかりません: {pdf_dir}")
            return
        
        pdf_files = list(pdf_dir.glob('*.pdf'))  # type: ignore
        pdf_files.extend(list(pdf_dir.glob('*.PDF')))  # type: ignore
        
        if not pdf_files:
            logger.error("PDFファイルが見つかりません")
            return
        
        # 件数制限
        if limit:
            pdf_files = pdf_files[:limit]
        
        print(f"\n{'='*60}")
        print("📁 一括変換開始 (Gemini Vision)")
        print(f"{'='*60}")
        print(f"PDFファイル数: {len(pdf_files)}")
        print(f"出力先: {self.output_dir}")
        print(f"API待機時間: {delay}秒/ファイル\n")
        
        results = {'total': len(pdf_files), 'success': 0, 'failed': 0}
        
        for i, pdf_file in enumerate(pdf_files, 1):
            print(f"[{i}/{len(pdf_files)}] {pdf_file.name}")
            
            result = self.convert_pdf(str(pdf_file))
            
            if result['success']:
                results['success'] += 1
            else:
                results['failed'] += 1
                print(f"   ❌ 失敗: {result.get('error', '不明')}\n")
            
            # API制限対策（待機）
            if i < len(pdf_files):
                print(f"   ⏳ 待機 {delay}秒...\n")
                time.sleep(delay)
        
        # サマリー
        print(f"\n{'='*60}")
        print("📊 一括変換完了 (Gemini Vision)")
        print(f"{'='*60}")
        print(f"総ファイル数: {results['total']}")
        print(f"✅ 成功: {results['success']}")
        print(f"❌ 失敗: {results['failed']}")
        print(f"📁 出力先: {self.output_dir}")
        print(f"{'='*60}\n")


def main():
    """メイン処理"""
    
    if len(sys.argv) < 2:
        print("使い方:")
        print("  python3 nippo_pdf_to_excel_gemini.py [PDFファイル]")
        print("  python3 nippo_pdf_to_excel_gemini.py [フォルダ] [件数制限]")
        print("\n例:")
        print("  python3 nippo_pdf_to_excel_gemini.py test.pdf")
        print("  python3 nippo_pdf_to_excel_gemini.py /path/to/pdfs 10")
        sys.exit(1)
    
    converter = NippoPDFToExcelGemini()
    
    target = sys.argv[1]
    target_path = Path(target)
    
    if target_path.is_file():
        converter.convert_pdf(target)
    elif target_path.is_dir():
        limit = int(sys.argv[2]) if len(sys.argv) > 2 else None
        converter.batch_convert(target, limit=limit, delay=2)  # type: ignore
    else:
        print(f"❌ パスが見つかりません: {target}")
        sys.exit(1)


if __name__ == '__main__':
    main()

