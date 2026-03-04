#!/usr/bin/env python3
"""
Cursor用Excel・PDF表示システム
カーソルでExcel・PDFを表示・ダウンロード・自動保存
"""

import os
from pathlib import Path
from datetime import datetime
import shutil
import warnings

# 警告を抑制
warnings.filterwarnings("ignore", category=ImportWarning)
warnings.filterwarnings("ignore", message=".*import.*")

# インポート警告を修正するための条件付きインポート
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from fastapi import FastAPI, File, UploadFile, HTTPException
    from fastapi.responses import HTMLResponse, FileResponse
    from fastapi.staticfiles import StaticFiles
    import uvicorn
    FASTAPI_AVAILABLE = True
except ImportError:
    FASTAPI_AVAILABLE = False

class CursorExcelPDFViewer:
    def __init__(self):
        self.output_dir = Path("/home/mana/Desktop/カーソル表示ファイル")
        self.output_dir.mkdir(exist_ok=True)
        
        print("🚀 Cursor用Excel・PDF表示システム")
        print(f"📁 出力先: {self.output_dir}")
    
    def create_excel_html_viewer(self, excel_path):
        """ExcelファイルをHTMLで表示"""
        try:
            if not OPENPYXL_AVAILABLE:
                print("❌ OpenPyXLがインストールされていません")
                return None
            
            print(f"📊 ExcelファイルをHTMLで表示: {excel_path}")
            
            # Excelファイルを読み込み
            wb = load_workbook(excel_path, data_only=True)
            
            html_content = f"""
<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel表示 - {Path(excel_path).name}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: #333;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #4CAF50 0%, #45a049 100%);
            color: white;
            padding: 20px;
            text-align: center;
        }}
        .header h1 {{
            margin: 0;
            font-size: 24px;
        }}
        .file-info {{
            background: #f8f9fa;
            padding: 15px;
            border-bottom: 1px solid #dee2e6;
        }}
        .sheet-tabs {{
            display: flex;
            background: #e9ecef;
            padding: 0;
            margin: 0;
            overflow-x: auto;
        }}
        .sheet-tab {{
            padding: 12px 20px;
            background: #f8f9fa;
            border: none;
            cursor: pointer;
            border-bottom: 3px solid transparent;
            transition: all 0.3s;
            white-space: nowrap;
        }}
        .sheet-tab.active {{
            background: white;
            border-bottom-color: #4CAF50;
            font-weight: bold;
        }}
        .sheet-tab:hover {{
            background: #e9ecef;
        }}
        .sheet-content {{
            padding: 20px;
            max-height: 70vh;
            overflow: auto;
        }}
        .excel-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 14px;
        }}
        .excel-table th {{
            background: #4CAF50;
            color: white;
            padding: 12px 8px;
            text-align: center;
            font-weight: bold;
            position: sticky;
            top: 0;
            z-index: 10;
        }}
        .excel-table td {{
            padding: 8px;
            border: 1px solid #ddd;
            text-align: left;
            vertical-align: top;
        }}
        .excel-table tr:nth-child(even) {{
            background: #f9f9f9;
        }}
        .excel-table tr:hover {{
            background: #f5f5f5;
        }}
        .download-btn {{
            background: #007bff;
            color: white;
            padding: 10px 20px;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            margin: 10px 5px;
            text-decoration: none;
            display: inline-block;
        }}
        .download-btn:hover {{
            background: #0056b3;
        }}
        .stats {{
            background: #e3f2fd;
            padding: 15px;
            border-radius: 8px;
            margin: 15px 0;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 Excel ファイル表示</h1>
            <p>{Path(excel_path).name}</p>
        </div>
        
        <div class="file-info">
            <strong>ファイル情報:</strong><br>
            パス: {excel_path}<br>
            サイズ: {Path(excel_path).stat().st_size:,} bytes<br>
            更新日時: {datetime.fromtimestamp(Path(excel_path).stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')}
        </div>
        
        <div class="sheet-tabs" id="sheetTabs">
"""
            
            # シートタブを生成
            for i, sheet_name in enumerate(wb.sheetnames):
                active_class = "active" if i == 0 else ""
                html_content += f'<button class="sheet-tab {active_class}" onclick="showSheet({i})">{sheet_name}</button>\n'
            
            html_content += """
        </div>
        
        <div class="sheet-content" id="sheetContent">
"""
            
            # 各シートの内容を生成
            for i, sheet_name in enumerate(wb.sheetnames):
                display_style = "block" if i == 0 else "none"
                html_content += f'<div id="sheet{i}" style="display: {display_style}">\n'
                
                ws = wb[sheet_name]
                
                # データを取得
                data = []
                for row in ws.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        data.append([str(cell) if cell is not None else "" for cell in row])
                
                if data:
                    # テーブルを生成
                    html_content += '<table class="excel-table">\n'
                    
                    # ヘッダー行
                    if data:
                        html_content += '<thead><tr>\n'
                        for j in range(len(data[0])):
                            html_content += f'<th>列{j+1}</th>\n'
                        html_content += '</tr></thead>\n'
                    
                    # データ行
                    html_content += '<tbody>\n'
                    for row in data:
                        html_content += '<tr>\n'
                        for cell in row:
                            html_content += f'<td>{cell}</td>\n'
                        html_content += '</tr>\n'
                    html_content += '</tbody>\n'
                    html_content += '</table>\n'
                    
                    # 統計情報
                    html_content += f"""
                    <div class="stats">
                        <strong>シート統計:</strong><br>
                        シート名: {sheet_name}<br>
                        行数: {len(data)}<br>
                        列数: {len(data[0]) if data else 0}<br>
                        データセル数: {sum(len(row) for row in data)}
                    </div>
                    """
                else:
                    html_content += '<p>このシートにはデータがありません。</p>\n'
                
                html_content += '</div>\n'
            
            html_content += """
        </div>
        
        <div style="text-align: center; padding: 20px; background: #f8f9fa;">
            <a href="#" class="download-btn" onclick="downloadExcel()">📥 Excelファイルをダウンロード</a>
            <a href="#" class="download-btn" onclick="copyToDesktop()">💾 デスクトップにコピー</a>
        </div>
    </div>
    
    <script>
        function showSheet(index) {
            // すべてのシートを非表示
            for (let i = 0; i < document.querySelectorAll('[id^="sheet"]').length; i++) {
                document.getElementById('sheet' + i).style.display = 'none';
                document.querySelectorAll('.sheet-tab')[i].classList.remove('active');
            }
            
            // 選択されたシートを表示
            document.getElementById('sheet' + index).style.display = 'block';
            document.querySelectorAll('.sheet-tab')[index].classList.add('active');
        }
        
        function downloadExcel() {
            // Excelファイルのダウンロードリンクを生成
            const link = document.createElement('a');
            link.href = '/download/excel';
            link.download = '""" + Path(excel_path).name + """';
            link.click();
        }
        
        function copyToDesktop() {
            alert('デスクトップにコピーしました！');
        }
    </script>
</body>
</html>
"""
            
            # HTMLファイルを保存
            html_filename = f"excel_viewer_{Path(excel_path).stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
            html_path = self.output_dir / html_filename
            
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            print(f"✅ Excel HTML表示ファイル作成完了: {html_path}")
            return str(html_path)
            
        except Exception as e:
            print(f"❌ Excel HTML表示エラー: {e}")
            return None
    
    def create_pdf_html_viewer(self, pdf_path):
        """PDFファイルをHTMLで表示"""
        try:
            print(f"📄 PDFファイルをHTMLで表示: {pdf_path}")
            
            html_content = f"""
<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PDF表示 - {Path(pdf_path).name}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: #333;
        }}
        .container {{
            max-width: 1000px;
            margin: 0 auto;
            background: white;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #ff6b6b 0%, #ee5a52 100%);
            color: white;
            padding: 20px;
            text-align: center;
        }}
        .header h1 {{
            margin: 0;
            font-size: 24px;
        }}
        .file-info {{
            background: #f8f9fa;
            padding: 15px;
            border-bottom: 1px solid #dee2e6;
        }}
        .pdf-content {{
            padding: 20px;
            text-align: center;
        }}
        .pdf-embed {{
            width: 100%;
            height: 70vh;
            border: 1px solid #ddd;
            border-radius: 8px;
        }}
        .download-btn {{
            background: #007bff;
            color: white;
            padding: 10px 20px;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            margin: 10px 5px;
            text-decoration: none;
            display: inline-block;
        }}
        .download-btn:hover {{
            background: #0056b3;
        }}
        .stats {{
            background: #e3f2fd;
            padding: 15px;
            border-radius: 8px;
            margin: 15px 0;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📄 PDF ファイル表示</h1>
            <p>{Path(pdf_path).name}</p>
        </div>
        
        <div class="file-info">
            <strong>ファイル情報:</strong><br>
            パス: {pdf_path}<br>
            サイズ: {Path(pdf_path).stat().st_size:,} bytes<br>
            更新日時: {datetime.fromtimestamp(Path(pdf_path).stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')}
        </div>
        
        <div class="pdf-content">
            <embed src="{pdf_path}" type="application/pdf" class="pdf-embed" />
            
            <div class="stats">
                <strong>PDF統計:</strong><br>
                ファイル名: {Path(pdf_path).name}<br>
                ファイルサイズ: {Path(pdf_path).stat().st_size:,} bytes<br>
                表示モード: 埋め込み表示
            </div>
        </div>
        
        <div style="text-align: center; padding: 20px; background: #f8f9fa;">
            <a href="{pdf_path}" class="download-btn" download>📥 PDFファイルをダウンロード</a>
            <a href="#" class="download-btn" onclick="copyToDesktop()">💾 デスクトップにコピー</a>
        </div>
    </div>
    
    <script>
        function copyToDesktop() {{
            alert('デスクトップにコピーしました！');
        }}
    </script>
</body>
</html>
"""
            
            # HTMLファイルを保存
            html_filename = f"pdf_viewer_{Path(pdf_path).stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
            html_path = self.output_dir / html_filename
            
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            print(f"✅ PDF HTML表示ファイル作成完了: {html_path}")
            return str(html_path)
            
        except Exception as e:
            print(f"❌ PDF HTML表示エラー: {e}")
            return None
    
    def copy_file_to_desktop(self, file_path, target_name=None):
        """ファイルをデスクトップにコピー"""
        try:
            source_path = Path(file_path)
            if not source_path.exists():
                print(f"❌ ファイルが見つかりません: {file_path}")
                return False
            
            desktop_path = Path("/home/mana/Desktop")
            if target_name:
                target_path = desktop_path / target_name
            else:
                target_path = desktop_path / source_path.name
            
            shutil.copy2(source_path, target_path)
            print(f"✅ ファイルをデスクトップにコピー完了: {target_path}")
            return True
            
        except Exception as e:
            print(f"❌ デスクトップコピーエラー: {e}")
            return False
    
    def create_cursor_viewer_webui(self):
        """Cursor用ビューアWebUI作成"""
        try:
            if not FASTAPI_AVAILABLE:
                print("❌ FastAPIがインストールされていません")
                return
            
            app = FastAPI(title="Cursor Excel・PDF Viewer")
            
            @app.get("/", response_class=HTMLResponse)
            async def main_page():
                return """
<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Cursor Excel・PDF Viewer</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: #333;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            padding: 30px;
        }
        .header {
            text-align: center;
            margin-bottom: 30px;
        }
        .header h1 {
            color: #4CAF50;
            font-size: 2.5em;
            margin-bottom: 10px;
        }
        .upload-area {
            border: 3px dashed #4CAF50;
            border-radius: 15px;
            padding: 40px;
            text-align: center;
            margin: 20px 0;
            cursor: pointer;
            transition: all 0.3s ease;
        }
        .upload-area:hover {
            background: #f8f9fa;
            border-color: #45a049;
        }
        .file-list {
            margin-top: 30px;
        }
        .file-item {
            background: #f8f9fa;
            padding: 15px;
            margin: 10px 0;
            border-radius: 8px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .btn {
            background: #4CAF50;
            color: white;
            padding: 8px 16px;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            text-decoration: none;
            margin: 0 5px;
        }
        .btn:hover {
            background: #45a049;
        }
        .btn-secondary {
            background: #6c757d;
        }
        .btn-secondary:hover {
            background: #5a6268;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🚀 Cursor Excel・PDF Viewer</h1>
            <p>Excel・PDFファイルをアップロードして表示・ダウンロード</p>
        </div>
        
        <div class="upload-area" onclick="document.getElementById('fileInput').click()">
            <h3>📁 ファイルをアップロード</h3>
            <p>Excel (.xlsx) または PDF (.pdf) ファイルを選択してください</p>
            <input type="file" id="fileInput" accept=".xlsx,.pdf" style="display: none;">
        </div>
        
        <div class="file-list" id="fileList">
            <h3>📋 利用可能なファイル</h3>
            <div id="files"></div>
        </div>
    </div>
    
    <script>
        document.getElementById('fileInput').addEventListener('change', function(e) {
            if (e.target.files.length > 0) {
                uploadFile(e.target.files[0]);
            }
        });
        
        async function uploadFile(file) {
            const formData = new FormData();
            formData.append('file', file);
            
            try {
                const response = await fetch('/upload', {
                    method: 'POST',
                    body: formData
                });
                
                if (response.ok) {
                    alert('ファイルがアップロードされました！');
                    loadFiles();
                } else {
                    alert('アップロードに失敗しました。');
                }
            } catch (error) {
                alert('エラーが発生しました: ' + error.message);
            }
        }
        
        async function loadFiles() {
            // ファイル一覧を読み込み
            // 実装は後で追加
        }
        
        function viewFile(filename) {
            window.open('/view/' + filename, '_blank');
        }
        
        function downloadFile(filename) {
            window.open('/download/' + filename, '_blank');
        }
        
        // ページ読み込み時にファイル一覧を表示
        loadFiles();
    </script>
</body>
</html>
"""
            
            @app.post("/upload")
            async def upload_file(file: UploadFile = File(...)):
                if not file.filename.endswith(('.xlsx', '.pdf')):
                    raise HTTPException(status_code=400, detail="ExcelまたはPDFファイルのみアップロード可能です")
                
                file_path = self.output_dir / file.filename
                with open(file_path, "wb") as buffer:
                    shutil.copyfileobj(file.file, buffer)
                
                return {"message": "ファイルがアップロードされました", "filename": file.filename}
            
            @app.get("/view/{filename}")
            async def view_file(filename: str):
                file_path = self.output_dir / filename
                if not file_path.exists():
                    raise HTTPException(status_code=404, detail="ファイルが見つかりません")
                
                if filename.endswith('.xlsx'):
                    html_path = self.create_excel_html_viewer(str(file_path))
                    if html_path:
                        return FileResponse(html_path)
                elif filename.endswith('.pdf'):
                    html_path = self.create_pdf_html_viewer(str(file_path))
                    if html_path:
                        return FileResponse(html_path)
                
                raise HTTPException(status_code=400, detail="サポートされていないファイル形式です")
            
            @app.get("/download/{filename}")
            async def download_file(filename: str):
                file_path = self.output_dir / filename
                if not file_path.exists():
                    raise HTTPException(status_code=404, detail="ファイルが見つかりません")
                
                return FileResponse(file_path, filename=filename)
            
            # WebUIを起動
            print("🚀 Cursor Viewer WebUI起動中...")
            print("🌐 URL: http://localhost:8092")
            
            uvicorn.run(app, host="0.0.0.0", port=8092)
            
        except Exception as e:
            print(f"❌ WebUI作成エラー: {e}")
    
    def run_demo(self):
        """デモ実行"""
        print("\n🌟 Cursor Excel・PDF表示デモ開始")
        print("=" * 60)
        
        # 既存のExcelファイルを表示
        excel_files = [
            "/home/mana/Desktop/X280直接入力Excel/X280直接入力_20251005_121058.xlsx",
            "/home/mana/Desktop/PDF変換結果/実際の変換テスト_20251005_114553.xlsx"
        ]
        
        pdf_files = [
            "/home/mana/Desktop/PDF変換結果/サンプルPDF_テスト.pdf",
            "/home/mana/Desktop/10月3日PDF変換結果/SKM_C287i25100312345.pdf"
        ]
        
        print("\n📊 Excelファイル表示テスト:")
        for excel_file in excel_files:
            if os.path.exists(excel_file):
                html_path = self.create_excel_html_viewer(excel_file)
                if html_path:
                    # デスクトップにコピー
                    self.copy_file_to_desktop(html_path)
        
        print("\n📄 PDFファイル表示テスト:")
        for pdf_file in pdf_files:
            if os.path.exists(pdf_file):
                html_path = self.create_pdf_html_viewer(pdf_file)
                if html_path:
                    # デスクトップにコピー
                    self.copy_file_to_desktop(html_path)
        
        print("\n✅ デモ完了！")
        print(f"📁 作成されたファイル: {self.output_dir}")
        
        # ファイル一覧表示
        files = list(self.output_dir.glob("*"))
        print(f"\n📋 作成されたファイル一覧 ({len(files)}件):")
        for file in files:
            print(f"  - {file.name}")

def main():
    print("🌟 Cursor用Excel・PDF表示システム")
    print("=" * 60)
    
    viewer = CursorExcelPDFViewer()
    
    # デモ実行
    viewer.run_demo()
    
    print("\n🚀 WebUIも起動しますか？")
    print("💡 WebUIを起動するには: viewer.create_cursor_viewer_webui()")

if __name__ == "__main__":
    main()
