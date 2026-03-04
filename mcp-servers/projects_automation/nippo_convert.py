#!/usr/bin/env python3
"""
日報PDF→Excel/スプレッドシート変換（最終版）
Gemini Vision使用・表構造完全保持

使い方:
  python3 nippo_convert.py [PDFファイル]
  python3 nippo_convert.py [フォルダ] [件数]
"""

import os
import sys
from pathlib import Path
import logging
from datetime import datetime
import time
import json

try:
    import fitz
    from PIL import Image
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import google.generativeai as genai
except ImportError as e:
    print(f"❌ ライブラリ不足: {e}")
    sys.exit(1)

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger("NippoConvert")


class NippoConverter:
    """日報PDF→Excel変換"""
    
    def __init__(self):
        self.output_dir = Path('/root/daily_reports/excel_final')
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Gemini初期化
        api_key = os.getenv("GEMINI_API_KEY", "AIzaSyC5SA9-bXkB_nlAux9Fve47HvvDm1jx0Y0")
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel('gemini-2.5-flash')
    
    def pdf_to_image(self, pdf_path: str, dpi: int = 200):
        """PDF→画像"""
        doc = fitz.open(pdf_path)
        page = doc[0]
        zoom = dpi / 72
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        doc.close()
        return img
    
    def gemini_extract_table(self, image):
        """Gemini Visionで表抽出"""
        prompt = """
この画像のPDFを、表形式でそのまま抽出してください。

【出力形式】
{
  "表データ": [
    ["列1", "列2", "列3", ...],
    ["値1-1", "値1-2", "値1-3", ...],
    ["値2-1", "値2-2", "値2-3", ...],
    ...
  ]
}

【重要】
- PDFの表構造を完全に保持
- 行・列の配置はそのまま
- 全ての文字を正確に認識
- 空セルは ""
- JSON のみ出力（説明不要）
"""
        
        response = self.model.generate_content([prompt, image])
        text = response.text.strip()
        
        if '```' in text:
            text = text.split('```')[1]
            if text.startswith('json'):
                text = text[4:]
            text = text.strip()
        
        data = json.loads(text)
        
        if isinstance(data, dict) and '表データ' in data:
            return data['表データ']
        elif isinstance(data, list):
            return data
        
        return None
    
    def create_excel(self, table_data, output_path: str, pdf_name: str):
        """Excel作成"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        
        # 表データをそのまま書き込み
        for r_idx, row_data in enumerate(table_data, start=1):
            for c_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = Font(name='メイリオ', size=10)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # 1行目は強調
                if r_idx == 1:
                    cell.font = Font(name='メイリオ', size=10, bold=True)
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 列幅調整
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
        
        wb.save(output_path)
        return True
    
    def convert(self, pdf_path: str):
        """変換実行"""
        pdf_path = Path(pdf_path)
        
        print(f"\n📄 {pdf_path.name}")
        
        try:
            # 1. 画像化
            image = self.pdf_to_image(str(pdf_path))
            print(f"   🖼️  画像: {image.size[0]}x{image.size[1]}")
            
            # 2. Gemini抽出
            print("   🤖 Gemini解析中...")
            table_data = self.gemini_extract_table(image)
            
            if not table_data:
                print("   ❌ 抽出失敗")
                return {'success': False}
            
            rows = len(table_data)
            cols = len(table_data[0]) if table_data else 0
            print(f"   ✅ 表抽出: {rows}行 x {cols}列")
            
            # 3. Excel作成
            output_name = f"{pdf_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = self.output_dir / output_name
            
            self.create_excel(table_data, str(output_path), pdf_path.name)
            print(f"   📊 Excel: {output_path.name}\n")
            
            return {'success': True, 'rows': rows, 'cols': cols}
            
        except Exception as e:
            print(f"   ❌ エラー: {e}\n")
            return {'success': False}
    
    def batch(self, pdf_dir: str, limit: int = None):
        """一括変換"""
        pdf_dir = Path(pdf_dir)
        pdfs = list(pdf_dir.glob('*.pdf')) + list(pdf_dir.glob('*.PDF'))
        
        if limit:
            pdfs = pdfs[:limit]
        
        print(f"\n{'='*60}")
        print("📁 日報PDF→Excel一括変換")
        print(f"{'='*60}")
        print(f"ファイル数: {len(pdfs)}")
        print(f"出力先: {self.output_dir}\n")
        
        results = {'total': len(pdfs), 'success': 0, 'failed': 0}
        
        for i, pdf in enumerate(pdfs, 1):
            print(f"[{i}/{len(pdfs)}]", end=' ')
            result = self.convert(str(pdf))
            
            if result['success']:
                results['success'] += 1
            else:
                results['failed'] += 1
            
            if i < len(pdfs):
                time.sleep(1)
        
        print(f"{'='*60}")
        print(f"✅ 成功: {results['success']}/{results['total']}")
        print(f"❌ 失敗: {results['failed']}")
        print(f"📁 {self.output_dir}")
        print(f"{'='*60}\n")


def main():
    if len(sys.argv) < 2:
        print("使い方:")
        print("  python3 nippo_convert.py [PDF]")
        print("  python3 nippo_convert.py [フォルダ] [件数]")
        sys.exit(1)
    
    converter = NippoConverter()
    target = Path(sys.argv[1])
    
    if target.is_file():
        converter.convert(str(target))
    elif target.is_dir():
        limit = int(sys.argv[2]) if len(sys.argv) > 2 else None
        converter.batch(str(target), limit=limit)


if __name__ == '__main__':
    main()

