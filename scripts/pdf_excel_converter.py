#!/usr/bin/env python3
"""
PDF-Excel変換システム
ManaOS統合システムに組み込まれた高機能PDF処理エンジン
"""

import json
import asyncio
import logging
import uuid
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, Any
import threading

# データ処理ライブラリ
try:
    import pandas as pd
    import numpy as np
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# PDF処理ライブラリ
try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

try:
    import PyPDF2
    PYPDF2_AVAILABLE = True
except ImportError:
    PYPDF2_AVAILABLE = False

# OCR処理（オプション）
try:
    import pytesseract
    from PIL import Image
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

class PDFExcelConverter:
    """PDF-Excel変換システム"""
    
    def __init__(self, config: Dict[str, Any] = None):  # type: ignore
        self.config = config or self.get_default_config()
        self.logger = self.setup_logger()
        self.temp_dir = Path("/tmp/pdf_excel_converter")
        self.temp_dir.mkdir(exist_ok=True)
        
        # 処理中タスクの管理
        self.active_tasks: Dict[str, Dict[str, Any]] = {}
        self.task_lock = threading.Lock()
        
        self.logger.info("PDF-Excel変換システム初期化完了")
    
    def get_default_config(self) -> Dict[str, Any]:
        """デフォルト設定取得"""
        return {
            "output_format": "excel",  # excel, csv, both
            "ocr_enabled": True,
            "table_detection": True,
            "page_range": "all",
            "google_drive_save": False,
            "template_style": "default",
            "language": "jpn",
            "max_file_size_mb": 50,
            "timeout_seconds": 300,
            "parallel_processing": True,
            "temp_cleanup": True
        }
    
    def setup_logger(self) -> logging.Logger:
        """ロガー設定"""
        logger = logging.getLogger("PDFExcelConverter")
        logger.setLevel(logging.INFO)
        
        # ファイルハンドラー
        log_file = f"/root/logs/pdf_excel_converter_{datetime.now().strftime('%Y%m%d')}.log"
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        
        # コンソールハンドラー
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        
        # フォーマッター
        formatter = logging.Formatter(
            '%(asctime)s [%(levelname)s] %(name)s: %(message)s'
        )
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)
        
        logger.addHandler(file_handler)
        logger.addHandler(console_handler)
        
        return logger
    
    def validate_pdf_file(self, file_path: str) -> Dict[str, Any]:
        """PDFファイル検証"""
        try:
            file_path = Path(file_path)  # type: ignore
            
            if not file_path.exists():  # type: ignore
                return {"valid": False, "error": "ファイルが存在しません"}
            
            if not file_path.suffix.lower() == '.pdf':  # type: ignore
                return {"valid": False, "error": "PDFファイルではありません"}
            
            # ファイルサイズチェック
            file_size_mb = file_path.stat().st_size / (1024 * 1024)  # type: ignore
            if file_size_mb > self.config["max_file_size_mb"]:
                return {
                    "valid": False, 
                    "error": f"ファイルサイズが上限を超えています ({file_size_mb:.1f}MB > {self.config['max_file_size_mb']}MB)"
                }
            
            # PDFファイルの基本検証
            if PYMUPDF_AVAILABLE:
                try:
                    doc = fitz.open(file_path)  # type: ignore[possibly-unbound]
                    page_count = len(doc)
                    doc.close()
                    
                    return {
                        "valid": True,
                        "file_size_mb": file_size_mb,
                        "page_count": page_count,
                        "file_path": str(file_path)
                    }
                except Exception as e:
                    return {"valid": False, "error": f"PDFファイルの読み込みに失敗: {e}"}
            
            return {"valid": True, "file_size_mb": file_size_mb}
            
        except Exception as e:
            return {"valid": False, "error": f"ファイル検証エラー: {e}"}
    
    def extract_text_from_pdf(self, pdf_path: str) -> Dict[str, Any]:
        """PDFからテキスト抽出"""
        try:
            if not PYMUPDF_AVAILABLE:
                return {"success": False, "error": "PyMuPDFが利用できません"}
            
            doc = fitz.open(pdf_path)  # type: ignore[possibly-unbound]
            extracted_data = {
                "pages": [],
                "full_text": "",
                "metadata": doc.metadata,
                "page_count": len(doc)
            }
            
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                text = page.get_text()
                
                page_data = {
                    "page_number": page_num + 1,
                    "text": text,
                    "char_count": len(text)
                }
                
                extracted_data["pages"].append(page_data)
                extracted_data["full_text"] += text + "\n"  # type: ignore
            
            doc.close()
            
            self.logger.info(f"テキスト抽出完了: {len(extracted_data['pages'])}ページ")
            return {"success": True, "data": extracted_data}
            
        except Exception as e:
            self.logger.error(f"テキスト抽出エラー: {e}")
            return {"success": False, "error": str(e)}
    
    def extract_tables_from_pdf(self, pdf_path: str) -> Dict[str, Any]:
        """PDFから表データ抽出"""
        try:
            if not PDFPLUMBER_AVAILABLE:
                return {"success": False, "error": "pdfplumberが利用できません"}
            
            tables_data = {
                "tables": [],
                "total_tables": 0
            }
            
            with pdfplumber.open(pdf_path) as pdf:  # type: ignore[possibly-unbound]
                for page_num, page in enumerate(pdf.pages):
                    tables = page.extract_tables()
                    
                    for table_num, table in enumerate(tables):
                        if table and len(table) > 0:
                            # 表データのクリーニング
                            cleaned_table = []
                            for row in table:
                                cleaned_row = []
                                for cell in row:
                                    if cell is None:
                                        cleaned_row.append("")
                                    else:
                                        cleaned_row.append(str(cell).strip())
                                cleaned_table.append(cleaned_row)
                            
                            table_data = {
                                "page_number": page_num + 1,
                                "table_number": table_num + 1,
                                "data": cleaned_table,
                                "rows": len(cleaned_table),
                                "columns": len(cleaned_table[0]) if cleaned_table else 0
                            }
                            
                            tables_data["tables"].append(table_data)
            
            tables_data["total_tables"] = len(tables_data["tables"])
            
            self.logger.info(f"表データ抽出完了: {tables_data['total_tables']}個の表")
            return {"success": True, "data": tables_data}
            
        except Exception as e:
            self.logger.error(f"表データ抽出エラー: {e}")
            return {"success": False, "error": str(e)}
    
    def extract_images_from_pdf(self, pdf_path: str) -> Dict[str, Any]:
        """PDFから画像抽出（OCR処理含む）"""
        try:
            if not PYMUPDF_AVAILABLE:
                return {"success": False, "error": "PyMuPDFが利用できません"}
            
            images_data = {
                "images": [],
                "total_images": 0,
                "ocr_text": ""
            }
            
            doc = fitz.open(pdf_path)  # type: ignore[possibly-unbound]
            
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                image_list = page.get_images()
                
                for img_num, img in enumerate(image_list):
                    try:
                        # 画像データ取得
                        xref = img[0]
                        pix = fitz.Pixmap(doc, xref)  # type: ignore[possibly-unbound]
                        
                        if pix.n - pix.alpha < 4:  # GRAY or RGB
                            img_filename = f"page_{page_num+1}_img_{img_num+1}.png"
                            img_path = self.temp_dir / img_filename
                            pix.save(img_path)
                            
                            image_info = {
                                "page_number": page_num + 1,
                                "image_number": img_num + 1,
                                "filename": img_filename,
                                "path": str(img_path),
                                "width": pix.width,
                                "height": pix.height
                            }
                            
                            # OCR処理
                            if OCR_AVAILABLE and self.config["ocr_enabled"]:
                                try:
                                    ocr_text = pytesseract.image_to_string(  # type: ignore[possibly-unbound]
                                        Image.open(img_path),  # type: ignore[possibly-unbound]
                                        lang='jpn+eng'
                                    )
                                    image_info["ocr_text"] = ocr_text
                                    images_data["ocr_text"] += ocr_text + "\n"
                                except Exception as ocr_error:
                                    self.logger.warning(f"OCR処理エラー: {ocr_error}")
                                    image_info["ocr_text"] = ""
                            
                            images_data["images"].append(image_info)
                        
                        pix = None
                        
                    except Exception as img_error:
                        self.logger.warning(f"画像抽出エラー: {img_error}")
                        continue
            
            doc.close()
            images_data["total_images"] = len(images_data["images"])
            
            self.logger.info(f"画像抽出完了: {images_data['total_images']}個の画像")
            return {"success": True, "data": images_data}
            
        except Exception as e:
            self.logger.error(f"画像抽出エラー: {e}")
            return {"success": False, "error": str(e)}
    
    def create_excel_from_data(self, text_data: Dict, tables_data: Dict, 
                              images_data: Dict, output_path: str) -> Dict[str, Any]:
        """抽出データからExcelファイル生成"""
        try:
            if not OPENPYXL_AVAILABLE or not PANDAS_AVAILABLE:
                return {"success": False, "error": "openpyxlまたはpandasが利用できません"}
            
            workbook = openpyxl.Workbook()  # type: ignore[possibly-unbound]
            
            # デフォルトシート削除
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])
            
            # スタイル設定
            header_font = Font(bold=True, color="FFFFFF")  # type: ignore[possibly-unbound]
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # type: ignore[possibly-unbound]
            header_alignment = Alignment(horizontal="center", vertical="center")  # type: ignore[possibly-unbound]
            border = Border(  # type: ignore[possibly-unbound]
                left=Side(style='thin'),  # type: ignore[possibly-unbound]
                right=Side(style='thin'),  # type: ignore[possibly-unbound]
                top=Side(style='thin'),  # type: ignore[possibly-unbound]
                bottom=Side(style='thin')  # type: ignore[possibly-unbound]
            )
            
            sheet_index = 1
            
            # 1. 概要シート（見やすく改善）
            summary_sheet = workbook.create_sheet("📋 概要", 0)
            
            # タイトル
            summary_sheet.append(["📄 PDF-Excel変換結果"])
            summary_sheet.append([])  # 空行
            
            # 基本情報
            summary_sheet.append(["📊 基本情報"])
            summary_data = [
                ["処理日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["総ページ数", text_data.get("page_count", 0)],
                ["抽出テキスト数", len(text_data.get("pages", []))],
                ["抽出表数", tables_data.get("total_tables", 0)],
                ["抽出画像数", images_data.get("total_images", 0)]
            ]
            
            for row in summary_data:
                summary_sheet.append(row)
            
            summary_sheet.append([])  # 空行
            
            # シート構成
            summary_sheet.append(["📑 シート構成"])
            sheet_info = [
                ["シート名", "内容"],
                ["概要", "このシート - 変換結果の概要"],
                ["ページ_1〜N", "各ページのテキスト内容（1ページ1シート）"],
                ["OCR_ページ_X", "各ページのOCR抽出テキスト（画像から抽出）"],
                ["表_X_Y", "抽出された表データ（ページXの表Y）"]
            ]
            
            for row in sheet_info:
                summary_sheet.append(row)
            
            # スタイル適用
            # タイトル行
            for cell in summary_sheet.iter_rows(min_row=1, max_row=1):
                for cell_obj in cell:
                    if cell_obj.value:
                        cell_obj.font = Font(bold=True, size=16, color="FFFFFF")  # type: ignore[possibly-unbound]
                        cell_obj.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")  # type: ignore[possibly-unbound]
                        cell_obj.alignment = Alignment(horizontal="center", vertical="center")  # type: ignore[possibly-unbound]
                        cell_obj.border = border
            
            # セクションヘッダー
            for row_num in [3, 8]:  # "基本情報"と"シート構成"の行
                for cell in summary_sheet.iter_rows(min_row=row_num, max_row=row_num):
                    for cell_obj in cell:
                        if cell_obj.value:
                            cell_obj.font = Font(bold=True, size=12, color="2F5597")  # type: ignore[possibly-unbound]
                            cell_obj.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")  # type: ignore[possibly-unbound]
                            cell_obj.border = border
            
            # データ行のヘッダー
            for row_num in [4, 9]:  # データヘッダー行
                for cell in summary_sheet.iter_rows(min_row=row_num, max_row=row_num):
                    for cell_obj in cell:
                        if cell_obj.value:
                            cell_obj.font = Font(bold=True, color="FFFFFF")  # type: ignore[possibly-unbound]
                            cell_obj.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # type: ignore[possibly-unbound]
                            cell_obj.alignment = Alignment(horizontal="center", vertical="center")  # type: ignore[possibly-unbound]
                            cell_obj.border = border
            
            # 列幅の自動調整
            for column in summary_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                summary_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # 2. 各ページのテキストシート（1ページ1シート）
            if text_data.get("pages"):
                for page_data in text_data["pages"]:
                    # 制御文字を除去
                    clean_text = page_data["text"]
                    import re
                    clean_text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', clean_text)
                    
                    # ページ番号でシート名を作成
                    sheet_name = f"ページ_{page_data['page_number']}"
                    page_sheet = workbook.create_sheet(sheet_name)
                    
                    # ページ情報ヘッダー
                    page_sheet.append(["ページ情報"])
                    page_sheet.append(["ページ番号", page_data["page_number"]])
                    page_sheet.append(["文字数", page_data["char_count"]])
                    page_sheet.append(["処理日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                    page_sheet.append([])  # 空行
                    
                    # テキスト内容
                    page_sheet.append(["テキスト内容"])
                    
                    # テキストを行ごとに分割して書き込み
                    text_lines = clean_text.split('\n')
                    for line in text_lines:
                        if line.strip():  # 空行はスキップ
                            page_sheet.append([line.strip()])
                    
                    # スタイル適用
                    # ヘッダー行のスタイル
                    for row_num in [1, 2, 3, 4, 6]:  # ヘッダー行
                        for cell in page_sheet.iter_rows(min_row=row_num, max_row=row_num):
                            for cell_obj in cell:
                                if cell_obj.value:
                                    cell_obj.font = Font(bold=True, color="2F5597")  # type: ignore[possibly-unbound]
                                    cell_obj.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")  # type: ignore[possibly-unbound]
                                    cell_obj.border = border
                    
                    # 列幅の自動調整
                    for column in page_sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except Exception:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        page_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # 3. 表データシート
            if tables_data.get("tables"):
                for table_data in tables_data["tables"]:
                    sheet_name = f"表_{table_data['page_number']}_{table_data['table_number']}"
                    if len(sheet_name) > 31:  # Excelのシート名制限
                        sheet_name = f"Table_{table_data['page_number']}_{table_data['table_number']}"
                    
                    table_sheet = workbook.create_sheet(sheet_name)
                    
                    # 表データの書き込み
                    for row in table_data["data"]:
                        # 各行の制御文字を除去
                        clean_row = []
                        for cell in row:
                            if isinstance(cell, str):
                                import re
                                clean_cell = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', cell)
                                clean_row.append(clean_cell)
                            else:
                                clean_row.append(cell)
                        table_sheet.append(clean_row)
                    
                    # 表シートのスタイル適用
                    if table_data["data"]:
                        for row in table_sheet.iter_rows(min_row=1, max_row=1):
                            for cell in row:
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = header_alignment
                                cell.border = border
                        
                        # 全セルにボーダー適用
                        for row in table_sheet.iter_rows():
                            for cell in row:
                                cell.border = border
            
            # 4. OCRテキストシート（ページごと）
            if images_data.get("ocr_text") and self.config["ocr_enabled"]:
                # ページごとにOCRテキストをグループ化
                page_ocr_data = {}
                for img_data in images_data["images"]:
                    if img_data.get("ocr_text"):
                        page_num = img_data['page_number']
                        if page_num not in page_ocr_data:
                            page_ocr_data[page_num] = []
                        
                        # 制御文字を除去
                        clean_ocr_text = img_data["ocr_text"]
                        import re
                        clean_ocr_text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', clean_ocr_text)
                        
                        page_ocr_data[page_num].append({
                            'image_number': img_data['image_number'],
                            'text': clean_ocr_text
                        })
                
                # 各ページのOCRシートを作成
                for page_num, ocr_list in page_ocr_data.items():
                    sheet_name = f"OCR_ページ_{page_num}"
                    ocr_sheet = workbook.create_sheet(sheet_name)
                    
                    # ヘッダー情報
                    ocr_sheet.append(["OCR抽出情報"])
                    ocr_sheet.append(["ページ番号", page_num])
                    ocr_sheet.append(["画像数", len(ocr_list)])
                    ocr_sheet.append(["処理日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                    ocr_sheet.append([])  # 空行
                    
                    # OCRテキスト内容
                    ocr_sheet.append(["OCR抽出テキスト"])
                    
                    for i, ocr_item in enumerate(ocr_list, 1):
                        ocr_sheet.append([f"画像{i} (No.{ocr_item['image_number']})"])
                        # テキストを行ごとに分割
                        text_lines = ocr_item['text'].split('\n')
                        for line in text_lines:
                            if line.strip():
                                ocr_sheet.append([f"  {line.strip()}"])
                        ocr_sheet.append([])  # 画像間の空行
                    
                    # スタイル適用
                    for row_num in [1, 2, 3, 4, 6]:  # ヘッダー行
                        for cell in ocr_sheet.iter_rows(min_row=row_num, max_row=row_num):
                            for cell_obj in cell:
                                if cell_obj.value:
                                    cell_obj.font = Font(bold=True, color="8B4513")  # type: ignore[possibly-unbound]
                                    cell_obj.fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")  # type: ignore[possibly-unbound]
                                    cell_obj.border = border
                    
                    # 列幅の自動調整
                    for column in ocr_sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except Exception:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ocr_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # 列幅の自動調整
            for sheet in workbook.worksheets:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter  # type: ignore
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            # ファイル保存
            workbook.save(output_path)
            
            self.logger.info(f"Excelファイル生成完了: {output_path}")
            return {"success": True, "output_path": output_path}
            
        except Exception as e:
            self.logger.error(f"Excel生成エラー: {e}")
            return {"success": False, "error": str(e)}
    
    def create_csv_from_tables(self, tables_data: Dict, output_path: str) -> Dict[str, Any]:
        """表データからCSVファイル生成"""
        try:
            if not PANDAS_AVAILABLE:
                return {"success": False, "error": "pandasが利用できません"}
            
            csv_files = []
            
            for table_data in tables_data["tables"]:
                if table_data["data"]:
                    # DataFrame作成
                    df = pd.DataFrame(table_data["data"][1:], columns=table_data["data"][0])  # type: ignore[possibly-unbound]
                    
                    # CSVファイル名生成
                    csv_filename = f"table_{table_data['page_number']}_{table_data['table_number']}.csv"
                    csv_path = output_path.replace('.csv', f'_{csv_filename}')
                    
                    # CSV保存
                    df.to_csv(csv_path, index=False, encoding='utf-8')
                    csv_files.append(csv_path)
            
            self.logger.info(f"CSVファイル生成完了: {len(csv_files)}個のファイル")
            return {"success": True, "csv_files": csv_files}
            
        except Exception as e:
            self.logger.error(f"CSV生成エラー: {e}")
            return {"success": False, "error": str(e)}
    
    async def convert_pdf_to_excel(self, pdf_path: str, output_path: str = None,  # type: ignore
                                 config: Dict[str, Any] = None) -> Dict[str, Any]:  # type: ignore
        """PDFからExcelへの変換（メイン処理）"""
        task_id = str(uuid.uuid4())
        
        try:
            # 設定更新
            if config:
                self.config.update(config)
            
            # タスク登録
            with self.task_lock:
                self.active_tasks[task_id] = {
                    "status": "processing",
                    "progress": 0,
                    "start_time": datetime.now().isoformat(),
                    "pdf_path": pdf_path,
                    "output_path": output_path
                }
            
            # 出力パス設定
            if not output_path:
                pdf_name = Path(pdf_path).stem
                output_path = str(self.temp_dir / f"{pdf_name}_converted.xlsx")
            
            # ファイル検証
            validation = self.validate_pdf_file(pdf_path)
            if not validation["valid"]:
                return {"success": False, "error": validation["error"]}
            
            # 進捗更新
            with self.task_lock:
                self.active_tasks[task_id]["progress"] = 10
            
            # 1. テキスト抽出
            self.logger.info("テキスト抽出開始")
            text_result = self.extract_text_from_pdf(pdf_path)
            if not text_result["success"]:
                return {"success": False, "error": f"テキスト抽出エラー: {text_result['error']}"}
            
            with self.task_lock:
                self.active_tasks[task_id]["progress"] = 30
            
            # 2. 表データ抽出
            self.logger.info("表データ抽出開始")
            tables_result = self.extract_tables_from_pdf(pdf_path)
            if not tables_result["success"]:
                return {"success": False, "error": f"表データ抽出エラー: {tables_result['error']}"}
            
            with self.task_lock:
                self.active_tasks[task_id]["progress"] = 60
            
            # 3. 画像・OCR抽出
            self.logger.info("画像・OCR抽出開始")
            images_result = self.extract_images_from_pdf(pdf_path)
            if not images_result["success"]:
                self.logger.warning(f"画像抽出エラー: {images_result['error']}")
                images_result = {"success": True, "data": {"images": [], "total_images": 0, "ocr_text": ""}}
            
            with self.task_lock:
                self.active_tasks[task_id]["progress"] = 80
            
            # 4. Excel生成
            self.logger.info("Excel生成開始")
            excel_result = self.create_excel_from_data(
                text_result["data"],
                tables_result["data"],
                images_result["data"],
                output_path
            )
            
            if not excel_result["success"]:
                return {"success": False, "error": f"Excel生成エラー: {excel_result['error']}"}
            
            # CSV生成（必要に応じて）
            csv_files = []
            if self.config["output_format"] in ["csv", "both"]:
                csv_result = self.create_csv_from_tables(tables_result["data"], output_path)
                if csv_result["success"]:
                    csv_files = csv_result["csv_files"]
            
            # 完了処理
            processing_time = (datetime.now() - datetime.fromisoformat(
                self.active_tasks[task_id]["start_time"]
            )).total_seconds()
            
            with self.task_lock:
                self.active_tasks[task_id].update({
                    "status": "completed",
                    "progress": 100,
                    "end_time": datetime.now().isoformat(),
                    "processing_time": processing_time,
                    "output_files": {
                        "excel": output_path if self.config["output_format"] in ["excel", "both"] else None,
                        "csv": csv_files if csv_files else None
                    }
                })
            
            result = {
                "success": True,
                "task_id": task_id,
                "output_files": {
                    "excel": output_path if self.config["output_format"] in ["excel", "both"] else None,
                    "csv": csv_files if csv_files else None
                },
                "processing_time": processing_time,
                "extraction_summary": {
                    "pages_processed": text_result["data"]["page_count"],
                    "tables_extracted": tables_result["data"]["total_tables"],
                    "images_extracted": images_result["data"]["total_images"],
                    "text_length": len(text_result["data"]["full_text"])
                }
            }
            
            self.logger.info(f"PDF-Excel変換完了: {task_id}")
            return result
            
        except Exception as e:
            self.logger.error(f"PDF-Excel変換エラー: {e}")
            with self.task_lock:
                if task_id in self.active_tasks:
                    self.active_tasks[task_id].update({
                        "status": "failed",
                        "error": str(e),
                        "end_time": datetime.now().isoformat()
                    })
            
            return {"success": False, "error": str(e), "task_id": task_id}
    
    def get_task_status(self, task_id: str) -> Dict[str, Any]:
        """タスク状態取得"""
        with self.task_lock:
            return self.active_tasks.get(task_id, {"error": "タスクが見つかりません"})
    
    def cleanup_temp_files(self):
        """一時ファイルクリーンアップ"""
        try:
            if self.config["temp_cleanup"] and self.temp_dir.exists():
                shutil.rmtree(self.temp_dir)
                self.temp_dir.mkdir(exist_ok=True)
                self.logger.info("一時ファイルクリーンアップ完了")
        except Exception as e:
            self.logger.error(f"クリーンアップエラー: {e}")
    
    def get_system_status(self) -> Dict[str, Any]:
        """システム状態取得"""
        return {
            "libraries": {
                "pandas": PANDAS_AVAILABLE,
                "openpyxl": OPENPYXL_AVAILABLE,
                "pymupdf": PYMUPDF_AVAILABLE,
                "pdfplumber": PDFPLUMBER_AVAILABLE,
                "pypdf2": PYPDF2_AVAILABLE,
                "ocr": OCR_AVAILABLE
            },
            "config": self.config,
            "active_tasks": len(self.active_tasks),
            "temp_dir": str(self.temp_dir)
        }

# 使用例
async def main():
    """使用例"""
    converter = PDFExcelConverter()
    
    # システム状態確認
    status = converter.get_system_status()
    print("システム状態:", json.dumps(status, indent=2, ensure_ascii=False))
    
    # サンプル変換（実際のPDFファイルパスを指定）
    # result = await converter.convert_pdf_to_excel("/path/to/sample.pdf")
    # print("変換結果:", json.dumps(result, indent=2, ensure_ascii=False))

if __name__ == "__main__":
    asyncio.run(main())



