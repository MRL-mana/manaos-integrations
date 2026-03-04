#!/usr/bin/env python3
"""
タイヤ交換広告をExcelで作成
Excel自動化でプロ仕様の広告画像を生成
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class TireAdExcelCreator:
    """タイヤ交換広告Excel作成クラス"""
    
    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "タイヤ交換広告"
        
        # カスタムサイズ設定 (1200x800px相当)
        self.worksheet.row_dimensions[1].height = 600  # 上部60%
        self.worksheet.row_dimensions[2].height = 400  # 下部40%
        
        # 列幅調整
        for col in range(1, 21):  # A-T列
            self.worksheet.column_dimensions[get_column_letter(col)].width = 6
        
    def create_tire_ad(self):
        """タイヤ交換広告を作成"""
        
        # 上部60% - タイヤ画像エリア
        self.create_tire_image_area()
        
        # 下部40% - テキストエリア
        self.create_text_area()
        
        # ファイル保存
        filename = "/root/tire_exchange_ad.xlsx"
        self.workbook.save(filename)
        print(f"🎉 タイヤ交換広告を作成しました: {filename}")
        
        return filename
    
    def create_tire_image_area(self):
        """上部60% - タイヤ画像エリア"""
        
        # 上部エリアの背景色（雪の白）
        tire_area = "A1:T1"
        fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        self.worksheet.merge_cells(tire_area)
        self.worksheet['A1'].fill = fill
        
        # タイヤ画像の代わりにテキストで表現
        self.worksheet['A1'].value = "❄️ タイヤの履き替えは当店で ❄️"
        self.worksheet['A1'].font = Font(name='メイリオ', size=24, bold=True, color="2C3E50")
        self.worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 雪の装飾
        self.worksheet['D1'] = "❄"
        self.worksheet['D1'].font = Font(size=20)
        self.worksheet['Q1'] = "❄"
        self.worksheet['Q1'].font = Font(size=20)
        
    def create_text_area(self):
        """下部40% - テキストエリア"""
        
        # ダークブルー背景
        text_area = "A2:T2"
        fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        self.worksheet.merge_cells(text_area)
        self.worksheet['A2'].fill = fill
        
        # メインテキスト
        main_text = "タイヤ交換は\nかんたんネット予約"
        self.worksheet['K2'] = main_text
        self.worksheet['K2'].font = Font(name='メイリオ', size=32, bold=True, color="e11d48")
        self.worksheet['K2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # CTAボタン（セルで表現）
        button_area = "I3:L3"
        button_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        self.worksheet.merge_cells(button_area)
        
        # ボーダー
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        self.worksheet['I3'].border = thin_border
        self.worksheet['I3'].fill = button_fill
        
        # ボタンテキスト
        self.worksheet['K3'] = "予約はこちら"
        self.worksheet['K3'].font = Font(name='メイリオ', size=16, bold=True, color="000000")
        self.worksheet['K3'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 指差しアイコン
        self.worksheet['M3'] = "👉"
        self.worksheet['M3'].font = Font(size=16)
        self.worksheet['M3'].alignment = Alignment(horizontal='center', vertical='center')

def main():
    """メイン実行関数"""
    print("🚗 タイヤ交換広告をExcelで作成中...")
    
    try:
        creator = TireAdExcelCreator()
        filename = creator.create_tire_ad()
        
        print("✅ 作成完了!")
        print(f"📁 ファイル: {filename}")
        print("🎨 デザイン:")
        print("  - 上部60%: 雪道タイヤエリア（白背景）")
        print("  - 下部40%: ダークブルー背景")
        print("  - メインテキスト: ピンク/赤色")
        print("  - CTAボタン: 白背景")
        
        return filename
        
    except Exception as e:
        print(f"❌ エラー: {e}")
        return None

if __name__ == "__main__":
    main()
