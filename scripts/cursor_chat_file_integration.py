#!/usr/bin/env python3
"""
Cursor チャットファイル統合システム
チャットに直接Excel・PDFを貼り付け・表示・ダウンロード
"""

import os
from pathlib import Path
from datetime import datetime
import shutil
import warnings

# 警告を抑制
warnings.filterwarnings("ignore", category=ImportWarning)
warnings.filterwarnings("ignore", message=".*import.*")

# インポート警告を修正するための条件付きインポート
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from fastapi import FastAPI, File, UploadFile, HTTPException, Request
    from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
    import uvicorn
    FASTAPI_AVAILABLE = True
except ImportError:
    FASTAPI_AVAILABLE = False

class CursorChatFileIntegration:
    def __init__(self):
        desktop_dir = Path("/home/mana/Desktop")
        desktop_dir.mkdir(parents=True, exist_ok=True)
        self.output_dir = desktop_dir / "チャットファイル表示"
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        print("🚀 Cursor チャットファイル統合システム")
        print(f"📁 出力先: {self.output_dir}")
    
    def create_chat_embed_excel(self, excel_path):
        """チャット埋め込み用Excel表示"""
        try:
            if not OPENPYXL_AVAILABLE:
                print("❌ OpenPyXLがインストールされていません")
                return None, None
            
            print(f"📊 チャット埋め込みExcel表示: {excel_path}")
            
            # Excelファイルを読み込み
            wb = load_workbook(excel_path, data_only=True)
            
            # チャット埋め込み用HTML
            chat_html = f"""
<div style="border: 2px solid #4CAF50; border-radius: 10px; padding: 15px; margin: 10px 0; background: #f8f9fa;">
    <div style="background: #4CAF50; color: white; padding: 10px; margin: -15px -15px 15px -15px; border-radius: 8px 8px 0 0;">
        <h3 style="margin: 0; font-size: 16px;">📊 Excel ファイル: {Path(excel_path).name}</h3>
    </div>
    
    <div style="font-size: 12px; color: #666; margin-bottom: 10px;">
        📁 {excel_path}<br>
        💾 {Path(excel_path).stat().st_size:,} bytes | 
        📅 {datetime.fromtimestamp(Path(excel_path).stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')}
    </div>
"""
            
            # 各シートの内容を生成
            for i, sheet_name in enumerate(wb.sheetnames[:3]):  # 最初の3シートのみ表示
                ws = wb[sheet_name]
                
                # データを取得（最初の10行のみ）
                data = []
                for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_num > 10:  # 最初の10行のみ
                        break
                    if any(cell is not None for cell in row):
                        data.append([str(cell) if cell is not None else "" for cell in row])
                
                if data:
                    chat_html += f"""
    <div style="margin: 15px 0;">
        <h4 style="color: #333; margin-bottom: 10px;">📋 シート: {sheet_name}</h4>
        <div style="overflow-x: auto; max-height: 300px; border: 1px solid #ddd; border-radius: 5px;">
            <table style="width: 100%; border-collapse: collapse; font-size: 12px;">
"""
                    
                    # ヘッダー行
                    if data:
                        chat_html += '<thead><tr style="background: #e9ecef;">'
                        for j in range(min(len(data[0]), 8)):  # 最初の8列のみ表示
                            chat_html += f'<th style="padding: 6px; border: 1px solid #ddd; text-align: center;">列{j+1}</th>'
                        chat_html += '</tr></thead>'
                    
                    # データ行
                    chat_html += '<tbody>'
                    for row_idx, row in enumerate(data):
                        if row_idx % 2 == 0:
                            bg_color = "#ffffff"
                        else:
                            bg_color = "#f8f9fa"
                        
                        chat_html += f'<tr style="background: {bg_color};">'
                        for col_idx, cell in enumerate(row[:8]):  # 最初の8列のみ表示
                            chat_html += f'<td style="padding: 4px; border: 1px solid #ddd; max-width: 100px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;" title="{cell}">{cell}</td>'
                        chat_html += '</tr>'
                    chat_html += '</tbody>'
                    
                    chat_html += """
            </table>
        </div>
        <div style="font-size: 11px; color: #666; margin-top: 5px;">
            📊 表示: 最初の10行 × 8列 | 全データ: {len(data)}行 × {len(data[0]) if data else 0}列
        </div>
    </div>
"""
                
                if i < len(wb.sheetnames) - 1 and i < 2:  # 3シート目以降は省略
                    chat_html += '<hr style="margin: 10px 0; border: none; border-top: 1px solid #ddd;">'
            
            # シート数が多い場合の表示
            if len(wb.sheetnames) > 3:
                chat_html += f"""
    <div style="background: #fff3cd; padding: 10px; border-radius: 5px; margin-top: 10px;">
        <strong>📋 その他のシート:</strong> {', '.join(wb.sheetnames[3:])}
        <br><small>（全{len(wb.sheetnames)}シート）</small>
    </div>
"""
            
            # ダウンロードボタン
            chat_html += f"""
    <div style="text-align: center; margin-top: 15px; padding-top: 10px; border-top: 1px solid #ddd;">
        <button onclick="downloadFile('{excel_path}')" style="background: #007bff; color: white; border: none; padding: 8px 16px; border-radius: 5px; cursor: pointer; margin: 0 5px;">
            📥 ダウンロード
        </button>
        <button onclick="copyToDesktop('{excel_path}')" style="background: #28a745; color: white; border: none; padding: 8px 16px; border-radius: 5px; cursor: pointer; margin: 0 5px;">
            💾 デスクトップにコピー
        </button>
        <button onclick="openInWebUI('{excel_path}')" style="background: #ffc107; color: black; border: none; padding: 8px 16px; border-radius: 5px; cursor: pointer; margin: 0 5px;">
            🌐 WebUIで開く
        </button>
    </div>
</div>

<script>
function downloadFile(filePath) {{
    // ダウンロード処理
    const link = document.createElement('a');
    link.href = '/download/' + encodeURIComponent(filePath);
    link.download = '{Path(excel_path).name}';
    link.click();
}}

function copyToDesktop(filePath) {{
    fetch('/copy-to-desktop', {{
        method: 'POST',
        headers: {{'Content-Type': 'application/json'}},
        body: JSON.stringify({{filePath: filePath}})
    }})
    .then(response => response.json())
    .then(data => {{
        if (data.success) {{
            alert('✅ デスクトップにコピーしました！');
        }} else {{
            alert('❌ コピーに失敗しました: ' + data.error);
        }}
    }});
}}

function openInWebUI(filePath) {{
    window.open('/view/' + encodeURIComponent(filePath), '_blank');
}}
</script>
"""
            
            # HTMLファイルを保存
            html_filename = f"chat_excel_{Path(excel_path).stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
            html_path = self.output_dir / html_filename
            
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(chat_html)
            
            print(f"✅ チャット埋め込みExcel表示完了: {html_path}")
            return str(html_path), chat_html
            
        except Exception as e:
            print(f"❌ チャット埋め込みExcel表示エラー: {e}")
            return None, None
    
    def create_chat_embed_pdf(self, pdf_path):
        """チャット埋め込み用PDF表示"""
        try:
            print(f"📄 チャット埋め込みPDF表示: {pdf_path}")
            
            # チャット埋め込み用HTML
            chat_html = f"""
<div style="border: 2px solid #ff6b6b; border-radius: 10px; padding: 15px; margin: 10px 0; background: #f8f9fa;">
    <div style="background: #ff6b6b; color: white; padding: 10px; margin: -15px -15px 15px -15px; border-radius: 8px 8px 0 0;">
        <h3 style="margin: 0; font-size: 16px;">📄 PDF ファイル: {Path(pdf_path).name}</h3>
    </div>
    
    <div style="font-size: 12px; color: #666; margin-bottom: 10px;">
        📁 {pdf_path}<br>
        💾 {Path(pdf_path).stat().st_size:,} bytes | 
        📅 {datetime.fromtimestamp(Path(pdf_path).stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')}
    </div>
    
    <div style="text-align: center; margin: 20px 0;">
        <div style="border: 2px dashed #ff6b6b; border-radius: 10px; padding: 30px; background: white;">
            <div style="font-size: 48px; color: #ff6b6b; margin-bottom: 15px;">📄</div>
            <h4 style="margin: 0; color: #333;">PDF プレビュー</h4>
            <p style="margin: 10px 0; color: #666;">
                {Path(pdf_path).name}<br>
                <small>ファイルサイズ: {Path(pdf_path).stat().st_size:,} bytes</small>
            </p>
        </div>
    </div>
    
    <div style="text-align: center; margin-top: 15px; padding-top: 10px; border-top: 1px solid #ddd;">
        <button onclick="downloadFile('{pdf_path}')" style="background: #007bff; color: white; border: none; padding: 8px 16px; border-radius: 5px; cursor: pointer; margin: 0 5px;">
            📥 ダウンロード
        </button>
        <button onclick="copyToDesktop('{pdf_path}')" style="background: #28a745; color: white; border: none; padding: 8px 16px; border-radius: 5px; cursor: pointer; margin: 0 5px;">
            💾 デスクトップにコピー
        </button>
        <button onclick="openInWebUI('{pdf_path}')" style="background: #ffc107; color: black; border: none; padding: 8px 16px; border-radius: 5px; cursor: pointer; margin: 0 5px;">
            🌐 WebUIで開く
        </button>
        <button onclick="previewPDF('{pdf_path}')" style="background: #6f42c1; color: white; border: none; padding: 8px 16px; border-radius: 5px; cursor: pointer; margin: 0 5px;">
            👁️ プレビュー
        </button>
    </div>
</div>

<script>
function downloadFile(filePath) {{
    const link = document.createElement('a');
    link.href = '/download/' + encodeURIComponent(filePath);
    link.download = '{Path(pdf_path).name}';
    link.click();
}}

function copyToDesktop(filePath) {{
    fetch('/copy-to-desktop', {{
        method: 'POST',
        headers: {{'Content-Type': 'application/json'}},
        body: JSON.stringify({{filePath: filePath}})
    }})
    .then(response => response.json())
    .then(data => {{
        if (data.success) {{
            alert('✅ デスクトップにコピーしました！');
        }} else {{
            alert('❌ コピーに失敗しました: ' + data.error);
        }}
    }});
}}

function openInWebUI(filePath) {{
    window.open('/view/' + encodeURIComponent(filePath), '_blank');
}}

function previewPDF(filePath) {{
    // PDFプレビューウィンドウを開く
    const previewWindow = window.open('', '_blank', 'width=800,height=600');
    previewWindow.document.write(`
        <html>
        <head><title>PDF プレビュー</title></head>
        <body style="margin:0; padding:0;">
            <embed src="${{filePath}}" type="application/pdf" width="100%" height="100%" />
        </body>
        </html>
    `);
}}
</script>
"""
            
            # HTMLファイルを保存
            html_filename = f"chat_pdf_{Path(pdf_path).stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
            html_path = self.output_dir / html_filename
            
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(chat_html)
            
            print(f"✅ チャット埋め込みPDF表示完了: {html_path}")
            return str(html_path), chat_html
            
        except Exception as e:
            print(f"❌ チャット埋め込みPDF表示エラー: {e}")
            return None, None
    
    def create_chat_file_dropzone(self):
        """チャット用ファイルドロップゾーン"""
        dropzone_html = """
<div id="chatFileDropzone" style="border: 2px dashed #4CAF50; border-radius: 10px; padding: 20px; margin: 10px 0; background: #f8f9fa; text-align: center; cursor: pointer;" onclick="document.getElementById('chatFileInput').click()">
    <div style="font-size: 24px; color: #4CAF50; margin-bottom: 10px;">📁</div>
    <h4 style="margin: 0; color: #333;">ファイルをドラッグ&ドロップ または クリック</h4>
    <p style="margin: 10px 0; color: #666; font-size: 14px;">
        Excel (.xlsx) または PDF (.pdf) ファイルを選択してください
    </p>
    <input type="file" id="chatFileInput" accept=".xlsx,.pdf" style="display: none;" multiple>
</div>

<script>
document.getElementById('chatFileInput').addEventListener('change', function(e) {
    if (e.target.files.length > 0) {
        for (let file of e.target.files) {
            uploadToChat(file);
        }
    }
});

document.getElementById('chatFileDropzone').addEventListener('dragover', function(e) {
    e.preventDefault();
    this.style.background = '#e8f5e8';
    this.style.borderColor = '#45a049';
});

document.getElementById('chatFileDropzone').addEventListener('dragleave', function(e) {
    e.preventDefault();
    this.style.background = '#f8f9fa';
    this.style.borderColor = '#4CAF50';
});

document.getElementById('chatFileDropzone').addEventListener('drop', function(e) {
    e.preventDefault();
    this.style.background = '#f8f9fa';
    this.style.borderColor = '#4CAF50';
    
    for (let file of e.dataTransfer.files) {
        if (file.type.includes('sheet') || file.type.includes('pdf')) {
            uploadToChat(file);
        }
    }
});

async function uploadToChat(file) {
    const formData = new FormData();
    formData.append('file', file);
    
    try {
        const response = await fetch('/chat/upload', {
            method: 'POST',
            body: formData
        });
        
        if (response.ok) {
            const result = await response.json();
            if (result.html) {
                // チャットにHTMLを挿入
                insertChatMessage(result.html);
            }
        } else {
            alert('ファイルのアップロードに失敗しました。');
        }
    } catch (error) {
        alert('エラーが発生しました: ' + error.message);
    }
}

function insertChatMessage(html) {
    // チャットメッセージエリアに挿入
    const chatArea = document.getElementById('chat-messages') || document.querySelector('.chat-messages') || document.body;
    const messageDiv = document.createElement('div');
    messageDiv.innerHTML = html;
    messageDiv.style.margin = '10px 0';
    chatArea.appendChild(messageDiv);
    
    // スクロールを最下部に
    chatArea.scrollTop = chatArea.scrollHeight;
}
</script>
"""
        return dropzone_html
    
    def create_chat_file_integration_webui(self):
        """チャットファイル統合WebUI作成"""
        try:
            if not FASTAPI_AVAILABLE:
                print("❌ FastAPIがインストールされていません")
                return
            
            app = FastAPI(title="Cursor Chat File Integration")
            
            @app.get("/", response_class=HTMLResponse)
            async def chat_page():
                return f"""
<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Cursor Chat File Integration</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: #333;
        }}
        .container {{
            max-width: 1000px;
            margin: 0 auto;
            background: white;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #4CAF50 0%, #45a049 100%);
            color: white;
            padding: 20px;
            text-align: center;
        }}
        .chat-area {{
            height: 500px;
            overflow-y: auto;
            padding: 20px;
            border-bottom: 1px solid #ddd;
        }}
        .input-area {{
            padding: 20px;
            background: #f8f9fa;
        }}
        .file-dropzone {{
            border: 2px dashed #4CAF50;
            border-radius: 10px;
            padding: 20px;
            text-align: center;
            margin: 10px 0;
            background: #f8f9fa;
            cursor: pointer;
        }}
        .file-dropzone:hover {{
            background: #e8f5e8;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🚀 Cursor Chat File Integration</h1>
            <p>Excel・PDFファイルをチャットに直接貼り付け・表示</p>
        </div>
        
        <div class="chat-area" id="chatMessages">
            <div style="text-align: center; color: #666; margin-top: 200px;">
                <h3>💬 チャットエリア</h3>
                <p>ファイルをドロップゾーンにドラッグ&ドロップしてください</p>
            </div>
        </div>
        
        <div class="input-area">
            {self.create_chat_file_dropzone()}
        </div>
    </div>
</body>
</html>
"""
            
            @app.post("/chat/upload")
            async def chat_upload(file: UploadFile = File(...)):
                if not file.filename.endswith(('.xlsx', '.pdf')):
                    raise HTTPException(status_code=400, detail="ExcelまたはPDFファイルのみアップロード可能です")
                
                # ファイルを一時保存
                temp_path = self.output_dir / file.filename
                with open(temp_path, "wb") as buffer:
                    shutil.copyfileobj(file.file, buffer)
                
                # チャット埋め込みHTMLを生成
                if file.filename.endswith('.xlsx'):
                    _, chat_html = self.create_chat_embed_excel(str(temp_path))
                else:
                    _, chat_html = self.create_chat_embed_pdf(str(temp_path))
                
                return JSONResponse({
                    "success": True,
                    "filename": file.filename,
                    "html": chat_html
                })
            
            @app.post("/copy-to-desktop")
            async def copy_to_desktop(request: Request):
                data = await request.json()
                file_path = data.get('filePath')
                
                try:
                    source_path = Path(file_path)
                    if not source_path.exists():
                        return JSONResponse({"success": False, "error": "ファイルが見つかりません"})
                    
                    desktop_path = Path("/home/mana/Desktop") / source_path.name
                    shutil.copy2(source_path, desktop_path)
                    
                    return JSONResponse({"success": True, "message": "デスクトップにコピーしました"})
                except Exception as e:
                    return JSONResponse({"success": False, "error": str(e)})
            
            # WebUIを起動
            print("🚀 Chat File Integration WebUI起動中...")
            print("🌐 URL: http://localhost:8093")
            
            uvicorn.run(app, host="0.0.0.0", port=8093)
            
        except Exception as e:
            print(f"❌ WebUI作成エラー: {e}")
    
    def run_chat_demo(self):
        """チャット統合デモ実行"""
        print("\n🌟 Cursor チャットファイル統合デモ開始")
        print("=" * 60)
        
        # 既存のExcelファイルをチャット表示
        excel_files = [
            "/home/mana/Desktop/X280直接入力Excel/X280直接入力_20251005_121058.xlsx",
            "/home/mana/Desktop/PDF変換結果/実際の変換テスト_20251005_114553.xlsx"
        ]
        
        pdf_files = [
            "/home/mana/Desktop/PDF変換結果/サンプルPDF_テスト.pdf",
            "/home/mana/Desktop/10月3日PDF変換結果/SKM_C287i25100312345.pdf"
        ]
        
        print("\n📊 Excelファイルをチャット表示:")
        for excel_file in excel_files:
            if os.path.exists(excel_file):
                html_path, chat_html = self.create_chat_embed_excel(excel_file)
                if html_path and chat_html:
                    print(f"✅ チャット表示用HTML作成: {Path(html_path).name}")
                    # デスクトップにもコピー
                    self.copy_file_to_desktop(html_path)
        
        print("\n📄 PDFファイルをチャット表示:")
        for pdf_file in pdf_files:
            if os.path.exists(pdf_file):
                html_path, chat_html = self.create_chat_embed_pdf(pdf_file)
                if html_path and chat_html:
                    print(f"✅ チャット表示用HTML作成: {Path(html_path).name}")
                    # デスクトップにもコピー
                    self.copy_file_to_desktop(html_path)
        
        print("\n✅ チャット統合デモ完了！")
        print(f"📁 作成されたファイル: {self.output_dir}")
        
        # ファイル一覧表示
        files = list(self.output_dir.glob("*"))
        print(f"\n📋 作成されたチャット表示ファイル ({len(files)}件):")
        for file in files:
            print(f"  - {file.name}")
    
    def copy_file_to_desktop(self, file_path):
        """ファイルをデスクトップにコピー"""
        try:
            source_path = Path(file_path)
            desktop_path = Path("/home/mana/Desktop") / source_path.name
            shutil.copy2(source_path, desktop_path)
            print(f"✅ デスクトップにコピー: {desktop_path}")
            return True
        except Exception as e:
            print(f"❌ デスクトップコピーエラー: {e}")
            return False

def main():
    print("🌟 Cursor チャットファイル統合システム")
    print("=" * 60)
    
    integration = CursorChatFileIntegration()
    
    # デモ実行
    integration.run_chat_demo()
    
    print("\n🚀 チャット統合WebUIも起動しますか？")
    print("💡 WebUIを起動するには: integration.create_chat_file_integration_webui()")

if __name__ == "__main__":
    main()
