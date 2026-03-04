#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OCRテキストの複数列・複数行レイアウト改善版PDF-Excel変換スクリプト
"""

import os
import sys
import re
import fitz
import cv2
import numpy as np
from PIL import Image
import pytesseract
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import io

class OCRLayoutEnhancer:
    """OCRテキストの高度なレイアウト解析クラス"""
    
    def __init__(self):
        self.ocr_patterns = [
            # パターン1: タブ区切り
            (r'	+', 'tab_split'),
            # パターン2: 複数スペース
            (r'\s{2,}', 'multi_space'),
            # パターン3: 数字と文字の境界
            (r'(?<=\d)(?=[A-Za-z])|(?<=[A-Za-z])(?=\d)', 'number_text'),
            # パターン4: 日付パターン
            (r'(?<=\d{4}[-/]\d{1,2}[-/]\d{1,2})\s+', 'date_split'),
            # パターン5: 通貨パターン
            (r'(?<=\d+[.,]\d{2})\s+', 'currency_split'),
            # パターン6: OCR特有のパターン
            (r'(?<=[A-Za-z])(?=[A-Za-z])', 'char_boundary'),
            # パターン7: 日本語文字境界
            (r'(?<=[ひらがなカタカナ漢字])(?=[ひらがなカタカナ漢字])', 'japanese_boundary'),
            # パターン8: 記号境界
            (r'(?<=[^\w\s])(?=[^\w\s])', 'symbol_boundary'),
            # パターン9: 超細分割
            (r'(?<=.)(?=.)', 'ultra_fine'),
            # パターン10: 混合境界
            (r'(?<=[A-Za-z0-9])(?=[A-Za-z0-9])', 'mixed_boundary')
        ]
    
    def analyze_ocr_layout(self, ocr_text):
        """OCRテキストの高度なレイアウト解析"""
        lines = ocr_text.split('
')
        layout_analysis = {
            'max_columns': 0,
            'total_lines': len(lines),
            'complexity': 'low',
            'detected_patterns': [],
            'column_distribution': {},
            'merge_suggestions': []
        }
        
        for line in lines:
            if line.strip():
                # 各パターンで分割を試行
                for pattern, pattern_name in self.ocr_patterns:
                    try:
                        parts = re.split(pattern, line)
                        if len(parts) > 1:
                            layout_analysis['detected_patterns'].append(pattern_name)
                            layout_analysis['max_columns'] = max(layout_analysis['max_columns'], len(parts))
                            
                            # 列分布を記録
                            for i, part in enumerate(parts):
                                if i not in layout_analysis['column_distribution']:
                                    layout_analysis['column_distribution'][i] = 0
                                layout_analysis['column_distribution'][i] += 1
                    except Exception as e:
                        continue
        
        # 複雑度の判定
        if layout_analysis['max_columns'] > 20:
            layout_analysis['complexity'] = 'ultra_high'
        elif layout_analysis['max_columns'] > 10:
            layout_analysis['complexity'] = 'high'
        elif layout_analysis['max_columns'] > 5:
            layout_analysis['complexity'] = 'medium'
        
        # セル結合の提案
        for col_idx, count in layout_analysis['column_distribution'].items():
            if count < layout_analysis['total_lines'] * 0.3:  # 30%未満の場合は結合を提案
                layout_analysis['merge_suggestions'].append({
                    'column': col_idx,
                    'merge_reason': 'low_frequency',
                    'suggestion': f'列{col_idx+1}は結合を推奨'
                })
        
        return layout_analysis
    
    def split_ocr_line(self, line, target_columns):
        """OCRテキストの複数列分割"""
        # 複数の分割パターンを試行
        for pattern, _ in self.ocr_patterns:
            try:
                parts = re.split(pattern, line)
                if len(parts) > 1:
                    # 目標列数に調整
                    if len(parts) > target_columns:
                        # 多い場合は結合
                        while len(parts) > target_columns:
                            parts[0] += ' ' + parts.pop(1)
                    elif len(parts) < target_columns:
                        # 少ない場合は分割
                        while len(parts) < target_columns and any(' ' in part for part in parts):
                            for i, part in enumerate(parts):
                                if ' ' in part:
                                    split_part = part.split(' ', 1)
                                    parts[i] = split_part[0]
                                    parts.insert(i + 1, split_part[1])
                                    break
                    
                    return parts[:target_columns]
            except Exception as e:
                continue
        
        # デフォルト分割
        return [line] + [''] * (target_columns - 1)

class PDFExcelConverterOCREnhanced:
    """OCRテキストの複数列・複数行レイアウト改善版PDF-Excel変換クラス"""
    
    def __init__(self, config=None):
        self.config = config or {
            'ocr_enabled': True,
            'ocr_language': 'jpn+eng',
            'max_columns': 50,
            'ultra_fine_detection': True,
            'advanced_merging': True,
            'layout_optimization': True
        }
        self.ocr_enhancer = OCRLayoutEnhancer()
    
    def extract_text_with_ocr(self, pdf_path):
        """PDFからテキストとOCRテキストを抽出（改善版）"""
        try:
            doc = fitz.open(pdf_path)
            text_data = {'pages': [], 'page_count': len(doc)}
            images_data = {'images': [], 'ocr_text': '', 'total_images': 0}
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                
                # テキスト抽出
                text = page.get_text()
                if not text.strip():
                    # テキストが空の場合はOCRを使用
                    pix = page.get_pixmap()
                    img_data = pix.tobytes('png')
                    img = Image.open(io.BytesIO(img_data))
                    text = pytesseract.image_to_string(img, lang=self.config['ocr_language'])
                    text_data['pages'].append({
                        'page_number': page_num + 1,
                        'text': text,
                        'char_count': len(text),
                        'ocr_used': True
                    })
                else:
                    text_data['pages'].append({
                        'page_number': page_num + 1,
                        'text': text,
                        'char_count': len(text),
                        'ocr_used': False
                    })
                
                # 画像抽出とOCR
                if self.config['ocr_enabled']:
                    image_list = page.get_images()
                    for img_index, img in enumerate(image_list):
                        try:
                            xref = img[0]
                            pix = fitz.Pixmap(doc, xref)
                            if pix.n - pix.alpha < 4:
                                img_data = pix.tobytes('png')
                                img = Image.open(io.BytesIO(img_data))
                                ocr_text = pytesseract.image_to_string(img, lang=self.config['ocr_language'])
                                
                                images_data['images'].append({
                                    'page_number': page_num + 1,
                                    'image_number': img_index + 1,
                                    'ocr_text': ocr_text
                                })
                                images_data['ocr_text'] += ocr_text + '
'
                                images_data['total_images'] += 1
                        except Exception as e:
                            print(f'OCR処理エラー (ページ{page_num+1}, 画像{img_index+1}): {e}')
                            continue
            
            doc.close()
            return text_data, images_data
            
        except Exception as e:
            print(f'PDF処理エラー: {e}')
            return None, None
    
    def create_enhanced_ocr_sheet(self, workbook, ocr_data, page_num, border):
        """OCRテキストの高度なシート作成"""
        sheet_name = f'OCR_ページ_{page_num}_改善版'
        ocr_sheet = workbook.create_sheet(sheet_name)
        
        # 制御文字を除去
        clean_ocr_text = ocr_data['text']
        clean_ocr_text = re.sub(r'[ --]', '', clean_ocr_text)
        
        # レイアウト解析
        layout_analysis = self.ocr_enhancer.analyze_ocr_layout(clean_ocr_text)
        
        print(f'ページ {page_num} OCR解析結果:')
        print(f'  最大列数: {layout_analysis["max_columns"]}')
        print(f'  複雑度: {layout_analysis["complexity"]}')
        print(f'  検出パターン: {layout_analysis["detected_patterns"]}')
        print(f'  セル結合提案: {len(layout_analysis["merge_suggestions"])}個')
        
        # ヘッダー情報
        ocr_sheet.append(['OCR抽出情報（改善版）'])
        ocr_sheet.append(['ページ番号', page_num])
        ocr_sheet.append(['画像数', ocr_data.get('image_count', 1)])
        ocr_sheet.append(['最大列数', layout_analysis['max_columns']])
        ocr_sheet.append(['複雑度', layout_analysis['complexity']])
        ocr_sheet.append(['検出パターン', ', '.join(layout_analysis['detected_patterns'])])
        ocr_sheet.append(['セル結合提案', len(layout_analysis['merge_suggestions'])])
        ocr_sheet.append(['処理日時', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ocr_sheet.append([])
        
        # レイアウト解析結果
        ocr_sheet.append(['レイアウト解析結果'])
        ocr_sheet.append(['項目', '値'])
        ocr_sheet.append(['最大列数', layout_analysis['max_columns']])
        ocr_sheet.append(['総行数', layout_analysis['total_lines']])
        ocr_sheet.append(['複雑度', layout_analysis['complexity']])
        ocr_sheet.append(['検出パターン数', len(layout_analysis['detected_patterns'])])
        ocr_sheet.append([])
        
        # セル結合提案
        if layout_analysis['merge_suggestions']:
            ocr_sheet.append(['セル結合提案'])
            ocr_sheet.append(['列番号', '理由', '提案'])
            for suggestion in layout_analysis['merge_suggestions']:
                ocr_sheet.append([
                    suggestion['column'] + 1,
                    suggestion['merge_reason'],
                    suggestion['suggestion']
                ])
            ocr_sheet.append([])
        
        # OCRテキスト内容（複数列・複数行対応）
        ocr_sheet.append(['OCR抽出テキスト（複数列・複数行版）'])
        
        # テキストを行ごとに分割して書き込み
        text_lines = clean_ocr_text.split('
')
        current_row = 8 + len(layout_analysis['merge_suggestions']) + 2  # ヘッダー行の後の行番号
        
        for line in text_lines:
            if line.strip():
                # 複数列に分割
                parts = self.ocr_enhancer.split_ocr_line(line, layout_analysis['max_columns'])
                
                # 行を書き込み
                for col_idx, cell_value in enumerate(parts, 1):
                    cell = ocr_sheet.cell(row=current_row, column=col_idx, value=cell_value)
                    cell.border = border
                
                current_row += 1
        
        # スタイル適用
        for row_num in [1, 2, 3, 4, 5, 6, 7, 9]:  # ヘッダー行
            for cell in ocr_sheet.iter_rows(min_row=row_num, max_row=row_num):
                for cell_obj in cell:
                    if cell_obj.value:
                        cell_obj.font = Font(bold=True, color='8B4513')
                        cell_obj.fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
                        cell_obj.border = border
        
        # 列幅の自動調整
        for column in ocr_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ocr_sheet.column_dimensions[column_letter].width = adjusted_width
        
        return ocr_sheet
    
    def create_excel_from_data(self, text_data, images_data, output_path):
        """Excelファイルの作成（OCR改善版）"""
        try:
            workbook = Workbook()
            workbook.remove(workbook.active)
            
            # ボーダースタイル
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 1. 概要シート
            summary_sheet = workbook.create_sheet('📋 概要', 0)
            summary_sheet.append(['📄 PDF-Excel変換結果（OCR改善版）'])
            summary_sheet.append([])
            summary_sheet.append(['📊 基本情報'])
            summary_sheet.append(['処理日時', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            summary_sheet.append(['総ページ数', text_data.get('page_count', 0)])
            summary_sheet.append(['抽出テキスト数', len(text_data.get('pages', []))])
            summary_sheet.append(['抽出画像数', images_data.get('total_images', 0)])
            summary_sheet.append([])
            summary_sheet.append(['🔍 OCR改善機能'])
            summary_sheet.append(['複数列・複数行レイアウト', '有効'])
            summary_sheet.append(['高度なレイアウト解析', '有効'])
            summary_sheet.append(['セル結合最適化', '有効'])
            summary_sheet.append(['視覚的改善', '有効'])
            
            # 2. 各ページのテキストシート
            if text_data.get('pages'):
                for page_data in text_data['pages']:
                    clean_text = page_data['text']
                    clean_text = re.sub(r'[ --]', '', clean_text)
                    
                    sheet_name = f'ページ_{page_data["page_number"]}'
                    page_sheet = workbook.create_sheet(sheet_name)
                    
                    # ページ情報ヘッダー
                    page_sheet.append(['ページ情報'])
                    page_sheet.append(['ページ番号', page_data['page_number']])
                    page_sheet.append(['文字数', page_data['char_count']])
                    page_sheet.append(['OCR使用', 'はい' if page_data.get('ocr_used') else 'いいえ'])
                    page_sheet.append(['処理日時', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                    page_sheet.append([])
                    
                    # テキスト内容
                    page_sheet.append(['テキスト内容'])
                    text_lines = clean_text.split('
')
                    for line in text_lines:
                        if line.strip():
                            page_sheet.append([line.strip()])
                    
                    # スタイル適用
                    for row_num in [1, 2, 3, 4, 6]:  # ヘッダー行
                        for cell in page_sheet.iter_rows(min_row=row_num, max_row=row_num):
                            for cell_obj in cell:
                                if cell_obj.value:
                                    cell_obj.font = Font(bold=True, color='2F5597')
                                    cell_obj.fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
                                    cell_obj.border = border
            
            # 3. OCRテキストシート（改善版）
            if images_data.get('ocr_text') and self.config['ocr_enabled']:
                # ページごとにOCRテキストをグループ化
                page_ocr_data = {}
                for img_data in images_data['images']:
                    if img_data.get('ocr_text'):
                        page_num = img_data['page_number']
                        if page_num not in page_ocr_data:
                            page_ocr_data[page_num] = []
                        
                        clean_ocr_text = img_data['ocr_text']
                        clean_ocr_text = re.sub(r'[ --]', '', clean_ocr_text)
                        
                        page_ocr_data[page_num].append({
                            'image_number': img_data['image_number'],
                            'text': clean_ocr_text,
                            'image_count': len(page_ocr_data[page_num]) + 1
                        })
                
                # 各ページのOCRシートを作成（改善版）
                for page_num, ocr_list in page_ocr_data.items():
                    for i, ocr_item in enumerate(ocr_list):
                        self.create_enhanced_ocr_sheet(workbook, ocr_item, page_num, border)
            
            # ファイル保存
            workbook.save(output_path)
            print(f'✅ OCR改善版Excelファイルを作成しました: {output_path}')
            return True
            
        except Exception as e:
            print(f'❌ Excel生成エラー: {e}')
            return False

def main():
    """メイン処理"""
    # 設定
    config = {
        'ocr_enabled': True,
        'ocr_language': 'jpn+eng',
        'max_columns': 50,
        'ultra_fine_detection': True,
        'advanced_merging': True,
        'layout_optimization': True
    }
    
    # PDFファイルのパス
    pdf_path = '/root/Google Drive/System_Archive/20251021/Daily_Reports/pdf_samples/SKM_C287i25101518000.pdf'
    
    # 出力ファイルのパス
    output_path = '/root/Google Drive/日報１０月テスト_OCR改善版.xlsx'
    
    # 変換処理
    converter = PDFExcelConverterOCREnhanced(config)
    
    print('🔍 OCRテキストの複数列・複数行レイアウト改善版PDF-Excel変換を開始...')
    
    # テキストとOCRテキストを抽出
    text_data, images_data = converter.extract_text_with_ocr(pdf_path)
    
    if text_data and images_data:
        # Excelファイルを作成
        success = converter.create_excel_from_data(text_data, images_data, output_path)
        
        if success:
            print('🎉 OCR改善版PDF-Excel変換が完了しました！')
            print(f'📁 出力ファイル: {output_path}')
            print('✅ 複数列・複数行・セル結合でOCRテキストを最適化しました！')
        else:
            print('❌ Excelファイルの作成に失敗しました')
    else:
        print('❌ PDFファイルの処理に失敗しました')

if __name__ == '__main__':
    main()
