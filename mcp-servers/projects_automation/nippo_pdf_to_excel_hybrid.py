#!/usr/bin/env python3
"""
日報PDF→Excel変換システム (ハイブリッド版)
Google Drive OCR (無料) + Gemini (賢い整形)

特徴：
- Google Drive OCRで高精度テキスト抽出（無料）
- Geminiで表構造化（フリーティア・低コスト）
- 表形式のExcel出力
- 最高精度 × 最低コスト
"""

import os
import sys
from pathlib import Path
import logging
from datetime import datetime
import time
import json
import io

# 必要なライブラリ
try:
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
    import google.generativeai as genai
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError as e:
    print(f"❌ 必要なライブラリが不足: {e}")
    sys.exit(1)

# ログ設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/root/logs/nippo_pdf_to_excel_hybrid.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("NippoPDFToExcelHybrid")


class NippoPDFToExcelHybrid:
    """日報PDF→Excel変換 (Google Drive OCR + Gemini ハイブリッド版)"""
    
    def __init__(self):
        self.output_dir = Path('/root/daily_reports/excel_hybrid')
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Google Drive API初期化
        self.drive_service = None
        self.temp_folder_id = None
        self.initialize_drive()
        
        # Gemini API初期化
        self.gemini_api_key = os.getenv("GEMINI_API_KEY", "AIzaSyC5SA9-bXkB_nlAux9Fve47HvvDm1jx0Y0")
        genai.configure(api_key=self.gemini_api_key)
        self.gemini_model = genai.GenerativeModel('gemini-2.5-flash')
        logger.info("✅ Gemini API初期化完了")
    
    def initialize_drive(self):
        """Google Drive API初期化"""
        try:
            token_path = '/root/token.json'
            
            if not os.path.exists(token_path):
                logger.error("❌ /root/token.jsonが見つかりません")
                sys.exit(1)
            
            # トークン読み込み
            with open(token_path, 'r') as f:
                token_data = json.load(f)
            
            creds = Credentials(
                token=token_data.get('token'),
                refresh_token=token_data.get('refresh_token'),
                token_uri=token_data.get('token_uri', 'https://oauth2.googleapis.com/token'),
                client_id=token_data.get('client_id'),
                client_secret=token_data.get('client_secret'),
                scopes=token_data.get('scopes', ['https://www.googleapis.com/auth/drive'])
            )
            
            if creds.expired and creds.refresh_token:
                creds.refresh(Request())
            
            self.drive_service = build('drive', 'v3', credentials=creds)
            logger.info("✅ Google Drive API接続成功")
            
            # 一時フォルダ
            self.temp_folder_id = self.get_or_create_folder('日報OCR処理用_一時')
            
        except Exception as e:
            logger.error(f"❌ Google Drive初期化エラー: {e}")
            sys.exit(1)
    
    def get_or_create_folder(self, folder_name: str):
        """フォルダ取得/作成"""
        try:
            query = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
            results = self.drive_service.files().list(q=query, spaces='drive', fields='files(id)').execute()  # type: ignore[union-attr]
            files = results.get('files', [])
            
            if files:
                return files[0]['id']
            
            file_metadata = {
                'name': folder_name,
                'mimeType': 'application/vnd.google-apps.folder'
            }
            folder = self.drive_service.files().create(body=file_metadata, fields='id').execute()  # type: ignore[union-attr]
            return folder.get('id')
            
        except Exception as e:
            logger.error(f"フォルダ作成エラー: {e}")
            return None
    
    def upload_pdf_with_ocr(self, pdf_path: str):
        """PDFをアップロード＆OCR"""
        try:
            file_metadata = {
                'name': os.path.basename(pdf_path),
                'parents': [self.temp_folder_id],
                'mimeType': 'application/vnd.google-apps.document'
            }
            
            media = MediaFileUpload(pdf_path, mimetype='application/pdf', resumable=True)
            
            file = self.drive_service.files().create(  # type: ignore[union-attr]
                body=file_metadata,
                media_body=media,
                fields='id',
                ocrLanguage='ja'
            ).execute()
            
            time.sleep(2)  # OCR処理待ち
            return file.get('id')
            
        except Exception as e:
            logger.error(f"PDFアップロードエラー: {e}")
            return None
    
    def export_text(self, file_id: str):
        """Google Docsからテキストエクスポート"""
        try:
            request = self.drive_service.files().export_media(fileId=file_id, mimeType='text/plain')  # type: ignore[union-attr]
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            
            done = False
            while not done:
                status, done = downloader.next_chunk()
            
            return fh.getvalue().decode('utf-8')
            
        except Exception as e:
            logger.error(f"テキストエクスポートエラー: {e}")
            return None
    
    def delete_temp_file(self, file_id: str):
        """一時ファイル削除"""
        try:
            self.drive_service.files().delete(fileId=file_id).execute()  # type: ignore[union-attr]
        except IOError:
            pass
    
    def gemini_structure_text(self, text: str):
        """Geminiでテキストを表形式に構造化"""
        try:
            prompt = f"""
以下は日報または見積書のOCRテキストです。
このテキストから重要な情報を抽出し、表形式のJSON配列として出力してください。

【テキスト】
{text}

【出力形式】
JSON配列で、各行を1つのオブジェクトとして表現：
[
  {{"項目": "項目名1", "値": "値1"}},
  {{"項目": "項目名2", "値": "値2"}},
  ...
]

【抽出する項目例】
- 日付
- 見積番号
- 会社名/顧客名
- 住所
- 担当者
- 電話番号/FAX
- 金額
- その他重要な情報

【ルール】
1. JSON配列のみを返す（説明文不要）
2. コードブロック不要
3. 項目名は日本語
4. 値は元のテキストを正確に
5. 金額は数値として認識
6. 不要な文字（BOMなど）は除去

JSON配列のみを出力：
"""
            
            logger.info("🤖 Geminiで表構造化中...")
            response = self.gemini_model.generate_content(prompt)
            response_text = response.text.strip()
            
            # JSONコードブロック削除
            if '```' in response_text:
                response_text = response_text.split('```')[1]
                if response_text.startswith('json'):
                    response_text = response_text[4:]
                response_text = response_text.strip()
            
            # JSON解析
            data = json.loads(response_text)
            logger.info(f"✅ Gemini構造化成功: {len(data)}項目")
            
            return data
            
        except json.JSONDecodeError as e:
            logger.error(f"JSON解析エラー: {e}")
            logger.error(f"レスポンス: {response_text[:500]}")  # type: ignore[possibly-unbound]
            return None
        except Exception as e:
            logger.error(f"Gemini構造化エラー: {e}")
            return None
    
    def create_excel(self, df, output_path: str, pdf_name: str):
        """Excel作成"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '日報データ'  # type: ignore[union-attr]
            
            # ヘッダー
            for c_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=1, column=c_idx, value=col_name)  # type: ignore[union-attr]
                cell.font = Font(name='メイリオ', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            
            # データ
            for r_idx, row in enumerate(df.itertuples(index=False), start=2):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)  # type: ignore[union-attr]
                    cell.font = Font(name='メイリオ', size=11)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    if r_idx % 2 == 0:
                        cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # 列幅調整
            for col in ws.columns:  # type: ignore[union-attr]
                max_length = max((len(str(cell.value or '')) for cell in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)  # type: ignore[union-attr]
            
            # サマリーシート
            summary_ws = wb.create_sheet('変換情報', 0)
            summary_data = [
                ['項目', '値'],
                ['元PDFファイル', pdf_name],
                ['変換日時', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['抽出項目数', len(df)],
                ['システム', '日報PDF→Excel (Google OCR + Gemini ハイブリッド)'],
                ['OCR', 'Google Drive OCR (無料)'],
                ['AI整形', 'Google Gemini 1.5 Flash (フリーティア)']
            ]
            
            for r_idx, row_data in enumerate(summary_data, start=1):
                for c_idx, value in enumerate(row_data, start=1):
                    cell = summary_ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.font = Font(name='メイリオ', size=11)
                    if r_idx == 1:
                        cell.font = Font(name='メイリオ', size=11, bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
            
            summary_ws.column_dimensions['A'].width = 25
            summary_ws.column_dimensions['B'].width = 50
            
            wb.save(output_path)
            logger.info(f"✅ Excel作成完了: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Excel作成エラー: {e}")
            return False
    
    def convert_pdf(self, pdf_path: str):
        """PDFをExcelに変換"""
        pdf_path = Path(pdf_path)  # type: ignore
        
        if not pdf_path.exists():  # type: ignore
            return {'success': False, 'error': 'File not found'}
        
        print(f"\n{'='*60}")
        print(f"📄 変換: {pdf_path.name}")  # type: ignore
        print(f"{'='*60}")
        
        file_id = None
        
        try:
            # 1. Google Drive OCR
            print("☁️  STEP 1: Google Drive OCR (無料)...")
            file_id = self.upload_pdf_with_ocr(str(pdf_path))
            if not file_id:
                return {'success': False, 'error': 'OCR failed'}
            
            # 2. テキスト抽出
            print("📝 STEP 2: テキスト抽出...")
            text = self.export_text(file_id)
            if not text:
                return {'success': False, 'error': 'Text export failed'}
            print(f"   抽出文字数: {len(text)}文字")
            
            # 3. Geminiで表構造化
            print("🤖 STEP 3: Gemini AI整形 (フリーティア)...")
            structured_data = self.gemini_structure_text(text)
            
            if not structured_data:
                return {'success': False, 'error': 'Gemini structuring failed'}
            
            # 4. DataFrame変換
            print("📊 STEP 4: DataFrame変換...")
            df = pd.DataFrame(structured_data)
            print(f"   項目数: {len(df)}")
            
            # 5. Excel作成
            print("📝 STEP 5: Excel作成...")
            output_filename = f"{pdf_path.stem}_hybrid_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"  # type: ignore
            output_path = self.output_dir / output_filename
            
            if self.create_excel(df, str(output_path), pdf_path.name):  # type: ignore
                print("\n✅ 変換完了！")
                print(f"📁 出力: {output_path}")
                print(f"{'='*60}\n")
                
                return {
                    'success': True,
                    'output_file': str(output_path),
                    'items': len(df)
                }
            else:
                return {'success': False, 'error': 'Excel creation failed'}
            
        except Exception as e:
            logger.error(f"変換エラー: {e}")
            return {'success': False, 'error': str(e)}
        
        finally:
            if file_id:
                print("🗑️  一時ファイル削除...")
                self.delete_temp_file(file_id)
    
    def batch_convert(self, pdf_dir: str, limit: int = None):  # type: ignore
        """一括変換"""
        pdf_dir = Path(pdf_dir)  # type: ignore
        pdf_files = list(pdf_dir.glob('*.pdf')) + list(pdf_dir.glob('*.PDF'))  # type: ignore
        
        if limit:
            pdf_files = pdf_files[:limit]
        
        print(f"\n{'='*60}")
        print("📁 一括変換 (Google OCR + Gemini)")
        print(f"{'='*60}")
        print(f"ファイル数: {len(pdf_files)}")
        print(f"出力先: {self.output_dir}\n")
        
        results = {'total': len(pdf_files), 'success': 0, 'failed': 0}
        
        for i, pdf_file in enumerate(pdf_files, 1):
            print(f"\n[{i}/{len(pdf_files)}] {pdf_file.name}")
            
            result = self.convert_pdf(str(pdf_file))
            
            if result['success']:
                results['success'] += 1
            else:
                results['failed'] += 1
                print(f"   ❌ 失敗: {result.get('error')}")
            
            # API制限対策
            if i < len(pdf_files):
                time.sleep(1)
        
        print(f"\n{'='*60}")
        print("📊 一括変換完了")
        print(f"{'='*60}")
        print(f"総数: {results['total']}")
        print(f"✅ 成功: {results['success']}")
        print(f"❌ 失敗: {results['failed']}")
        print(f"📁 出力: {self.output_dir}")
        print(f"{'='*60}\n")


def main():
    if len(sys.argv) < 2:
        print("使い方:")
        print("  python3 nippo_pdf_to_excel_hybrid.py [PDF]")
        print("  python3 nippo_pdf_to_excel_hybrid.py [フォルダ] [件数]")
        print("\n例:")
        print("  python3 nippo_pdf_to_excel_hybrid.py test.pdf")
        print("  python3 nippo_pdf_to_excel_hybrid.py /path/to/pdfs 5")
        sys.exit(1)
    
    converter = NippoPDFToExcelHybrid()
    target = Path(sys.argv[1])
    
    if target.is_file():
        converter.convert_pdf(str(target))
    elif target.is_dir():
        limit = int(sys.argv[2]) if len(sys.argv) > 2 else None
        converter.batch_convert(str(target), limit=limit)  # type: ignore
    else:
        print(f"❌ パスが見つかりません: {target}")
        sys.exit(1)


if __name__ == '__main__':
    main()

