#!/usr/bin/env python3
"""
PDF→Excel変換 簡単版
Google Drive OCRを使った高精度で簡単な変換
使い方: python3 easy_pdf_to_excel.py [PDFファイルパスまたはフォルダID]
"""

import os
import sys
import json
import time
from pathlib import Path
from datetime import datetime

# Google Drive API
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
import io

# Excel処理
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

class EasyPDFToExcel:
    """簡単PDF→Excel変換（Google Drive OCR使用）"""

    def __init__(self):
        print("🚀 簡単PDF→Excel変換システム初期化中...")

        # Google Drive API初期化
        self.service = None
        self.temp_folder_id = None
        self.init_google_drive()

        # 出力ディレクトリ
        self.output_dir = Path("/root/excel_output_easy")
        self.output_dir.mkdir(exist_ok=True)

        print("✅ 初期化完了\n")

    def init_google_drive(self):
        """Google Drive API初期化"""
        try:
            token_path = "/mnt/storage500/credentials/token.pickle"
            creds_file = "/mnt/storage500/credentials/credentials.json"

            # token.json形式にも対応
            token_json_path = "/root/token.json"

            creds = None

            # token.pickleから読み込み
            if os.path.exists(token_path):
                import pickle
                with open(token_path, 'rb') as token:
                    creds = pickle.load(token)

            # token.jsonから読み込み（フォールバック）
            elif os.path.exists(token_json_path):
                with open(token_json_path, 'r') as f:
                    token_data = json.load(f)

                creds = Credentials(
                    token=token_data.get('token'),
                    refresh_token=token_data.get('refresh_token'),
                    token_uri=token_data.get('token_uri', 'https://oauth2.googleapis.com/token'),
                    client_id=token_data.get('client_id'),
                    client_secret=token_data.get('client_secret'),
                    scopes=token_data.get('scopes', [
                        'https://www.googleapis.com/auth/drive',
                        'https://www.googleapis.com/auth/drive.file'
                    ])
                )

            if not creds:
                print("❌ Google Drive認証情報が見つかりません")
                print("   /mnt/storage500/credentials/token.pickle または /root/token.json が必要です")
                sys.exit(1)

            # スコープチェック
            required_scopes = [
                'https://www.googleapis.com/auth/drive',
                'https://www.googleapis.com/auth/drive.file'
            ]

            current_scopes = creds.scopes if hasattr(creds, 'scopes') and creds.scopes else []

            # 書き込み権限がない場合は警告
            has_write_scope = any(scope in current_scopes for scope in required_scopes)

            if not has_write_scope:
                print("⚠️  現在の認証は読み取り専用です")
                print("📝 Google Drive OCRを使うには書き込み権限が必要です")
                print("\n🔧 解決方法:")
                print("   1. 新しい認証で書き込み権限を取得する")
                print("   2. または、既存の方法（pytesseract OCR）を使う")
                print("\n   既存の方法で続行しますか？ (y/n)")

                # デフォルトで既存方法を使用
                print("   → 既存の方法（pytesseract OCR）を使用します...")
                return None  # 後で既存方法を呼び出す

            # トークン更新
            if creds.expired and creds.refresh_token:
                creds.refresh(Request())

            self.service = build('drive', 'v3', credentials=creds)

            # 一時フォルダ作成
            self.temp_folder_id = self.get_or_create_folder('PDF→Excel変換一時')

            print("✅ Google Drive API接続成功")

        except Exception as e:
            print(f"❌ Google Drive初期化エラー: {e}")
            sys.exit(1)

    def get_or_create_folder(self, folder_name: str):
        """フォルダを取得または作成"""
        try:
            query = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
            results = self.service.files().list(q=query, fields='files(id)').execute()  # type: ignore[union-attr]
            files = results.get('files', [])

            if files:
                return files[0]['id']

            # フォルダ作成
            file_metadata = {'name': folder_name, 'mimeType': 'application/vnd.google-apps.folder'}
            folder = self.service.files().create(body=file_metadata, fields='id').execute()  # type: ignore[union-attr]
            return folder.get('id')

        except Exception as e:
            print(f"❌ フォルダ作成エラー: {e}")
            return None

    def convert_pdf_via_drive_ocr(self, pdf_path: str):
        """Google Drive OCRを使ってPDF→Excel変換"""
        pdf_path = Path(pdf_path)  # type: ignore

        if not pdf_path.exists():  # type: ignore
            print(f"❌ PDFファイルが見つかりません: {pdf_path}")
            return None

        print(f"\n📄 変換開始: {pdf_path.name}")  # type: ignore
        print("-" * 60)

        try:
            # 1. PDFをGoogle Driveにアップロード（OCR付き）
            print("📤 Step 1: PDFをGoogle Driveにアップロード中（OCR処理）...")

            file_metadata = {
                'name': pdf_path.name,  # type: ignore
                'parents': [self.temp_folder_id],
                'mimeType': 'application/vnd.google-apps.document'  # Google Docs形式に変換
            }

            media = MediaFileUpload(
                str(pdf_path),
                mimetype='application/pdf',
                resumable=True
            )

            file = self.service.files().create(  # type: ignore[union-attr]
                body=file_metadata,
                media_body=media,
                fields='id, name',
                ocrLanguage='ja'  # 日本語OCR
            ).execute()

            file_id = file.get('id')
            print("✅ アップロード完了（OCR処理中...）")

            # OCR処理待ち
            print("⏳ OCR処理を待機中...（5-10秒）")
            time.sleep(8)

            # 2. テキストとしてエクスポート
            print("📥 Step 2: OCR結果をテキストとして取得中...")

            request = self.service.files().export_media(  # type: ignore[union-attr]
                fileId=file_id,
                mimeType='text/plain'
            )

            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()

            ocr_text = fh.getvalue().decode('utf-8')
            print(f"✅ テキスト抽出完了（{len(ocr_text)}文字）")

            # 3. 一時ファイル削除
            print("🗑️  Step 3: 一時ファイルを削除中...")
            try:
                self.service.files().delete(fileId=file_id).execute()  # type: ignore[union-attr]
            except Exception:
                pass

            # 4. OCRテキストの精度改善（誤字修正）
            print("🔧 Step 4: OCR精度改善中...")
            from improve_ocr_accuracy import correct_ocr_text
            ocr_text = correct_ocr_text(ocr_text)

            # 5. テキストをExcel形式に変換
            print("📊 Step 5: Excelファイルを生成中...")
            excel_path = self.text_to_excel(ocr_text, pdf_path.stem)  # type: ignore

            if excel_path:
                print("✅ 変換完了！")
                print(f"📁 保存先: {excel_path}")
                return excel_path
            else:
                print("❌ Excel変換に失敗しました")
                return None

        except Exception as e:
            print(f"❌ エラー発生: {e}")
            import traceback
            traceback.print_exc()
            return None

    def text_to_excel(self, text: str, base_name: str):
        """OCRテキストをExcelに変換"""
        try:
            # ワークブック作成
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "OCR抽出結果"  # type: ignore[union-attr]

            # ヘッダー行のスタイル
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # テキストを行ごとに分割
            lines = [line.strip() for line in text.split('\n') if line.strip()]

            if not lines:
                ws.append(["OCR結果が空でした"])  # type: ignore[union-attr]
                ws['A1'].font = Font(italic=True, color="808080")  # type: ignore[index]
            else:
                # 各行を1セルに
                ws.append(["行番号", "内容"])  # type: ignore[union-attr]

                # ヘッダー行のスタイル適用
                for cell in ws[1]:  # type: ignore[index]
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # データ行
                for i, line in enumerate(lines, 1):
                    ws.append([i, line])  # type: ignore[union-attr]

                # 列幅の自動調整
                ws.column_dimensions['A'].width = 12  # type: ignore[union-attr]
                ws.column_dimensions['B'].width = 80  # type: ignore[union-attr]

            # 保存
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_filename = f"{base_name}_{timestamp}.xlsx"
            excel_path = self.output_dir / excel_filename

            wb.save(excel_path)

            return str(excel_path)

        except Exception as e:
            print(f"❌ Excel生成エラー: {e}")
            import traceback
            traceback.print_exc()
            return None

    def upload_to_google_drive(self, excel_path: str, folder_id: str = None, file_name: str = None):  # type: ignore
        """ExcelファイルをGoogle Driveにアップロード"""
        try:
            if not self.service:
                print("❌ Google Drive APIが初期化されていません")
                return None

            excel_path = Path(excel_path)  # type: ignore
            if not excel_path.exists():  # type: ignore
                print(f"❌ ファイルが見つかりません: {excel_path}")
                return None

            # ファイル名決定
            if not file_name:
                file_name = excel_path.name  # type: ignore

            print(f"\n📤 Google Driveにアップロード中: {file_name}")

            # メタデータ
            file_metadata = {
                'name': file_name
            }

            # フォルダ指定がある場合は追加
            if folder_id:
                file_metadata['parents'] = [folder_id]

            # ファイルアップロード
            from googleapiclient.http import MediaFileUpload

            media = MediaFileUpload(
                str(excel_path),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                resumable=True
            )

            file = self.service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id, name, webViewLink'
            ).execute()

            file_id = file.get('id')
            file_link = file.get('webViewLink')

            print("✅ アップロード完了！")
            print(f"   ファイルID: {file_id}")
            print(f"   リンク: {file_link}")

            return {
                'id': file_id,
                'name': file.get('name'),
                'link': file_link
            }

        except Exception as e:
            print(f"❌ アップロードエラー: {e}")
            import traceback
            traceback.print_exc()
            return None

    def convert_from_google_drive(self, folder_id: str, max_files: int = 1, upload_to_drive: bool = True):
        """Google DriveフォルダからPDFを取得して変換"""
        try:
            print(f"\n📁 Google Driveフォルダから取得: {folder_id}")

            # PDFファイル取得
            results = self.service.files().list(  # type: ignore[union-attr]
                q=f"'{folder_id}' in parents and mimeType='application/pdf'",
                fields="files(id,name)",
                pageSize=max_files,
                orderBy="name"
            ).execute()

            files = results.get('files', [])

            if not files:
                print("❌ PDFファイルが見つかりませんでした")
                return []

            print(f"✅ {len(files)}個のPDFファイルを発見")

            # Excel保存用フォルダID
            excel_folder_id = folder_id  # 同じフォルダに保存

            results = []
            for i, file_info in enumerate(files[:max_files], 1):
                print(f"\n[{i}/{len(files[:max_files])}] {file_info['name']}")

                # ダウンロード
                temp_path = Path("/tmp") / f"temp_{file_info['id']}.pdf"
                request = self.service.files().get_media(fileId=file_info['id'])  # type: ignore[union-attr]

                with open(temp_path, 'wb') as f:
                    downloader = MediaIoBaseDownload(f, request)
                    done = False
                    while not done:
                        status, done = downloader.next_chunk()

                # 変換
                excel_path = self.convert_pdf_via_drive_ocr(str(temp_path))

                if excel_path:
                    result = {
                        'pdf_name': file_info['name'],
                        'excel_path': excel_path
                    }

                    # Google Driveにアップロード
                    if upload_to_drive:
                        excel_file_name = file_info['name'].replace('.pdf', '.xlsx')
                        upload_result = self.upload_to_google_drive(
                            excel_path,
                            folder_id=excel_folder_id,
                            file_name=excel_file_name
                        )

                        if upload_result:
                            result['drive_file_id'] = upload_result['id']
                            result['drive_link'] = upload_result['link']

                    results.append(result)

                # 一時ファイル削除
                if temp_path.exists():
                    temp_path.unlink()

            return results

        except Exception as e:
            print(f"❌ Google Driveからの変換エラー: {e}")
            import traceback
            traceback.print_exc()
            return []

def fallback_to_local_method(pdf_path_or_folder_id, is_folder=False):
    """既存のローカルOCR方法を使用（フォールバック）"""
    print("\n" + "=" * 60)
    print("📄 ローカルOCR方法で変換します（pytesseract）")
    print("=" * 60)

    # test_single_pdf_conversion.py のロジックを使用
    import asyncio
    sys.path.append('/root/scripts')
    from pdf_excel_converter import PDFExcelConverter
    from pathlib import Path
    import pickle
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload

    class SimpleConverter:
        def __init__(self):
            self.pdf_converter = PDFExcelConverter()
            self.service = None
            self.init_drive()

        def init_drive(self):
            token_path = "/mnt/storage500/credentials/token.pickle"
            if os.path.exists(token_path):
                with open(token_path, 'rb') as token:
                    creds = pickle.load(token)
                if creds.expired and creds.refresh_token:
                    creds.refresh(Request())
                self.service = build('drive', 'v3', credentials=creds)

        def get_pdf_files(self, folder_id):
            results = self.service.files().list(  # type: ignore[union-attr]
                q=f"'{folder_id}' in parents and mimeType='application/pdf'",
                fields="files(id,name)", pageSize=1
            ).execute()
            return results.get('files', [])

        def download(self, file_id, output_path):
            request = self.service.files().get_media(fileId=file_id)  # type: ignore[union-attr]
            with open(output_path, 'wb') as f:
                downloader = MediaIoBaseDownload(f, request)
                done = False
                while not done:
                    status, done = downloader.next_chunk()

    async def convert():
        converter = SimpleConverter()

        if is_folder:
            files = converter.get_pdf_files(pdf_path_or_folder_id)
            if not files:
                print("❌ PDFファイルが見つかりません")
                return
            file_id = files[0]['id']
            file_name = files[0]['name']

            temp_path = Path("/tmp") / f"temp_{file_id}.pdf"
            converter.download(file_id, str(temp_path))
            pdf_path = str(temp_path)
        else:
            pdf_path = pdf_path_or_folder_id
            file_name = Path(pdf_path).name

        print(f"\n📄 変換中: {file_name}")
        result = await converter.pdf_converter.convert_pdf_to_excel(pdf_path)

        if result.get("success"):
            # 一時ディレクトリからExcelファイルを探す
            import glob
            temp_excel = glob.glob(str(converter.pdf_converter.temp_dir / "*_converted.xlsx"))
            if temp_excel:
                excel_path = temp_excel[0]
                output_dir = Path("/root/excel_output_easy")
                output_dir.mkdir(exist_ok=True)
                output_file = output_dir / f"{Path(file_name).stem}.xlsx"
                import shutil
                shutil.copy2(excel_path, output_file)
                print("\n✅ 変換完了！")
                print(f"📊 Excelファイル: {output_file}")
                return output_file

        print("❌ 変換に失敗しました")
        return None

    return asyncio.run(convert())

def main():
    print("=" * 60)
    print("📄 PDF→Excel変換 簡単版")
    print("=" * 60)

    converter = EasyPDFToExcel()

    # 認証が失敗した場合はフォールバック
    if converter.service is None:
        if len(sys.argv) >= 2:
            input_path = sys.argv[1]
            is_folder = '--folder' in sys.argv
            fallback_to_local_method(input_path, is_folder)
        return

    if len(sys.argv) < 2:
        print("\n使い方:")
        print("  python3 easy_pdf_to_excel.py [PDFファイルパス]")
        print("  python3 easy_pdf_to_excel.py [Google DriveフォルダID] --folder")
        print("\n例:")
        print("  python3 easy_pdf_to_excel.py /path/to/file.pdf")
        print("  python3 easy_pdf_to_excel.py 1bJlfAI0QeO4KxPrJ6i38dRI9oeyXPnj8 --folder")
        sys.exit(1)

    input_path = sys.argv[1]

    # Google DriveフォルダIDの場合
    if '--folder' in sys.argv or len(sys.argv[1]) > 20:
        folder_id = sys.argv[1]
        results = converter.convert_from_google_drive(folder_id, max_files=1)

        if results:
            print("\n" + "=" * 60)
            print("✅ 変換完了！")
            print("=" * 60)
            for result in results:
                print(f"\n📄 {result['pdf_name']}")
                print(f"📊 {result['excel_path']}")
        else:
            print("\n❌ 変換に失敗しました")

    # ローカルファイルの場合
    else:
        excel_path = converter.convert_pdf_via_drive_ocr(input_path)

        if excel_path:
            print("\n" + "=" * 60)
            print("✅ 変換完了！")
            print(f"📊 Excelファイル: {excel_path}")
            print("=" * 60)
        else:
            print("\n❌ 変換に失敗しました")

if __name__ == "__main__":
    main()

