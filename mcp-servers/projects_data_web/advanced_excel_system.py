#!/usr/bin/env python3
"""
Advanced Excel System for ManaOS
高度なExcel生成・分析・可視化システム
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path

class AdvancedExcelSystem:
    def __init__(self):
        self.output_dir = Path("/root/excel_advanced_output")
        self.output_dir.mkdir(exist_ok=True)
        
        self.templates_dir = Path("/root/excel_templates")
        self.templates_dir.mkdir(exist_ok=True)
        
        self.data_dir = Path("/root/excel_data")
        self.data_dir.mkdir(exist_ok=True)
        
        self.charts_dir = Path("/root/excel_charts")
        self.charts_dir.mkdir(exist_ok=True)
        
        print("🚀 Advanced Excel System for ManaOS")
        print(f"📁 出力ディレクトリ: {self.output_dir}")
        print(f"📁 テンプレートディレクトリ: {self.templates_dir}")
        print(f"📁 データディレクトリ: {self.data_dir}")
        print(f"📁 チャートディレクトリ: {self.charts_dir}")
    
    def create_advanced_excel_template(self, template_name, config):
        """高度なExcelテンプレート作成"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "メインデータ"  # type: ignore[union-attr]
            
            # テンプレート設定に基づいてシート作成
            for sheet_config in config.get("sheets", []):
                sheet_name = sheet_config["name"]
                sheet_type = sheet_config.get("type", "data")
                
                if sheet_type == "data":
                    self._create_data_sheet(wb, sheet_name, sheet_config)
                elif sheet_type == "dashboard":
                    self._create_dashboard_sheet(wb, sheet_name, sheet_config)
                elif sheet_type == "chart":
                    self._create_chart_sheet(wb, sheet_name, sheet_config)
                elif sheet_type == "summary":
                    self._create_summary_sheet(wb, sheet_name, sheet_config)
            
            # テンプレート保存
            template_file = self.templates_dir / f"{template_name}.xlsx"
            wb.save(template_file)
            
            return {
                "success": True,
                "template_file": str(template_file),
                "message": f"Advanced template '{template_name}' created successfully"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def _create_data_sheet(self, wb, sheet_name, config):
        """データシート作成"""
        ws = wb.create_sheet(title=sheet_name)
        
        # ヘッダー設定
        headers = config.get("headers", [])
        if headers:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(name='Meiryo UI', size=12, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # サンプルデータ挿入
        sample_data = config.get("sample_data", [])
        for row, data_row in enumerate(sample_data, 2):
            for col, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col, value=value)
        
        # テーブル形式設定
        if headers and sample_data:
            table = Table(displayName=f"Table_{sheet_name}", ref=f"A1:{get_column_letter(len(headers))}{len(sample_data) + 1}")
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            table.tableStyleInfo = style
            ws.add_table(table)
        
        # 列幅自動調整
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_dashboard_sheet(self, wb, sheet_name, config):
        """ダッシュボードシート作成"""
        ws = wb.create_sheet(title=sheet_name)
        
        # ダッシュボードタイトル
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"{sheet_name} - ダッシュボード"
        title_cell.font = Font(name='Meiryo UI', size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # KPI指標エリア
        kpis = config.get("kpis", [])
        for i, kpi in enumerate(kpis):
            row = 3 + (i // 2) * 3
            col = 1 + (i % 2) * 3
            
            # KPIタイトル
            ws.cell(row=row, column=col, value=kpi["title"]).font = Font(bold=True)
            ws.cell(row=row+1, column=col, value=kpi["value"]).font = Font(size=20, bold=True, color="2E8B57")
            ws.cell(row=row+2, column=col, value=kpi["change"]).font = Font(size=10, color="696969")
        
        # チャートエリア（プレースホルダー）
        ws.cell(row=10, column=1, value="チャートエリア").font = Font(bold=True)
        ws.merge_cells('A10:F15')
        chart_cell = ws['A10']
        chart_cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        chart_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _create_chart_sheet(self, wb, sheet_name, config):
        """チャートシート作成"""
        ws = wb.create_sheet(title=sheet_name)
        
        # チャートデータ準備
        chart_data = config.get("chart_data", {})
        if chart_data:
            # データ挿入
            data = chart_data.get("data", [])
            headers = chart_data.get("headers", [])
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            for row, data_row in enumerate(data, 2):
                for col, value in enumerate(data_row, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # チャート作成
            chart_type = chart_data.get("type", "bar")
            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            
            data_range = Reference(ws, min_col=2, min_row=1, max_row=len(data) + 1, max_col=len(headers))
            categories = Reference(ws, min_col=1, min_row=2, max_row=len(data) + 1)
            
            chart.add_data(data_range, titles_from_data=True)  # type: ignore[possibly-unbound]
            chart.set_categories(categories)  # type: ignore[possibly-unbound]
            chart.title = chart_data.get("title", f"{sheet_name} チャート")  # type: ignore[possibly-unbound]
            
            # チャート配置
            ws.add_chart(chart, "E2")  # type: ignore[possibly-unbound]
    
    def _create_summary_sheet(self, wb, sheet_name, config):
        """サマリーシート作成"""
        ws = wb.create_sheet(title=sheet_name)
        
        # サマリータイトル
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = f"{sheet_name} - サマリー"
        title_cell.font = Font(name='Meiryo UI', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # サマリーデータ
        summary_items = config.get("summary_items", [])
        for i, item in enumerate(summary_items, 3):
            ws.cell(row=i, column=1, value=item["label"]).font = Font(bold=True)
            ws.cell(row=i, column=2, value=item["value"])
            ws.cell(row=i, column=3, value=item.get("description", ""))
    
    def create_data_analysis_excel(self, data_source, analysis_config):
        """データ分析Excel作成"""
        try:
            # データ読み込み
            if data_source.endswith('.csv'):
                df = pd.read_csv(data_source)
            elif data_source.endswith('.json'):
                df = pd.read_json(data_source)
            else:
                return {"success": False, "error": "Unsupported data format"}
            
            # Excelファイル作成
            wb = Workbook()
            
            # 1. 生データシート
            ws_data = wb.active
            ws_data.title = "生データ"  # type: ignore[union-attr]
            
            # データをExcelに書き込み
            for col, column_name in enumerate(df.columns, 1):
                ws_data.cell(row=1, column=col, value=column_name).font = Font(bold=True)  # type: ignore[union-attr]
            
            for row, (_, data_row) in enumerate(df.iterrows(), 2):
                for col, value in enumerate(data_row, 1):
                    ws_data.cell(row=row, column=col, value=value)  # type: ignore[union-attr]
            
            # 2. 基本統計シート
            ws_stats = wb.create_sheet("基本統計")
            stats = df.describe()
            
            ws_stats.cell(row=1, column=1, value="統計項目").font = Font(bold=True)
            for col, column_name in enumerate(stats.columns, 2):
                ws_stats.cell(row=1, column=col, value=column_name).font = Font(bold=True)
            
            for row, (stat_name, stat_values) in enumerate(stats.iterrows(), 2):
                ws_stats.cell(row=row, column=1, value=stat_name)
                for col, value in enumerate(stat_values, 2):
                    ws_stats.cell(row=row, column=col, value=value)
            
            # 3. データ分析シート
            ws_analysis = wb.create_sheet("データ分析")
            
            # 欠損値分析
            missing_data = df.isnull().sum()
            ws_analysis.cell(row=1, column=1, value="欠損値分析").font = Font(size=14, bold=True)
            ws_analysis.cell(row=2, column=1, value="カラム名").font = Font(bold=True)
            ws_analysis.cell(row=2, column=2, value="欠損数").font = Font(bold=True)
            ws_analysis.cell(row=2, column=3, value="欠損率(%)").font = Font(bold=True)
            
            for i, (col_name, missing_count) in enumerate(missing_data.items(), 3):
                ws_analysis.cell(row=i, column=1, value=col_name)
                ws_analysis.cell(row=i, column=2, value=missing_count)
                ws_analysis.cell(row=i, column=3, value=f"{missing_count/len(df)*100:.2f}")
            
            # 4. 可視化チャート
            self._create_visualization_charts(df, wb)
            
            # ファイル保存
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"data_analysis_{timestamp}.xlsx"
            filepath = self.output_dir / filename
            wb.save(filepath)
            
            return {
                "success": True,
                "filepath": str(filepath),
                "filename": filename,
                "message": "Data analysis Excel created successfully"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def _create_visualization_charts(self, df, wb):
        """可視化チャート作成"""
        try:
            # 数値カラムのみ選択
            numeric_columns = df.select_dtypes(include=[np.number]).columns
            
            if len(numeric_columns) > 0:
                # チャートシート作成
                ws_chart = wb.create_sheet("チャート")
                
                # 相関行列ヒートマップ
                if len(numeric_columns) > 1:
                    plt.figure(figsize=(10, 8))
                    correlation_matrix = df[numeric_columns].corr()
                    sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', center=0)
                    plt.title('相関行列ヒートマップ')
                    
                    chart_path = self.charts_dir / f"correlation_heatmap_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
                    plt.savefig(chart_path, dpi=300, bbox_inches='tight')
                    plt.close()
                    
                    # Excelに画像挿入
                    img = Image(str(chart_path))
                    img.width = 400
                    img.height = 300
                    ws_chart.add_image(img, 'A1')
                
                # 分布ヒストグラム
                for i, column in enumerate(numeric_columns[:4]):  # 最大4つのカラム
                    plt.figure(figsize=(8, 6))
                    df[column].hist(bins=20)
                    plt.title(f'{column} の分布')
                    plt.xlabel(column)
                    plt.ylabel('頻度')
                    
                    chart_path = self.charts_dir / f"histogram_{column}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
                    plt.savefig(chart_path, dpi=300, bbox_inches='tight')
                    plt.close()
                    
                    # Excelに画像挿入
                    img = Image(str(chart_path))
                    img.width = 300
                    img.height = 200
                    row = 1 + (i // 2) * 15
                    col = 1 + (i % 2) * 20
                    ws_chart.add_image(img, f'{get_column_letter(col)}{row}')
        
        except Exception as e:
            print(f"チャート作成エラー: {e}")
    
    def create_automated_report_excel(self, report_config):
        """自動レポートExcel作成"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "自動レポート"  # type: ignore[union-attr]
            
            # レポートヘッダー
            ws.merge_cells('A1:F1')  # type: ignore[union-attr]
            header_cell = ws['A1']  # type: ignore[index]
            header_cell.value = f"{report_config.get('title', '自動レポート')} - {datetime.now().strftime('%Y年%m月%d日')}"
            header_cell.font = Font(name='Meiryo UI', size=16, bold=True, color="FFFFFF")
            header_cell.fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 35  # type: ignore[union-attr]
            
            # レポートセクション
            sections = report_config.get("sections", [])
            current_row = 3
            
            for section in sections:
                # セクションタイトル
                ws.cell(row=current_row, column=1, value=section["title"]).font = Font(size=14, bold=True, color="2E8B57")  # type: ignore[union-attr]
                current_row += 1
                
                # セクション内容
                if "table" in section:
                    table_data = section["table"]
                    headers = table_data.get("headers", [])
                    data = table_data.get("data", [])
                    
                    # ヘッダー
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col, value=header)  # type: ignore[union-attr]
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    current_row += 1
                    
                    # データ
                    for data_row in data:
                        for col, value in enumerate(data_row, 1):
                            ws.cell(row=current_row, column=col, value=value)  # type: ignore[union-attr]
                        current_row += 1
                    current_row += 1
                
                elif "text" in section:
                    ws.cell(row=current_row, column=1, value=section["text"])  # type: ignore[union-attr]
                    current_row += 2
                
                elif "metrics" in section:
                    metrics = section["metrics"]
                    for metric in metrics:
                        ws.cell(row=current_row, column=1, value=metric["label"]).font = Font(bold=True)  # type: ignore[union-attr]
                        ws.cell(row=current_row, column=2, value=metric["value"]).font = Font(size=12, color="2E8B57")  # type: ignore[union-attr]
                        current_row += 1
                    current_row += 1
            
            # ファイル保存
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"automated_report_{timestamp}.xlsx"
            filepath = self.output_dir / filename
            wb.save(filepath)
            
            return {
                "success": True,
                "filepath": str(filepath),
                "filename": filename,
                "message": "Automated report Excel created successfully"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def create_trinity_integrated_excel(self, trinity_config):
        """Trinity統合Excel作成"""
        try:
            wb = Workbook()
            
            # Trinity AI シート
            ws_trinity_ai = wb.active
            ws_trinity_ai.title = "Trinity AI"  # type: ignore[union-attr]
            
            ws_trinity_ai.merge_cells('A1:D1')  # type: ignore[union-attr]
            title_cell = ws_trinity_ai['A1']  # type: ignore[index]
            title_cell.value = "Trinity AI - データ分析レポート"
            title_cell.font = Font(size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Trinity Automation シート
            ws_trinity_auto = wb.create_sheet("Trinity Automation")
            
            ws_trinity_auto.merge_cells('A1:D1')
            title_cell = ws_trinity_auto['A1']
            title_cell.value = "Trinity Automation - ワークフロー管理"
            title_cell.font = Font(size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Trinity Monitor シート
            ws_trinity_monitor = wb.create_sheet("Trinity Monitor")
            
            ws_trinity_monitor.merge_cells('A1:D1')
            title_cell = ws_trinity_monitor['A1']
            title_cell.value = "Trinity Monitor - システム監視"
            title_cell.font = Font(size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Trinity MCP シート
            ws_trinity_mcp = wb.create_sheet("Trinity MCP")
            
            ws_trinity_mcp.merge_cells('A1:D1')
            title_cell = ws_trinity_mcp['A1']
            title_cell.value = "Trinity MCP - 統合管理"
            title_cell.font = Font(size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 統合ダッシュボードシート
            ws_dashboard = wb.create_sheet("統合ダッシュボード")
            
            ws_dashboard.merge_cells('A1:F1')
            title_cell = ws_dashboard['A1']
            title_cell.value = "ManaOS Trinity 統合ダッシュボード"
            title_cell.font = Font(size=18, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 各Trinityのステータス
            trinity_status = [
                ("Trinity AI", "アクティブ", "PDF処理・データ分析中"),
                ("Trinity Automation", "アクティブ", "ワークフロー管理中"),
                ("Trinity Monitor", "アクティブ", "システム監視中"),
                ("Trinity MCP", "アクティブ", "統合管理中")
            ]
            
            for i, (name, status, description) in enumerate(trinity_status, 3):
                ws_dashboard.cell(row=i, column=1, value=name).font = Font(bold=True)
                ws_dashboard.cell(row=i, column=2, value=status).font = Font(color="008000")
                ws_dashboard.cell(row=i, column=3, value=description)
            
            # ファイル保存
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"trinity_integrated_excel_{timestamp}.xlsx"
            filepath = self.output_dir / filename
            wb.save(filepath)
            
            return {
                "success": True,
                "filepath": str(filepath),
                "filename": filename,
                "message": "Trinity integrated Excel created successfully"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def create_realtime_data_excel(self, data_source, update_interval=60):
        """リアルタイムデータ更新Excel作成"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "リアルタイムデータ"  # type: ignore[union-attr]
            
            # ヘッダー設定
            headers = ["タイムスタンプ", "データソース", "値", "ステータス", "更新時刻"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)  # type: ignore[union-attr]
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
                cell.font = Font(color="FFFFFF")
            
            # 初期データ挿入
            current_time = datetime.now()
            sample_data = [
                [current_time.strftime('%Y-%m-%d %H:%M:%S'), "システム1", "100", "正常", current_time.strftime('%H:%M:%S')],
                [current_time.strftime('%Y-%m-%d %H:%M:%S'), "システム2", "250", "正常", current_time.strftime('%H:%M:%S')],
                [current_time.strftime('%Y-%m-%d %H:%M:%S'), "システム3", "75", "注意", current_time.strftime('%H:%M:%S')]
            ]
            
            for row, data_row in enumerate(sample_data, 2):
                for col, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row, column=col, value=value)  # type: ignore[union-attr]
                    if col == 4:  # ステータス列
                        if value == "正常":
                            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        elif value == "注意":
                            cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # 条件付き書式設定
            from openpyxl.formatting.rule import CellIsRule
            red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
            green_fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
            
            # 値による条件付き書式
            ws.conditional_formatting.add('C2:C100', CellIsRule(operator='greaterThan', formula=[200], fill=red_fill))  # type: ignore[union-attr]
            ws.conditional_formatting.add('C2:C100', CellIsRule(operator='between', formula=[100, 200], fill=yellow_fill))  # type: ignore[union-attr]
            ws.conditional_formatting.add('C2:C100', CellIsRule(operator='lessThan', formula=[100], fill=green_fill))  # type: ignore[union-attr]
            
            # ファイル保存
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"realtime_data_excel_{timestamp}.xlsx"
            filepath = self.output_dir / filename
            wb.save(filepath)
            
            return {
                "success": True,
                "filepath": str(filepath),
                "filename": filename,
                "message": "Real-time data Excel created successfully"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def get_system_status(self):
        """システム状態取得"""
        try:
            status = {
                "timestamp": datetime.now().isoformat(),
                "output_directory": str(self.output_dir),
                "templates_directory": str(self.templates_dir),
                "data_directory": str(self.data_dir),
                "charts_directory": str(self.charts_dir),
                "available_templates": len(list(self.templates_dir.glob("*.xlsx"))),
                "generated_files": len(list(self.output_dir.glob("*.xlsx"))),
                "system_health": "healthy"
            }
            
            return {
                "success": True,
                "status": status
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }

def main():
    print("🌟 Advanced Excel System for ManaOS")
    print("=" * 60)
    
    excel_system = AdvancedExcelSystem()
    
    # システム状態確認
    status = excel_system.get_system_status()
    print(f"システム状態: {status}")
    
    # 高度なテンプレート作成テスト
    template_config = {
        "sheets": [
            {
                "name": "データ入力",
                "type": "data",
                "headers": ["ID", "名前", "部署", "役職", "入社日"],
                "sample_data": [
                    [1, "山田太郎", "営業部", "部長", "2020-04-01"],
                    [2, "鈴木花子", "開発部", "エンジニア", "2021-07-15"],
                    [3, "田中健太", "人事部", "主任", "2022-01-10"]
                ]
            },
            {
                "name": "ダッシュボード",
                "type": "dashboard",
                "kpis": [
                    {"title": "総従業員数", "value": "150", "change": "+5%"},
                    {"title": "月間売上", "value": "¥2,500,000", "change": "+12%"}
                ]
            }
        ]
    }
    
    result = excel_system.create_advanced_excel_template("employee_template", template_config)
    print(f"テンプレート作成結果: {result}")
    
    # Trinity統合Excel作成テスト
    trinity_result = excel_system.create_trinity_integrated_excel({})
    print(f"Trinity統合Excel作成結果: {trinity_result}")
    
    print("\n✅ Advanced Excel System テスト完了！")

if __name__ == "__main__":
    main()
