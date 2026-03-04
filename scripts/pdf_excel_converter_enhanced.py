#!/usr/bin/env python3
"""
Enhanced PDF to Excel Converter with Dynamic Column Detection
PDFtk + 動的列数検出 + レイアウト再構築の組み合わせ
"""

import os
import sys
import subprocess
import re
from datetime import datetime

# PDF処理ライブラリ
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class EnhancedPDFExcelConverter:
    def __init__(self, config=None):
        self.config = config or {
            "ocr_enabled": True,
            "table_detection": True,
            "dynamic_columns": True,
            "layout_reconstruction": True,
            "max_columns": 15,
            "min_columns": 3
        }
        
    def analyze_pdf_structure_with_pdftk(self, pdf_path):
        """PDFtkを使ってPDFの構造を詳細に解析"""
        try:
            # PDFtkでPDFの基本情報を取得
            result = subprocess.run(
                ['pdftk', pdf_path, 'dump_data'],
                capture_output=True,
                text=True,
                check=True
            )
            
            # ページ情報を解析
            pages_info = []
            lines = result.stdout.split('\n')
            current_page = None
            
            for line in lines:
                if line.startswith('PageMediaBegin'):
                    if current_page:
                        pages_info.append(current_page)
                    current_page = {}
                elif line.startswith('PageMediaNumber:'):
                    current_page['page_number'] = int(line.split(':')[1].strip())
                elif line.startswith('PageMediaDimensions:'):
                    dims = line.split(':')[1].strip().split()
                    current_page['width'] = float(dims[0])
                    current_page['height'] = float(dims[1])
                elif line.startswith('PageMediaRotation:'):
                    current_page['rotation'] = int(line.split(':')[1].strip())
            
            if current_page:
                pages_info.append(current_page)
            
            return {
                'total_pages': len(pages_info),
                'pages': pages_info,
                'pdftk_info': result.stdout
            }
            
        except subprocess.CalledProcessError as e:
            print(f"PDFtk解析エラー: {e}")
            return None
        except Exception as e:
            print(f"PDFtk解析エラー: {e}")
            return None
    
    def analyze_text_layout_enhanced(self, text, page_info=None):
        """改良版テキストレイアウト解析（OCR対応の高度な列数検出）"""
        lines = text.split('\n')
        max_columns = 0
        column_patterns = []
        
        for line in lines:
            if line.strip():
                # 複数の分割パターンを試す
                patterns = []
                
                # パターン1: タブで分割
                if '\t' in line:
                    parts = [p.strip() for p in line.split('\t') if p.strip()]
                    patterns.append(('tab', parts))
                
                # パターン2: 複数スペースで分割
                parts = [p.strip() for p in re.split(r'\s{2,}', line) if p.strip()]
                if len(parts) > 1:
                    patterns.append(('spaces', parts))
                
                # パターン3: 特定の文字で分割（数字、記号など）
                parts = [p.strip() for p in re.split(r'[|;:]+', line) if p.strip()]
                if len(parts) > 1:
                    patterns.append(('delimiter', parts))
                
                # パターン4: 数字と文字の境界で分割
                parts = [p.strip() for p in re.split(r'(?<=\d)\s+(?=[A-Za-z])|(?<=[A-Za-z])\s+(?=\d)', line) if p.strip()]
                if len(parts) > 1:
                    patterns.append(('number_text', parts))
                
                # パターン5: 日付形式で分割（簡易版）
                parts = [p.strip() for p in re.split(r'\s+(?=\d{4}[-/]\d{1,2}[-/]\d{1,2})|\s+(?=\d{1,2}[-/]\d{1,2}[-/]\d{4})', line) if p.strip()]
                if len(parts) > 1:
                    patterns.append(('date', parts))
                
                # パターン6: 金額形式で分割（簡易版）
                parts = [p.strip() for p in re.split(r'\s+(?=[¥$€£])|\s+(?=\d)', line) if p.strip()]
                if len(parts) > 1:
                    patterns.append(('currency', parts))
                
                # パターン7: より細かいスペース分割
                parts = [p.strip() for p in re.split(r'\s+', line) if p.strip()]
                if len(parts) > 1:
                    patterns.append(('single_space', parts))
                
                # パターン8: OCR特有のパターン（短い文字列の連続）
                parts = [p.strip() for p in re.split(r'(?<=[A-Za-z])\s+(?=[A-Za-z])', line) if p.strip()]
                if len(parts) > 1:
                    patterns.append(('ocr_short', parts))
                
                # パターン9: 数字の連続パターン
                parts = [p.strip() for p in re.split(r'(?<=\d)\s+(?=\d)', line) if p.strip()]
                if len(parts) > 1:
                    patterns.append(('ocr_number', parts))
                
                # パターン10: 混合パターン（文字と数字の混合）
                parts = [p.strip() for p in re.split(r'(?<=[A-Za-z])\s+(?=\d)|(?<=\d)\s+(?=[A-Za-z])', line) if p.strip()]
                if len(parts) > 1:
                    patterns.append(('ocr_mixed', parts))
                
                # パターン11: 日本語文字の境界分割
                parts = [p.strip() for p in re.split(r'(?<=[ひらがなカタカナ漢字])\s+(?=[ひらがなカタカナ漢字])', line) if p.strip()]
                if len(parts) > 1:
                    patterns.append(('japanese', parts))
                
                # パターン12: より細かい分割（1文字ずつ）
                parts = [p.strip() for p in re.split(r'(?<=[A-Za-z0-9])\s+(?=[A-Za-z0-9])', line) if p.strip()]
                if len(parts) > 1:
                    patterns.append(('fine_split', parts))
                
                # パターン13: 極細分割（すべてのスペースで分割）
                parts = [p.strip() for p in line.split() if p.strip()]
                if len(parts) > 1:
                    patterns.append(('ultra_fine', parts))
                
                # パターン14: 文字単位分割
                parts = [p.strip() for p in re.split(r'(?<=.)\s+(?=.)', line) if p.strip()]
                if len(parts) > 1:
                    patterns.append(('char_level', parts))
                
                # パターン15: 超細分割（すべての文字で分割）
                parts = [p.strip() for p in re.split(r'(?<=.)(?=.)', line) if p.strip()]
                if len(parts) > 1:
                    patterns.append(('ultra_char', parts))
                
                # パターン16: 数字と文字の超細分割
                parts = [p.strip() for p in re.split(r'(?<=\d)(?=[A-Za-z])|(?<=[A-Za-z])(?=\d)', line) if p.strip()]
                if len(parts) > 1:
                    patterns.append(('number_char', parts))
                
                # パターン17: 記号分割
                parts = [p.strip() for p in re.split(r'(?<=[^\w\s])(?=[^\w\s])', line) if p.strip()]
                if len(parts) > 1:
                    patterns.append(('symbol_split', parts))
                
                # パターン18: 混合超細分割
                parts = [p.strip() for p in re.split(r'(?<=[A-Za-z0-9])(?=[A-Za-z0-9])', line) if p.strip()]
                if len(parts) > 1:
                    patterns.append(('mixed_ultra', parts))
                
                # 最適なパターンを選択
                if patterns:
                    best_pattern = max(patterns, key=lambda x: len(x[1]))
                    column_patterns.append(best_pattern)
                    if len(best_pattern[1]) > max_columns:
                        max_columns = len(best_pattern[1])
        
        # 列数の統計を取る
        column_counts = {}
        for pattern_type, parts in column_patterns:
            count = len(parts)
            column_counts[count] = column_counts.get(count, 0) + 1
        
        # 最も頻繁に現れる列数を選択
        if column_counts:
            optimal_columns = max(column_counts, key=column_counts.get)
        else:
            optimal_columns = self.config["min_columns"]
        
        # 設定された範囲内に制限（上限を上げる）
        optimal_columns = max(self.config["min_columns"], 
                            min(optimal_columns, self.config["max_columns"]))
        
        return {
            'detected_columns': optimal_columns,
            'column_patterns': column_patterns,
            'column_counts': column_counts,
            'max_columns_found': max_columns
        }
    
    def extract_text_with_layout(self, pdf_path):
        """レイアウト情報を含むテキスト抽出（OCR対応 + 表構造解析）"""
        text_data = {"pages": [], "page_count": 0}
        
        try:
            # PyMuPDFでレイアウト情報を含むテキスト抽出
            doc = fitz.open(pdf_path)
            text_data["page_count"] = len(doc)
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                
                # テキストブロックを取得（位置情報付き）
                blocks = page.get_text("dict")
                
                # テキストを抽出
                text = page.get_text()
                
                # テキストが空の場合はOCRを使用
                if not text.strip():
                    print(f"ページ {page_num + 1}: テキストが空のためOCRを使用")
                    text = self.extract_text_with_ocr(page)
                
                # レイアウト情報を解析
                layout_info = self.analyze_page_layout(blocks)
                
                # 表構造を詳細解析
                table_structure = self.analyze_table_structure(page, page_num + 1)
                
                page_data = {
                    "page_number": page_num + 1,
                    "text": text,
                    "char_count": len(text),
                    "layout_info": layout_info,
                    "table_structure": table_structure,
                    "ocr_used": not page.get_text().strip()
                }
                
                text_data["pages"].append(page_data)
            
            doc.close()
            return text_data
            
        except Exception as e:
            print(f"テキスト抽出エラー: {e}")
            return text_data
    
    def analyze_table_structure(self, page, page_num):
        """表構造の詳細解析（修正版）"""
        try:
            import pdfplumber
            
            # PDFファイルパスを取得
            pdf_path = page.parent.name
            
            # pdfplumberで表を抽出
            with pdfplumber.open(pdf_path) as pdf:
                if page_num <= len(pdf.pages):
                    pdf_page = pdf.pages[page_num - 1]
                    
                    # 表を抽出
                    tables = pdf_page.extract_tables()
                    
                    # 表構造を解析
                    table_structures = []
                    for i, table in enumerate(tables):
                        if table:
                            structure = {
                                'table_number': i + 1,
                                'rows': len(table),
                                'columns': len(table[0]) if table else 0,
                                'data': table,
                                'cell_positions': self.analyze_cell_positions(table)
                            }
                            table_structures.append(structure)
                    
                    return {
                        'total_tables': len(tables),
                        'tables': table_structures
                    }
        except Exception as e:
            print(f"表構造解析エラー: {e}")
            return {'total_tables': 0, 'tables': []}
    
    def analyze_cell_positions(self, table):
        """セルの位置情報を解析"""
        cell_positions = []
        
        for row_idx, row in enumerate(table):
            for col_idx, cell in enumerate(row):
                if cell and cell.strip():
                    cell_positions.append({
                        'row': row_idx,
                        'column': col_idx,
                        'content': cell.strip(),
                        'length': len(cell.strip())
                    })
        
        return cell_positions
    
    def extract_text_with_ocr(self, page):
        """OCRでテキストを抽出"""
        try:
            import pytesseract
            from PIL import Image
            import io
            
            # ページを画像に変換
            pix = page.get_pixmap()
            img_data = pix.tobytes('png')
            img = Image.open(io.BytesIO(img_data))
            
            # OCRでテキストを抽出
            text = pytesseract.image_to_string(img, lang='jpn+eng')
            return text
            
        except Exception as e:
            print(f"OCR抽出エラー: {e}")
            return ""
    
    def analyze_page_layout(self, blocks):
        """ページのレイアウト情報を解析"""
        layout_info = {
            'blocks': len(blocks.get('blocks', [])),
            'text_blocks': 0,
            'image_blocks': 0,
            'table_blocks': 0,
            'columns_detected': 0
        }
        
        for block in blocks.get('blocks', []):
            if 'lines' in block:
                layout_info['text_blocks'] += 1
            elif 'image' in block:
                layout_info['image_blocks'] += 1
            elif 'table' in block:
                layout_info['table_blocks'] += 1
        
        return layout_info
    
    def create_excel_with_dynamic_layout(self, text_data, tables_data, images_data, output_path):
        """動的レイアウト対応のExcel作成"""
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        # ボーダースタイル
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. 概要シート
        summary_sheet = workbook.create_sheet("📋 概要", 0)
        self.create_summary_sheet(summary_sheet, text_data, tables_data, images_data, border)
        
        # 2. 各ページのテキストシート（動的列数対応）
        if text_data.get("pages"):
            for page_data in text_data["pages"]:
                self.create_page_sheet(workbook, page_data, border)
        
        # 3. OCRテキストシート
        if images_data.get("ocr_text") and self.config["ocr_enabled"]:
            self.create_ocr_sheets(workbook, images_data, border)
        
        # 4. 表データシート
        if tables_data.get("tables"):
            self.create_table_sheets(workbook, tables_data, border)
        
        # Excelファイルを保存
        workbook.save(output_path)
        print(f"Excelファイルを作成しました: {output_path}")
    
    def create_summary_sheet(self, sheet, text_data, tables_data, images_data, border):
        """概要シートの作成"""
        # タイトル
        sheet.append(["📄 PDF-Excel変換結果（改良版）"])
        sheet.append([])
        
        # 基本情報
        sheet.append(["📊 基本情報"])
        summary_data = [
            ["処理日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["総ページ数", text_data.get("page_count", 0)],
            ["抽出テキスト数", len(text_data.get("pages", []))],
            ["抽出表数", tables_data.get("total_tables", 0)],
            ["抽出画像数", images_data.get("total_images", 0)],
            ["動的列数検出", "有効" if self.config["dynamic_columns"] else "無効"],
            ["レイアウト再構築", "有効" if self.config["layout_reconstruction"] else "無効"]
        ]
        
        for row in summary_data:
            sheet.append(row)
        
        sheet.append([])
        
        # シート構成
        sheet.append(["📑 シート構成"])
        sheet_info = [
            ["シート名", "内容"],
            ["概要", "このシート - 変換結果の概要"],
            ["ページ_1〜N", "各ページのテキスト内容（動的列数対応）"],
            ["OCR_ページ_X", "各ページのOCR抽出テキスト"],
            ["表_X_Y", "抽出された表データ"]
        ]
        
        for row in sheet_info:
            sheet.append(row)
        
        # スタイル適用
        self.apply_summary_styles(sheet, border)
    
    def create_page_sheet(self, workbook, page_data, border):
        """ページシートの作成（超高度なレイアウト再構築）"""
        sheet_name = f"ページ_{page_data['page_number']}"
        page_sheet = workbook.create_sheet(sheet_name)
        
        # 制御文字を除去
        clean_text = page_data["text"]
        clean_text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', clean_text)
        
        # レイアウト解析
        layout_analysis = self.analyze_text_layout_enhanced(clean_text)
        detected_columns = layout_analysis['detected_columns']
        
        # 表構造情報を取得
        table_structure = page_data.get('table_structure', {})
        total_tables = table_structure.get('total_tables', 0)
        
        # 高度なレイアウト解析
        advanced_layout = self.analyze_advanced_layout(clean_text, detected_columns)
        
        print(f"ページ {page_data['page_number']}: 検出された列数 = {detected_columns}, 表数 = {total_tables}, 高度レイアウト = {advanced_layout['complexity']}")
        
        # ページ情報ヘッダー
        page_sheet.append(["ページ情報"])
        page_sheet.append(["ページ番号", page_data["page_number"]])
        page_sheet.append(["文字数", page_data["char_count"]])
        page_sheet.append(["検出列数", detected_columns])
        page_sheet.append(["レイアウト解析", f"最大{layout_analysis['max_columns_found']}列検出"])
        page_sheet.append(["表数", total_tables])
        page_sheet.append(["高度レイアウト", advanced_layout['complexity']])
        page_sheet.append(["セル結合数", advanced_layout['merge_count']])
        page_sheet.append(["OCR使用", "はい" if page_data.get("ocr_used") else "いいえ"])
        page_sheet.append(["処理日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        page_sheet.append([])
        
        # 表構造情報を表示
        if total_tables > 0:
            page_sheet.append(["表構造情報"])
            for table_info in table_structure.get('tables', []):
                page_sheet.append([
                    f"表{table_info['table_number']}",
                    f"{table_info['rows']}行 x {table_info['columns']}列",
                    f"セル数: {len(table_info['cell_positions'])}"
                ])
            page_sheet.append([])
        
        # テキスト内容
        page_sheet.append(["テキスト内容"])
        
        # テキストを行ごとに分割して書き込み（超高度なセル結合対応）
        text_lines = clean_text.split('\n')
        current_row = 9 + (total_tables * 2) if total_tables > 0 else 9  # ヘッダー行の後の行番号
        
        for line in text_lines:
            if line.strip():
                # 列に分割
                parts = self.split_line_into_columns(line, detected_columns)
                
                # 超高度なセル結合の判定
                merged_cells = self.detect_advanced_merged_cells(parts, detected_columns, advanced_layout)
                
                # 行を書き込み
                for col_idx, cell_value in enumerate(parts, 1):
                    cell = page_sheet.cell(row=current_row, column=col_idx, value=cell_value)
                    cell.border = border
                
                # セル結合を適用
                for merge_info in merged_cells:
                    start_col, end_col = merge_info
                    if end_col > start_col:
                        page_sheet.merge_cells(
                            start_row=current_row,
                            end_row=current_row,
                            start_column=start_col,
                            end_column=end_col
                        )
                
                current_row += 1
        
        # スタイル適用
        self.apply_page_styles(page_sheet, border, detected_columns)
    
    def analyze_advanced_layout(self, text, detected_columns):
        """高度なレイアウト解析"""
        lines = text.split('\n')
        complexity = "低"
        merge_count = 0
        
        # 複雑度の判定
        if detected_columns > 20:
            complexity = "超高"
        elif detected_columns > 10:
            complexity = "高"
        elif detected_columns > 5:
            complexity = "中"
        
        # セル結合数の推定
        for line in lines:
            if line.strip():
                parts = self.split_line_into_columns(line, detected_columns)
                merged_cells = self.detect_merged_cells(parts, detected_columns)
                merge_count += len(merged_cells)
        
        return {
            'complexity': complexity,
            'merge_count': merge_count
        }
    
    def detect_advanced_merged_cells(self, parts, total_columns, advanced_layout):
        """超高度なセル結合の判定"""
        merged_cells = []
        
        # 基本のセル結合判定
        merged_cells.extend(self.detect_merged_cells(parts, total_columns))
        
        # 高度なセル結合判定
        if self.config.get("advanced_merging", False):
            # 複雑度に応じた結合
            if advanced_layout['complexity'] == "超高":
                # 超複雑なレイアウトの場合、より積極的に結合
                for i in range(len(parts) - 1):
                    if parts[i].strip() and parts[i+1].strip():
                        if len(parts[i].strip()) <= 3 and len(parts[i+1].strip()) <= 3:
                            merged_cells.append((i+1, i+2))
        
        return merged_cells
    
    def detect_merged_cells(self, parts, total_columns):
        """高度なセル結合の判定"""
        merged_cells = []
        
        # 空のセルが連続している場合、結合候補とする
        start_col = None
        for i, part in enumerate(parts, 1):
            if not part.strip():  # 空のセル
                if start_col is None:
                    start_col = i
            else:  # 非空のセル
                if start_col is not None and i - 1 > start_col:
                    merged_cells.append((start_col, i - 1))
                start_col = None
        
        # 最後のセルが空の場合
        if start_col is not None and len(parts) > start_col:
            merged_cells.append((start_col, len(parts)))
        
        # 高度なセル結合判定
        if self.config.get("advanced_merging", False):
            # 短い文字列の連続を結合
            short_cells = []
            for i, part in enumerate(parts, 1):
                if part.strip() and len(part.strip()) <= 2:
                    short_cells.append(i)
            
            # 短いセルが連続している場合、結合
            if len(short_cells) > 1:
                start = short_cells[0]
                for i in range(1, len(short_cells)):
                    if short_cells[i] - short_cells[i-1] == 1:
                        continue
                    else:
                        if short_cells[i-1] - start > 0:
                            merged_cells.append((start, short_cells[i-1]))
                        start = short_cells[i]
                
                # 最後のグループ
                if short_cells[-1] - start > 0:
                    merged_cells.append((start, short_cells[-1]))
        
        return merged_cells
    
    def split_line_into_columns(self, line, target_columns):
        """行を指定された列数に分割（OCR対応の高度な分割パターン）"""
        # 複数の分割パターンを試す
        patterns = []
        
        # パターン1: タブで分割
        if '\t' in line:
            parts = [p.strip() for p in line.split('\t') if p.strip()]
            patterns.append(('tab', parts))
        
        # パターン2: 複数スペースで分割
        parts = [p.strip() for p in re.split(r'\s{2,}', line) if p.strip()]
        if len(parts) > 1:
            patterns.append(('spaces', parts))
        
        # パターン3: 特定の文字で分割
        parts = [p.strip() for p in re.split(r'[|;:]+', line) if p.strip()]
        if len(parts) > 1:
            patterns.append(('delimiter', parts))
        
        # パターン4: 数字と文字の境界で分割
        parts = [p.strip() for p in re.split(r'(?<=\d)\s+(?=[A-Za-z])|(?<=[A-Za-z])\s+(?=\d)', line) if p.strip()]
        if len(parts) > 1:
            patterns.append(('number_text', parts))
        
        # パターン5: 日付形式で分割（簡易版）
        parts = [p.strip() for p in re.split(r'\s+(?=\d{4}[-/]\d{1,2}[-/]\d{1,2})|\s+(?=\d{1,2}[-/]\d{1,2}[-/]\d{4})', line) if p.strip()]
        if len(parts) > 1:
            patterns.append(('date', parts))
        
        # パターン6: 金額形式で分割（簡易版）
        parts = [p.strip() for p in re.split(r'\s+(?=[¥$€£])|\s+(?=\d)', line) if p.strip()]
        if len(parts) > 1:
            patterns.append(('currency', parts))
        
        # パターン7: より細かいスペース分割
        parts = [p.strip() for p in re.split(r'\s+', line) if p.strip()]
        if len(parts) > 1:
            patterns.append(('single_space', parts))
        
        # パターン8: OCR特有のパターン（短い文字列の連続）
        parts = [p.strip() for p in re.split(r'(?<=[A-Za-z])\s+(?=[A-Za-z])', line) if p.strip()]
        if len(parts) > 1:
            patterns.append(('ocr_short', parts))
        
        # パターン9: 数字の連続パターン
        parts = [p.strip() for p in re.split(r'(?<=\d)\s+(?=\d)', line) if p.strip()]
        if len(parts) > 1:
            patterns.append(('ocr_number', parts))
        
        # パターン10: 混合パターン（文字と数字の混合）
        parts = [p.strip() for p in re.split(r'(?<=[A-Za-z])\s+(?=\d)|(?<=\d)\s+(?=[A-Za-z])', line) if p.strip()]
        if len(parts) > 1:
            patterns.append(('ocr_mixed', parts))
        
        # パターン11: 日本語文字の境界分割
        parts = [p.strip() for p in re.split(r'(?<=[ひらがなカタカナ漢字])\s+(?=[ひらがなカタカナ漢字])', line) if p.strip()]
        if len(parts) > 1:
            patterns.append(('japanese', parts))
        
        # パターン12: より細かい分割（1文字ずつ）
        parts = [p.strip() for p in re.split(r'(?<=[A-Za-z0-9])\s+(?=[A-Za-z0-9])', line) if p.strip()]
        if len(parts) > 1:
            patterns.append(('fine_split', parts))
        
        # パターン13: 極細分割（すべてのスペースで分割）
        parts = [p.strip() for p in line.split() if p.strip()]
        if len(parts) > 1:
            patterns.append(('ultra_fine', parts))
        
        # パターン14: 文字単位分割
        parts = [p.strip() for p in re.split(r'(?<=.)\s+(?=.)', line) if p.strip()]
        if len(parts) > 1:
            patterns.append(('char_level', parts))
        
        # パターン15: 超細分割（すべての文字で分割）
        parts = [p.strip() for p in re.split(r'(?<=.)(?=.)', line) if p.strip()]
        if len(parts) > 1:
            patterns.append(('ultra_char', parts))
        
        # パターン16: 数字と文字の超細分割
        parts = [p.strip() for p in re.split(r'(?<=\d)(?=[A-Za-z])|(?<=[A-Za-z])(?=\d)', line) if p.strip()]
        if len(parts) > 1:
            patterns.append(('number_char', parts))
        
        # パターン17: 記号分割
        parts = [p.strip() for p in re.split(r'(?<=[^\w\s])(?=[^\w\s])', line) if p.strip()]
        if len(parts) > 1:
            patterns.append(('symbol_split', parts))
        
        # パターン18: 混合超細分割
        parts = [p.strip() for p in re.split(r'(?<=[A-Za-z0-9])(?=[A-Za-z0-9])', line) if p.strip()]
        if len(parts) > 1:
            patterns.append(('mixed_ultra', parts))
        
        # 最適なパターンを選択
        if patterns:
            best_pattern = max(patterns, key=lambda x: len(x[1]))
            parts = best_pattern[1]
        else:
            parts = [line.strip()]
        
        # 目標列数に調整
        if len(parts) < target_columns:
            parts.extend([''] * (target_columns - len(parts)))
        elif len(parts) > target_columns:
            parts = parts[:target_columns]
        
        return parts
    
    def create_ocr_sheets(self, workbook, images_data, border):
        """OCRテキストシートの作成"""
        # ページごとにOCRテキストをグループ化
        page_ocr_data = {}
        for img_data in images_data["images"]:
            if img_data.get("ocr_text"):
                page_num = img_data['page_number']
                if page_num not in page_ocr_data:
                    page_ocr_data[page_num] = []
                
                clean_ocr_text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', img_data["ocr_text"])
                page_ocr_data[page_num].append({
                    'image_number': img_data['image_number'],
                    'text': clean_ocr_text
                })
        
        # 各ページのOCRシートを作成
        for page_num, ocr_list in page_ocr_data.items():
            sheet_name = f"OCR_ページ_{page_num}"
            ocr_sheet = workbook.create_sheet(sheet_name)
            
            # ヘッダー情報
            ocr_sheet.append(["OCR抽出情報"])
            ocr_sheet.append(["ページ番号", page_num])
            ocr_sheet.append(["画像数", len(ocr_list)])
            ocr_sheet.append(["処理日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            ocr_sheet.append([])
            
            # OCRテキスト内容
            ocr_sheet.append(["OCR抽出テキスト"])
            
            for i, ocr_item in enumerate(ocr_list, 1):
                ocr_sheet.append([f"画像{i} (No.{ocr_item['image_number']})"])
                text_lines = ocr_item['text'].split('\n')
                for line in text_lines:
                    if line.strip():
                        ocr_sheet.append([f"  {line.strip()}"])
                ocr_sheet.append([])
            
            # スタイル適用
            self.apply_ocr_styles(ocr_sheet, border)
    
    def create_table_sheets(self, workbook, tables_data, border):
        """表データシートの作成"""
        for table in tables_data["tables"]:
            sheet_name = f"表_{table['page']}_{table['table_number']}"
            table_sheet = workbook.create_sheet(sheet_name)
            
            # ヘッダー情報
            table_sheet.append(["表データ情報"])
            table_sheet.append(["ページ番号", table['page']])
            table_sheet.append(["表番号", table['table_number']])
            table_sheet.append(["行数", len(table['data'])])
            table_sheet.append(["列数", len(table['data'][0]) if table['data'] else 0])
            table_sheet.append([])
            
            # 表データ
            table_sheet.append(["表データ"])
            for row in table['data']:
                clean_row = []
                for cell in row:
                    if isinstance(cell, str):
                        clean_cell = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', cell)
                        clean_row.append(clean_cell)
                    else:
                        clean_row.append(cell)
                table_sheet.append(clean_row)
            
            # スタイル適用
            self.apply_table_styles(table_sheet, border)
    
    def apply_summary_styles(self, sheet, border):
        """概要シートのスタイル適用"""
        # タイトル行
        for cell in sheet.iter_rows(min_row=1, max_row=1):
            for cell_obj in cell:
                if cell_obj.value:
                    cell_obj.font = Font(bold=True, size=16, color="FFFFFF")
                    cell_obj.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
                    cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                    cell_obj.border = border
        
        # セクションヘッダー
        for row_num in [3, 8]:
            for cell in sheet.iter_rows(min_row=row_num, max_row=row_num):
                for cell_obj in cell:
                    if cell_obj.value:
                        cell_obj.font = Font(bold=True, size=12, color="2F5597")
                        cell_obj.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
                        cell_obj.border = border
        
        # データ行のヘッダー
        for row_num in [4, 9]:
            for cell in sheet.iter_rows(min_row=row_num, max_row=row_num):
                for cell_obj in cell:
                    if cell_obj.value:
                        cell_obj.font = Font(bold=True, color="FFFFFF")
                        cell_obj.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                        cell_obj.border = border
        
        # 列幅の自動調整
        self.auto_adjust_columns(sheet)
    
    def apply_page_styles(self, sheet, border, columns):
        """ページシートのスタイル適用"""
        # ヘッダー行のスタイル
        for row_num in [1, 2, 3, 4, 5, 6, 8]:
            for cell in sheet.iter_rows(min_row=row_num, max_row=row_num):
                for cell_obj in cell:
                    if cell_obj.value:
                        cell_obj.font = Font(bold=True, color="2F5597")
                        cell_obj.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
                        cell_obj.border = border
        
        # 列幅の自動調整
        for col_idx in range(1, columns + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in sheet.iter_cols(min_col=col_idx, max_col=col_idx):
                for cell_obj in cell:
                    try:
                        if len(str(cell_obj.value)) > max_length:
                            max_length = len(str(cell_obj.value))
                    except Exception:
                        pass
            adjusted_width = min(max_length + 2, 30)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def apply_ocr_styles(self, sheet, border):
        """OCRシートのスタイル適用"""
        for row_num in [1, 2, 3, 4, 6]:
            for cell in sheet.iter_rows(min_row=row_num, max_row=row_num):
                for cell_obj in cell:
                    if cell_obj.value:
                        cell_obj.font = Font(bold=True, color="8B4513")
                        cell_obj.fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
                        cell_obj.border = border
        
        self.auto_adjust_columns(sheet)
    
    def apply_table_styles(self, sheet, border):
        """表シートのスタイル適用"""
        for row_num in [1, 2, 3, 4, 5, 7]:
            for cell in sheet.iter_rows(min_row=row_num, max_row=row_num):
                for cell_obj in cell:
                    if cell_obj.value:
                        cell_obj.font = Font(bold=True, color="2F5597")
                        cell_obj.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
                        cell_obj.border = border
        
        self.auto_adjust_columns(sheet)
    
    def auto_adjust_columns(self, sheet):
        """列幅の自動調整"""
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

def main():
    """メイン処理"""
    if len(sys.argv) < 2:
        print("使用方法: python3 pdf_excel_converter_enhanced.py <PDFファイルパス>")
        sys.exit(1)
    
    pdf_path = sys.argv[1]
    
    if not os.path.exists(pdf_path):
        print(f"PDFファイルが見つかりません: {pdf_path}")
        sys.exit(1)
    
    # 設定（最終最適化版）
    config = {
        "ocr_enabled": True,
        "table_detection": True,
        "dynamic_columns": True,
        "layout_reconstruction": True,
        "max_columns": 100,  # 上限を大幅に上げる
        "min_columns": 3,
        "ultra_fine_detection": True,  # 超細かい検出
        "advanced_merging": True,  # 高度なセル結合
        "layout_optimization": True  # レイアウト最適化
    }
    
    # コンバーター初期化
    converter = EnhancedPDFExcelConverter(config)
    
    print("=== 改良版PDF-Excel変換開始 ===")
    print(f"PDFファイル: {pdf_path}")
    print(f"動的列数検出: {config['dynamic_columns']}")
    print(f"レイアウト再構築: {config['layout_reconstruction']}")
    print()
    
    try:
        # 1. PDFtkでPDFの構造を解析
        print("1. PDFtkでPDFの構造を解析中...")
        pdf_structure = converter.analyze_pdf_structure_with_pdftk(pdf_path)
        if pdf_structure:
            print(f"   - 総ページ数: {pdf_structure['total_pages']}")
            print(f"   - ページ情報: {len(pdf_structure['pages'])}ページ")
        
        # 2. テキスト抽出
        print("2. テキスト抽出中...")
        text_data = converter.extract_text_with_layout(pdf_path)
        print(f"   - 抽出ページ数: {text_data['page_count']}")
        
        # 3. 表データ抽出（簡易版）
        print("3. 表データ抽出中...")
        tables_data = {"tables": [], "total_tables": 0}
        
        # 4. 画像データ抽出（簡易版）
        print("4. 画像データ抽出中...")
        images_data = {"images": [], "total_images": 0}
        
        # 5. Excelファイル作成
        print("5. Excelファイル作成中...")
        output_path = f"{pdf_path}_enhanced_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        converter.create_excel_with_dynamic_layout(text_data, tables_data, images_data, output_path)
        
        print(f"✅ 変換完了: {output_path}")
        
    except Exception as e:
        print(f"❌ 変換エラー: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()