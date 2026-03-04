#!/usr/bin/env python3
"""
日報PDF→Excel表形式変換システム v2.0
画像ベースPDFから表構造を認識してExcel化

特徴：
- 高解像度画像変換
- 高精度OCR（日本語・英語）
- 表構造の自動認識
- Excel表形式で出力
"""

import sys
from pathlib import Path
import logging
from datetime import datetime
import re

# 必要なライブラリ
try:
    import fitz  # PyMuPDF
    import pytesseract
    from PIL import Image
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import cv2
    import numpy as np
except ImportError as e:
    print(f"❌ 必要なライブラリが不足しています: {e}")
    print("インストール: pip3 install PyMuPDF pytesseract Pillow pandas openpyxl opencv-python-headless")
    sys.exit(1)

# ログ設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/root/logs/nippo_pdf_to_excel_v2.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("NippoPDFToExcelV2")


class NippoPDFToExcelV2:
    """日報PDF→Excel変換システム v2.0"""
    
    def __init__(self):
        self.output_dir = Path('/root/daily_reports/excel_v2')
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Tesseract設定（日本語＋英語）
        self.tesseract_config = '--oem 3 --psm 6 -l jpn+eng'
    
    def pdf_to_high_res_image(self, pdf_path: str, page_num: int = 0, dpi: int = 300):
        """PDFを高解像度画像に変換"""
        try:
            doc = fitz.open(pdf_path)
            page = doc[page_num]
            
            # 高解像度で画像化（DPI 300）
            zoom = dpi / 72  # 72 DPI がデフォルト
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat)
            
            # PIL Imageに変換
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            
            doc.close()
            return img
            
        except Exception as e:
            logger.error(f"PDF→画像変換エラー: {e}")
            return None
    
    def detect_table_structure(self, image):
        """画像から表構造を検出"""
        try:
            # PIL Image → OpenCV形式
            img_array = np.array(image)
            gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
            
            # 二値化
            _, binary = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)
            
            # 水平・垂直線を検出
            horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
            vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))
            
            horizontal_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, horizontal_kernel)
            vertical_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, vertical_kernel)
            
            # 表の格子を合成
            table_mask = cv2.add(horizontal_lines, vertical_lines)
            
            # 輪郭検出
            contours, _ = cv2.findContours(table_mask, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            
            # セル候補を抽出（面積でフィルタ）
            cells = []
            for contour in contours:
                x, y, w, h = cv2.boundingRect(contour)
                area = w * h
                
                # 小さすぎる・大きすぎるものを除外
                if 1000 < area < 100000:
                    cells.append({'x': x, 'y': y, 'w': w, 'h': h, 'area': area})
            
            # Y座標でソート（行ごと）
            cells_sorted = sorted(cells, key=lambda c: (c['y'], c['x']))
            
            return cells_sorted, img_array
            
        except Exception as e:
            logger.warning(f"表構造検出エラー: {e}")
            return [], np.array(image)
    
    def extract_text_from_region(self, image_array, x, y, w, h):
        """画像の特定領域からテキスト抽出"""
        try:
            # 領域を切り出し
            region = image_array[y:y+h, x:x+w]
            
            # PIL Imageに変換
            region_img = Image.fromarray(region)
            
            # OCRでテキスト抽出
            text = pytesseract.image_to_string(region_img, config=self.tesseract_config)
            
            # クリーンアップ
            text = text.strip()
            text = re.sub(r'\s+', ' ', text)  # 余分な空白を削除
            
            return text
            
        except Exception as e:
            logger.warning(f"領域テキスト抽出エラー: {e}")
            return ""
    
    def simple_ocr_extraction(self, image):
        """シンプルなOCR抽出（表構造が検出できない場合）"""
        try:
            # 全体をOCR
            text = pytesseract.image_to_string(image, config=self.tesseract_config)
            
            # 行ごとに分割
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            
            # データフレーム化（1列）
            df = pd.DataFrame({'内容': lines})
            
            return df
            
        except Exception as e:
            logger.error(f"シンプルOCRエラー: {e}")
            return pd.DataFrame()
    
    def cells_to_dataframe(self, cells, image_array):
        """セル情報からDataFrameを作成"""
        try:
            if not cells:
                return None
            
            # 行をグループ化（Y座標が近いものをまとめる）
            rows = []
            current_row = []
            last_y = cells[0]['y'] if cells else 0
            
            for cell in cells:
                # Y座標が大きく変わったら新しい行
                if abs(cell['y'] - last_y) > 20:
                    if current_row:
                        rows.append(current_row)
                    current_row = [cell]
                    last_y = cell['y']
                else:
                    current_row.append(cell)
            
            if current_row:
                rows.append(current_row)
            
            # 各セルからテキスト抽出
            data = []
            for row_cells in rows:
                row_data = []
                for cell in sorted(row_cells, key=lambda c: c['x']):
                    text = self.extract_text_from_region(
                        image_array,
                        cell['x'], cell['y'], cell['w'], cell['h']
                    )
                    row_data.append(text)
                data.append(row_data)
            
            # DataFrameに変換
            if data:
                # 最大列数を取得
                max_cols = max(len(row) for row in data)
                
                # 列数を統一
                normalized_data = []
                for row in data:
                    while len(row) < max_cols:
                        row.append('')
                    normalized_data.append(row[:max_cols])
                
                df = pd.DataFrame(normalized_data)
                return df
            
            return None
            
        except Exception as e:
            logger.error(f"DataFrame変換エラー: {e}")
            return None
    
    def create_excel(self, df, output_path: str, pdf_name: str):
        """DataFrameからExcelファイルを作成"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '日報データ'
            
            # データを書き込み
            for r_idx, row in enumerate(df.itertuples(index=False), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # スタイル設定
                    cell.font = Font(name='メイリオ', size=11)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # ヘッダー行（1行目）を強調
                    if r_idx == 1:
                        cell.font = Font(name='メイリオ', size=11, bold=True)
                        cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
            
            # 列幅自動調整
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # サマリーシート追加
            summary_ws = wb.create_sheet('変換情報', 0)
            summary_data = [
                ['項目', '値'],
                ['元PDFファイル', pdf_name],
                ['変換日時', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['抽出行数', len(df)],
                ['抽出列数', len(df.columns)],
                ['システム', '日報PDF→Excel変換システム v2.0']
            ]
            
            for r_idx, row_data in enumerate(summary_data, start=1):
                for c_idx, value in enumerate(row_data, start=1):
                    cell = summary_ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.font = Font(name='メイリオ', size=11)
                    if r_idx == 1:
                        cell.font = Font(name='メイリオ', size=11, bold=True)
                        cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
            
            summary_ws.column_dimensions['A'].width = 20
            summary_ws.column_dimensions['B'].width = 50
            
            # 保存
            wb.save(output_path)
            logger.info(f"✅ Excel作成完了: {output_path}")
            
            return True
            
        except Exception as e:
            logger.error(f"Excel作成エラー: {e}")
            return False
    
    def convert_pdf(self, pdf_path: str):
        """PDFをExcelに変換（メイン処理）"""
        pdf_path = Path(pdf_path)
        
        if not pdf_path.exists():
            logger.error(f"PDFファイルが見つかりません: {pdf_path}")
            return {'success': False, 'error': 'File not found'}
        
        print(f"\n{'='*60}")
        print(f"📄 変換開始: {pdf_path.name}")
        print(f"{'='*60}")
        
        try:
            # 1. PDFを高解像度画像に変換
            print("🖼️  STEP 1: PDF→高解像度画像変換...")
            image = self.pdf_to_high_res_image(str(pdf_path), dpi=300)
            
            if image is None:
                return {'success': False, 'error': 'Image conversion failed'}
            
            print(f"   画像サイズ: {image.size[0]} x {image.size[1]}")
            
            # 2. 表構造を検出
            print("🔍 STEP 2: 表構造検出...")
            cells, img_array = self.detect_table_structure(image)
            print(f"   検出セル数: {len(cells)}")
            
            # 3. データ抽出
            print("📊 STEP 3: データ抽出...")
            
            if len(cells) > 10:
                # 表構造が検出できた場合
                print("   方法: 表構造認識")
                df = self.cells_to_dataframe(cells, img_array)
            else:
                # 表構造が検出できない場合はシンプルOCR
                print("   方法: シンプルOCR")
                df = self.simple_ocr_extraction(image)
            
            if df is None or df.empty:
                logger.warning("データ抽出できませんでした")
                return {'success': False, 'error': 'No data extracted'}
            
            print(f"   抽出データ: {len(df)}行 x {len(df.columns)}列")
            
            # 4. Excel作成
            print("📝 STEP 4: Excel作成...")
            output_filename = f"{pdf_path.stem}_v2_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = self.output_dir / output_filename
            
            success = self.create_excel(df, str(output_path), pdf_path.name)
            
            if success:
                print("\n✅ 変換完了！")
                print(f"📁 出力: {output_path}")
                print(f"{'='*60}\n")
                
                return {
                    'success': True,
                    'output_file': str(output_path),
                    'rows': len(df),
                    'cols': len(df.columns)
                }
            else:
                return {'success': False, 'error': 'Excel creation failed'}
            
        except Exception as e:
            logger.error(f"変換エラー: {e}")
            return {'success': False, 'error': str(e)}
    
    def batch_convert(self, pdf_dir: str):
        """フォルダ内の全PDFを一括変換"""
        pdf_dir = Path(pdf_dir)
        
        if not pdf_dir.exists():
            logger.error(f"フォルダが見つかりません: {pdf_dir}")
            return
        
        pdf_files = list(pdf_dir.glob('*.pdf'))
        pdf_files.extend(list(pdf_dir.glob('*.PDF')))
        
        if not pdf_files:
            logger.error("PDFファイルが見つかりません")
            return
        
        print(f"\n{'='*60}")
        print("📁 一括変換開始")
        print(f"{'='*60}")
        print(f"PDFファイル数: {len(pdf_files)}")
        print(f"出力先: {self.output_dir}\n")
        
        results = {'total': len(pdf_files), 'success': 0, 'failed': 0}
        
        for i, pdf_file in enumerate(pdf_files, 1):
            print(f"[{i}/{len(pdf_files)}] {pdf_file.name}")
            
            result = self.convert_pdf(str(pdf_file))
            
            if result['success']:
                results['success'] += 1
            else:
                results['failed'] += 1
                print(f"   ❌ 失敗: {result.get('error', '不明')}\n")
        
        # サマリー
        print(f"\n{'='*60}")
        print("📊 一括変換完了")
        print(f"{'='*60}")
        print(f"総ファイル数: {results['total']}")
        print(f"✅ 成功: {results['success']}")
        print(f"❌ 失敗: {results['failed']}")
        print(f"📁 出力先: {self.output_dir}")
        print(f"{'='*60}\n")


def main():
    """メイン処理"""
    converter = NippoPDFToExcelV2()
    
    if len(sys.argv) < 2:
        print("使い方:")
        print("  python3 nippo_pdf_to_excel_v2.py [PDFファイル]")
        print("  python3 nippo_pdf_to_excel_v2.py [フォルダ]")
        sys.exit(1)
    
    target = sys.argv[1]
    target_path = Path(target)
    
    if target_path.is_file():
        converter.convert_pdf(target)
    elif target_path.is_dir():
        converter.batch_convert(target)
    else:
        print(f"❌ パスが見つかりません: {target}")
        sys.exit(1)


if __name__ == '__main__':
    main()



