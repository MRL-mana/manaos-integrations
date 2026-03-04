#!/usr/bin/env python3
"""
Mana PDF to Excel Advanced System
ManaOS統合用の高機能PDF→Excel変換システム
"""

import json
import os
import asyncio
import logging
import uuid
import shutil
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional
import threading

# 共通学習レイヤー
try:
    sys.path.insert(0, '/root')
    from manaos_learning import get_learning_api, apply_rules
    LEARNING_AVAILABLE = True
except ImportError:
    LEARNING_AVAILABLE = False

# データ処理ライブラリ
try:
    import pandas as pd
    import numpy as np
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# PDF処理ライブラリ
try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

# OCR処理（オプション）
try:
    import pytesseract
    from PIL import Image
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

# PaddleOCR（中国最強OCR、日本語に強い）
# メモリ不足対策: デフォルトで無効化（必要に応じて有効化）
PADDLEOCR_AVAILABLE = False
try:
    # メモリが十分な場合のみ有効化
    from paddleocr import PaddleOCR
    PADDLEOCR_AVAILABLE = True
except (ImportError, MemoryError, Exception):
    PADDLEOCR_AVAILABLE = False

# OpenCV（画像処理・傾き補正用）
try:
    import cv2
    import numpy as np
    OPENCV_AVAILABLE = True
except ImportError:
    OPENCV_AVAILABLE = False

# Ollama（ローカルLLM、OCR結果の整形・表構造化用）
try:
    import requests
    OLLAMA_AVAILABLE = True
except ImportError:
    OLLAMA_AVAILABLE = False

# Camelot（高精度表抽出、オプション）
try:
    import camelot
    CAMELOT_AVAILABLE = True
except (ImportError, MemoryError, Exception):
    CAMELOT_AVAILABLE = False


class ManaPDFExcelAdvanced:
    """ManaOS統合用PDF→Excel変換システム（超強化版）"""

    def __init__(self, config: Dict[str, Any] = None):
        self.config = config or self.get_default_config()
        self.logger = self.setup_logger()
        self.output_dir = Path("/root/excel_output_advanced")
        self.output_dir.mkdir(exist_ok=True)
        self.temp_dir = Path("/tmp/mana_pdf_excel")
        self.temp_dir.mkdir(exist_ok=True)

        # Ollama設定（ローカルLLM）
        self.ollama_url = os.getenv("OLLAMA_URL", "http://localhost:11434")
        self.ollama_model = os.getenv("OLLAMA_MODEL", "llama3:8b")  # デフォルトモデル
        self.ollama_enabled = self.config.get("ollama_enabled", False)

        # 共通学習レイヤー
        self.learning_enabled = LEARNING_AVAILABLE and self.config.get(
            "learning_enabled", True)
        if self.learning_enabled:
            try:
                self.learning_api = get_learning_api()
                self.logger.info("共通学習レイヤーを有効化しました")
            except Exception as e:
                self.logger.warning(f"学習レイヤー初期化エラー: {e}")
                self.learning_enabled = False
        else:
            self.learning_api = None

        # 統計情報
        self.stats = {
            'total_conversions': 0,
            'successful_conversions': 0,
            'failed_conversions': 0,
            'total_pages_processed': 0,
            'total_tables_extracted': 0,
            'high_quality_tables': 0,
            'ocr_used_count': 0
        }

        # 処理中タスクの管理
        self.active_tasks: Dict[str, Dict[str, Any]] = {}
        self.task_lock = threading.Lock()

        self.logger.info("Mana PDF Excel Advanced システム初期化完了")

    def get_default_config(self) -> Dict[str, Any]:
        """デフォルト設定取得"""
        return {
            "output_format": "excel",
            "ocr_enabled": False,  # デフォルトはFalse（必要時のみ有効化）
            "table_detection": True,
            "quality_analysis": True,
            "data_type_detection": True,
            "merge_tables": True,
            "page_range": "all",
            "google_drive_save": False,
            "template_style": "default",
            "language": "jpn",
            "max_file_size_mb": 50,
            "timeout_seconds": 300,
            "parallel_processing": True,
            "temp_cleanup": True,
            "use_camelot": CAMELOT_AVAILABLE,  # Camelotが利用可能なら使用
            "ollama_enabled": False,  # デフォルトは無効（必要時のみ有効化）
            "ollama_url": os.getenv("OLLAMA_URL", "http://localhost:11434"),
            "ollama_model": os.getenv("OLLAMA_MODEL", "llama3:8b")
        }

    def setup_logger(self) -> logging.Logger:
        """ロガー設定"""
        logger = logging.getLogger("ManaPDFExcelAdvanced")
        logger.setLevel(logging.INFO)

        # ファイルハンドラー
        log_file = f"/root/logs/mana_pdf_excel_advanced_{datetime.now().strftime('%Y%m%d')}.log"
        Path(log_file).parent.mkdir(parents=True, exist_ok=True)
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.INFO)

        # コンソールハンドラー
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)

        # フォーマッター
        formatter = logging.Formatter(
            '%(asctime)s [%(levelname)s] %(name)s: %(message)s'
        )
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)

        logger.addHandler(file_handler)
        logger.addHandler(console_handler)

        return logger

    def validate_pdf_file(self, file_path: str) -> Dict[str, Any]:
        """PDFファイル検証"""
        try:
            file_path = Path(file_path)

            if not file_path.exists():
                return {"valid": False, "error": "ファイルが存在しません"}

            if not file_path.suffix.lower() == '.pdf':
                return {"valid": False, "error": "PDFファイルではありません"}

            # ファイルサイズチェック
            file_size_mb = file_path.stat().st_size / (1024 * 1024)
            if file_size_mb > self.config["max_file_size_mb"]:
                return {
                    "valid": False,
                    "error": f"ファイルサイズが上限を超えています ({file_size_mb:.1f}MB > {self.config['max_file_size_mb']}MB)"
                }

            # PDFファイルの基本検証
            if PYMUPDF_AVAILABLE:
                try:
                    doc = fitz.open(file_path)
                    page_count = len(doc)
                    doc.close()

                    return {
                        "valid": True,
                        "file_size_mb": file_size_mb,
                        "page_count": page_count,
                        "file_path": str(file_path)
                    }
                except Exception as e:
                    return {"valid": False, "error": f"PDFファイルの読み込みに失敗: {e}"}

            return {"valid": True, "file_size_mb": file_size_mb}

        except Exception as e:
            return {"valid": False, "error": f"ファイル検証エラー: {e}"}

    def extract_text_from_pdf(self, pdf_path: str) -> Dict[str, Any]:
        """PDFからテキスト抽出"""
        try:
            if not PYMUPDF_AVAILABLE:
                return {"success": False, "error": "PyMuPDFが利用できません"}

            doc = fitz.open(pdf_path)
            extracted_data = {
                "pages": [],
                "full_text": "",
                "metadata": doc.metadata,
                "page_count": len(doc)
            }

            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                text = page.get_text()

                page_data = {
                    "page_number": page_num + 1,
                    "text": text,
                    "char_count": len(text)
                }

                extracted_data["pages"].append(page_data)
                extracted_data["full_text"] += text + "\n"

            doc.close()

            self.logger.info(f"テキスト抽出完了: {len(extracted_data['pages'])}ページ")
            return {"success": True, "data": extracted_data}

        except Exception as e:
            self.logger.error(f"テキスト抽出エラー: {e}")
            return {"success": False, "error": str(e)}

    def is_image_based_pdf(self, pdf_path: str) -> bool:
        """PDFが画像ベースかどうかを判定"""
        try:
            if not PYMUPDF_AVAILABLE:
                return False

            doc = fitz.open(pdf_path)
            image_based_count = 0
            total_pages = min(len(doc), 5)  # 最初の5ページをチェック

            for page_num in range(total_pages):
                page = doc.load_page(page_num)
                text = page.get_text().strip()
                images = page.get_images()

                # テキストが少なく、画像が多い場合は画像ベースと判定
                if len(text) < 100 and len(images) > 0:
                    image_based_count += 1

            doc.close()

            # 50%以上が画像ベースなら画像ベースPDFと判定
            return image_based_count >= total_pages * 0.5

        except Exception as e:
            self.logger.warning(f"画像ベースPDF判定エラー: {e}")
            return False

    def extract_tables_from_pdf(self, pdf_path: str, use_camelot: bool = None) -> Dict[str, Any]:
        """PDFから表データ抽出（複数エンジン対応）"""
        if use_camelot is None:
            use_camelot = self.config.get("use_camelot", False)

        tables_data = {
            "tables": [],
            "total_tables": 0,
            "extraction_methods": [],
            "is_image_based": False
        }

        # 画像ベースPDFかどうかを判定
        is_image_based = self.is_image_based_pdf(pdf_path)
        tables_data["is_image_based"] = is_image_based

        if is_image_based:
            self.logger.info("画像ベースPDFを検出。OCRを使用することを推奨します。")

        # 方法1: pdfplumber（基本）
        if PDFPLUMBER_AVAILABLE:
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    for page_num, page in enumerate(pdf.pages):
                        tables = page.extract_tables()

                        for table_num, table in enumerate(tables):
                            if table and len(table) > 0:
                                cleaned_table = []
                                for row in table:
                                    cleaned_row = []
                                    for cell in row:
                                        if cell is None:
                                            cleaned_row.append("")
                                        else:
                                            cleaned_row.append(
                                                str(cell).strip())
                                    cleaned_table.append(cleaned_row)

                                if cleaned_table:
                                    table_data = {
                                        "page_number": page_num + 1,
                                        "table_number": table_num + 1,
                                        "data": cleaned_table,
                                        "rows": len(cleaned_table),
                                        "columns": len(cleaned_table[0]) if cleaned_table else 0,
                                        "method": "pdfplumber",
                                        "quality_score": self.calculate_table_quality(cleaned_table)
                                    }

                                    tables_data["tables"].append(table_data)
                                    tables_data["extraction_methods"].append(
                                        "pdfplumber")

                self.logger.info(
                    f"pdfplumberで{len(tables_data['tables'])}個の表を抽出")
            except Exception as e:
                self.logger.warning(f"pdfplumber抽出エラー: {e}")

        # 方法2: Camelot（高精度、罫線ベース）
        # 画像ベースPDFの場合はCamelotをスキップ（警告が出るため）
        if use_camelot and CAMELOT_AVAILABLE and not is_image_based:
            try:
                # Lattice方式（罫線ベース）
                camelot_tables = camelot.read_pdf(
                    pdf_path, flavor='lattice', pages='all')

                for table_idx, table in enumerate(camelot_tables):
                    table_df = table.df
                    table_list = table_df.values.tolist()

                    if table_list:
                        # ページ番号を推定（Camelotはページ情報を直接提供しない場合がある）
                        page_num = getattr(table, 'page', table_idx + 1)

                        table_data = {
                            "page_number": int(page_num),
                            "table_number": table_idx + 1,
                            "data": table_list,
                            "rows": len(table_list),
                            "columns": len(table_list[0]) if table_list else 0,
                            "method": "camelot_lattice",
                            "quality_score": self.calculate_table_quality(table_list),
                            "accuracy": table.accuracy
                        }

                        tables_data["tables"].append(table_data)
                        tables_data["extraction_methods"].append(
                            "camelot_lattice")

                self.logger.info(f"Camelotで{len(camelot_tables)}個の表を抽出")
            except Exception as e:
                self.logger.warning(f"Camelot抽出エラー: {e}")

        tables_data["total_tables"] = len(tables_data["tables"])

        # 高品質表のカウント
        high_quality_count = sum(1 for t in tables_data["tables"]
                                 if t.get("quality_score", 0) >= 70)
        tables_data["high_quality_tables"] = high_quality_count

        self.logger.info(
            f"表データ抽出完了: {tables_data['total_tables']}個（高品質: {high_quality_count}個）")
        return {"success": True, "data": tables_data}

    def extract_tables_from_ocr(self, ocr_pages: List[Dict[str, Any]]) -> Dict[str, Any]:
        """OCR結果から表を抽出（改善版：1行OCRにも対応）"""
        tables_data = {
            "tables": [],
            "total_tables": 0,
            "high_quality_tables": 0,
            "extraction_methods": []
        }

        for page_data in ocr_pages:
            page_num = page_data.get("page_number", 0)
            ocr_text = page_data.get("ocr_text", "")

            if not ocr_text or len(ocr_text.strip()) < 50:
                continue

            # 改行で分割
            lines = re.split(r'\n+|\r+|\r\n+', ocr_text)
            lines = [line.strip() for line in lines if line.strip()]

            # 1行にまとまっている場合の処理
            if len(lines) <= 1 and len(ocr_text) > 100:
                # キーワードや数値パターンで分割を試行
                self.logger.info(f"ページ{page_num}: 1行OCRテキストを分割処理")

                # 方法1: キーワードで分割（表のヘッダーや区切り）
                keywords = ['金額', '数量', '合計', '計', '現金', 'ブリカ',
                            'クレジット', '品名', '商品', 'F/O', 'f/o', '粗利']
                split_patterns = []
                for kw in keywords:
                    # キーワードの前後で分割
                    split_patterns.append(rf'\s+(?={re.escape(kw)})')
                    split_patterns.append(rf'(?<={re.escape(kw)})\s+')

                # 方法2: 数値パターンで分割（カンマ区切りの数値の前後）
                split_patterns.extend([
                    r'(?<=\d)\s{3,}(?=\d)',  # 数値間の大きなスペース
                    r'(?<=[0-9,.])\s{4,}(?=[0-9])',  # 数値間の4文字以上のスペース
                    r'(?<=[\u4E00-\u9FAF])\s{3,}(?=[0-9])',  # 日本語と数値の間
                    r'(?<=[0-9])\s{3,}(?=[\u4E00-\u9FAF])',  # 数値と日本語の間
                ])

                # 分割を試行
                for pattern in split_patterns:
                    temp_lines = re.split(pattern, ocr_text)
                    if len(temp_lines) > len(lines):
                        lines = [line.strip()
                                 for line in temp_lines if line.strip()]
                        if len(lines) >= 5:  # 5行以上に分割できたら採用
                            self.logger.info(
                                f"ページ{page_num}: 1行テキストを{len(lines)}行に分割")
                            break

                # それでも1行の場合は、数値の位置を基準に列を検出
                if len(lines) <= 1:
                    from collections import Counter
                    number_positions = []
                    for match in re.finditer(r'\d+[,\d]*\.?\d*', ocr_text):
                        number_positions.append(match.start())

                    if len(number_positions) >= 5:
                        # 位置の間隔を分析
                        intervals = []
                        for i in range(len(number_positions) - 1):
                            interval = number_positions[i +
                                                        1] - number_positions[i]
                            if 10 <= interval <= 100:  # 10-100文字の間隔
                                intervals.append(interval)

                        if intervals:
                            interval_counter = Counter(intervals)
                            common_interval = interval_counter.most_common(1)[
                                0][0]

                            # 列の区切り位置を決定
                            column_positions = [0]
                            current_pos = common_interval
                            while current_pos < len(ocr_text):
                                column_positions.append(current_pos)
                                current_pos += common_interval

                            # 列ごとに分割
                            split_cells = []
                            for i in range(len(column_positions)):
                                start = column_positions[i]
                                end = column_positions[i + 1] if i + \
                                    1 < len(column_positions) else len(ocr_text)
                                cell = ocr_text[start:end].strip()
                                if cell:
                                    split_cells.append(cell)

                            if len(split_cells) >= 3:
                                lines = split_cells
                                self.logger.info(
                                    f"ページ{page_num}: 数値位置基準で{len(split_cells)}列に分割")

            if len(lines) < 3:  # 3行未満は表とみなさない
                continue

            # 行を解析して表データを作成
            parsed_rows = []
            max_columns = 0

            for line in lines:
                cells = self.parse_ocr_line(line)

                # 「金額 数量」のようなキーワード行を適切に分割
                if len(cells) == 1 and any(kw in line for kw in ['金額', '数量', '粗利', 'F/O', 'f/o', '合計', '計']):
                    # キーワードで分割
                    keywords = ['金額', '数量', '粗利金額', '粗利',
                                'F/O', 'f/o', 'FノO', '合計', '計']
                    split_cells = []
                    remaining = line

                    for keyword in keywords:
                        if keyword in remaining:
                            parts = remaining.split(keyword, 1)
                            if parts[0].strip():
                                split_cells.append(parts[0].strip())
                            split_cells.append(keyword)
                            remaining = parts[1] if len(parts) > 1 else ""

                    if remaining.strip():
                        split_cells.append(remaining.strip())

                    if len(split_cells) >= 2:
                        cells = split_cells

                # 列数が多すぎる場合（100列以上）は分割が失敗しているとみなす
                if len(cells) >= 2 and len(cells) < 100:  # 100列未満の場合のみ
                    parsed_rows.append(cells)
                    max_columns = max(max_columns, len(cells))

            # 列数が多すぎる場合は、より適切な分割方法を試行
            if max_columns >= 100 or (len(parsed_rows) > 0 and max_columns >= 50):
                self.logger.warning(
                    f"ページ{page_num}: 列数が多すぎます（{max_columns}列）。分割方法を変更します。")
                # より積極的な分割を試行
                parsed_rows = []
                max_columns = 0

                for line in lines:
                    # 複数スペース（3文字以上）で強制的に分割
                    cells = re.split(r'\s{3,}', line.strip())
                    cells = [c.strip() for c in cells if c.strip()]
                    if len(cells) >= 2 and len(cells) < 50:  # 50列未満
                        parsed_rows.append(cells)
                        max_columns = max(max_columns, len(cells))

            # まだ列数が多い場合は、さらに積極的に分割
            if len(parsed_rows) > 0 and max_columns >= 50:
                self.logger.info(
                    f"ページ{page_num}: さらに積極的な分割を試行（現在{max_columns}列）")
                parsed_rows = []
                max_columns = 0

                for line in lines:
                    # 2文字以上のスペースで分割
                    cells = re.split(r'\s{2,}', line.strip())
                    cells = [c.strip() for c in cells if c.strip()]
                    # 数値のみのセルを結合（カンマ区切りの数値の可能性）
                    merged_cells = []
                    i = 0
                    while i < len(cells):
                        if i < len(cells) - 1 and re.match(r'^\d+$', cells[i]) and re.match(r'^\d+$', cells[i+1]):
                            merged_cells.append(f"{cells[i]},{cells[i+1]}")
                            i += 2
                        else:
                            merged_cells.append(cells[i])
                            i += 1

                    if len(merged_cells) >= 2 and len(merged_cells) < 30:  # 30列未満
                        parsed_rows.append(merged_cells)
                        max_columns = max(max_columns, len(merged_cells))

            # 3行以上、2列以上、30列未満の場合のみ表として扱う
            if len(parsed_rows) >= 3 and max_columns >= 2 and max_columns < 30:
                # 列数を統一（最大列数に合わせる）
                normalized_rows = []
                for row in parsed_rows:
                    normalized_row = row[:max_columns]  # 最大列数まで
                    # 不足分を空文字で埋める
                    while len(normalized_row) < max_columns:
                        normalized_row.append("")
                    normalized_rows.append(normalized_row)

                # 品質スコアを計算
                quality_score = self.calculate_table_quality(normalized_rows)

                table_data = {
                    "page_number": page_num,
                    "table_number": 1,  # OCRの場合は1ページ1表として扱う
                    "data": normalized_rows,
                    "rows": len(normalized_rows),
                    "columns": max_columns,
                    "method": "ocr",
                    "quality_score": quality_score
                }

                tables_data["tables"].append(table_data)
                tables_data["extraction_methods"].append("ocr")

                if quality_score >= 70:
                    tables_data["high_quality_tables"] += 1

        tables_data["total_tables"] = len(tables_data["tables"])
        return tables_data

    def calculate_table_quality(self, table_data: List[List[str]]) -> float:
        """表の品質スコアを計算（0-100）"""
        if not table_data or not table_data[0]:
            return 0.0

        score = 100.0

        # 空セル率による減点
        total_cells = len(table_data) * len(table_data[0])
        empty_cells = sum(
            1 for row in table_data for cell in row if not cell or cell.strip() == "")
        empty_ratio = empty_cells / total_cells if total_cells > 0 else 0
        score -= empty_ratio * 30  # 空セルが多いと減点

        # 行数が少ないと減点
        if len(table_data) < 2:
            score -= 20

        # 列数が少ないと減点
        if len(table_data[0]) < 2:
            score -= 20

        return max(0.0, min(100.0, score))

    def detect_data_types(self, table_data: List[List[str]]) -> Dict[str, Any]:
        """データ型を自動判別"""
        if not table_data or len(table_data) < 2:
            return {}

        type_info = {}
        headers = table_data[0]

        for col_idx, header in enumerate(headers):
            if col_idx >= len(table_data[0]):
                continue

            # 列のデータを取得（ヘッダー除く）
            column_data = [row[col_idx] if col_idx < len(row) else ""
                           for row in table_data[1:]]

            # データ型判定
            detected_type = "text"
            numeric_count = 0
            date_count = 0

            for cell in column_data:
                if not cell or cell.strip() == "":
                    continue

                # 数値判定
                try:
                    # カンマ区切り数値対応
                    cleaned = cell.replace(',', '').replace(
                        '¥', '').replace('$', '').strip()
                    float(cleaned)
                    numeric_count += 1
                except:
                    pass

                # 日付判定（簡易）
                if '/' in cell or '-' in cell:
                    date_count += 1

            if numeric_count > len(column_data) * 0.7:
                detected_type = "numeric"
            elif date_count > len(column_data) * 0.5:
                detected_type = "date"

            type_info[header] = detected_type

        return type_info

    def fix_encoding_errors(self, text: str, event_id: Optional[str] = None) -> str:
        """文字エンコーディングエラーを修正（NFKC正規化 + 文脈ベース修正版）"""
        if not text:
            return text

        import re

        # 修正前のテキストを保存（自動記録用）
        original_text = text

        # 共通学習レイヤーのルールを適用（最初に適用）
        if self.learning_enabled:
            try:
                text = apply_rules(text, "pdf_excel")
            except Exception as e:
                self.logger.debug(f"共通ルール適用エラー（無視）: {e}")

        # まずは Unicode 正規化（全角→半角含む）
        fixed = unicodedata.normalize("NFKC", text)

        # 置換文字・怪しいコードポイントの除去
        for ch in ['\ufffd', '\ufffe', '\uffff', '\uFFFD']:
            fixed = fixed.replace(ch, '')

        # 全角スペース → 半角
        fixed = fixed.replace('　', ' ')

        # --- ここから文字化けパターン潰し ---

        # 1) 先頭の「<」が単なるノイズっぽいケースを削除
        #   例: "< 量", "< 合計", "< 金額" みたいなやつ
        fixed = re.sub(
            r'(?m)^\s*<\s*(?=[\u4E00-\u9FFF])',  # 行頭 + < + 漢字 なら < を削除
            '',
            fixed
        )

        # 2) セルっぽい場所での「; ,」「, ;」混在を修正
        fixed = re.sub(r';\s*,\s*', ',', fixed)  # "; ," → ","
        fixed = re.sub(r',\s*;\s*', ',', fixed)  # ", ;" → ","

        # 3) 数字の間に入ってしまった ; を , に変換
        #   例: "1;234" → "1,234"
        fixed = re.sub(r'(?<=\d);\s*(?=\d)', ',', fixed)

        # 4) 数字の間に変な記号があるやつを削る（保守的に）
        fixed = re.sub(r'(?<=\d)[<>\[\]{}|\\/]+(?=\d)', '', fixed)

        # 5) 不要記号のまとめ削除（日本語・英数字・基本記号以外を除去）
        fixed = re.sub(
            r'[^\w\s\u3040-\u30FF\u4E00-\u9FAF\u3400-\u4DBF.,;:()（）【】「」『』〈〉《》\d\-]+',
            '',
            fixed
        )

        # 制御文字削除
        fixed = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]', '', fixed)

        # 連続スペース → 1つ
        fixed = re.sub(r'\s+', ' ', fixed)

        result = fixed.strip()

        # 修正が発生した場合、自動記録（設定で有効な場合のみ）
        if (self.learning_enabled and
            self.config.get("auto_record_corrections", False) and
            original_text != result and
            len(original_text) < 500 and  # 長すぎるテキストは記録しない
                event_id):  # event_idがある場合のみ記録
            try:
                from learning_api import log_correction
                log_correction(
                    tool="pdf_excel",
                    task="ocr_to_table",
                    source_event_id=event_id,
                    raw_output=original_text[:500],
                    corrected_output=result[:500],
                    feedback="corrected",
                    tags=["encoding_fix", "auto_correction"],
                    meta={"method": "fix_encoding_errors"}
                )
            except Exception as e:
                self.logger.debug(f"修正ログ記録エラー（無視）: {e}")

        return result

    def clean_ocr_text(self, text: str) -> str:
        """OCRエラー文字を修正（より詳細に）"""
        if not text:
            return text

        # 修正前のテキストを保存（自動記録用）
        original_text = text

        # 共通学習レイヤーのルールを適用（最初に適用）
        if self.learning_enabled:
            try:
                text = apply_rules(text, "pdf_excel")
            except Exception as e:
                self.logger.debug(f"共通ルール適用エラー（無視）: {e}")

        # よくあるOCRエラーを修正
        corrections = {
            '和': '年',
            '朋': '月',
            'P0S': 'POS',  # P0S → POS
            '年則': '和則',  # 年則 → 和則
            'O': '0',  # 英字Oと数字0の混同を修正（ただし、文脈に応じて慎重に）
            'OO': '00',
            'Bs': '売',
            'apy': '計',
            '天井ば': '',
            'ASS': '',
            'Bell': '',
            '(HaHap': '',
            '天候:': '',
            'べべ': '',
            'Maat': '',
            '{a a': '',
            '(Gs': '',
            'Ate:': '',
            '気温:': '',
            '5t': '',
            'F/0': 'F/O',  # F/0 → F/O
            'M0': 'MO',
            'Ca?': '',
            'tested': '',
            'Fe sm': '',
            # 追加の修正パターン（順序重要：長いパターンから先に）
            'FノO': 'F/O',  # F/Oの誤認識（Fノより先に）
            'Fノ': 'F/',  # F/の誤認識
            '式会社': '株式会社',  # 「株式」の誤認識（「株式」の前に処理）
            'H0e': 'HO',  # HOの誤認識（H0より先に）
            'H0': 'HO',  # HOの誤認識
            '0e': 'O',  # Oeの誤認識
            '数恒': '数量',  # 数量の誤認識
        }

        cleaned = text
        for error, correct in corrections.items():
            # 大文字小文字を区別せずに置換（ただし、完全一致のみ）
            import re
            # 単語境界を考慮した置換
            cleaned = re.sub(r'\b' + re.escape(error) + r'\b',
                             correct, cleaned, flags=re.IGNORECASE)
            # 単語境界がない場合も置換（より積極的に）
            cleaned = cleaned.replace(error, correct)

        # 余分な記号を削除（数値の前後）
        import re
        # OCR記号を削除（【「」】『』〈〉《》など）
        cleaned = re.sub(r'[【「」】『』〈〉《》]', '', cleaned)
        # 連続するカンマやピリオドを修正（より積極的に）
        cleaned = re.sub(r',\s*,\s*', ',', cleaned)  # ", ," → ","
        cleaned = re.sub(r'\.\s*\.\s*', '.', cleaned)  # ". ." → "."
        cleaned = re.sub(r',\s*\.', '.', cleaned)  # ", ." → "."
        cleaned = re.sub(r'\.\s*,', ',', cleaned)  # ". ," → ","
        # 数値の後の余分な記号を削除（より積極的に）
        cleaned = re.sub(r'(\d+)\s*,\s*,\s*', r'\1,', cleaned)  # "1, ," → "1,"
        cleaned = re.sub(r'(\d+\.\d+)\s*\.\s*\.\s*', r'\1',
                         cleaned)  # "1.23. ." → "1.23"
        # カンマ区切り数値の修正（例: "1, , 676" → "1,676"）
        cleaned = re.sub(r'(\d+)\s*,\s*,\s*(\d+)', r'\1,\2', cleaned)
        # 小数点の修正（例: "1.23. ." → "1.23"）
        cleaned = re.sub(r'(\d+\.\d+)\s*\.\s*\.\s*', r'\1', cleaned)
        # 数値と記号の間の余分なスペースを削除
        cleaned = re.sub(r'(\d+)\s*,\s*,\s*(\d+)', r'\1,\2', cleaned)
        cleaned = re.sub(r'(\d+)\s*\.\s*\.\s*', r'\1.', cleaned)
        # セミコロンの連続を修正
        cleaned = re.sub(r';\s*;\s*', ';', cleaned)

        # 数値内のスペースを小数点に変換（例: "124 10" → "124.10"）
        # ただし、2桁以上の数値の間のスペースのみ（小数の可能性が高い）
        cleaned = re.sub(r'(\d{2,})\s+(\d{2})', r'\1.\2',
                         cleaned)  # "124 10" → "124.10"

        # 小数点の後の余分な記号を削除（複数回適用）
        # パターン1: "19.473. ." → "19.473"
        cleaned = re.sub(r'(\d+\.\d+)\s*\.\s*\.\s*', r'\1', cleaned)
        # パターン2: "19.473." → "19.473"
        cleaned = re.sub(r'(\d+\.\d+)\s*\.\s*([^\d])', r'\1\2', cleaned)
        cleaned = re.sub(r'(\d+\.\d+)\s*\.\s*$', r'\1', cleaned)
        # パターン3: 連続するピリオドを削除
        cleaned = re.sub(r'\.\s*\.\s*', '', cleaned)

        # カンマの後の余分な記号を削除（複数回適用）
        # パターン1: "1, ," → "1"（行末または非数字の前）
        # "1, , " → "1 "
        cleaned = re.sub(r'(\d+)\s*,\s*,\s*([^\d])', r'\1\2', cleaned)
        cleaned = re.sub(r'(\d+)\s*,\s*,\s*$', r'\1',
                         cleaned)  # "1, ," (行末) → "1"
        # パターン2: 数値の間にある「, ,」を修正
        cleaned = re.sub(r'(\d+)\s*,\s*,\s*(\d+)', r'\1,\2',
                         cleaned)  # "266, , 662" → "266,662"
        # パターン3: 連続するカンマを削除（一般的なケース）
        cleaned = re.sub(r',\s*,\s*', ',', cleaned)
        # パターン4: 数値の後に続く「, ,」を削除（残存するケース）
        cleaned = re.sub(r'(\d+)\s*,\s*,\s*', r'\1,', cleaned)

        result = cleaned

        # 修正が発生した場合、自動記録（設定で有効な場合のみ）
        if (self.learning_enabled and
            self.config.get("auto_record_corrections", False) and
            original_text != result and
                len(original_text) < 500):  # 長すぎるテキストは記録しない
            try:
                from manaos_learning import register_correction
                register_correction(
                    tool="pdf_excel",
                    input_data=original_text[:200],
                    raw_output=original_text[:200],
                    corrected_output=result[:200],
                    feedback="needs_review",
                    tags=["自動記録", "ocr_clean"],
                    meta={"method": "clean_ocr_text"}
                )
            except Exception as e:
                self.logger.debug(f"自動記録エラー（無視）: {e}")

        return result

    def clean_numeric_value(self, value: str) -> str:
        """数値セルの値をクリーンアップ（強化版：数字再フォーマット対応）"""
        if not value or not isinstance(value, str):
            return value

        cleaned = str(value).strip()

        # 空の場合はそのまま返す
        if not cleaned:
            return cleaned

        # 数値が含まれていない場合はそのまま返す
        if not re.search(r'\d', cleaned):
            return cleaned

        # セル内がほぼ数字＋カンマ＋ピリオドだけなら、数字と区切り記号以外は全部削除
        numeric_only_pattern = r'^[\d,.\s\-]+$'
        if re.match(numeric_only_pattern, cleaned):
            # 数字と区切り記号以外を削除
            cleaned = re.sub(r'[^\d,.\-]', '', cleaned)

        # カンマが2個以上入ってるのに桁が変な場合の修正
        # 例: "1,23,456" → "123,456" → 数値として再フォーマット
        comma_count = cleaned.count(',')
        if comma_count >= 2:
            # 数字だけを抽出
            digits_only = re.sub(r'[^\d]', '', cleaned)
            if digits_only:
                try:
                    # 数値としてパースして再フォーマット
                    num_value = int(digits_only)
                    cleaned = f"{num_value:,}"
                except ValueError:
                    # 整数化できない場合は小数点を含む可能性がある
                    try:
                        num_value = float(digits_only)
                        # 小数点以下がある場合はそのまま、ない場合はカンマ区切り
                        if num_value == int(num_value):
                            cleaned = f"{int(num_value):,}"
                        else:
                            cleaned = f"{num_value:,.2f}".replace('.00', '')
                    except ValueError:
                        pass  # パースできない場合はそのまま

        # パターン1: 連続するカンマやピリオドを修正
        cleaned = re.sub(r',\s*,\s*', ',', cleaned)  # ", ," → ","
        cleaned = re.sub(r'\.\s*\.\s*', '.', cleaned)  # ". ." → "."
        cleaned = re.sub(r',\s*\.', '.', cleaned)  # ", ." → "."
        cleaned = re.sub(r'\.\s*,', ',', cleaned)  # ". ," → ","

        # パターン2: 数値の後の余分な記号を削除
        # "19.473." → "19.473"
        cleaned = re.sub(r'(\d+\.\d+)\s*\.\s*$', r'\1', cleaned)
        cleaned = re.sub(r'(\d+\.\d+)\s*\.\s*\.\s*', r'\1', cleaned)
        # "1," → "1"（行末または非数字の前、かつ小数点がない場合）
        cleaned = re.sub(r'(\d+)\s*,\s*$', r'\1', cleaned)  # 行末のカンマを削除
        # 非数字の前のカンマを削除
        cleaned = re.sub(r'(\d+)\s*,\s*([^\d\s]|$)', r'\1\2', cleaned)
        # "1, ," → "1"（行末または非数字の前）
        cleaned = re.sub(r'(\d+)\s*,\s*,\s*([^\d\s]|$)', r'\1\2', cleaned)
        # "266, , 662" → "266,662"
        cleaned = re.sub(r'(\d+)\s*,\s*,\s*(\d+)', r'\1,\2', cleaned)

        # パターン3: 数値の前後の余分な記号を削除
        # 括弧、セミコロン、アンダースコア、ハイフン、イコール、パイプを削除
        cleaned = re.sub(r'[\[\]()_;=\|]', '', cleaned)

        # パターン4: 数値の前後の余分なスペースを削除
        cleaned = re.sub(r'^\s+|\s+$', '', cleaned)

        # パターン5: 連続するハイフンを削除
        cleaned = re.sub(r'-\s*-', '', cleaned)

        # パターン6: 数値の後の余分な文字を削除（例: "139]" → "139"）
        cleaned = re.sub(r'(\d+)\s*[\]\)\}\;\:\|]', r'\1', cleaned)

        # パターン7: 複数の数値が1つのセルに残っている場合、最初の数値のみを保持
        # 例: "5,520.15  428,090  5" → "5,520.15"
        num_matches = list(re.finditer(
            r'\d{1,3}(?:,\d{3})*(?:\.\d+)?|\d+\.\d+|\d+', cleaned))
        if len(num_matches) > 1:
            # 最初の数値のみを保持
            first_num = num_matches[0].group()
            # 最初の数値の前後のテキストも保持（必要に応じて）
            cleaned = first_num

        # パターン8: カンマとピリオドの統一（日本の数値表記ルールに従う）
        # カンマは千の位区切り、ピリオドは小数点として統一
        # ルール: 小数点以下が3桁以上の場合、千の位区切りとして扱う
        # 例: "125.281" → "125,281"（小数点以下3桁以上）
        # 例: "125.28" → "125.28"（小数点以下2桁以下は小数点として保持）
        # 例: "125,281" → "125,281"（そのまま）

        # ピリオドが含まれている場合
        if '.' in cleaned:
            # 小数点パターンを検出
            decimal_match = re.search(r'(\d+)\.(\d+)', cleaned)
            if decimal_match:
                integer_part = decimal_match.group(1)
                decimal_part = decimal_match.group(2)

                # 小数点以下が3桁以上の場合、千の位区切りとして扱う
                if len(decimal_part) >= 3:
                    # ピリオドをカンマに変換（例: "125.281" → "125,281"）
                    cleaned = re.sub(r'(\d+)\.(\d{3,})', r'\1,\2', cleaned)
                # 小数点以下が2桁以下の場合、小数点として保持（例: "125.28" → "125.28"）
                # そのまま保持するので何もしない

        # パターン9: 小数点が欠けている問題を修正
        # "00" → "0.00"（数値のみで、小数点がない場合）
        if re.match(r'^\d{2,}$', cleaned) and len(cleaned) >= 2:
            # 2桁以上の数値で、小数点がない場合
            # ただし、千の位区切りの可能性もあるので、慎重に判定
            if len(cleaned) == 2 and cleaned.startswith('0'):
                # "00" → "0.00"
                cleaned = '0.00'
            elif len(cleaned) == 2 and not re.search(r'[,\d]{3,}', cleaned):
                # 2桁の数値で、千の位区切りでない場合、小数点の可能性を考慮
                # ただし、既にカンマがある場合は除外
                pass  # そのまま保持

        # パターン10: 小数点だけが残っている問題を修正
        # "0." → "0" または "0.00"
        if cleaned == '0.' or cleaned.endswith('.0'):
            cleaned = '0'
        elif cleaned.endswith('.'):
            # 末尾がピリオドだけの場合、削除
            cleaned = cleaned.rstrip('.')

        return cleaned.strip()

    def merge_split_numbers(self, row: List[str]) -> List[str]:
        """分割された数値を結合する（例：「266」「662」→「266,662」）"""
        if not row or len(row) < 2:
            return row

        merged_row = []
        i = 0
        while i < len(row):
            current_cell = str(row[i]).strip() if row[i] else ""

            # 現在のセルが数値（カンマを含む可能性あり）の場合
            # カンマを含む数値も対象にする（例：「2,265」）
            if current_cell and (re.match(r'^\d+$', current_cell) or re.match(r'^\d{1,3}(?:,\d{3})+$', current_cell)):
                # 次のセルも数値のみの場合、結合を試行
                if i + 1 < len(row):
                    next_cell = str(row[i + 1]).strip() if row[i + 1] else ""
                    if next_cell and re.match(r'^\d+$', next_cell):
                        # 誤結合を防ぐ：現在のセルが「0」のみの場合は結合しない
                        if current_cell == '0':
                            merged_row.append(current_cell)
                            i += 1
                            continue

                        # 小数点の可能性を考慮（次のセルが2桁以下の場合）
                        if len(next_cell) <= 2:
                            # 小数点として結合（例：「2,265」「83」→「2,265.83」）
                            # 現在のセルにカンマが含まれている場合、または3桁以上の場合
                            if ',' in current_cell or len(re.sub(r',', '', current_cell)) >= 3:
                                combined = f"{current_cell}.{next_cell}"
                                merged_row.append(combined)
                                i += 2
                                continue

                        # 千の位区切りの可能性を考慮
                        # 両方とも3桁以上の場合
                        if len(current_cell) >= 3 and len(next_cell) >= 3:
                            combined = f"{current_cell},{next_cell}"
                            merged_row.append(combined)
                            i += 2
                            continue
                        # 現在のセルが短く、次のセルが長い場合（例：「13」「731」→「13,731」）
                        # ただし、現在のセルが1桁で、次のセルが3桁以上の場合
                        elif len(current_cell) == 1 and len(next_cell) >= 3:
                            # 「1」と「676」のような場合は結合しない（誤結合の可能性が高い）
                            # ただし、現在のセルが2桁以上で、次のセルが3桁以上の場合のみ結合
                            pass  # 結合しない
                        elif len(current_cell) >= 2 and len(current_cell) <= 3 and len(next_cell) >= 3:
                            combined = f"{current_cell},{next_cell}"
                            merged_row.append(combined)
                            i += 2
                            continue

            # 結合しない場合はそのまま追加
            merged_row.append(current_cell)
            i += 1

        return merged_row

    def check_ollama_available(self) -> bool:
        """Ollamaが利用可能かチェック"""
        if not OLLAMA_AVAILABLE:
            return False

        try:
            response = requests.get(f"{self.ollama_url}/api/tags", timeout=2)
            return response.status_code == 200
        except Exception:
            return False

    def ollama_clean_table(self, ocr_text: str, table_type: str = "売上レポート") -> str:
        """Ollamaを使ってOCR結果を整形・表構造化（強化版）"""
        if not self.ollama_enabled or not self.check_ollama_available():
            return ocr_text

        try:
            # より詳細なプロンプトで表構造を認識
            prompt = f"""あなたは表データ抽出の専門家です。以下のOCR結果から表を抽出し、構造化されたCSV形式に変換してください。

【入力OCR結果】
{ocr_text[:3000]}

【タスク】
1. OCR結果から表データを正確に抽出
2. 列の構造を自動認識（ヘッダー行を特定）
3. 数値データを正規化（カンマ区切り、小数点の修正）
4. 行と列を適切に分割
5. CSV形式で出力（カンマ区切り、改行で行を区切る）

【重要なルール】
- ヘッダー行は最初の行に配置
- 数値はカンマ区切りで統一（例: 1,234,567）
- 不要な文字や記号は削除
- 空行は削除
- 表の構造を保持（列数は統一）

【出力形式】
CSV形式のみを出力してください。説明やコメントは不要です。
各行は改行で区切り、各列はカンマで区切ってください。

出力:"""

            response = requests.post(
                f"{self.ollama_url}/api/generate",
                json={
                    "model": self.ollama_model,
                    "prompt": prompt,
                    "stream": False,
                    "options": {
                        "temperature": 0.1,  # 低い温度で一貫性のある出力
                        "top_p": 0.9,
                        "num_predict": 2000  # 最大出力トークン数
                    }
                },
                timeout=60  # タイムアウトを延長
            )

            if response.status_code == 200:
                result = response.json()
                cleaned_text = result.get("response", ocr_text)

                # CSV形式の検証とクリーンアップ
                cleaned_text = self.validate_and_clean_csv(cleaned_text)

                self.logger.info("Ollamaによる整形完了")
                return cleaned_text
            else:
                self.logger.warning(f"Ollama API エラー: {response.status_code}")
                return ocr_text

        except Exception as e:
            self.logger.warning(f"Ollama整形エラー: {e}")
            return ocr_text

    def validate_and_clean_csv(self, csv_text: str) -> str:
        """CSV形式のテキストを検証・クリーンアップ"""
        if not csv_text:
            return csv_text

        lines = csv_text.strip().split('\n')
        cleaned_lines = []

        for line in lines:
            line = line.strip()
            if not line:
                continue

            # CSV形式の検証（カンマが含まれているか）
            if ',' in line or '\t' in line:
                # タブをカンマに変換
                if '\t' in line:
                    line = line.replace('\t', ',')
                cleaned_lines.append(line)
            elif len(line) > 10:  # 長い行は保持（後で処理）
                cleaned_lines.append(line)

        return '\n'.join(cleaned_lines)

    def ollama_extract_table_structure(self, ocr_text: str) -> Dict[str, Any]:
        """Ollamaを使って表構造を自動認識"""
        if not self.ollama_enabled or not self.check_ollama_available():
            return {"type": "single_line", "columns": 0}

        try:
            prompt = f"""以下のOCR結果から表の構造を分析してください。

【OCR結果】
{ocr_text[:2000]}

【タスク】
1. 表の列数を特定
2. ヘッダー行を特定
3. 表のタイプを判定（固定幅/可変幅/区切り文字）
4. 列の区切り位置を特定

【出力形式】
JSON形式で出力してください：
{{
  "columns": 列数（整数）,
  "header_row": ヘッダー行のインデックス（0始まり）,
  "type": "fixed_width" または "delimiter" または "variable",
  "delimiter": 区切り文字（"|" または "," または "tab" または "space"）,
  "column_positions": [列の開始位置のリスト]（固定幅の場合のみ）
}}

出力:"""

            response = requests.post(
                f"{self.ollama_url}/api/generate",
                json={
                    "model": self.ollama_model,
                    "prompt": prompt,
                    "stream": False,
                    "options": {
                        "temperature": 0.1,
                        "num_predict": 500
                    }
                },
                timeout=30
            )

            if response.status_code == 200:
                result = response.json()
                response_text = result.get("response", "")

                # JSONを抽出（```json```ブロックから）
                import json
                json_match = re.search(
                    r'```json\s*(\{.*?\})\s*```', response_text, re.DOTALL)
                if json_match:
                    response_text = json_match.group(1)
                else:
                    # JSONブロックがない場合、最初の{}を探す
                    json_match = re.search(
                        r'\{.*?\}', response_text, re.DOTALL)
                    if json_match:
                        response_text = json_match.group(0)

                try:
                    structure = json.loads(response_text)
                    self.logger.info(f"Ollamaによる表構造認識: {structure}")
                    return structure
                except json.JSONDecodeError:
                    self.logger.warning("OllamaのJSON解析エラー")
                    return {"type": "single_line", "columns": 0}
            else:
                return {"type": "single_line", "columns": 0}
        except Exception as e:
            self.logger.warning(f"Ollama表構造認識エラー: {e}")
            return {"type": "single_line", "columns": 0}

    def merge_split_numbers_improved(self, row: List[str]) -> List[str]:
        """分割された数値を結合する（改善版：より積極的に結合、カンマ区切り数値対応）"""
        if not row or len(row) < 2:
            return row

        import re
        merged_row = []
        i = 0
        while i < len(row):
            current_cell = str(row[i]).strip() if row[i] else ""

            # 現在のセルが数値の一部（数字を含む）の場合
            if current_cell and re.search(r'\d', current_cell):
                # 連続する数値セルを探す
                numeric_parts = [current_cell]
                j = i + 1

                # 次のセルも数値の一部かどうかを確認
                while j < len(row):
                    next_cell = str(row[j]).strip() if row[j] else ""
                    if not next_cell:
                        j += 1
                        continue

                    # 数値の一部（数字、カンマ、ピリオドのみ）かどうか
                    if re.match(r'^[\d,.\s\-]+$', next_cell):
                        # 前のセルと結合できるか判定
                        prev_part = numeric_parts[-1]

                        # カンマ区切りの数値の続きか、小数点の続きか
                        if re.search(r'\d$', prev_part) and re.match(r'^[\d,.\s]+', next_cell):
                            # 数値の続きとして結合
                            next_clean = re.sub(r'[^\d,.\-]', '', next_cell)
                            if next_clean:
                                numeric_parts.append(next_clean)
                                j += 1
                                continue

                    # カンマだけのセル（「,」）の場合、前後の数値と結合
                    if next_cell == ',' or next_cell == '，':
                        # 前が数値で、次も数値の場合のみ結合
                        if j + 1 < len(row):
                            next_next_cell = str(
                                row[j + 1]).strip() if row[j + 1] else ""
                            if re.match(r'^\d+', next_next_cell):
                                numeric_parts.append(next_cell)
                                j += 1
                                continue

                        # 小数点の可能性（2桁以下の数字）
                        if len(re.sub(r'[^\d]', '', next_cell)) <= 2:
                            next_clean = re.sub(r'[^\d]', '0', next_cell) if not re.search(
                                r'\d', next_cell) else re.sub(r'[^\d]', '', next_cell)
                            if next_clean:
                                numeric_parts.append(next_clean)
                                j += 1
                                continue

                    # それ以外の場合は結合しない
                    break

                # 数値パーツを結合
                if len(numeric_parts) > 1:
                    # 各パーツから数字だけを抽出
                    digits_only = ''.join(
                        [re.sub(r'[^\d]', '', part) for part in numeric_parts])

                    # 桁数が異常に大きい場合は結合しない（過剰結合を防ぐ）
                    if len(digits_only) > 15:  # 15桁を超える場合は分割された数値の可能性が高い
                        # 元のセルをそのまま保持
                        merged_row.append(current_cell)
                        i += 1
                    else:
                        # 数字だけを抽出して、3桁ごとにカンマを入れて正規化
                        try:
                            num_value = int(digits_only)
                            # 3桁ごとにカンマ区切りでフォーマット
                            combined = f"{num_value:,}"
                            merged_row.append(combined)
                            i = j
                        except ValueError:
                            # パースできない場合は元のセルを保持
                            merged_row.append(current_cell)
                            i += 1
                else:
                    merged_row.append(current_cell)
                    i += 1
            else:
                # 数値でない場合はそのまま追加
                merged_row.append(current_cell)
                i += 1

        return merged_row if merged_row else row

    def apply_pdf_like_styles(self, worksheet, max_columns: int, max_rows: int):
        """PDFの見た目に近づけるスタイルを適用"""
        if not OPENPYXL_AVAILABLE:
            return

        try:
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            # 罫線のスタイル（細い線）
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            # ヘッダー行のスタイル（背景色、太字）
            header_fill = PatternFill(
                start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            header_font = Font(bold=True, size=10)

            # 通常セルのスタイル
            normal_font = Font(size=9)
            center_alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(
                horizontal='left', vertical='center', wrap_text=True)
            right_alignment = Alignment(
                horizontal='right', vertical='center', wrap_text=True)

            # すべてのセルにスタイルを適用
            for row_idx in range(1, min(max_rows + 1, worksheet.max_row + 1)):
                for col_idx in range(1, min(max_columns + 1, worksheet.max_column + 1)):
                    cell = worksheet.cell(row=row_idx, column=col_idx)

                    # 罫線を適用
                    cell.border = thin_border

                    # 1行目はヘッダー行として扱う
                    if row_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_alignment
                    else:
                        cell.font = normal_font
                        # 数値の場合は右寄せ、テキストの場合は左寄せ
                        cell_value = str(cell.value) if cell.value else ""
                        if cell_value and (cell_value.replace(',', '').replace('.', '').replace('-', '').isdigit() or
                                           any(kw in cell_value for kw in ['金額', '数量', '合計', '計'])):
                            cell.alignment = right_alignment
                        else:
                            cell.alignment = left_alignment

                    # 列幅を自動調整（最小幅を設定）
                    col_letter = cell.column_letter
                    if col_idx <= max_columns:
                        # 列幅を適切に設定（文字数に応じて）
                        current_width = worksheet.column_dimensions[col_letter].width
                        if not current_width or current_width < 10:
                            worksheet.column_dimensions[col_letter].width = 12

            # 行の高さを調整
            for row_idx in range(1, min(max_rows + 1, worksheet.max_row + 1)):
                worksheet.row_dimensions[row_idx].height = 18

            self.logger.info(f"スタイル適用完了: {max_rows}行 x {max_columns}列")

        except Exception as e:
            self.logger.warning(f"スタイル適用エラー: {e}")

    def create_excel_from_position_data(self, worksheet, position_data: List[Dict], ocr_text: str):
        """OCR位置情報を使ってPDFの見た目を完コピしてExcelに配置（改善版）"""
        if not position_data or len(position_data) == 0:
            return

        try:
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from collections import defaultdict
            import statistics

            # 位置情報をY座標（行）でグループ化（改善版：より柔軟なグループ化）
            rows_by_y = defaultdict(list)
            y_positions = []

            for item in position_data:
                # Y座標の中央値を計算
                y_center = (item['top'] + item['bottom']) / 2
                y_positions.append(y_center)
                rows_by_y[y_center].append(item)

            # Y座標の間隔を分析して、適切なグループ化の閾値を決定
            if len(y_positions) > 1:
                y_positions_sorted = sorted(set(y_positions))
                y_intervals = [y_positions_sorted[i+1] - y_positions_sorted[i]
                               for i in range(len(y_positions_sorted)-1)]
                if y_intervals:
                    # 中央値の間隔を使用（外れ値を除外）
                    median_interval = statistics.median(y_intervals)
                    # グループ化の閾値は間隔の50%とする
                    group_threshold = median_interval * 0.5
                else:
                    group_threshold = 10.0
            else:
                group_threshold = 10.0

            # Y座標でグループ化（閾値内のY座標は同じ行とみなす）
            grouped_rows = defaultdict(list)
            sorted_y_positions = sorted(set(y_positions))

            current_group_y = None
            for y_pos in sorted_y_positions:
                if current_group_y is None or abs(y_pos - current_group_y) > group_threshold:
                    current_group_y = y_pos
                grouped_rows[current_group_y].extend(rows_by_y[y_pos])

            # 行をY座標順にソート
            sorted_rows = sorted(grouped_rows.items(), key=lambda x: x[0])

            # X座標の列位置を分析（改善版：実際の列位置を検出）
            all_x_positions = []
            for row_y, items in sorted_rows:
                for item in items:
                    all_x_positions.append(item['left'])

            # X座標の列位置を決定（クラスタリング風のアプローチ）
            if len(all_x_positions) > 0:
                sorted_x_positions = sorted(set(all_x_positions))
                # 列の区切り位置を決定（間隔が大きいところで区切る）
                column_positions = []
                if len(sorted_x_positions) > 1:
                    x_intervals = [sorted_x_positions[i+1] - sorted_x_positions[i]
                                   for i in range(len(sorted_x_positions)-1)]
                    if x_intervals:
                        median_x_interval = statistics.median(x_intervals)
                        # 間隔が中央値の2倍以上の場合、新しい列とみなす
                        threshold_x = median_x_interval * 2.0

                        column_positions.append(sorted_x_positions[0])
                        for i in range(len(sorted_x_positions)-1):
                            if x_intervals[i] > threshold_x:
                                column_positions.append(
                                    sorted_x_positions[i+1])
            else:
                column_positions = []

            # 各行内でX座標順にソートして列に割り当て
            table_data = []
            for row_y, items in sorted_rows:
                # X座標順にソート
                sorted_items = sorted(items, key=lambda x: x['left'])

                # 列位置に基づいてセルを配置
                row_cells = [""] * \
                    max(len(column_positions), len(sorted_items))

                for item in sorted_items:
                    text = item.get('text', '').strip()
                    if not text:
                        continue

                    # 最も近い列位置を決定
                    x_pos = item['left']
                    if column_positions:
                        # 最も近い列位置を見つける
                        min_dist = float('inf')
                        best_col = 0
                        for col_idx, col_x in enumerate(column_positions):
                            dist = abs(x_pos - col_x)
                            if dist < min_dist:
                                min_dist = dist
                                best_col = col_idx

                        # 列位置の範囲内かチェック
                        if best_col < len(row_cells):
                            # 既にセルに値がある場合は結合（スペースで区切る）
                            if row_cells[best_col]:
                                row_cells[best_col] += " " + text
                            else:
                                row_cells[best_col] = text
                    else:
                        # 列位置がない場合は順番に配置
                        for col_idx in range(len(row_cells)):
                            if not row_cells[col_idx]:
                                row_cells[col_idx] = text
                                break

                # 空でないセルのみを保持
                non_empty_cells = [cell for cell in row_cells if cell.strip()]
                if non_empty_cells:
                    table_data.append(non_empty_cells)

            # 最大列数を決定
            max_columns = max(len(row)
                              for row in table_data) if table_data else 0

            # Excelに書き込み
            for row_idx, row_cells in enumerate(table_data, start=1):
                # 列数を統一
                normalized_row = row_cells[:max_columns]
                while len(normalized_row) < max_columns:
                    normalized_row.append("")

                # OCRエラーを修正
                cleaned_row = []
                for cell in normalized_row:
                    cell_str = str(cell) if cell else ""
                    cell_str = self.clean_ocr_text(cell_str)
                    cell_str = self.clean_numeric_value(cell_str)
                    cleaned_row.append(cell_str)

                worksheet.append(cleaned_row)

            # スタイルを適用
            self.apply_pdf_like_styles(worksheet, max_columns, len(table_data))

            # 列幅を位置情報から計算して調整（改善版）
            if position_data and sorted_rows:
                # 各列のX座標範囲を計算
                column_x_ranges = defaultdict(
                    lambda: {'min': float('inf'), 'max': float('-inf')})

                for row_y, items in sorted_rows:
                    sorted_items = sorted(items, key=lambda x: x['left'])
                    for col_idx, item in enumerate(sorted_items):
                        if col_idx < max_columns:
                            x_left = item['left']
                            x_right = item.get(
                                'right', x_left + item.get('width', 0))
                            if col_idx not in column_x_ranges:
                                column_x_ranges[col_idx] = {
                                    'min': x_left, 'max': x_right}
                            else:
                                column_x_ranges[col_idx]['min'] = min(
                                    column_x_ranges[col_idx]['min'], x_left)
                                column_x_ranges[col_idx]['max'] = max(
                                    column_x_ranges[col_idx]['max'], x_right)

                # 列幅を設定
                for col_idx in range(max_columns):
                    col_letter = worksheet.cell(
                        row=1, column=col_idx + 1).column_letter
                    if col_idx in column_x_ranges:
                        col_width_px = column_x_ranges[col_idx]['max'] - \
                            column_x_ranges[col_idx]['min']
                        # ピクセルをExcelの列幅に変換（より正確な換算）
                        # 1ピクセル ≈ 0.75ポイント、Excelの列幅は文字数ベース
                        # 7.5ピクセル ≈ 1文字幅
                        excel_width = max(8, min(60, col_width_px / 7.5))
                        worksheet.column_dimensions[col_letter].width = excel_width
                    else:
                        worksheet.column_dimensions[col_letter].width = 12

            self.logger.info(
                f"位置情報ベースでExcel作成完了: {len(table_data)}行 x {max_columns}列（列幅自動調整済み）")

        except Exception as e:
            self.logger.warning(f"位置情報ベースExcel作成エラー: {e}")
            import traceback
            self.logger.debug(traceback.format_exc())

    def detect_anomalies(self, rows: List[List[str]]) -> Dict[str, Any]:
        """異常値検出エンジン（AI監査）"""
        anomalies = {
            'warnings': [],
            'errors': [],
            'suspicious_rows': []
        }

        if not rows or len(rows) < 2:
            return anomalies

        import re

        # ヘッダー行を検出
        header_row_idx = None
        for i, row in enumerate(rows[:5]):  # 最初の5行をチェック
            row_text = ' '.join(str(cell) for cell in row if cell)
            header_keywords = ['金額', '数量', '合計', '計', '品名']
            if sum(1 for kw in header_keywords if kw in row_text) >= 2:
                header_row_idx = i
                break

        if header_row_idx is None:
            return anomalies

        # データ行を分析
        data_rows = rows[header_row_idx + 1:]

        # 列ごとの型を推定
        column_types = {}
        if len(rows) > header_row_idx + 1:
            sample_row = rows[header_row_idx + 1]
            for col_idx in range(len(sample_row)):
                column_data = [str(row[col_idx]) if col_idx < len(row) else ""
                               for row in data_rows[:20]]  # 最初の20行をサンプル

                # 数値列かどうかを判定
                numeric_count = 0
                for cell in column_data:
                    if cell and re.search(r'\d+[,\d]*\.?\d*', cell):
                        numeric_count += 1

                if numeric_count >= len([c for c in column_data if c]) * 0.7:
                    column_types[col_idx] = 'numeric'

        # 異常値チェック
        for row_idx, row in enumerate(data_rows):
            row_anomalies = []

            for col_idx, cell in enumerate(row):
                if not cell or not str(cell).strip():
                    continue

                cell_str = str(cell).strip()

                # 数値列の場合
                if col_idx in column_types and column_types[col_idx] == 'numeric':
                    # 数字を抽出
                    numbers = re.findall(r'\d+[,\d]*\.?\d*', cell_str)
                    if numbers:
                        for num_str in numbers:
                            # カンマを除去して数値化
                            try:
                                num_clean = num_str.replace(
                                    ',', '').replace('.', '')
                                if num_clean:
                                    num_value = float(num_clean)

                                    # 異常値チェック
                                    # 1. 金額が1,000倍変（例: 1000円が1000000円になっている）
                                    if num_value > 1000000:
                                        row_anomalies.append(
                                            f"列{col_idx+1}: 異常に大きな数値 ({num_str})")

                                    # 2. 負の値（金額が負になることは通常ない）
                                    if num_value < 0:
                                        row_anomalies.append(
                                            f"列{col_idx+1}: 負の値 ({num_str})")

                                    # 3. 数量が100を超える（通常の売上では稀）
                                    if '数量' in str(rows[header_row_idx][col_idx] if col_idx < len(rows[header_row_idx]) else ""):
                                        if num_value > 100:
                                            row_anomalies.append(
                                                f"列{col_idx+1}: 数量が異常に大きい ({num_str})")
                            except ValueError:
                                pass

            if row_anomalies:
                anomalies['suspicious_rows'].append({
                    'row_index': header_row_idx + 1 + row_idx,
                    'anomalies': row_anomalies
                })
                anomalies['warnings'].extend(row_anomalies)

        return anomalies

    def detect_column_positions(self, lines: List[str]) -> List[int]:
        """固定幅フォントの表を想定した列位置を自動検出"""
        if not lines:
            return []

        import re
        from collections import Counter

        # 各行の文字位置を分析（数値やキーワードの位置を記録）
        position_markers = []  # 各行の重要な位置（数値の開始位置など）

        for line in lines[:20]:  # 最初の20行を分析
            if not line.strip():
                continue

            # 数値の開始位置を記録
            for match in re.finditer(r'\d+[,\d]*\.?\d*', line):
                position_markers.append(match.start())

            # キーワードの位置を記録
            keywords = ['金額', '数量', '合計', '計', '品名', '現金', 'ブリカ', 'クレジット']
            for keyword in keywords:
                pos = line.find(keyword)
                if pos >= 0:
                    position_markers.append(pos)

        if not position_markers:
            return []

        # 位置の分布を分析（最も多く出現する位置を列の区切りとみなす）
        position_counter = Counter(position_markers)
        # 10文字以内の誤差を許容してグループ化
        threshold = 10
        column_positions = []
        sorted_positions = sorted(
            position_counter.items(), key=lambda x: x[1], reverse=True)

        for pos, count in sorted_positions:
            # 既存の列位置と近すぎる場合はスキップ
            if any(abs(pos - cp) < threshold for cp in column_positions):
                continue
            # 出現回数が3回以上の場合のみ列位置として採用
            if count >= 3:
                column_positions.append(pos)

        return sorted(column_positions)

    def parse_fixed_width_table(self, line: str, column_positions: List[int]) -> List[str]:
        """固定幅フォントの表を解析して列に分割"""
        if not column_positions:
            return [line]

        cells = []
        prev_pos = 0

        for col_pos in column_positions:
            if col_pos > prev_pos:
                cell = line[prev_pos:col_pos].strip()
                cells.append(cell)
                prev_pos = col_pos

        # 最後の列
        if prev_pos < len(line):
            cell = line[prev_pos:].strip()
            cells.append(cell)

        return cells if cells else [line]

    def detect_table_structure(self, ocr_text: str) -> Dict[str, Any]:
        """表の構造を自動検出（列数、行数、列の区切り方法など）"""
        import re
        from collections import Counter

        lines = [line.strip() for line in ocr_text.split('\n') if line.strip()]

        if not lines:
            return {'type': 'unknown', 'columns': 0, 'rows': 0}

        # 1. パイプ区切りの表かどうか
        pipe_lines = [line for line in lines if '|' in line]
        if len(pipe_lines) >= len(lines) * 0.5:  # 50%以上がパイプ区切り
            # パイプ区切りの列数を分析
            column_counts = [len(line.split('|')) for line in pipe_lines]
            if column_counts:
                col_counter = Counter(column_counts)
                most_common_cols = col_counter.most_common(1)[0][0]
                return {
                    'type': 'pipe_delimited',
                    'columns': most_common_cols,
                    'rows': len(lines),
                    'delimiter': '|'
                }

        # 2. 固定幅フォントの表かどうか（列位置を検出）
        column_positions = self.detect_column_positions(lines)
        if column_positions and len(column_positions) >= 3:
            return {
                'type': 'fixed_width',
                'columns': len(column_positions) + 1,
                'rows': len(lines),
                'column_positions': column_positions
            }

        # 3. スペース区切りの表かどうか
        # 数値が多く含まれる行を分析
        numeric_lines = []
        for line in lines:
            numbers = re.findall(r'\d+[,\d]*\.?\d*', line)
            if len(numbers) >= 3:  # 3つ以上の数値が含まれる行
                numeric_lines.append(line)

        if numeric_lines:
            # スペースで分割した場合の列数を分析
            space_split_counts = []
            for line in numeric_lines:
                # 3文字以上のスペースで分割
                parts = re.split(r'\s{3,}', line)
                if len(parts) >= 3:
                    space_split_counts.append(len(parts))

            if space_split_counts:
                col_counter = Counter(space_split_counts)
                most_common_cols = col_counter.most_common(1)[0][0]
                return {
                    'type': 'space_delimited',
                    'columns': most_common_cols,
                    'rows': len(lines),
                    'delimiter': r'\s{3,}'
                }

        # 4. その他（1行にまとまっている可能性）
        if len(lines) <= 1 and len(ocr_text) > 100:
            return {
                'type': 'single_line',
                'columns': 0,  # 未確定
                'rows': 1,
                'needs_parsing': True
            }

        # デフォルト: 不明
        return {
            'type': 'unknown',
            'columns': 0,
            'rows': len(lines)
        }

    def score_ocr_text(self, ocr_text: str, confidence: float = 0.0) -> float:
        """OCR結果の品質を評価（ドメイン寄りスコアリング、0-100）"""
        if not ocr_text or not ocr_text.strip():
            return 0.0

        import re

        total_len = len(ocr_text)

        # 1) 置換文字（）が少ないほど高スコア
        bad_chars = ocr_text.count('\ufffd') + ocr_text.count('\uFFFD')
        score = 100.0 - bad_chars * 5.0

        # 2) 有効文字率（日本語 + 英数字 + 基本記号）
        valid = re.sub(
            r'[^\w\s\u3040-\u30FF\u4E00-\u9FAF\u3400-\u4DBF.,;:()（）【】「」『』〈〉《》\d\-]+',
            '',
            ocr_text
        )
        valid_ratio = len(valid) / max(total_len, 1)
        score += valid_ratio * 30.0

        # 3) 数値のヒット率（レシートなら数字多めが自然）
        numbers = re.findall(r'\d+[,\d]*\.?\d*', ocr_text)
        if numbers:
            score += min(len(numbers), 50)  # 数量に応じて +α

        # 4) キーワードの含有（金額、数量など）
        keywords = ['金額', '数量', '合計', '計', 'POS', '売上', '品名']
        keyword_count = sum(1 for keyword in keywords if keyword in ocr_text)
        score += min(keyword_count * 5, 20)  # 最大20点

        # 5) 行数の適切さ（多すぎず少なすぎず）
        lines = [l for l in ocr_text.split('\n') if l.strip()]
        line_count = len(lines)
        if 5 <= line_count <= 50:
            score += 20  # 最大20点
        elif line_count < 5:
            score += line_count * 4
        else:
            score += (50 / line_count) * 20

        # 6) エンジンの信頼度を補正（Vision API の confidence など）
        if confidence > 0:
            score += confidence * 0.3  # 信頼度をスコアに反映

        return min(max(score, 0.0), 100.0)

    def _is_table_like_layout(self, text: str) -> bool:
        """レイアウトが表っぽいかどうかを簡易判定"""
        if not text or not text.strip():
            return False

        import re

        lines = [l for l in text.split('\n') if l.strip()]
        line_count = len(lines)

        # 行数が少なすぎる場合は表ではない
        if line_count < 3:
            return False

        # 各行に数字 or カンマが含まれる行数をカウント
        numeric_line_count = 0
        for line in lines:
            if re.search(r'\d+[,\d]*\.?\d*', line):
                numeric_line_count += 1

        # 行の50%以上に数字が含まれていれば表っぽい
        numeric_ratio = numeric_line_count / line_count if line_count > 0 else 0

        return numeric_ratio >= 0.5 and line_count >= 5

    def evaluate_ocr_quality(self, ocr_text: str) -> float:
        """OCR結果の品質を評価（後方互換性のため残す）"""
        return self.score_ocr_text(ocr_text)

    def is_header_row(self, row: List[str]) -> bool:
        """ヘッダー行かどうかを判定"""
        if not row or len(row) == 0:
            return False

        # 空行は除外
        if not any(cell.strip() for cell in row if cell):
            return False

        # ヘッダー行の特徴
        header_keywords = ['金額', '数量', '合計', '計',
                           '日', '月', '年', '担当', 'POS', '売上', '在庫']
        row_text = ' '.join(str(cell) for cell in row if cell)

        # キーワードが含まれているか
        keyword_count = sum(
            1 for keyword in header_keywords if keyword in row_text)

        # 数字が少ない（ヘッダーは数字が少ない）
        digit_count = sum(
            1 for cell in row if cell and re.search(r'\d', str(cell)))
        total_cells = sum(1 for cell in row if cell and str(cell).strip())

        # ヘッダーの可能性が高い
        if keyword_count >= 2 and digit_count < total_cells * 0.3:
            return True

        # 最初の数行で、数字が少ない場合はヘッダーの可能性
        if keyword_count >= 1 and digit_count < total_cells * 0.2:
            return True

        return False

    def split_number_and_text(self, text: str) -> List[str]:
        """数値とテキストを分離（改善版：数値はまとめて保持）"""
        # 数値パターン（カンマ区切り、小数点含む、より柔軟に）
        # カンマ区切りの数値、小数点付き数値、単純な数値を認識
        number_pattern = r'\d{1,3}(?:,\d{3})*(?:\.\d+)?|\d+\.\d+|\d+'

        parts = []
        last_end = 0

        for match in re.finditer(number_pattern, text):
            # 数値の前のテキスト
            if match.start() > last_end:
                prefix = text[last_end:match.start()].strip()
                if prefix:
                    parts.append(prefix)

            # 数値（まとめて保持）
            parts.append(match.group())
            last_end = match.end()

        # 残りのテキスト
        if last_end < len(text):
            suffix = text[last_end:].strip()
            if suffix:
                parts.append(suffix)

        return parts if parts else [text]

    def parse_ocr_line(self, line: str) -> List[str]:
        """OCRテキスト行を解析して表形式に変換（改善版：数値位置基準の列検出追加）"""
        if not line.strip():
            return []

        # OCRエラーを修正（ただし、数値の結合を防ぐため、分割後に適用）
        original_line = line.strip()
        line = original_line  # まずは元の行を使用

        # パターン0: 数値の位置を基準にした列の自動検出（改善版：数値自体を分割しない）
        import re
        from collections import Counter

        # 数値の範囲（開始位置と終了位置）を全て記録
        number_ranges = []
        for match in re.finditer(r'\d{1,3}(?:,\d{3})*(?:\.\d+)?|\d+\.\d+|\d+', line):
            number_ranges.append((match.start(), match.end(), match.group()))

        # キーワードの位置も記録
        keyword_positions = []
        keywords = ['金額', '数量', '合計', '計', '現金',
                    'ブリカ', 'クレジット', '品名', 'F/O', 'f/o']
        for keyword in keywords:
            for match in re.finditer(re.escape(keyword), line, re.IGNORECASE):
                keyword_positions.append(match.start())

        # 数値が3つ以上ある場合、数値の間のスペースを分析して列の区切りを検出
        if len(line) >= 50 and len(number_ranges) >= 3:
            # 数値の間のスペース位置を記録（数値の終了位置から次の数値の開始位置まで）
            space_ranges = []
            for i in range(len(number_ranges) - 1):
                num1_end = number_ranges[i][1]
                num2_start = number_ranges[i + 1][0]
                space_length = num2_start - num1_end
                # スペースが1文字以上の場合に列の区切りとみなす（改善：1文字でも検出）
                if space_length >= 1:
                    space_ranges.append((num1_end, num2_start, space_length))

            if space_ranges:
                # スペースの長さの分布を分析
                space_lengths = [sp[2] for sp in space_ranges]
                space_counter = Counter(space_lengths)
                # 最も多いスペース長を採用（ただし1-20文字の範囲、改善：1文字でも検出）
                valid_spaces = [
                    length for length, count in space_counter.most_common() if 1 <= length <= 20]

                if valid_spaces:
                    common_space_length = valid_spaces[0]
                    # 列の区切り位置を決定（数値の終了位置を基準に）
                    column_positions = [0]
                    for num_start, num_end, num_text in number_ranges:
                        # 数値の終了位置を列の区切りとして追加（ただし、既存の位置と近すぎる場合はスキップ）
                        if not any(abs(num_end - cp) < 5 for cp in column_positions):
                            column_positions.append(num_end)

                    # 最後の数値の終了位置も追加
                    if number_ranges:
                        last_num_end = number_ranges[-1][1]
                        if last_num_end not in column_positions:
                            column_positions.append(last_num_end)

                    # 列位置に基づいて分割（数値自体は分割しない）
                    if len(column_positions) >= 2:
                        split_cells = []
                        for i in range(len(column_positions)):
                            start = column_positions[i]
                            end = column_positions[i + 1] if i + \
                                1 < len(column_positions) else len(line)
                            cell = line[start:end].strip()
                            if cell:
                                # 数値のクリーンアップは行わない（数値が結合されるのを防ぐ）
                                split_cells.append(cell)

                        # 列数が適切な範囲（2-15列）の場合のみ採用
                        if len(split_cells) >= 2 and len(split_cells) <= 15:
                            return split_cells

        # パターン1: パイプ（|）で分割（最も一般的な表の区切り）
        if '|' in line:
            parts = [part.strip() for part in line.split('|') if part.strip()]
            if len(parts) > 1:
                # パイプで分割した後、各パート内をさらに分割する
                final_cells = []
                for part in parts:
                    # まず複数スペース（3文字以上）で分割を試行（最も確実な区切り）
                    sub_parts = [p.strip()
                                 for p in re.split(r'\s{3,}', part) if p.strip()]

                    # 3文字以上のスペースで分割できた場合は、それを優先
                    if len(sub_parts) > 1:
                        refined_parts = []
                        for sub in sub_parts:
                            # 各サブパート内にさらにスペースがある場合、2文字以上のスペースで分割
                            if re.search(r'\s{2,}', sub):
                                further_parts = [p.strip() for p in re.split(
                                    r'\s{2,}', sub) if p.strip()]
                                refined_parts.extend(further_parts)
                            else:
                                refined_parts.append(sub)
                        final_cells.extend(refined_parts)
                    else:
                        # 3文字以上のスペースがない場合、2文字以上のスペースで分割
                        sub_parts = [p.strip() for p in re.split(
                            r'\s{2,}', part) if p.strip()]
                        if len(sub_parts) > 1:
                            final_cells.extend(sub_parts)
                        else:
                            # 2文字以上のスペースもない場合、そのまま追加
                            final_cells.append(part)

                return [c for c in final_cells if c]  # 空要素を除外

        # パターン2: タブで分割
        if '\t' in line:
            parts = [part.strip() for part in line.split('\t')]
            if len(parts) > 1:
                return [p for p in parts if p]

        # パターン3: 複数スペース（4文字以上）で分割（より厳密に）
        parts = [part.strip()
                 for part in re.split(r'\s{4,}', line) if part.strip()]
        if len(parts) > 1:
            return parts  # 既に十分な分割なので、追加分割はしない

        # パターン4: 複数スペース（3文字以上）で分割
        parts = [part.strip()
                 for part in re.split(r'\s{3,}', line) if part.strip()]
        if len(parts) > 1:
            return parts  # 既に十分な分割なので、追加分割はしない

        # パターン5: 複数スペース（2文字以上）で分割
        parts = [part.strip()
                 for part in re.split(r'\s{2,}', line) if part.strip()]
        if len(parts) > 1:
            return parts

        # パターン6: 数字と文字の境界で分割を試行
        parts = [part.strip() for part in re.split(
            r'(?<=\d)\s+(?=[A-Za-z])|(?<=[A-Za-z])\s+(?=\d)', line) if part.strip()]
        if len(parts) > 1:
            return parts

        # パターン7: 数値とテキストを分離
        if re.search(r'\d', line):
            parts = self.split_number_and_text(line)
            if len(parts) > 1:
                return parts

        # パターン8: キーワード行の分割（「金額 数量 金額 数量」など）
        # キーワードで分割を試行
        keyword_pattern = r'(金額|数量|粗利|F/O|f/o|FノO|合計|計|現金|ブリカ|クレジット|品名|商品)'
        keyword_matches = list(re.finditer(keyword_pattern, line))
        if len(keyword_matches) >= 2:
            # キーワードの位置で分割
            cells = []
            for i, match in enumerate(keyword_matches):
                start = match.start()
                end = keyword_matches[i + 1].start() if i + \
                    1 < len(keyword_matches) else len(line)
                cell = line[start:end].strip()
                if cell:
                    cells.append(cell)
            if len(cells) >= 2:
                return cells

        # パターン9: 数値と数値の間で分割（「1,234 56 7,890 12」など）
        # 数値パターンで分割
        number_pattern = r'\d+[,\d]*\.?\d*'
        number_matches = list(re.finditer(number_pattern, line))
        if len(number_matches) >= 2:
            # 数値の間のスペースで分割
            cells = []
            for i, match in enumerate(number_matches):
                start = match.start()
                end = number_matches[i + 1].start() if i + \
                    1 < len(number_matches) else len(line)
                cell = line[start:end].strip()
                if cell:
                    # 数値だけのセルは結合しない
                    if re.match(r'^\d+[,\d]*\.?\d*$', cell):
                        cells.append(cell)
                    else:
                        # 数値とテキストが混在する場合は分割
                        parts = re.split(r'\s+', cell)
                        cells.extend([p for p in parts if p])
            if len(cells) >= 2:
                return cells

        # パターン10: すべてを1列として返す（分割できない場合）
        # 最後にOCRエラーを修正（各セルに対して適用）
        result = [line] if line else []
        if result:
            result = [self.clean_ocr_text(cell) for cell in result]
        return result

    def get_skew_angle(self, image_path: str) -> float:
        """画像の傾き角度を検出（OpenCV使用）"""
        if not OPENCV_AVAILABLE:
            return 0.0

        try:
            img = cv2.imread(str(image_path))
            if img is None:
                return 0.0

            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            # エッジ検出
            edges = cv2.Canny(gray, 50, 150, apertureSize=3)
            # 直線検出
            lines = cv2.HoughLines(edges, 1, np.pi / 180, 200)

            if lines is None or len(lines) == 0:
                return 0.0

            angles = []
            for line in lines:
                rho, theta = line[0]
                angle = (theta * 180 / np.pi) - 90
                # -45度から45度の範囲のみ考慮
                if -45 <= angle <= 45:
                    angles.append(angle)

            if not angles:
                return 0.0

            # 中央値で角度を決定（外れ値に強い）
            angles = sorted(angles)
            median_angle = angles[len(angles) // 2]

            # 1度未満の傾きは無視
            if abs(median_angle) < 1.0:
                return 0.0

            return median_angle
        except Exception as e:
            self.logger.warning(f"傾き検出エラー: {e}")
            return 0.0

    def correct_skew(self, image_path: str, angle: float) -> str:
        """画像の傾きを補正"""
        if not OPENCV_AVAILABLE or abs(angle) < 0.5:
            return str(image_path)

        try:
            img = cv2.imread(str(image_path))
            if img is None:
                return str(image_path)

            # 画像の中心を取得
            h, w = img.shape[:2]
            center = (w // 2, h // 2)

            # 回転行列を計算
            M = cv2.getRotationMatrix2D(center, angle, 1.0)

            # 回転後の画像サイズを計算
            cos = np.abs(M[0, 0])
            sin = np.abs(M[0, 1])
            new_w = int((h * sin) + (w * cos))
            new_h = int((h * cos) + (w * sin))

            # 回転行列を調整
            M[0, 2] += (new_w / 2) - center[0]
            M[1, 2] += (new_h / 2) - center[1]

            # 画像を回転
            rotated = cv2.warpAffine(img, M, (new_w, new_h),
                                     flags=cv2.INTER_CUBIC,
                                     borderMode=cv2.BORDER_REPLICATE)

            # 補正済み画像を保存
            corrected_path = str(image_path).replace('.png', '_corrected.png')
            cv2.imwrite(corrected_path, rotated)

            return corrected_path
        except Exception as e:
            self.logger.warning(f"傾き補正エラー: {e}")
            return str(image_path)

    def extract_images_with_ocr(self, pdf_path: str) -> Dict[str, Any]:
        """PDFから画像抽出とOCR処理（ページ単位でOCRを実行、複数エンジン対応）"""
        if not OCR_AVAILABLE or not PYMUPDF_AVAILABLE:
            return {"success": True, "data": {"images": [], "total_images": 0, "ocr_text": "", "pages": []}}

        try:
            images_data = {
                "images": [],
                "total_images": 0,
                "ocr_text": "",
                "pages": []  # ページごとのOCR結果
            }

            # Vision API Engineの初期化（利用可能な場合）
            vision_engine = None
            try:
                sys.path.append('/root/ocr_system/engines')
                from vision_api_engine import VisionAPIEngine
                vision_engine = VisionAPIEngine()
                if vision_engine.available and vision_engine.check_quota()['quota_ok']:
                    self.logger.info("Vision API Engine利用可能: 複数OCRエンジンで精度向上")
                else:
                    vision_engine = None
            except Exception as e:
                self.logger.debug(
                    f"Vision API Engine初期化失敗（Tesseractのみ使用）: {e}")
                vision_engine = None

            # PaddleOCRの初期化（利用可能な場合）
            paddle_ocr = None
            if PADDLEOCR_AVAILABLE:
                try:
                    # 新しいバージョンではshow_logパラメータが使えない
                    try:
                        paddle_ocr = PaddleOCR(
                            use_textline_orientation=True, lang='japan')
                    except TypeError:
                        # 古いバージョンの場合
                        try:
                            paddle_ocr = PaddleOCR(
                                use_angle_cls=True, lang='japan')
                        except Exception:
                            paddle_ocr = PaddleOCR(lang='japan')
                    self.logger.info("PaddleOCR利用可能: 日本語OCR精度向上（位置情報付き）")
                except Exception as e:
                    self.logger.info("PaddleOCR利用可能: 日本語OCR精度向上（位置情報付き）")
                except Exception as e:
                    self.logger.debug(f"PaddleOCR初期化失敗: {e}")
                    paddle_ocr = None

            # PDFファイルを開く（エラーハンドリング強化）
            try:
                doc = fitz.open(pdf_path)
            except Exception as e:
                self.logger.error(f"PDFファイルを開けませんでした: {e}")
                return {"success": True, "data": {"images": [], "total_images": 0, "ocr_text": "", "pages": []}}

            # ページごとにOCR処理（推奨方法）
            total_pages = len(doc)
            self.logger.info(f"ページごとにOCR処理を開始（全{total_pages}ページ、より正確な結果を得るため）")

            for page_num in range(total_pages):
                # 進捗表示
                progress = (page_num + 1) / total_pages * 100
                self.logger.info(
                    f"📄 OCR処理中: {page_num + 1}/{total_pages}ページ ({progress:.1f}%)")
                try:
                    page = doc.load_page(page_num)

                    # ページ全体を画像として変換（解像度を調整：位置情報取得のため適度な解像度に）
                    # 位置情報取得のため、メモリを考慮して解像度を調整
                    # 1.5倍の解像度で位置情報とメモリのバランスを取る（メモリ節約優先）
                    mat = fitz.Matrix(1.5, 1.5)  # 1.5倍の解像度（メモリ節約、位置情報も取得可能）
                    pix = page.get_pixmap(matrix=mat)

                    page_img_path = self.temp_dir / \
                        f"page_{page_num+1}_full.png"
                    pix.save(page_img_path)
                    # メモリ解放
                    del pix
                    pix = None

                    # 傾き補正（OpenCV使用）
                    if OPENCV_AVAILABLE:
                        angle = self.get_skew_angle(str(page_img_path))
                        if abs(angle) >= 0.5:
                            self.logger.info(
                                f"ページ {page_num + 1}: 傾き検出 {angle:.2f}度、補正中...")
                            page_img_path = Path(
                                self.correct_skew(str(page_img_path), angle))

                    # 複数のOCR設定を試して、最良の結果を選択
                    from PIL import Image, ImageEnhance, ImageFilter
                    img = Image.open(page_img_path)

                    # 画像前処理: 複数のパターンを試して最良の結果を選択
                    img_variants = []

                    # パターン1: グレースケール + コントラスト + シャープネス
                    img1 = img.convert('L') if img.mode != 'L' else img.copy()
                    enhancer = ImageEnhance.Contrast(img1)
                    img1 = enhancer.enhance(1.5)  # コントラストを上げる
                    enhancer = ImageEnhance.Sharpness(img1)
                    img1 = enhancer.enhance(1.8)  # シャープネスを上げる
                    img_variants.append(('contrast_sharp', img1))

                    # パターン2: グレースケール + 二値化風（コントラスト極大）
                    img2 = img.convert('L') if img.mode != 'L' else img.copy()
                    enhancer = ImageEnhance.Contrast(img2)
                    img2 = enhancer.enhance(2.0)  # より高いコントラスト
                    enhancer = ImageEnhance.Brightness(img2)
                    img2 = enhancer.enhance(1.1)  # 明るさを少し上げる
                    img_variants.append(('high_contrast', img2))

                    # パターン3: 元の画像（カラー/グレー）
                    img3 = img.copy()
                    if img3.mode != 'L':
                        img3 = img3.convert('L')
                    img_variants.append(('original', img3))

                    # ページのレイアウトを簡易判定してPSM優先順を決定
                    # まず軽量なOCRでプレビューを取得（レイアウト判定用）
                    preview_text = ""
                    try:
                        preview_text = pytesseract.image_to_string(
                            img_variants[0][1] if img_variants else img,
                            lang='jpn+eng',
                            config='--psm 6 --oem 3'
                        )
                    except:
                        pass

                    # レイアウト判定：表っぽいかどうか
                    is_table_like = self._is_table_like_layout(preview_text)

                    # レイアウトに応じてPSM優先順を変更
                    if is_table_like:
                        # 表系：PSM 4 を優先（単一列の可変サイズテキスト）
                        psm_modes = [
                            ('psm4', '--psm 4 --oem 3'),
                            ('psm6', '--psm 6 --oem 3'),
                            ('psm11', '--psm 11 --oem 3'),
                        ]
                    else:
                        # タイトル/単票系：PSM 6 を優先
                        psm_modes = [
                            ('psm6', '--psm 6 --oem 3'),
                            ('psm11', '--psm 11 --oem 3'),
                            ('psm4', '--psm 4 --oem 3'),
                        ]

                    # 複数OCRエンジンで実行して比較
                    ocr_results = []

                    # 0. PaddleOCRを最優先で実行（位置情報が確実に取得できる）
                    if paddle_ocr:
                        try:
                            self.logger.info(
                                f"ページ {page_num + 1}: PaddleOCR実行中（位置情報付き）...")
                            paddle_result = paddle_ocr.ocr(
                                str(page_img_path), cls=True)
                            if paddle_result and paddle_result[0]:
                                # PaddleOCRの結果をテキストに変換（位置情報も保存）
                                paddle_text_lines = []
                                paddle_position_data = []  # 位置情報を保存
                                for line in paddle_result[0]:
                                    if line and len(line) >= 2:
                                        # 位置情報（座標）を取得
                                        # [[x1,y1], [x2,y2], [x3,y3], [x4,y4]]
                                        box = line[0]
                                        text_info = line[1]
                                        text = text_info[0] if isinstance(
                                            text_info, (list, tuple)) else str(text_info)
                                        confidence = text_info[1] if isinstance(
                                            text_info, (list, tuple)) and len(text_info) > 1 else 0.0

                                        # 位置情報を計算（左上と右下の座標）
                                        if box and len(box) >= 4:
                                            x_coords = [point[0]
                                                        for point in box]
                                            y_coords = [point[1]
                                                        for point in box]
                                            left = min(x_coords)
                                            top = min(y_coords)
                                            right = max(x_coords)
                                            bottom = max(y_coords)

                                            paddle_position_data.append({
                                                'text': text,
                                                'left': left,
                                                'top': top,
                                                'right': right,
                                                'bottom': bottom,
                                                'width': right - left,
                                                'height': bottom - top,
                                                'confidence': confidence
                                            })

                                        paddle_text_lines.append(text)

                                paddle_text = '\n'.join(paddle_text_lines)
                                # 文字化け修正
                                paddle_text = self.fix_encoding_errors(
                                    paddle_text)

                                # 平均信頼度を計算
                                avg_confidence = 0.0
                                conf_count = 0
                                for line in paddle_result[0]:
                                    if line and len(line) >= 2 and isinstance(line[1], (list, tuple)) and len(line[1]) > 1:
                                        avg_confidence += line[1][1]
                                        conf_count += 1
                                avg_confidence = (
                                    avg_confidence / conf_count * 100) if conf_count > 0 else 0.0

                                quality_score = self.score_ocr_text(
                                    paddle_text, confidence=avg_confidence)
                                ocr_results.append({
                                    'engine': 'paddleocr',
                                    'psm': 'paddle',
                                    'text': paddle_text,
                                    'quality': quality_score,
                                    'confidence': avg_confidence,
                                    'position_data': paddle_position_data  # 位置情報を追加
                                })
                                self.logger.info(
                                    f"ページ {page_num + 1}: PaddleOCR完了（信頼度: {avg_confidence:.1f}%, 位置情報: {len(paddle_position_data)}個）")
                        except Exception as e:
                            self.logger.warning(f"PaddleOCRエラー: {e}")

                    # 1. Tesseract OCR（複数画像前処理 × 複数PSMモード、位置情報取得はスキップ）
                    for img_variant_name, img_variant in img_variants:
                        for psm_name, ocr_config in psm_modes:
                            try:
                                # 文字化けを防ぐため、UTF-8エンコーディングを明示
                                ocr_text = pytesseract.image_to_string(
                                    img_variant,
                                    lang='jpn+eng',
                                    config=ocr_config
                                )

                                # 文字エンコーディングの確認と修正
                                if isinstance(ocr_text, bytes):
                                    try:
                                        ocr_text = ocr_text.decode('utf-8')
                                    except UnicodeDecodeError:
                                        try:
                                            ocr_text = ocr_text.decode(
                                                'shift_jis')
                                        except UnicodeDecodeError:
                                            ocr_text = ocr_text.decode(
                                                'utf-8', errors='replace')

                                # 文字化け修正（event_idは後で設定）
                                ocr_text = self.fix_encoding_errors(
                                    ocr_text, event_id=None)

                                # 結果の品質を評価（改善版スコアリング）
                                quality_score = self.score_ocr_text(ocr_text)
                                ocr_results.append({
                                    'engine': 'tesseract',
                                    'psm': psm_name,
                                    'img_variant': img_variant_name,
                                    'text': ocr_text,
                                    'quality': quality_score
                                })
                            except Exception as e:
                                self.logger.warning(
                                    f"Tesseract OCRエラー ({img_variant_name}/{psm_name}): {e}")
                                continue

                    # 2. PaddleOCRは既に実行済み（上記で最優先実行）

                    # 3. Vision API Engine（利用可能な場合）
                    if vision_engine:
                        try:
                            vision_result = vision_engine.extract_text(
                                str(page_img_path))
                            if vision_result.get('success'):
                                vision_text = vision_result.get('text', '')
                                vision_confidence = vision_result.get(
                                    'confidence', 0)
                                # Vision API の confidence をスコアに反映
                                quality_score = self.score_ocr_text(
                                    vision_text, confidence=vision_confidence)
                                ocr_results.append({
                                    'engine': 'vision_api',
                                    'psm': 'vision',
                                    'text': vision_text,
                                    'quality': quality_score,
                                    'confidence': vision_confidence
                                })
                                self.logger.info(
                                    f"ページ {page_num + 1}: Vision API OCR完了（信頼度: {vision_result.get('confidence', 0):.1f}%）")
                        except Exception as e:
                            self.logger.warning(f"Vision API OCRエラー: {e}")

                    # 最良の結果を選択（位置情報があるものを最優先、次に品質スコア）
                    ocr_position_data = None  # 位置情報を保存
                    if ocr_results:
                        # 位置情報がある結果を優先（PaddleOCRを最優先）
                        results_with_position = [
                            r for r in ocr_results if r.get('position_data')]
                        if results_with_position:
                            # PaddleOCRの位置情報がある結果を優先
                            paddle_results = [r for r in results_with_position if r.get(
                                'engine') == 'paddleocr']
                            if paddle_results:
                                best_result = max(
                                    paddle_results, key=lambda x: x['quality'])
                                self.logger.info(
                                    f"ページ {page_num + 1}: PaddleOCR位置情報付き結果を優先選択")
                            else:
                                # 位置情報がある結果から最良のものを選択
                                best_result = max(
                                    results_with_position, key=lambda x: x['quality'])
                                self.logger.info(
                                    f"ページ {page_num + 1}: 位置情報付きOCR結果を優先選択")
                        else:
                            # 位置情報がない場合は、品質スコアで選択
                        best_result = max(
                            ocr_results, key=lambda x: x['quality'])

                        ocr_text = best_result['text']
                        engine_name = best_result.get('engine', 'tesseract')
                        psm_name = best_result.get('psm', 'unknown')
                        img_variant_name = best_result.get(
                            'img_variant', 'unknown')
                        # 位置情報を取得（TesseractまたはPaddleOCR）
                        ocr_position_data = best_result.get('position_data')
                        self.logger.info(
                            f"ページ {page_num + 1}: 最良のOCR結果 ({engine_name}/{img_variant_name}/{psm_name}), 品質スコア: {best_result['quality']:.2f}, 位置情報: {'あり' if ocr_position_data else 'なし'}")
                    else:
                        # フォールバック: デフォルト設定
                        img_fallback = img_variants[0][1] if img_variants else img
                        ocr_text = pytesseract.image_to_string(
                            img_fallback,
                            lang='jpn+eng',
                            config='--psm 6 --oem 3'
                        )
                        # 文字化け修正（event_idは後で設定）
                        ocr_text = self.fix_encoding_errors(
                            ocr_text, event_id=None)
                        engine_name = 'tesseract'
                        psm_name = 'psm6'
                        img_variant_name = 'fallback'

                    # OCR直後に学習ログを記録
                    event_id = None
                    if self.learning_enabled and self.config.get("auto_record_corrections", False):
                        try:
                            from learning_api import log_event
                            event_id = log_event(
                                tool="pdf_excel",
                                task="ocr_to_table",
                                phase="post_ocr",
                                input_data={
                                    "source": "scan_folder",
                                    "raw_pdf_path": pdf_path,
                                    "page": page_num + 1,
                                    "params": {
                                        "ocr_engine": engine_name,
                                        "psm": psm_name,
                                        "img_variant": img_variant_name,
                                    }
                                },
                                raw_output=ocr_text[:500],  # 最初の500文字
                                tags=["ocr", "raw", f"page_{page_num + 1}"],
                                meta={
                                    "user": "mana",
                                    "notes": "初回OCR結果"
                                }
                            )
                        except Exception as e:
                            self.logger.debug(f"学習ログ記録エラー（無視）: {e}")

                    # Ollamaによる整形（有効な場合）
                    if self.ollama_enabled and len(ocr_text) > 50:
                        try:
                            original_text = ocr_text
                            # 過去の成功事例を取得してfew-shot学習に使う
                            if self.learning_enabled:
                                try:
                                    from learning_api import get_recent_examples
                                    examples = get_recent_examples(
                                        tool="pdf_excel",
                                        task="ocr_to_table",
                                        limit=3,
                                        require_corrected=True
                                    )
                                    # プロンプトに組み込む（ollama_clean_table内で使用）
                                    if examples:
                                        self.logger.debug(
                                            f"過去の成功事例を取得: {len(examples)}件")
                                except Exception as e:
                                    self.logger.debug(f"成功事例取得エラー（無視）: {e}")

                            ocr_text = self.ollama_clean_table(
                                ocr_text, "売上レポート")
                            if ocr_text != original_text:
                                self.logger.info(
                                    f"ページ {page_num + 1}: Ollamaによる整形適用")
                                # Ollama整形後の修正を記録
                                if self.learning_enabled and event_id:
                                    try:
                                        from learning_api import log_correction
                                        log_correction(
                                            tool="pdf_excel",
                                            task="ocr_to_table",
                                            source_event_id=event_id,
                                            raw_output=original_text[:500],
                                            corrected_output=ocr_text[:500],
                                            feedback="corrected",
                                            tags=["ollama_clean",
                                                  "llm_correction"],
                                            meta={
                                                "method": "ollama_clean_table"}
                                        )
                                    except Exception as e:
                                        self.logger.debug(
                                            f"Ollama修正ログ記録エラー（無視）: {e}")
                        except Exception as e:
                            self.logger.warning(f"Ollama整形スキップ: {e}")

                    page_info = {
                        "page_number": page_num + 1,
                        "ocr_text": ocr_text,
                        "image_path": str(page_img_path),
                        "event_id": event_id,  # 学習ログのイベントIDを保存
                        "position_data": ocr_position_data  # OCR位置情報を保存（PDFの見た目再現用）
                    }

                    images_data["pages"].append(page_info)
                    images_data["ocr_text"] += f"\n--- ページ {page_num + 1} ---\n{ocr_text}\n"

                    # メモリ解放: 画像変数をクリア
                    for img_var_name, img_var in img_variants:
                        if hasattr(img_var, 'close'):
                            try:
                                img_var.close()
                            except:
                                pass
                    img_variants = []
                    if 'img' in locals():
                        try:
                            img.close()
                        except:
                            pass
                    img = None
                    pix = None

                    # 処理済み画像ファイルを削除（メモリ節約）
                    try:
                        if page_img_path.exists():
                            page_img_path.unlink()
                    except:
                        pass

                except Exception as page_error:
                    self.logger.warning(
                        f"ページ {page_num + 1} OCRエラー: {page_error}")
                    continue

            # 個別画像の抽出も保持（必要に応じて）
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                image_list = page.get_images()

                for img_num, img in enumerate(image_list):
                    try:
                        xref = img[0]
                        pix = fitz.Pixmap(doc, xref)

                        if pix.n - pix.alpha < 4:  # GRAY or RGB
                            img_filename = f"page_{page_num+1}_img_{img_num+1}.png"
                            img_path = self.temp_dir / img_filename
                            pix.save(img_path)

                            image_info = {
                                "page_number": page_num + 1,
                                "image_number": img_num + 1,
                                "filename": img_filename,
                                "path": str(img_path)
                            }

                            images_data["images"].append(image_info)

                        pix = None

                    except Exception as img_error:
                        self.logger.warning(f"画像抽出エラー: {img_error}")
                        continue

            doc.close()
            images_data["total_images"] = len(images_data["images"])

            self.logger.info(
                f"画像・OCR抽出完了: {images_data['total_images']}個の画像、{len(images_data['pages'])}ページのOCR完了")
            return {"success": True, "data": images_data}

        except Exception as e:
            self.logger.error(f"画像・OCR抽出エラー: {e}")
            return {"success": True, "data": {"images": [], "total_images": 0, "ocr_text": "", "pages": []}}

    def process_pages_one_by_one(self, pdf_path: str, workbook, text_data: Dict, tables_data: Dict) -> Dict[str, Any]:
        """1ページずつOCR処理してExcelシートに追加（メモリ節約版）"""
        try:
            import fitz

            # PDFファイルを開く
            doc = fitz.open(pdf_path)
            total_pages = len(doc)

            self.logger.info(f"1ページずつ処理開始（全{total_pages}ページ）")

            # OCRエンジンの初期化（Tesseractのみ、PaddleOCRはスキップ）
            if not OCR_AVAILABLE:
                return {"success": False, "error": "OCRが利用できません"}

            # ページごとに処理
            for page_num in range(total_pages):
                progress = (page_num + 1) / total_pages * 100
                self.logger.info(
                    f"📄 処理中: {page_num + 1}/{total_pages}ページ ({progress:.1f}%)")

                try:
                    page = doc.load_page(page_num)

                    # ページを画像に変換（超低解像度でメモリ節約）
                    # 0.8倍の解像度でメモリ使用量を約36%削減（0.8^2 = 0.64）
                    mat = fitz.Matrix(0.8, 0.8)
                    pix = page.get_pixmap(matrix=mat)

                    page_img_path = self.temp_dir / \
                        f"page_{page_num+1}_temp.png"
                    pix.save(page_img_path)
                    del pix
                    pix = None

                    # Tesseract OCR実行（位置情報なし、メモリ節約）
                    from PIL import Image
                    img = Image.open(page_img_path)
                    img_gray = img.convert('L')

                    # Tesseract OCR実行（メモリ節約設定）
                    # --psm 11: スパーステキスト（メモリ節約、軽量）
                    # --oem 1: LSTMエンジンのみ（Legacyより軽量）
                    # 環境変数でメモリ制限を設定（プロセス全体に適用）
                    import os
                    original_omp_thread_limit = os.environ.get(
                        'OMP_THREAD_LIMIT')
                    original_omp_num_threads = os.environ.get(
                        'OMP_NUM_THREADS')
                    os.environ['OMP_THREAD_LIMIT'] = '1'  # スレッド数を1に制限（メモリ節約）
                    os.environ['OMP_NUM_THREADS'] = '1'

                    ocr_text = ""
                    try:
                        # 画像サイズを確認して、大きすぎる場合はリサイズ
                        max_size = 2000  # 最大サイズを2000pxに制限
                        if img_gray.width > max_size or img_gray.height > max_size:
                            ratio = min(max_size / img_gray.width,
                                        max_size / img_gray.height)
                            new_size = (int(img_gray.width * ratio),
                                        int(img_gray.height * ratio))
                            # PILのバージョンに応じてリサンプリング方法を選択
                            try:
                                img_gray = img_gray.resize(
                                    new_size, Image.Resampling.LANCZOS)
                            except AttributeError:
                                # 古いPILバージョンの場合
                                img_gray = img_gray.resize(
                                    new_size, Image.LANCZOS)
                            self.logger.info(f"   画像をリサイズ: {new_size}")

                        # PSM 6: 単一の均一なテキストブロック（表形式に適している）
                        # PSM 11はスパーステキスト用で、表形式には不向き
                        ocr_text = pytesseract.image_to_string(
                            img_gray,
                            lang='jpn+eng',
                            config='--psm 6 --oem 1'  # 表形式に適した設定
                        )
                    except Exception as ocr_error:
                        # OCRエラーが発生しても処理を続行
                        error_msg = str(ocr_error)
                        self.logger.warning(
                            f"   ページ{page_num + 1}のOCR処理でエラー: {error_msg}")
                        # 空のテキストで続行
                        ocr_text = ""
                    finally:
                        # 環境変数を元に戻す
                        if original_omp_thread_limit is not None:
                            os.environ['OMP_THREAD_LIMIT'] = original_omp_thread_limit
                        elif 'OMP_THREAD_LIMIT' in os.environ:
                            del os.environ['OMP_THREAD_LIMIT']
                        if original_omp_num_threads is not None:
                            os.environ['OMP_NUM_THREADS'] = original_omp_num_threads
                        elif 'OMP_NUM_THREADS' in os.environ:
                            del os.environ['OMP_NUM_THREADS']

                    # 文字化け修正
                    if ocr_text:
                        ocr_text = self.fix_encoding_errors(ocr_text)

                    # 画像を閉じる
                    img.close()
                    img_gray.close()
                    img = None
                    img_gray = None

                    # メモリを明示的に解放
                    import gc
                    gc.collect()

                    # シート名を作成
                    sheet_name = f"ページ{page_num + 1}"
                    if len(sheet_name) > 31:
                        sheet_name = f"P{page_num + 1}"

                    # Excelシートを作成
                    page_sheet = workbook.create_sheet(sheet_name)

                    # OCRテキストが空の場合は空行を追加
                    if not ocr_text or not ocr_text.strip():
                        page_sheet.append(["OCR処理でテキストを取得できませんでした"])
                        self.logger.warning(
                            f"   ページ{page_num + 1}: OCRテキストが空です")
                    else:
                        # OCRテキストを行ごとに解析
                        import re
                        # 改行で分割（空行も保持）
                        # まず、改行文字で分割
                        lines = ocr_text.split('\n')
                        # さらに、\rで分割された行も処理
                        all_lines = []
                        for line in lines:
                            all_lines.extend(line.split('\r'))

                        # 各行を処理
                        processed_lines = 0
                        for line in all_lines:
                            line = line.strip()
                            if not line:
                                # 空行の場合は空の行を追加
                                page_sheet.append([])
                                continue

                            # 行をセルに分割
                            cells = self.parse_ocr_line(line)
                            if cells and len(cells) > 0:
                                # 数値の分割を修正
                                cells = self.merge_split_numbers_improved(
                                    cells)
                                page_sheet.append(cells)
                                processed_lines += 1
                            else:
                                # セルに分割できない場合は、そのまま1セルとして追加
                                page_sheet.append([line])
                                processed_lines += 1

                        self.logger.info(
                            f"   ページ{page_num + 1}: {processed_lines}行を処理しました（総行数: {len(all_lines)}、空でない行: {len([l for l in all_lines if l.strip()])}）")

                    # PDF風のスタイルを適用
                    if page_sheet.max_row > 0 and page_sheet.max_column > 0:
                        self.apply_pdf_like_styles(
                            page_sheet, page_sheet.max_column, page_sheet.max_row)

                    # 一時画像ファイルを削除
                    try:
                        if page_img_path.exists():
                            page_img_path.unlink()
                    except:
                        pass

                    # ページ処理後にメモリを解放
                    import gc
                    gc.collect()

                    self.logger.info(f"✅ ページ{page_num + 1}処理完了")

                except Exception as page_error:
                    self.logger.warning(
                        f"ページ{page_num + 1}処理エラー: {page_error}")
                    continue

            doc.close()
            return {"success": True}

        except Exception as e:
            self.logger.error(f"1ページずつ処理エラー: {e}")
            return {"success": False, "error": str(e)}

    def create_excel_from_data(self, text_data: Dict, tables_data: Dict,
                               images_data: Dict, output_path: str) -> Dict[str, Any]:
        """抽出データからExcelファイル生成（メモリ最適化版）"""
        try:
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxlが利用できません"}

            workbook = openpyxl.Workbook()

            # デフォルトシート削除
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])

            # スタイル設定
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(
                horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 1. 変換サマリーシート
            summary_sheet = workbook.create_sheet("変換サマリー", 0)
            summary_data = [
                ["項目", "値"],
                ["変換日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["総ページ数", text_data.get("page_count", 0)],
                ["抽出文字数", len(text_data.get("full_text", ""))],
                ["抽出表数", tables_data.get("total_tables", 0)],
                ["高品質表数", tables_data.get("high_quality_tables", 0)],
                ["抽出画像数", images_data.get("total_images", 0)]
            ]

            for row in summary_data:
                summary_sheet.append(row)

            # サマリーシートのスタイル適用
            for row in summary_sheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border

            # 2. テキストシート
            if text_data.get("pages"):
                text_sheet = workbook.create_sheet("テキスト")
                text_sheet.append(["ページ番号", "テキスト内容", "文字数"])

                for page_data in text_data["pages"]:
                    text_sheet.append([
                        page_data["page_number"],
                        page_data["text"][:1000] +
                        "..." if len(page_data["text"]
                                     ) > 1000 else page_data["text"],
                        page_data["char_count"]
                    ])

                # テキストシートのスタイル適用
                for row in text_sheet.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = border

            # 3. 表データシート
            if tables_data.get("tables"):
                for table_data in tables_data["tables"]:
                    sheet_name = f"表_{table_data['page_number']}_{table_data['table_number']}"
                    if len(sheet_name) > 31:  # Excelのシート名制限
                        sheet_name = f"Table_{table_data['page_number']}_{table_data['table_number']}"

                    table_sheet = workbook.create_sheet(sheet_name)

                    # 表データの書き込み
                    for row in table_data["data"]:
                        table_sheet.append(row)

                    # 表シートのスタイル適用
                    if table_data["data"]:
                        for row in table_sheet.iter_rows(min_row=1, max_row=1):
                            for cell in row:
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = header_alignment
                                cell.border = border

                        # 全セルにボーダー適用
                        for row in table_sheet.iter_rows():
                            for cell in row:
                                cell.border = border

            # 4. 各ページを個別シートとして作成（1ページ1シート、PDFの見た目を完コピ）
            if images_data.get("pages"):
                for page_data in images_data["pages"]:
                    page_num = page_data["page_number"]
                    ocr_text = page_data.get("ocr_text", "")
                    position_data = page_data.get(
                        "position_data")  # OCR位置情報を取得

                    # シート名（Excelの制限: 31文字以内）
                    sheet_name = f"ページ{page_num}"
                    if len(sheet_name) > 31:
                        sheet_name = f"P{page_num}"

                    page_sheet = workbook.create_sheet(sheet_name)

                    # 位置情報がある場合は、それを使ってPDFの見た目を完コピ
                    if position_data and len(position_data) > 0:
                        self.logger.info(
                            f"ページ{page_num}: 位置情報を使用してPDFの見た目を完コピ（{len(position_data)}個の要素）")
                        self.create_excel_from_position_data(
                            page_sheet, position_data, ocr_text)
                        continue  # 位置情報ベースの処理が完了したので、次の処理をスキップ

                    # 位置情報がない場合は、従来の方法で処理
                    # 表の構造を自動検出
                    table_structure = self.detect_table_structure(ocr_text)
                    self.logger.info(
                        f"ページ{page_num}: 表構造検出 - タイプ: {table_structure.get('type')}, 列数: {table_structure.get('columns')}")

                    # OCRテキストを行ごとに解析して表形式で配置（改善版：行・列を正確に統一）
                    # 改行で分割（複数の改行パターンに対応）
                    import re  # モジュールをインポート
                    lines = re.split(r'\n+|\r+|\r\n+', ocr_text)
                    # 空行を除去
                    lines = [line.strip() for line in lines if line.strip()]

                    # 固定幅フォントの表かどうかを判定
                    is_fixed_width = False
                    column_positions = None
                    parsed_lines_fixed = None

                    # 固定幅フォントの表の場合
                    if table_structure.get('type') == 'fixed_width' and table_structure.get('column_positions'):
                        column_positions = table_structure['column_positions']
                        max_columns = table_structure.get(
                            'columns', len(column_positions) + 1)
                        is_fixed_width = True
                        self.logger.info(
                            f"ページ{page_num}: 固定幅フォントの表として解析（列位置: {column_positions}, 列数: {max_columns}）")
                        # 各行を固定幅で分割
                        parsed_lines_fixed = []
                        for line in lines:
                            cells = self.parse_fixed_width_table(
                                line, column_positions)
                            parsed_lines_fixed.append(cells)

                    # 改行がない場合（1行にまとまっている場合）の処理
                    if len(lines) <= 1 and len(ocr_text) > 100 and not is_fixed_width:
                        # 長いテキストが1行の場合は、より積極的に分割を試行
                        # 方法1: 数値の位置を基準にした列の自動検出
                        from collections import Counter

                        # 数値の位置を全て記録
                        number_positions = []
                        for match in re.finditer(r'\d+[,\d]*\.?\d*', ocr_text):
                            number_positions.append(match.start())

                        # キーワードの位置も記録
                        keyword_positions = []
                        keywords = ['金額', '数量', '合計', '計', '現金',
                                    'ブリカ', 'クレジット', '品名', '商品', 'F/O', 'f/o']
                        for keyword in keywords:
                            for match in re.finditer(re.escape(keyword), ocr_text, re.IGNORECASE):
                                keyword_positions.append(match.start())

                        # 位置を統合して列の区切りを推定
                        all_positions = sorted(
                            set(number_positions + keyword_positions))

                        # 位置の間隔を分析（10文字以内の誤差を許容）
                        if len(all_positions) >= 3:
                            # 間隔の分布を分析
                            intervals = []
                            for i in range(len(all_positions) - 1):
                                interval = all_positions[i +
                                                         1] - all_positions[i]
                                if interval > 5:  # 5文字以上の間隔のみを考慮
                                    intervals.append(interval)

                            if intervals:
                                # 最も多い間隔を基準に列の区切りを決定
                                interval_counter = Counter(intervals)
                                common_interval = interval_counter.most_common(1)[
                                    0][0]

                                # 列の区切り位置を決定
                                column_positions = [0]
                                current_pos = common_interval
                                while current_pos < len(ocr_text):
                                    column_positions.append(current_pos)
                                    current_pos += common_interval

                                # 列の区切りで行を分割
                                split_lines = []
                                for i in range(len(column_positions)):
                                    start = column_positions[i]
                                    end = column_positions[i + 1] if i + 1 < len(
                                        column_positions) else len(ocr_text)
                                    cell = ocr_text[start:end].strip()
                                    if cell:
                                        split_lines.append(cell)

                                if len(split_lines) >= 3:
                                    # 列ごとに分割された行を、行ごとに再構成
                                    # 複数の列を1つの行として扱う
                                    lines = split_lines
                                    self.logger.info(
                                        f"ページ{page_num}: 数値位置基準で{len(split_lines)}列に分割")

                        # 方法2: キーワードや数値パターンで分割を試行（フォールバック）
                        if len(lines) <= 1:
                            split_patterns = [
                                r'\s+(?=金額|数量|合計|計|現金|ブリカ|クレジット|品名|商品|F/O|f/o)',
                                r'\s{4,}',  # 4文字以上のスペース（より厳格に）
                                r'(?<=[0-9,.])\s{3,}(?=[0-9])',  # 数値間の大きなスペース
                                # 数値と日本語の間
                                r'(?<=[0-9])\s{3,}(?=[\u4E00-\u9FAF])',
                                # 日本語と数値の間
                                r'(?<=[\u4E00-\u9FAF])\s{3,}(?=[0-9])',
                            ]
                            for pattern in split_patterns:
                                temp_lines = re.split(pattern, ocr_text)
                                if len(temp_lines) > len(lines):
                                    lines = [line.strip()
                                             for line in temp_lines if line.strip()]
                                    if len(lines) > 5:  # 5行以上に分割できたら採用
                                        self.logger.info(
                                            f"ページ{page_num}: 1行テキストを{len(lines)}行に分割（パターン分割）")
                                        break

                    # 上部・左側の不要な行・列をフィルタリング
                    # 意味のあるデータ（キーワードや数値が含まれる行）から始まるようにする
                    filtered_lines = []
                    header_keywords = ['品名', '数量', '金額', '相利',
                                       'f/o', '現金', 'ブリカ', 'クレジット', '合計', '計']
                    data_started = False

                    for line in lines:
                        line_stripped = line.strip()
                        if not line_stripped:
                            # 空行は、データが開始した後は保持
                            if data_started:
                                filtered_lines.append(line)
                            continue

                        # キーワードが含まれている場合はデータ開始とみなす
                        if any(kw in line_stripped for kw in header_keywords):
                            data_started = True
                            filtered_lines.append(line)
                            continue

                        # 数値が含まれている場合もデータ開始とみなす
                        if re.search(r'\d+[,\d]*\.?\d*', line_stripped):
                            data_started = True
                            filtered_lines.append(line)
                            continue

                        # データが開始した後は、すべての行を保持
                        if data_started:
                            filtered_lines.append(line)

                    # フィルタリング後の行を使用（固定幅フォントの表の場合はスキップ）
                    if not is_fixed_width:
                        lines = filtered_lines if filtered_lines else lines

                    # 固定幅フォントの表の場合は、既に解析済みの行を使用
                    if is_fixed_width and parsed_lines_fixed:
                        parsed_rows = parsed_lines_fixed
                        max_columns = table_structure.get('columns', len(
                            column_positions) + 1) if column_positions else 0
                        header_row_index = None
                        # ヘッダー行を検出
                        for i, row in enumerate(parsed_rows[:5]):
                            if self.is_header_row(row):
                                header_row_index = i
                                break
                    else:
                        # 通常の解析処理
                        max_columns = 0
                        parsed_rows = []
                        header_row_index = None

                        # ステップ1: パイプ（|）で分割されている行を全て探して、最大列数を決定
                        # ヘッダー行を優先的に確認（列数が多く、キーワードを含む行）
                        candidate_headers = []
                        pipe_based_rows = []

                        for line in lines:
                            if '|' in line and line.strip():
                                parts = [part.strip()
                                         for part in line.split('|')]
                                if len(parts) > 1:
                                    pipe_based_rows.append((line, parts))
                                    # ヘッダー行の候補（キーワードが含まれる行）
                                    line_text = ' '.join(parts).lower()
                                    header_keywords = [
                                        '品名', '数量', '金額', '相利', 'f/o', '現金', 'ブリカ', 'クレジット']
                                    keyword_count = sum(
                                        1 for kw in header_keywords if kw in line_text)
                                    if keyword_count >= 3:  # 3つ以上のキーワードがあればヘッダー候補
                                        candidate_headers.append(
                                            (line, parts, len(parts)))
                                    max_columns = max(max_columns, len(parts))

                    # データ行（数値が多く含まれる行）の列数を優先的に分析
                    from collections import Counter
                    data_row_columns = []

                    for line, parts in pipe_based_rows:
                        row_text = ' '.join(parts)
                        # 数値が含まれるセルの数をカウント
                        num_cells = sum(1 for p in parts if re.search(
                            r'\d+[,\d]*\.?\d*', p))
                        # 数値が3つ以上含まれる行をデータ行とみなす
                        if num_cells >= 3:
                            data_row_columns.append(len(parts))

                    # データ行の列数の分布を分析
                    if data_row_columns:
                        col_counter = Counter(data_row_columns)
                        most_common_cols = col_counter.most_common(5)
                        # 10列以上を優先的に採用（表構造が複雑な場合）
                        high_col_rows = [col for col,
                                         count in most_common_cols if col >= 8]
                        if high_col_rows:
                            # 8列以上の行がある場合は、その中で最も多い列数を採用
                            optimal_columns = high_col_rows[0]
                        else:
                            # 8列未満の場合は、最も多い列数を採用
                            optimal_columns = most_common_cols[0][0]

                        # ヘッダー候補の列数も考慮（より多い方を採用）
                        if candidate_headers:
                            header_cols = max([h[2]
                                              for h in candidate_headers])
                            optimal_columns = max(optimal_columns, header_cols)

                        self.logger.info(
                            f"ページ{page_num}: データ行の列数分析 - {dict(col_counter)}, 最適列数: {optimal_columns}")
                        max_columns = optimal_columns
                    elif candidate_headers:
                        # データ行がない場合は、ヘッダー候補の列数を基準にする
                        candidate_headers.sort(
                            key=lambda x: x[2], reverse=True)
                        header_columns = candidate_headers[0][2]
                        # 全てのパイプ行の最大列数も確認
                        all_max_columns = max(
                            len(parts) for _, parts in pipe_based_rows) if pipe_based_rows else 0
                        max_columns = max(header_columns, all_max_columns)
                        self.logger.info(
                            f"ページ{page_num}: ヘッダー候補から列数決定: {max_columns}")
                    elif max_columns > 0:
                        # それもない場合は、最大列数を基準にする
                        max_columns = max(max_columns, max(
                            len(parts) for _, parts in pipe_based_rows) if pipe_based_rows else 0)
                        self.logger.info(
                            f"ページ{page_num}: 最大列数を使用: {max_columns}")

                    # パイプベースの行がある場合、それを基準にする
                    use_pipe_structure = len(
                        pipe_based_rows) > 0 and max_columns > 0

                    # ステップ2: 全ての行を順番に処理して、列数を統一した表データを作成
                    # 改善版：各行を適切に分割して、最大列数を決定
                    import re  # モジュールを最初にインポート
                    all_parsed_rows = []

                    for line in lines:
                        if not line.strip():  # 空行はスキップ
                            continue

                        # 行を解析（parse_ocr_lineを使用）
                        row_cells = self.parse_ocr_line(line)
                        if row_cells:
                            # 列数が多すぎる場合（50列以上）は分割が失敗しているとみなす
                            # その場合は、より積極的な分割を試行
                            if len(row_cells) >= 50:
                                # 複数スペースで再分割を試行
                                parts = re.split(r'\s{2,}', line.strip())
                                if len(parts) > len(row_cells) and len(parts) < 50:
                                    row_cells = [p.strip()
                                                 for p in parts if p.strip()]

                            # 列数が適切な範囲（2-30列）の場合のみ追加
                            if 2 <= len(row_cells) <= 30:
                                all_parsed_rows.append(row_cells)
                            elif len(row_cells) == 1 and len(line.strip()) > 20:
                                # 1列だが長い場合は、スペースで分割を試行
                                parts = re.split(r'\s+', line.strip())
                                if len(parts) >= 2:
                                    all_parsed_rows.append(parts)

                    # 最大列数を決定（実際に解析された行から）
                    if all_parsed_rows:
                        # 列数の分布を分析
                        from collections import Counter
                        col_counts = [len(row) for row in all_parsed_rows]
                        col_counter = Counter(col_counts)

                        # 最も多い列数を採用（ただし、2-15列の範囲で）
                        valid_cols = [
                            col for col, count in col_counter.most_common() if 2 <= col <= 15]
                        if valid_cols:
                            actual_max_columns = valid_cols[0]
                        else:
                            actual_max_columns = max(
                                col_counts) if col_counts else 5

                        # 既存のmax_columnsと比較して、大きい方を採用（ただし最大15列）
                        max_columns = min(max(
                            max_columns, actual_max_columns) if max_columns > 0 else actual_max_columns, 15)

                        self.logger.info(
                            f"ページ{page_num}: 列数分析 - {dict(col_counter)}, 採用列数: {max_columns}")

                        # 列数を統一した表データを作成（改善版：数値の結合処理を追加）
                        normalized_table = []
                        for row_cells in all_parsed_rows:
                            # まず、分割された数値を結合
                            merged_row = self.merge_split_numbers_improved(
                                row_cells)

                            # 最大列数に合わせて拡張
                            normalized_row = list(
                                merged_row[:max_columns]) if max_columns > 0 else list(merged_row)
                            # 不足分を空文字で埋める
                            while len(normalized_row) < max_columns:
                                normalized_row.append("")
                            # 超過分は切り詰め
                            if len(normalized_row) > max_columns:
                                normalized_row = normalized_row[:max_columns]
                            normalized_table.append(normalized_row)

                        # 表データをExcelに書き込み（スタイル適用）
                        if normalized_table:
                            for row_idx, row in enumerate(normalized_table, start=1):
                                # OCRエラーを修正してから書き込み
                                cleaned_row = []
                                for cell in row:
                                    cell_str = str(cell) if cell else ""
                                    # OCRエラーを修正
                                    cell_str = self.clean_ocr_text(cell_str)
                                    # 数値のクリーンアップ
                                    cell_str = self.clean_numeric_value(
                                        cell_str)
                                    cleaned_row.append(cell_str)
                                page_sheet.append(cleaned_row)

                            # スタイルを適用（PDFの見た目に近づける）
                            self.apply_pdf_like_styles(
                                page_sheet, max_columns, len(normalized_table))

                            self.logger.info(
                                f"ページ{page_num}: {len(normalized_table)}行 x {max_columns}列の表を作成（スタイル適用済み）")
                            # 既存の複雑な処理をスキップ
                            continue

                    # 既存の処理（フォールバック）
                    # ステップ2: 全ての行を順番に処理（行の順序を保持）
                    for line in lines:
                        if not line.strip():  # 空行はスキップ
                            continue

                        row_data = None

                        # パイプベースの構造がある場合は、パイプで分割を優先
                        if use_pipe_structure and '|' in line:
                            parts = [part.strip() for part in line.split('|')]

                            # 各セル内をさらに分割（より積極的に分割）
                            expanded_parts = []
                            for part in parts:
                                # セル内のOCRエラーを修正
                                part = self.clean_ocr_text(part)
                                if not part:
                                    expanded_parts.append("")
                                    continue

                                # 1文字セルは結合対象（次のセルと結合を試行）
                                if len(part) == 1 and part not in ['年', '月', '日', '時', '分', '秒']:
                                    # 次のセルと結合を試行（ただし、数値の場合は結合しない）
                                    if expanded_parts and not re.match(r'^\d+$', part):
                                        # 前のセルと結合
                                        expanded_parts[-1] = expanded_parts[-1] + part
                                        continue
                                # ヘッダー行の「金額」「数量」パターンを分割
                                elif '金額' in part and ('数量' in part or ('数' in part and '量' in part)):
                                    # 「金額  数量  金額  数量」のようなパターンを分割
                                    # まず、複数のスペースで分割してから、「金額」「数量」を検出
                                    # パターン: 「金額」または「数量」の前後で分割
                                    header_parts = []
                                    # 「金額」または「数量」（スペース入りも含む）で分割
                                    # 「数 量」（スペース入り）も検出できるように修正
                                    pattern = r'\s*(金額|数量|数\s+量)\s*'
                                    # 分割して、各マッチを個別のセルに
                                    split_parts = re.split(pattern, part)

                                    for i, split_part in enumerate(split_parts):
                                        split_part = split_part.strip()
                                        if split_part:
                                            # 「金額」または「数量」の場合はそのまま追加
                                            if split_part in ['金額', '数量']:
                                                header_parts.append(split_part)
                                            # 「数 量」（スペース入り）を「数量」に変換
                                            elif '数' in split_part and '量' in split_part:
                                                header_parts.append('数量')
                                            elif split_part == '数':
                                                # 次の要素が「量」か確認
                                                if i + 1 < len(split_parts) and '量' in str(split_parts[i + 1]):
                                                    header_parts.append('数量')
                                                else:
                                                    header_parts.append(
                                                        split_part)
                                            # それ以外のテキストも追加（空でなければ）
                                            else:
                                                header_parts.append(split_part)

                                    # デバッグログ（必要に応じて）
                                    if len(header_parts) > 1:
                                        self.logger.debug(
                                            f"ヘッダー分割: {part} -> {header_parts}")

                                    # もしくは、より単純に「金額」「数量」を検出して分割
                                    if not header_parts:
                                        # 「金額」と「数量」（スペースを含む）の位置を全て取得
                                        amount_positions = [
                                            (m.start(), '金額', 2) for m in re.finditer('金額', part)]
                                        # 「数量」は「数 量」や「数量」の両方に対応
                                        qty_positions = [
                                            (m.start(), '数量', 2) for m in re.finditer('数量', part)]
                                        # 「数 量」（スペース入り）も検出
                                        qty_space_positions = [
                                            (m.start(), '数量', 3) for m in re.finditer('数\s+量', part)]
                                        all_positions = sorted(
                                            amount_positions + qty_positions + qty_space_positions)

                                        if all_positions:
                                            last_end = 0
                                            for pos, label, length in all_positions:
                                                # 前のテキストを追加（空でなければ）
                                                if pos > last_end:
                                                    prev_text = part[last_end:pos].strip(
                                                    )
                                                    if prev_text:
                                                        header_parts.append(
                                                            prev_text)
                                                # 「金額」または「数量」を追加
                                                header_parts.append(label)
                                                last_end = pos + length

                                            # 残りのテキストを追加
                                            if last_end < len(part):
                                                remaining = part[last_end:].strip(
                                                )
                                                if remaining:
                                                    header_parts.append(
                                                        remaining)

                                    if len(header_parts) > 1:
                                        expanded_parts.extend(header_parts)
                                    else:
                                        expanded_parts.append(part)
                                elif re.search(r'\d', part):
                                    # 数値が含まれる場合は、より慎重に分割
                                    # 改善された数値パターン（カンマ、小数点を含む）
                                    # 例: 1,676.51, 125.281, 30.68 などを1つの数値として認識
                                    num_pattern = r'\d{1,3}(?:,\d{3})*(?:\.\d+)?|\d+\.\d+'
                                    num_matches = list(
                                        re.finditer(num_pattern, part))
                                    num_count = len(num_matches)

                                    if num_count >= 2:
                                        # 数値が2つ以上ある場合は、数値の間のスペースが大きい場合のみ分割
                                        sub_parts = []
                                        last_end = 0

                                        for i, match in enumerate(num_matches):
                                            # 数値の前のテキスト
                                            if match.start() > last_end:
                                                prefix = part[last_end:match.start()].strip(
                                                )
                                                # 前の数値との間隔を確認
                                                if prefix:
                                                    # スペースが3文字以上ある場合のみ、前のテキストを分離
                                                    if len(prefix) >= 3 or (i == 0 and len(prefix) > 1):
                                                        sub_parts.append(
                                                            prefix)
                                                    # スペースが小さい場合は、数値と一緒にする
                                                    elif i > 0:
                                                        # 前の数値に結合
                                                        if sub_parts:
                                                            sub_parts[-1] += " " + \
                                                                prefix

                                            # 数値とその後のスペース
                                            num_end = match.end()
                                            if i < len(num_matches) - 1:
                                                next_start = num_matches[i + 1].start()
                                                num_with_space = part[match.start(
                                                ):next_start].strip()
                                                # 次の数値との間隔を確認
                                                gap = next_start - num_end
                                                if gap >= 3:
                                                    # 間隔が大きい場合は分割
                                                    sub_parts.append(
                                                        num_with_space)
                                                else:
                                                    # 間隔が小さい場合は、次の数値と一緒にする
                                                    if sub_parts:
                                                        sub_parts[-1] += " " + \
                                                            num_with_space
                                                    else:
                                                        sub_parts.append(
                                                            num_with_space)
                                            else:
                                                num_with_space = part[match.start(
                                                ):].strip()
                                                sub_parts.append(
                                                    num_with_space)

                                            last_end = num_end

                                        # 残りのテキスト
                                        if last_end < len(part):
                                            suffix = part[last_end:].strip()
                                            if suffix and len(suffix) > 1:
                                                sub_parts.append(suffix)

                                        if len(sub_parts) > 1:
                                            expanded_parts.extend(sub_parts)
                                        else:
                                            # 分割できない場合は、複数スペースで分割
                                            sub_parts = [p.strip() for p in re.split(
                                                r'\s{3,}', part) if p.strip()]
                                            if len(sub_parts) > 1:
                                                expanded_parts.extend(
                                                    sub_parts)
                                            else:
                                                expanded_parts.append(part)
                                    else:
                                        # 数値が1つ以下の場合は、複数スペースで分割
                                        sub_parts = [p.strip() for p in re.split(
                                            r'\s{3,}', part) if p.strip()]
                                        if len(sub_parts) > 1:
                                            expanded_parts.extend(sub_parts)
                                        else:
                                            sub_parts = [p.strip() for p in re.split(
                                                r'\s{2,}', part) if p.strip()]
                                            if len(sub_parts) > 1:
                                                expanded_parts.extend(
                                                    sub_parts)
                                            else:
                                                expanded_parts.append(part)
                                else:
                                    expanded_parts.append(part)

                            parts = expanded_parts

                            # 列数を統一（最大列数に合わせる）
                            # 不足分は空文字で埋める
                            while len(parts) < max_columns:
                                parts.append("")
                            # 超過分は切り捨て
                            if len(parts) > max_columns:
                                parts = parts[:max_columns]

                            # データ行の各セルをクリーンアップ（文字化け修正も含む）
                            # まず、1文字セルを結合（後処理）
                            merged_parts = []
                            for i, cell in enumerate(parts):
                                cell_str = str(cell).strip()
                                if not cell_str:
                                    merged_parts.append("")
                                    continue

                                # 1文字セルは前のセルと結合（ただし、数値や日付関連は除く）
                                if len(cell_str) == 1 and cell_str not in ['年', '月', '日', '時', '分', '秒', ':', '-', '/']:
                                    if merged_parts and merged_parts[-1]:
                                        # 前のセルと結合
                                        merged_parts[-1] = merged_parts[-1] + cell_str
                                    else:
                                        merged_parts.append(cell_str)
                                else:
                                    merged_parts.append(cell_str)

                            parts = merged_parts

                            cleaned_parts = []
                            for cell in parts:
                                # まず文字化け修正
                                cell = self.fix_encoding_errors(str(cell))
                                # 次に数値クリーンアップ（より積極的に）
                                cell = self.clean_numeric_value(cell)
                                # セル内がほぼ数字のみの場合、数値として再フォーマット
                                if cell and re.match(r'^[\d,.\s\-]+$', cell):
                                    # 数字だけを抽出
                                    digits_only = re.sub(r'[^\d]', '', cell)
                                    if digits_only and len(digits_only) >= 3:
                                        try:
                                            # 整数としてパースして再フォーマット
                                            num_value = int(digits_only)
                                            # 3桁ごとにカンマ区切りでフォーマット
                                            cell = f"{num_value:,}"
                                        except ValueError:
                                            # 小数点を含む可能性がある
                                            try:
                                                num_value = float(digits_only)
                                                if num_value == int(num_value):
                                                    cell = f"{int(num_value):,}"
                                                else:
                                                    # 小数点以下2桁まで
                                                    cell = f"{num_value:,.2f}".rstrip(
                                                        '0').rstrip('.')
                                            except ValueError:
                                                pass  # パースできない場合はそのまま
                                cleaned_parts.append(cell)
                            row_data = cleaned_parts
                        elif use_pipe_structure:
                            # パイプがない行も、最大列数に合わせる（空セルで埋める）
                            # 意味のある内容がある場合は1列目に入れる
                            if line.strip():
                                row_data = [line.strip()] + [""] * \
                                    (max_columns - 1)
                                if len(row_data) > max_columns:
                                    row_data = row_data[:max_columns]
                                # データ行の各セルをクリーンアップ
                                row_data = [self.clean_numeric_value(
                                    str(cell)) if cell else "" for cell in row_data]
                        else:
                            # パイプがない場合は通常の解析
                            row_data = self.parse_ocr_line(line)
                            if row_data:
                                max_columns = max(max_columns, len(row_data))
                                # データ行の各セルをクリーンアップ
                                row_data = [self.clean_numeric_value(
                                    str(cell)) if cell else "" for cell in row_data]

                        if row_data:
                            # 1文字セルの後処理（最終チェック）
                            merged_row_data = []
                            for i, cell in enumerate(row_data):
                                cell_str = str(cell).strip() if cell else ""
                                if not cell_str:
                                    merged_row_data.append("")
                                    continue

                                # 1文字セルは前のセルと結合（ただし、数値や日付関連は除く）
                                if len(cell_str) == 1 and cell_str not in ['年', '月', '日', '時', '分', '秒', ':', '-', '/']:
                                    if merged_row_data and merged_row_data[-1]:
                                        # 前のセルと結合
                                        merged_row_data[-1] = str(
                                            merged_row_data[-1]) + cell_str
                                    else:
                                        merged_row_data.append(cell_str)
                                else:
                                    merged_row_data.append(cell_str)

                            row_data = merged_row_data

                            # 最終的な列数の統一（念のため）
                            while len(row_data) < max_columns:
                                row_data.append("")
                            if len(row_data) > max_columns and max_columns > 0:
                                row_data = row_data[:max_columns]
                            # 最終的なクリーンアップ（文字化け修正も含む）
                            cleaned_row = []
                            for cell in row_data:
                                cell_str = str(cell) if cell else ""
                                # まず文字化け修正
                                cell_str = self.fix_encoding_errors(cell_str)
                                # 次に数値クリーンアップ
                                cell_str = self.clean_numeric_value(cell_str)
                                # セル内がほぼ数字のみの場合、数値として再フォーマット（より積極的に）
                                if cell_str and re.match(r'^[\d,.\s\-]+$', cell_str):
                                    # 数字だけを抽出
                                    digits_only = re.sub(
                                        r'[^\d]', '', cell_str)
                                    if digits_only and len(digits_only) >= 3:
                                        try:
                                            # 整数としてパースして再フォーマット
                                            num_value = int(digits_only)
                                            # 3桁ごとにカンマ区切りでフォーマット
                                            cell_str = f"{num_value:,}"
                                        except ValueError:
                                            # 小数点を含む可能性がある
                                            try:
                                                num_value = float(digits_only)
                                                if num_value == int(num_value):
                                                    cell_str = f"{int(num_value):,}"
                                                else:
                                                    # 小数点以下2桁まで
                                                    cell_str = f"{num_value:,.2f}".rstrip(
                                                        '0').rstrip('.')
                                            except ValueError:
                                                pass  # パースできない場合はそのまま
                                cleaned_row.append(cell_str)
                            row_data = cleaned_row

                            # ヘッダー行の判定
                            if header_row_index is None and self.is_header_row(row_data):
                                header_row_index = len(parsed_rows)

                            parsed_rows.append(row_data)

                    # 空行やノイズが多い行を除外（改善版）
                    # 左側の不要な列もフィルタリング
                    filtered_rows = []
                    left_column_start = 0  # 左側の不要な列の開始位置

                    # 左側の不要な列を検出（最初の数行で、最初の列が意味のないデータの場合）
                    if len(parsed_rows) > 3:
                        first_col_samples = [str(row[0]).strip() if row and len(row) > 0 and row[0] else ""
                                             for row in parsed_rows[:5]]
                        # 最初の列が単独の数字（1-3桁）や1文字のみの場合、左側の不要な列とみなす
                        noise_pattern = re.compile(r'^(\d{1,3}|[A-Za-z]{1})$')
                        noise_count = sum(1 for sample in first_col_samples
                                          if sample and noise_pattern.match(sample))

                        # 最初の列の50%以上がノイズの場合、左側の不要な列とみなす
                        if noise_count >= len([s for s in first_col_samples if s]) * 0.5:
                            left_column_start = 1
                            self.logger.info(
                                f"ページ{page_num}: 左側の不要な列を検出、1列目をスキップ")

                    for row in parsed_rows:
                        # 左側の不要な列をスキップ
                        if left_column_start > 0 and len(row) > left_column_start:
                            row = row[left_column_start:]

                        # 有効なセルをカウント（1文字以上、かつ意味のある内容）
                        valid_cells = []
                        for cell in row:
                            if cell and str(cell).strip():
                                cell_str = str(cell).strip()
                                # 意味のあるセル（長さ2以上、または数字を含む）
                                if len(cell_str) >= 2 or re.search(r'\d', cell_str):
                                    valid_cells.append(cell)

                        # 有効なセルが1つ以上ある行を保持（条件を緩和）
                        # または数値が含まれる行、または日本語が含まれる行
                        has_japanese = any(re.search(
                            r'[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FAF]', str(c)) for c in row if c)
                        has_number = any(re.search(r'\d', str(c))
                                         for c in row if c)

                        if len(valid_cells) >= 1 or has_japanese or has_number:
                            # 分割された数値を結合する処理を追加（改善版）
                            row = self.merge_split_numbers_improved(row)
                            # 結合後の列数を最大列数に合わせる（列の整列を維持）
                            while len(row) < max_columns:
                                row.append("")
                            if len(row) > max_columns and max_columns > 0:
                                row = row[:max_columns]
                            filtered_rows.append(row)

                    # 異常値検出（AI監査）
                    anomalies = self.detect_anomalies(filtered_rows)
                    if anomalies['warnings']:
                        self.logger.info(
                            f"ページ{page_num}: 異常値検出 - {len(anomalies['warnings'])}件の警告")
                        for warning in anomalies['warnings'][:5]:  # 最初の5件のみログ
                            self.logger.warning(f"  警告: {warning}")

                    # 全ての行の列数を統一（最大列数に合わせる）
                    for row in filtered_rows:
                        # 列数を統一
                        while len(row) < max_columns:
                            row.append("")
                        while len(row) > max_columns and max_columns > 0:
                            row = row[:max_columns]
                        page_sheet.append(row)

                    # ヘッダー行にスタイルを適用
                    if header_row_index is not None and page_sheet.max_row > header_row_index:
                        header_row_num = header_row_index + 1  # 1ベースの行番号
                        if header_row_num <= page_sheet.max_row:
                            for cell in page_sheet[header_row_num]:
                                if cell.value:  # 値があるセルのみ
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = header_alignment
                                    cell.border = border

                    # 列幅を自動調整
                    if max_columns > 0:
                        for col_idx in range(1, max_columns + 1):
                            max_length = 15  # 最小幅
                            column_letter = openpyxl.utils.get_column_letter(
                                col_idx)

                            for row in page_sheet.iter_rows(min_row=1, max_row=page_sheet.max_row,
                                                            min_col=col_idx, max_col=col_idx):
                                for cell in row:
                                    try:
                                        if cell.value and len(str(cell.value)) > max_length:
                                            max_length = min(
                                                len(str(cell.value)), 50)  # 最大50文字
                                    except Exception:
                                        pass

                            page_sheet.column_dimensions[column_letter].width = max_length + 2

                    # ボーダーを適用（見やすくするため）
                    for row in page_sheet.iter_rows():
                        for cell in row:
                            cell.border = border

            # 列幅の自動調整
            for sheet in workbook.worksheets:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass

                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width

            # ファイル保存
            workbook.save(output_path)

            self.logger.info(f"Excelファイル生成完了: {output_path}")
            return {"success": True, "output_path": output_path}

        except Exception as e:
            self.logger.error(f"Excel生成エラー: {e}")
            return {"success": False, "error": str(e)}

    def convert_pdf_to_excel(self, pdf_path: str, use_ocr: bool = False,
                             custom_name: str = None) -> Dict[str, Any]:
        """
        PDFをExcelに変換（pdf_excel_agent.py用のインターフェース）

        Args:
            pdf_path: PDFファイルパス
            use_ocr: OCR使用フラグ
            custom_name: カスタムファイル名

        Returns:
            変換結果の辞書
        """
        self.stats['total_conversions'] += 1

        try:
            # ファイル検証
            validation = self.validate_pdf_file(pdf_path)
            if not validation["valid"]:
                self.stats['failed_conversions'] += 1
                return {"success": False, "error": validation["error"]}

            # 出力パス設定
            if custom_name:
                output_filename = f"{custom_name}.xlsx"
            else:
                pdf_name = Path(pdf_path).stem
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_filename = f"{pdf_name}_{timestamp}.xlsx"

            output_path = str(self.output_dir / output_filename)

            # OCR設定更新
            if use_ocr:
                self.config["ocr_enabled"] = True
                self.stats['ocr_used_count'] += 1

            # 1. テキスト抽出
            self.logger.info("テキスト抽出開始")
            text_result = self.extract_text_from_pdf(pdf_path)
            if not text_result["success"]:
                self.stats['failed_conversions'] += 1
                return {"success": False, "error": f"テキスト抽出エラー: {text_result['error']}"}

            self.stats['total_pages_processed'] += text_result["data"]["page_count"]

            # 2. 表データ抽出
            self.logger.info("表データ抽出開始")
            tables_result = self.extract_tables_from_pdf(pdf_path)
            if not tables_result["success"]:
                self.stats['failed_conversions'] += 1
                return {"success": False, "error": f"表データ抽出エラー: {tables_result['error']}"}

            self.stats['total_tables_extracted'] += tables_result["data"]["total_tables"]
            self.stats['high_quality_tables'] += tables_result["data"].get(
                "high_quality_tables", 0)

            # 3. Excelファイルを事前作成（1ページずつ処理するため）
            if not OPENPYXL_AVAILABLE:
                return {"success": False, "error": "openpyxlが利用できません"}

            workbook = openpyxl.Workbook()
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])

            # 変換サマリーシートを作成
            summary_sheet = workbook.create_sheet("変換サマリー", 0)
            summary_data = [
                ["項目", "値"],
                ["変換日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["総ページ数", text_result["data"]["page_count"]],
                ["抽出文字数", len(text_result["data"].get("full_text", ""))],
            ]
            for row in summary_data:
                summary_sheet.append(row)

            # 4. 1ページずつOCR処理してExcelシートに追加（メモリ節約）
            images_result = {"success": True, "data": {
                "images": [], "total_images": 0, "ocr_text": "", "pages": []}}

            if use_ocr:
                self.logger.info("1ページずつOCR処理を開始（メモリ節約モード）")
                try:
                    excel_result = self.process_pages_one_by_one(
                        pdf_path, workbook, text_result["data"], tables_result["data"]
                    )
                except AttributeError:
                    # 関数が存在しない場合は従来の方法にフォールバック
                    self.logger.warning(
                        "process_pages_one_by_one関数が見つかりません。従来の方法で処理します。")
                images_result = self.extract_images_with_ocr(pdf_path)
                excel_result = self.create_excel_from_data(
                    text_result["data"],
                    tables_result["data"],
                    images_result["data"],
                    output_path
                )
                workbook.save(output_path)
                excel_result = {"success": True}
            else:
                # OCRなしの場合は従来通り
                images_result = {"success": True, "data": {
                    "images": [], "total_images": 0, "ocr_text": ""}}
            self.logger.info("Excel生成開始")
            excel_result = self.create_excel_from_data(
                text_result["data"],
                tables_result["data"],
                images_result["data"],
                output_path
            )
            # 既存のworkbookを保存
            workbook.save(output_path)
            excel_result = {"success": True}

            if not excel_result.get("success", True):
                workbook.close()
                self.stats['failed_conversions'] += 1
                return {"success": False, "error": f"Excel生成エラー: {excel_result.get('error', 'Unknown error')}"}

            # Excelファイルを保存
            try:
                workbook.save(output_path)
                workbook.close()
                self.logger.info(f"Excelファイル保存完了: {output_path}")
            except Exception as e:
                self.logger.error(f"Excelファイル保存エラー: {e}")
                return {"success": False, "error": f"Excelファイル保存エラー: {e}"}

            if not excel_result["success"]:
                self.stats['failed_conversions'] += 1
                return {"success": False, "error": f"Excel生成エラー: {excel_result['error']}"}

            # 成功
            self.stats['successful_conversions'] += 1

            # 画像ベースPDFの場合はOCR推奨メッセージを追加
            recommendations = []
            if tables_result["data"].get("is_image_based", False):
                recommendations.append(
                    "このPDFは画像ベースです。OCRを使用すると表が抽出できる可能性があります。")

            if tables_result["data"]["total_tables"] == 0 and not use_ocr:
                recommendations.append("表が抽出されませんでした。OCRを使用してみてください。")

            result = {
                "success": True,
                "excel_file": output_path,
                "excel_path": output_path,  # 互換性のため
                "total_tables": tables_result["data"]["total_tables"],
                "high_quality_tables": tables_result["data"].get("high_quality_tables", 0),
                "pages_processed": text_result["data"]["page_count"],
                "is_image_based": tables_result["data"].get("is_image_based", False),
                "recommendations": recommendations,
                "extraction_summary": {
                    "pages_processed": text_result["data"]["page_count"],
                    "tables_extracted": tables_result["data"]["total_tables"],
                    "high_quality_tables": tables_result["data"].get("high_quality_tables", 0),
                    "images_extracted": images_result["data"].get("total_images", 0),
                    "is_image_based": tables_result["data"].get("is_image_based", False)
                }
            }

            self.logger.info(f"PDF-Excel変換完了: {output_path}")
            return result

        except Exception as e:
            self.logger.error(f"PDF-Excel変換エラー: {e}")
            self.stats['failed_conversions'] += 1
            return {"success": False, "error": str(e)}

    def batch_convert(self, pdf_paths: List[str], use_ocr: bool = False) -> Dict[str, Any]:
        """
        複数PDFを一括変換

        Args:
            pdf_paths: PDFファイルパスのリスト
            use_ocr: OCR使用フラグ

        Returns:
            バッチ変換結果
        """
        results = {
            "successful": [],
            "failed": [],
            "total_files": len(pdf_paths),
            "successful_count": 0,
            "failed_count": 0
        }

        for pdf_path in pdf_paths:
            result = self.convert_pdf_to_excel(pdf_path, use_ocr=use_ocr)

            if result["success"]:
                results["successful"].append({
                    "pdf_path": pdf_path,
                    "excel_path": result["excel_path"]
                })
                results["successful_count"] += 1
            else:
                results["failed"].append({
                    "pdf_path": pdf_path,
                    "error": result.get("error", "Unknown error")
                })
                results["failed_count"] += 1

        return results

    def get_stats(self) -> Dict[str, Any]:
        """統計情報を取得"""
        return self.stats.copy()


# 使用例
if __name__ == "__main__":
    import sys

    converter = ManaPDFExcelAdvanced()

    # システム状態確認
    print("=" * 60)
    print("Mana PDF Excel Advanced System")
    print("=" * 60)
    print(f"PyMuPDF: {PYMUPDF_AVAILABLE}")
    print(f"pdfplumber: {PDFPLUMBER_AVAILABLE}")
    print(f"Camelot: {CAMELOT_AVAILABLE}")
    print(f"openpyxl: {OPENPYXL_AVAILABLE}")
    print(f"OCR: {OCR_AVAILABLE}")
    print("=" * 60)

    # コマンドライン引数でPDFファイルを指定
    if len(sys.argv) > 1:
        pdf_path = sys.argv[1]
        use_ocr = "--ocr" in sys.argv or "-o" in sys.argv

        print(f"\n変換開始: {pdf_path}")
        print(f"OCR使用: {use_ocr}")

        result = converter.convert_pdf_to_excel(pdf_path, use_ocr=use_ocr)

        if result["success"]:
            print(f"\n✅ 変換完了!")
            print(f"Excelファイル: {result['excel_path']}")
            print(f"抽出表数: {result['total_tables']}")
            print(f"高品質表数: {result['high_quality_tables']}")
        else:
            print(f"\n❌ 変換失敗: {result.get('error', 'Unknown error')}")
    else:
        print("\n使用方法:")
        print("  python3 mana_pdf_excel_advanced.py <PDFファイルパス> [--ocr]")
        print("\n例:")
        print("  python3 mana_pdf_excel_advanced.py /path/to/file.pdf")
        print("  python3 mana_pdf_excel_advanced.py /path/to/file.pdf --ocr")
