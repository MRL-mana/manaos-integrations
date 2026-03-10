#!/usr/bin/env python3
"""
Excel生成エラー修正版PDF-Excel変換システム
特殊文字フィルタリング機能付き
"""

import asyncio
import logging
import os
import sys
import re
import uuid
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any

# 既存システムのインポート
sys.path.append('/root')
from pdf_excel_converter import PDFExcelConverter

class FixedExcelGenerator(PDFExcelConverter):
    """Excel生成エラー修正版変換システム"""
    
    def __init__(self):
        super().__init__()
        self.logger = logging.getLogger("FixedExcelGenerator")
        self.logger.setLevel(logging.INFO)
        
        # タスク管理
        self.tasks = {}
        
        # 出力設定
        self.output_config = {
            "base_directory": "/root/excel_output_fixed",
            "desktop_directory": "/home/mana/Desktop/修正版PDF-Excel変換結果",
            "create_timestamped_folder": True
        }
        
        # ディレクトリ作成
        self.setup_directories()
        
        self.logger.info("🚀 修正版Excel生成システム初期化完了")
    
    def setup_directories(self):
        """ディレクトリ設定"""
        # ベースディレクトリ作成
        Path(self.output_config["base_directory"]).mkdir(parents=True, exist_ok=True)
        
        # デスクトップディレクトリ作成
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        desktop_folder = f"{self.output_config['desktop_directory']}_{timestamp}"
        Path(desktop_folder).mkdir(parents=True, exist_ok=True)
        
        self.output_config["desktop_folder"] = desktop_folder
        self.logger.info(f"📁 出力ディレクトリ設定完了: {desktop_folder}")
    
    def clean_text_for_excel(self, text: str) -> str:
        """Excel用テキストクリーニング"""
        if not text:
            return ""
        
        # 制御文字を除去（ASCII 0-31、127-159）
        cleaned = ""
        for char in text:
            char_code = ord(char)
            # 許可する文字範囲
            if (32 <= char_code <= 126) or (char_code >= 160):  # 印刷可能文字 + 拡張文字
                cleaned += char
            else:
                # 制御文字は空白に置換
                cleaned += " "
        
        # 連続する空白を1つにまとめる
        cleaned = re.sub(r'\s+', ' ', cleaned)
        
        # 先頭と末尾の空白を除去
        cleaned = cleaned.strip()
        
        # 空の文字列の場合は "（空白）" に置換
        if not cleaned:
            cleaned = "（空白）"
        
        return cleaned
    
    def clean_table_data_for_excel(self, table_data: List[List[str]]) -> List[List[str]]:
        """表データをExcel用にクリーニング"""
        cleaned_table = []
        for row in table_data:
            cleaned_row = []
            for cell in row:
                cleaned_cell = self.clean_text_for_excel(str(cell))
                cleaned_row.append(cleaned_cell)
            cleaned_table.append(cleaned_row)
        return cleaned_table
    
    async def _generate_excel_file(self, output_path: str, extracted_data: Dict[str, Any]) -> Dict[str, Any]:
        """Excel生成（修正版）"""
        try:
            self.logger.info("Excel生成開始（修正版）")
            
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            # 新しいワークブック作成
            wb = Workbook()
            
            # デフォルトシートを削除
            wb.remove(wb.active)  # type: ignore
            
            # テキストデータシート作成
            if extracted_data.get("text_data"):
                text_sheet = wb.create_sheet("テキストデータ")
                text_data = extracted_data["text_data"]
                
                # テキストデータをクリーニング
                cleaned_text_data = []
                for page_num, page_text in text_data.items():
                    cleaned_text = self.clean_text_for_excel(page_text)
                    cleaned_text_data.append([f"ページ {page_num}", cleaned_text])
                
                # ヘッダー追加
                text_sheet.append(["ページ", "テキスト内容"])
                
                # データ追加
                for row in cleaned_text_data:
                    text_sheet.append(row)
                
                # スタイル適用
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                for cell in text_sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # 列幅調整
                text_sheet.column_dimensions['A'].width = 15
                text_sheet.column_dimensions['B'].width = 100
            
            # 表データシート作成
            if extracted_data.get("tables"):
                tables = extracted_data["tables"]
                for i, table in enumerate(tables, 1):
                    sheet_name = f"表{i}"
                    if len(sheet_name) > 31:  # Excelシート名の長さ制限
                        sheet_name = f"表{i}"
                    
                    table_sheet = wb.create_sheet(sheet_name)
                    
                    # 表データをクリーニング
                    cleaned_table = self.clean_table_data_for_excel(table)
                    
                    if cleaned_table:
                        # データ追加
                        for row in cleaned_table:
                            table_sheet.append(row)
                        
                        # スタイル適用
                        if cleaned_table:
                            header_font = Font(bold=True)
                            header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                            
                            # ヘッダー行のスタイル
                            for cell in table_sheet[1]:
                                cell.font = header_font
                                cell.fill = header_fill
                            
                            # 列幅調整
                            for col in table_sheet.columns:
                                max_length = 0
                                column = col[0].column_letter
                                for cell in col:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except Exception:
                                        pass
                                adjusted_width = min(max_length + 2, 50)
                                table_sheet.column_dimensions[column].width = adjusted_width
            
            # OCR結果シート作成
            if extracted_data.get("ocr_results"):
                ocr_sheet = wb.create_sheet("OCR結果")
                ocr_data = extracted_data["ocr_results"]
                
                # OCR結果をクリーニング
                cleaned_ocr_data = []
                for page_num, ocr_text in ocr_data.items():
                    cleaned_ocr_text = self.clean_text_for_excel(ocr_text)
                    cleaned_ocr_data.append([f"ページ {page_num}", cleaned_ocr_text])
                
                # ヘッダー追加
                ocr_sheet.append(["ページ", "OCR認識結果"])
                
                # データ追加
                for row in cleaned_ocr_data:
                    ocr_sheet.append(row)
                
                # スタイル適用
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                
                for cell in ocr_sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # 列幅調整
                ocr_sheet.column_dimensions['A'].width = 15
                ocr_sheet.column_dimensions['B'].width = 100
            
            # サマリーシート作成
            summary_sheet = wb.create_sheet("変換サマリー")
            summary_data = [
                ["項目", "値"],
                ["変換日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["処理ページ数", extracted_data.get("pages_processed", 0)],
                ["抽出テキスト長", len(str(extracted_data.get("text_data", "")))],
                ["抽出表数", len(extracted_data.get("tables", []))],
                ["OCR処理", "実行済み" if extracted_data.get("ocr_results") else "未実行"],
                ["システム", "ManaOS統合PDF-Excel変換システム v2.0"],
                ["修正版", "特殊文字フィルタリング機能付き"]
            ]
            
            for row in summary_data:
                summary_sheet.append(row)
            
            # サマリーシートのスタイル
            for cell in summary_sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            
            # 列幅調整
            summary_sheet.column_dimensions['A'].width = 25
            summary_sheet.column_dimensions['B'].width = 50
            
            # ファイル保存
            wb.save(output_path)
            
            self.logger.info(f"Excelファイル生成完了（修正版）: {output_path}")
            
            return {
                "success": True,
                "output_file": output_path,
                "sheets_created": len(wb.worksheets),
                "total_data_rows": sum(len(list(sheet.rows)) - 1 for sheet in wb.worksheets if len(list(sheet.rows)) > 1)
            }
            
        except Exception as e:
            self.logger.error(f"Excel生成エラー（修正版）: {e}")
            return {
                "success": False,
                "error": f"Excel生成エラー（修正版）: {str(e)}"
            }
    
    async def convert_pdf_to_excel_fixed(self, pdf_path: str, output_path: str, config: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """修正版PDF-Excel変換"""
        try:
            task_id = str(uuid.uuid4())
            self.tasks[task_id] = {"status": "processing", "progress": 0}
            
            start_time = time.time()
            
            # 1. テキスト抽出
            self.logger.info("テキスト抽出開始")
            text_result = self.extract_text_from_pdf(pdf_path)
            text_data = text_result.get("text_data", {})
            self.tasks[task_id]["progress"] = 25
            
            # 2. 表データ抽出
            self.logger.info("表データ抽出開始")
            tables_result = self.extract_tables_from_pdf(pdf_path)
            tables = tables_result.get("tables", [])
            self.tasks[task_id]["progress"] = 50
            
            # 3. 画像・OCR抽出
            self.logger.info("画像・OCR抽出開始")
            images_result = self.extract_images_from_pdf(pdf_path)
            images = images_result.get("images", [])
            ocr_results = {}  # 簡略化のため空の辞書
            self.tasks[task_id]["progress"] = 75
            
            # 4. 抽出データをまとめる
            extracted_data = {
                "text_data": text_data,
                "tables": tables,
                "images": images,
                "ocr_results": ocr_results,
                "pages_processed": len(text_data) if text_data else 0
            }
            
            # 5. Excel生成（修正版）
            excel_result = await self._generate_excel_file(output_path, extracted_data)
            self.tasks[task_id]["progress"] = 100
            
            processing_time = time.time() - start_time
            
            if excel_result.get("success"):
                self.tasks[task_id]["status"] = "completed"
                self.tasks[task_id]["result"] = {
                    "success": True,
                    "task_id": task_id,
                    "output_file": output_path,
                    "processing_time": processing_time,
                    "extraction_summary": {
                        "pages_processed": extracted_data["pages_processed"],
                        "tables_extracted": len(tables) if tables else 0,
                        "images_extracted": len(images) if images else 0,
                        "text_length": sum(len(str(page_text)) for page_text in text_data.values()) if text_data else 0,
                        "sheets_created": excel_result.get("sheets_created", 0),
                        "total_data_rows": excel_result.get("total_data_rows", 0)
                    }
                }
                
                # デスクトップにコピー
                desktop_output_path = os.path.join(self.output_config["desktop_folder"], os.path.basename(output_path))
                import shutil
                shutil.copy2(output_path, desktop_output_path)
                
                return self.tasks[task_id]["result"]
            else:
                self.tasks[task_id]["status"] = "failed"
                self.tasks[task_id]["error"] = excel_result.get("error")
                return {
                    "success": False,
                    "task_id": task_id,
                    "error": excel_result.get("error")
                }
                
        except Exception as e:
            self.logger.error(f"修正版PDF-Excel変換エラー: {e}")
            return {
                "success": False,
                "error": str(e)
            }
    
    async def test_single_pdf_conversion(self, pdf_path: str) -> Dict[str, Any]:
        """単一PDFファイル変換テスト"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"fixed_conversion_test_{timestamp}.xlsx"
            output_path = os.path.join(self.output_config["base_directory"], output_filename)
            
            # 変換設定
            conversion_config = {
                "table_detection": True,
                "ocr_enabled": True,
                "language": "jpn+eng",
                "fallback_language": "eng",
                "confidence_threshold": 60,
                "preprocessing": True
            }
            
            # 修正版変換実行
            result = await self.convert_pdf_to_excel_fixed(pdf_path, output_path, conversion_config)
            
            if result.get("success"):
                # デスクトップにコピー
                desktop_output_path = os.path.join(self.output_config["desktop_folder"], output_filename)
                import shutil
                shutil.copy2(output_path, desktop_output_path)
                
                return {
                    "success": True,
                    "pdf_file": os.path.basename(pdf_path),
                    "output_file": output_path,
                    "desktop_file": desktop_output_path,
                    "processing_time": result.get("processing_time", 0),
                    "extraction_summary": result.get("extraction_summary", {}),
                    "timestamp": datetime.now().isoformat()
                }
            else:
                return {
                    "success": False,
                    "pdf_file": os.path.basename(pdf_path),
                    "error": result.get("error"),
                    "timestamp": datetime.now().isoformat()
                }
                
        except Exception as e:
            return {
                "success": False,
                "pdf_file": os.path.basename(pdf_path),
                "error": str(e),
                "timestamp": datetime.now().isoformat()
            }

async def main():
    """メイン実行関数"""
    # ログ設定
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('fixed_excel_generator.log'),
            logging.StreamHandler()
        ]
    )
    
    converter = FixedExcelGenerator()
    
    print("🚀 修正版Excel生成システム")
    print("=" * 60)
    
    # テスト用PDF作成
    print("📄 テスト用PDF作成中...")
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        
        # テスト用PDF作成
        pdf_path = os.path.join(converter.output_config["base_directory"], "test_document_with_special_chars.pdf")
        doc = SimpleDocTemplate(pdf_path, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # ヘッダー
        header = Paragraph("修正版Excel生成テスト文書", styles['Title'])
        story.append(header)
        story.append(Spacer(1, 20))
        
        # テストテキスト（特殊文字を含む）
        test_text = Paragraph("""
        この文書には特殊文字が含まれています：<br/>
        • 制御文字: \x00\x01\x02<br/>
        • 改行文字: \n\r<br/>
        • タブ文字: \t<br/>
        • 日本語文字: こんにちは世界<br/>
        • 英語文字: Hello World<br/>
        • 数値: 123456789<br/>
        """, styles['Normal'])
        story.append(test_text)
        story.append(Spacer(1, 20))
        
        # テスト用表データ
        test_data = [
            ['項目', '値', '備考'],
            ['テスト1', '特殊文字\x00\x01', '制御文字テスト'],
            ['テスト2', '日本語テスト', '文字認識テスト'],
            ['テスト3', 'English Test', 'Character Recognition'],
            ['テスト4', '123456', '数値認識テスト']
        ]
        
        test_table = Table(test_data)
        test_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
        ]))
        
        story.append(test_table)
        
        doc.build(story)
        
        print(f"✅ テスト用PDF作成完了: {pdf_path}")
        
        # 修正版変換テスト
        print("\n🔄 修正版変換テスト実行中...")
        result = await converter.test_single_pdf_conversion(pdf_path)
        
        # 結果表示
        print("\n" + "=" * 60)
        print("🎉 **修正版Excel生成テスト完了！**")
        print("=" * 60)
        
        if result.get("success"):
            print("📊 **テスト結果サマリー:**")
            print("   ステータス: ✅ 成功")
            print(f"   処理時間: {result.get('processing_time', 0):.2f}秒")
            summary = result.get('extraction_summary', {})
            print(f"   処理ページ数: {summary.get('pages_processed', 0)}")
            print(f"   抽出表数: {summary.get('tables_extracted', 0)}")
            print(f"   作成シート数: {summary.get('sheets_created', 0)}")
            print(f"   総データ行数: {summary.get('total_data_rows', 0)}")
            print(f"   出力ファイル: {result.get('output_file', 'N/A')}")
            print(f"   デスクトップ: {result.get('desktop_file', 'N/A')}")
            
            print("\n✅ **修正版Excel生成システムが正常動作しています！**")
            
        else:
            print(f"❌ **テスト失敗:** {result.get('error', '不明なエラー')}")
        
        print("\n📋 **次のステップ:**")
        print("   1. X280デスクトップの修正版変換結果フォルダを確認")
        print("   2. 修正版Excelファイルを開いて特殊文字フィルタリングを確認")
        print("   3. 実際のGoogle Driveファイルで修正版変換実行")
        
    except Exception as e:
        print(f"❌ テスト実行エラー: {e}")

if __name__ == "__main__":
    asyncio.run(main())
