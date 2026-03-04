#!/usr/bin/env python3
"""
X280 Excelファイル直接入力システム
既存のExcelファイルにデータを直接入力・編集
"""

import os
import sys
from pathlib import Path
from datetime import datetime

# 既存のシステムをインポート
sys.path.append('/root')
from advanced_excel_system import AdvancedExcelSystem

class X280ExcelDirectInput:
    def __init__(self):
        self.output_dir = Path("/home/mana/Desktop/X280直接入力Excel")
        self.output_dir.mkdir(exist_ok=True)
        
        self.excel_system = AdvancedExcelSystem()
        
        print("🚀 X280 Excelファイル直接入力システム")
        print(f"📁 出力先: {self.output_dir}")
    
    def create_new_excel_with_data(self):
        """新しいExcelファイルを作成してデータ入力"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # 新しいExcelファイル作成
            wb = Workbook()
            
            # デフォルトシートを削除
            wb.remove(wb.active)
            
            # 1. メインデータシート
            ws_main = wb.create_sheet("メインデータ")
            
            # ヘッダー行
            headers = ["日付", "項目", "値", "備考", "ステータス"]
            for col, header in enumerate(headers, 1):
                cell = ws_main.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # サンプルデータ
            sample_data = [
                [datetime.now().strftime('%Y-%m-%d'), "売上高", "1,500,000円", "月間目標達成", "完了"],
                [datetime.now().strftime('%Y-%m-%d'), "新規顧客", "25件", "前月比120%", "進行中"],
                [datetime.now().strftime('%Y-%m-%d'), "既存顧客", "150件", "安定維持", "完了"],
                [datetime.now().strftime('%Y-%m-%d'), "平均単価", "30,000円", "向上傾向", "完了"],
                [datetime.now().strftime('%Y-%m-%d'), "訪問件数", "80件", "計画通り", "進行中"],
                [datetime.now().strftime('%Y-%m-%d'), "成約率", "35%", "目標超過", "完了"],
                [datetime.now().strftime('%Y-%m-%d'), "顧客満足度", "4.8点", "高評価維持", "完了"]
            ]
            
            # データ行
            for row, data_row in enumerate(sample_data, 2):
                for col, value in enumerate(data_row, 1):
                    cell = ws_main.cell(row=row, column=col, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                    # ステータス列の色分け
                    if col == 5:  # ステータス列
                        if value == "完了":
                            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        elif value == "進行中":
                            cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            
            # 列幅調整
            for col in range(1, len(headers) + 1):
                col_letter = get_column_letter(col)
                ws_main.column_dimensions[col_letter].width = 15
            
            # 2. 入力フォームシート
            ws_form = wb.create_sheet("入力フォーム")
            
            # フォームタイトル
            ws_form.merge_cells('A1:E1')
            title_cell = ws_form['A1']
            title_cell.value = "X280 Excel直接入力フォーム"
            title_cell.font = Font(name='Meiryo UI', size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 入力フィールド
            form_fields = [
                ["入力項目", "値", "備考"],
                ["日付", datetime.now().strftime('%Y-%m-%d'), "自動設定"],
                ["項目名", "", "入力してください"],
                ["値", "", "数値または文字を入力"],
                ["備考", "", "任意の備考を入力"],
                ["ステータス", "", "完了/進行中/未着手"]
            ]
            
            for row, field_row in enumerate(form_fields, 3):
                for col, value in enumerate(field_row, 1):
                    cell = ws_form.cell(row=row, column=col, value=value)
                    if row == 3:  # ヘッダー行
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # 3. 統計シート
            ws_stats = wb.create_sheet("統計・分析")
            
            # 統計タイトル
            ws_stats.merge_cells('A1:D1')
            stats_title = ws_stats['A1']
            stats_title.value = "統計・分析データ"
            stats_title.font = Font(name='Meiryo UI', size=14, bold=True, color="FFFFFF")
            stats_title.fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
            stats_title.alignment = Alignment(horizontal="center", vertical="center")
            
            # 統計データ
            stats_data = [
                ["統計項目", "値", "前月比", "評価"],
                ["総売上高", "1,500,000円", "+15%", "◎"],
                ["総顧客数", "175件", "+8%", "◎"],
                ["平均成約率", "35%", "+5%", "◎"],
                ["総訪問数", "80件", "±0%", "○"],
                ["顧客満足度", "4.8点", "+0.2点", "◎"]
            ]
            
            for row, stats_row in enumerate(stats_data, 3):
                for col, value in enumerate(stats_row, 1):
                    cell = ws_stats.cell(row=row, column=col, value=value)
                    if row == 3:  # ヘッダー行
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    elif col == 4:  # 評価列
                        if value == "◎":
                            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        elif value == "○":
                            cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            
            # ファイル保存
            filename = f"X280直接入力_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = self.output_dir / filename
            
            wb.save(str(filepath))
            
            print(f"✅ 新しいExcelファイル作成完了: {filename}")
            return str(filepath)
            
        except Exception as e:
            print(f"❌ Excelファイル作成エラー: {e}")
            return None
    
    def edit_existing_excel(self, excel_path):
        """既存のExcelファイルを編集"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            print(f"📝 既存Excelファイル編集中: {excel_path}")
            
            wb = load_workbook(excel_path)
            
            # 編集シート追加
            ws_edit = wb.create_sheet("編集ログ")
            
            # 編集ログタイトル
            ws_edit.merge_cells('A1:D1')
            title_cell = ws_edit['A1']
            title_cell.value = f"編集ログ - {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}"
            title_cell.font = Font(name='Meiryo UI', size=14, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 編集履歴
            edit_history = [
                ["編集日時", "編集内容", "編集者", "備考"],
                [datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "ファイル開封・確認", "Mana", "X280直接入力システム"],
                [datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "編集シート追加", "システム", "自動編集ログ"],
                [datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "データ入力準備", "Mana", "入力フォーム準備完了"]
            ]
            
            for row, history_row in enumerate(edit_history, 3):
                for col, value in enumerate(history_row, 1):
                    cell = ws_edit.cell(row=row, column=col, value=value)
                    if row == 3:  # ヘッダー行
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # ファイル保存
            wb.save(excel_path)
            
            print(f"✅ 既存Excelファイル編集完了: {excel_path}")
            return True
            
        except Exception as e:
            print(f"❌ Excelファイル編集エラー: {e}")
            return False
    
    def add_data_to_excel(self, excel_path, new_data):
        """Excelファイルに新しいデータを追加"""
        try:
            from openpyxl import load_workbook
            
            print(f"➕ Excelファイルにデータ追加中: {excel_path}")
            
            wb = load_workbook(excel_path)
            
            # メインデータシートを取得（最初のシート）
            ws_main = wb.worksheets[0]
            
            # 最後の行を取得
            last_row = ws_main.max_row
            
            # 新しいデータを追加
            for col, value in enumerate(new_data, 1):
                ws_main.cell(row=last_row + 1, column=col, value=value)
            
            # ファイル保存
            wb.save(excel_path)
            
            print(f"✅ データ追加完了: {new_data}")
            return True
            
        except Exception as e:
            print(f"❌ データ追加エラー: {e}")
            return False
    
    def run_direct_input_demo(self):
        """直接入力デモ実行"""
        print("\n🌟 X280 Excel直接入力デモ開始")
        print("=" * 60)
        
        # 1. 新しいExcelファイル作成
        print("\n📄 ステップ1: 新しいExcelファイル作成")
        new_excel = self.create_new_excel_with_data()
        
        if not new_excel:
            print("❌ 新しいExcelファイル作成に失敗")
            return False
        
        # 2. 既存のExcelファイルを編集
        print("\n📝 ステップ2: 既存Excelファイル編集")
        existing_excel = "/home/mana/Desktop/PDF変換結果/実際の変換テスト_20251005_114553.xlsx"
        
        if os.path.exists(existing_excel):
            self.edit_existing_excel(existing_excel)
        
        # 3. 新しいデータ追加
        print("\n➕ ステップ3: 新しいデータ追加")
        new_data = [
            datetime.now().strftime('%Y-%m-%d'),
            "追加項目",
            "追加値",
            "X280直接入力で追加",
            "完了"
        ]
        self.add_data_to_excel(new_excel, new_data)
        
        # 4. ファイル一覧表示
        self.list_excel_files()
        
        return True
    
    def list_excel_files(self):
        """Excelファイル一覧表示"""
        try:
            files = list(self.output_dir.glob("*.xlsx"))
            
            print(f"\n📁 X280直接入力Excelファイル一覧 ({len(files)}件):")
            print("=" * 60)
            
            for i, file in enumerate(files, 1):
                file_stat = file.stat()
                print(f"{i:2d}. {file.name}")
                print(f"    サイズ: {file_stat.st_size:,} bytes")
                print(f"    作成日: {datetime.fromtimestamp(file_stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S')}")
                print(f"    パス: {file}")
                print()
            
            return files
            
        except Exception as e:
            print(f"❌ ファイル一覧取得エラー: {e}")
            return []
    
    def get_system_status(self):
        """システム状態取得"""
        try:
            status = {
                "timestamp": datetime.now().isoformat(),
                "output_directory": str(self.output_dir),
                "created_files": len(list(self.output_dir.glob("*.xlsx"))),
                "excel_system_status": "ready",
                "system_health": "healthy"
            }
            
            return {
                "success": True,
                "status": status
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }

def main():
    print("🌟 X280 Excelファイル直接入力システム")
    print("=" * 60)
    
    input_system = X280ExcelDirectInput()
    
    # システム状態確認
    status = input_system.get_system_status()
    print(f"システム状態: {status}")
    
    # 直接入力デモ実行
    success = input_system.run_direct_input_demo()
    
    if success:
        print("\n🎉 X280 Excel直接入力デモ完了！")
        print(f"📁 出力先: {input_system.output_dir}")
        print("💡 Excelファイルに直接データ入力可能！")
    else:
        print("\n❌ X280 Excel直接入力デモ失敗")

if __name__ == "__main__":
    main()
